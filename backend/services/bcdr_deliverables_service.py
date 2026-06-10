"""
BCDR Deliverables Service
=========================
Provides structured data for:
  - Timeline / Action Plan (6 phases over 40 weeks)
  - DR Testing Plan (quarterly schedule, checklists, RACI)
  - Compliance Checklist (Qatar PDPPL, 7 categories)
  - BCDR Strategy Reference (7 DR patterns)
  - Executive Summary (aggregated from assessment data)
  - Excel report generation (multi-sheet BCDR assessment report)

Ported from Phase2-AddRecommendations.ps1 deliverable sheets.
"""

from __future__ import annotations

import io
import logging
from datetime import datetime
from typing import Optional

from .bcdr_assessment_service import ZoneAssessment
from .bcdr_recommendation_service import (
    BCDRRecommendation,
    generate_recommendation,
)

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Timeline / Action Plan
# ---------------------------------------------------------------------------
def get_timeline_action_plan() -> dict:
    """Return the 6-phase BCDR implementation timeline."""
    return {
        "title": "BCDR Implementation Timeline — Tentative",
        "disclaimer": {
            "timeline_nature": "This timeline is TENTATIVE and GENERAL GUIDANCE ONLY. It is based on typical Azure DR implementation patterns and does not account for customer-specific constraints, team capacity, or organizational approval processes.",
            "customer_responsibility": "The customer must define the actual implementation timeline based on business priorities, available resources, budget, compliance review timelines, and internal change management processes.",
            "dependencies": "Many actions have prerequisites (e.g., networking must be provisioned before VMs can be replicated). The sequence below reflects technical dependencies, but actual timing depends on customer execution.",
            "dr_region_assumption": "This timeline uses the customer's chosen DR region. Common Qatar patterns: Qatar Central → West Europe or North Europe (both NIA-certified). Replace with your chosen DR region throughout.",
        },
        "phases": [
            {
                "name": "Phase 1: Foundation",
                "weeks": "1–4",
                "actions": [
                    {"id": "1.1", "action": "Data Classification Workshop", "notes": "Categorize all resources by data sensitivity (Public, Internal, Confidential, Restricted). Required for Qatar PDPPL compliance approval."},
                    {"id": "1.2", "action": "DPO/InfoSec Review & Approval", "notes": "Submit DR replication plan to Data Protection Officer and Information Security team. Cross-border data transfer approval required per Qatar PDPPL."},
                    {"id": "1.3", "action": "Select DR Region (Customer Choice)", "notes": "Customer to confirm DR region: West Europe or North Europe (both NIA-certified), or other region based on compliance, latency, and data residency."},
                    {"id": "1.4", "action": "Provision DR VNet & Core Networking", "notes": "Foundation prerequisite. Deploy VNet, NSGs, route tables, firewall rules, ExpressRoute/VPN connectivity. All other DR actions depend on this."},
                    {"id": "1.5", "action": "Establish IaC Parity", "notes": "Ensure all Qatar Central infrastructure is fully defined in IaC (Bicep/Terraform). This is the MOST CRITICAL factor for RTO per Microsoft Qatar BCDR Plan."},
                ],
            },
            {
                "name": "Phase 2: Quick Wins",
                "weeks": "5–8",
                "actions": [
                    {"id": "2.1", "action": "Enable ACR Geo-Replication", "notes": "Upgrade ACR to Premium tier, enable geo-replication. Required for AKS and Container Apps DR. Low effort, high value."},
                    {"id": "2.2", "action": "Configure Object Replication for Storage", "notes": "GRS not available in Qatar Central. Configure Object Replication for block blobs; use AzCopy/ADF for Azure Files."},
                    {"id": "2.3", "action": "Enable Cosmos DB Multi-Region Write", "notes": "Add DR region as secondary for Cosmos DB accounts. Quick Win for globally distributed data."},
                    {"id": "2.4", "action": "Sync Automation Account + Key Vault", "notes": "Deploy runbooks and secrets to DR region. Ensures operational access during DR events."},
                    {"id": "2.5", "action": "Add DR Endpoints to Front Door / Traffic Manager", "notes": "Extend global routing to DR endpoints. Quick Win once DR services are provisioned."},
                ],
            },
            {
                "name": "Phase 3: Critical Workloads — P1",
                "weeks": "9–16",
                "actions": [
                    {"id": "3.1", "action": "Azure Backup Region of Choice (RoC)", "notes": "Engage Microsoft Engineering for RoC preview enablement. Supported: IaaS VM, SQL, SAP, AFS. Target vault: Sweden Central or Switzerland North ONLY."},
                    {"id": "3.2", "action": "Enable ASR for Production VMs (P1)", "notes": "Customer to confirm production VMs. Configure ASR replication from Qatar Central to DR region. Typically 24–48h initial sync per VM."},
                    {"id": "3.3", "action": "Enable MySQL/PostgreSQL Geo-Redundant Backup", "notes": "Enable geo-backups on Flexible Servers. Customer to confirm production scope."},
                    {"id": "3.4", "action": "Enable ANF Cross-Region Replication", "notes": "Replicate ANF volumes to DR region. Customer to confirm production workloads."},
                    {"id": "3.5", "action": "Deploy Critical PaaS Services to DR Region", "notes": "Deploy critical App Services, Function Apps, Logic Apps, AKS clusters in standby mode (scaled down). IaC deployment recommended."},
                ],
            },
            {
                "name": "Phase 4: High Priority — P2",
                "weeks": "17–24",
                "actions": [
                    {"id": "4.1", "action": "Enable ASR for P2 High-Priority VMs", "notes": "Continue ASR rollout for high-priority workloads identified by customer."},
                    {"id": "4.2", "action": "Deploy P2 PaaS Services to DR Region", "notes": "Deploy high-priority PaaS services. May be standby or active-passive based on cost vs. RTO tradeoff."},
                    {"id": "4.3", "action": "Configure SQL Geo-Replication / Failover Groups", "notes": "Enable geo-replication for critical SQL databases. Customer to confirm scope and acceptable data lag."},
                    {"id": "4.4", "action": "Test Failover for P1 Workloads", "notes": "Conduct non-disruptive failover tests for P1 workloads. Document RTO/RPO actuals vs. targets."},
                ],
            },
            {
                "name": "Phase 5: Optimization",
                "weeks": "25–32",
                "actions": [
                    {"id": "5.1", "action": "Enable DR for Medium-Priority Workloads", "notes": "Customer to confirm scope. May include lower-priority apps or dev/test if business justified."},
                    {"id": "5.2", "action": "Optimize DR Costs", "notes": "Review active-active vs. active-passive. Scale down standby resources. Evaluate Reserved Instances for DR region."},
                    {"id": "5.3", "action": "Automation & Orchestration", "notes": "Develop failover runbooks, automated triggers, post-failover validation scripts. Integrate with monitoring and alerting."},
                    {"id": "5.4", "action": "DR Documentation & Runbooks", "notes": "Create comprehensive DR playbooks: failover procedures, rollback, contact lists, escalation paths."},
                ],
            },
            {
                "name": "Phase 6: Testing & Readiness",
                "weeks": "33–40",
                "actions": [
                    {"id": "6.1", "action": "Full DR Test — Regional Failover", "notes": "Conduct full-scale DR test. Fail over P1+P2 workloads to DR region. Measure RTO/RPO actuals. Document lessons learned."},
                    {"id": "6.2", "action": "Post-Test Remediation", "notes": "Address gaps identified during DR test. Refine runbooks, fix configuration issues, adjust RPO/RTO targets if needed."},
                    {"id": "6.3", "action": "Quarterly DR Testing Schedule", "notes": "Establish recurring DR test schedule (quarterly recommended). Assign owners, define success criteria, track results over time."},
                    {"id": "6.4", "action": "DR Readiness Report to Leadership", "notes": "Present DR readiness status to executive leadership. Include RTO/RPO actuals, test results, outstanding risks."},
                ],
            },
        ],
        "ongoing": [
            {"action": "Quarterly DR Tests", "notes": "Execute DR failover tests every quarter. Rotate test scenarios (planned vs. unplanned, partial vs. full)."},
            {"action": "Configuration Drift Monitoring", "notes": "Ensure all changes in primary region are replicated to DR. Automated IaC pipelines recommended."},
            {"action": "Azure Service Updates", "notes": "Monitor Azure updates that may impact DR strategy (new geo-replication capabilities, region expansions, SKU changes)."},
            {"action": "Compliance Re-Certification", "notes": "Annual review of data classification, DPO approval, and cross-border data transfer compliance (Qatar PDPPL)."},
        ],
    }


# ---------------------------------------------------------------------------
# DR Testing Plan
# ---------------------------------------------------------------------------
def get_dr_testing_plan() -> dict:
    """Return the DR testing plan template."""
    return {
        "title": "BCDR Testing Plan Template",
        "purpose": {
            "objective": "Validate that disaster recovery procedures work as documented, measure actual RTO/RPO against targets, and identify gaps in DR readiness.",
            "frequency": "Quarterly testing recommended for critical workloads (P1). Semi-annual testing for high-priority workloads (P2).",
            "test_types": [
                "Tabletop Exercise (documentation review)",
                "Non-Disruptive Test (test failover, no production impact)",
                "Full Failover Test (actual production switch to DR region)",
            ],
        },
        "checklist": [
            {
                "phase": "1. Pre-Test Planning",
                "steps": [
                    {"id": "1.1", "step": "Define Test Scope", "notes": "Select which workloads will be tested. Document in-scope vs. out-of-scope systems."},
                    {"id": "1.2", "step": "Define Test Type", "notes": "Tabletop, Non-Disruptive, or Full Failover? Align with business risk tolerance."},
                    {"id": "1.3", "step": "Schedule Test Window", "notes": "Coordinate with stakeholders. Prefer low-traffic periods. Minimum 4-hour window for full tests."},
                    {"id": "1.4", "step": "Notify Stakeholders", "notes": "Send test notification to all affected teams: app owners, DBA, network, security, exec sponsors."},
                    {"id": "1.5", "step": "Confirm Change Freeze", "notes": "Ensure no concurrent changes to production or DR regions during test window."},
                    {"id": "1.6", "step": "Document Baseline Metrics", "notes": "Capture current production metrics: uptime, transaction volume, latency, error rates."},
                ],
            },
            {
                "phase": "2. Test Execution — Failover",
                "steps": [
                    {"id": "2.1", "step": "Start Time", "notes": "Record exact start time. This is T=0 for RTO measurement."},
                    {"id": "2.2", "step": "Initiate Failover", "notes": "Execute failover per DR runbook. For ASR: trigger planned failover. For PaaS: activate Traffic Manager, scale up standby."},
                    {"id": "2.3", "step": "Monitor Replication Status", "notes": "Verify all data replication is complete before failing over databases and storage."},
                    {"id": "2.4", "step": "Update DNS / Traffic Manager", "notes": "Redirect traffic to DR endpoints. Verify DNS propagation (may take 5–15 minutes)."},
                    {"id": "2.5", "step": "Start DR Services", "notes": "Scale up App Services, Function Apps, AKS in DR region. Verify all services start successfully."},
                    {"id": "2.6", "step": "Validate Connectivity", "notes": "Test network connectivity: VMs→databases, apps→storage, users→front-end. Check NSG/firewall rules."},
                    {"id": "2.7", "step": "Application Smoke Tests", "notes": "Execute critical business transactions end-to-end. Verify data integrity. Check application logs."},
                    {"id": "2.8", "step": "End Time", "notes": "Record time when app is fully functional in DR. Calculate RTO = End Time − Start Time."},
                ],
            },
            {
                "phase": "3. Validate Data Integrity",
                "steps": [
                    {"id": "3.1", "step": "Data Completeness Check", "notes": "Compare record counts in DR vs. production. Verify no data loss."},
                    {"id": "3.2", "step": "Data Consistency Check", "notes": "Validate referential integrity, check for orphaned records."},
                    {"id": "3.3", "step": "Measure RPO", "notes": "Identify timestamp of last replicated transaction. RPO = Failover Time − Last Replicated Transaction."},
                    {"id": "3.4", "step": "Storage Validation", "notes": "Verify all required files, blobs, and file shares are accessible in DR region."},
                ],
            },
            {
                "phase": "4. Rollback to Production",
                "steps": [
                    {"id": "4.1", "step": "Initiate Failback", "notes": "Fail back to production region. For ASR: trigger commit and reprotect."},
                    {"id": "4.2", "step": "Verify Production Services", "notes": "Ensure all production services running normally. Check for replication lag."},
                    {"id": "4.3", "step": "Monitor for Issues", "notes": "Monitor production for 24 hours post-test. Watch for delayed errors or configuration drift."},
                ],
            },
            {
                "phase": "5. Documentation & Reporting",
                "steps": [
                    {"id": "5.1", "step": "Document RTO Actuals", "notes": "Record actual RTO. Compare against target. Document root cause of any gap."},
                    {"id": "5.2", "step": "Document RPO Actuals", "notes": "Record actual RPO. Compare against target. Document root cause of any gap."},
                    {"id": "5.3", "step": "Document Issues & Failures", "notes": "List all issues: failed services, connectivity problems, data inconsistencies, procedure gaps."},
                    {"id": "5.4", "step": "Assign Remediation Actions", "notes": "For each issue, assign owner, priority, and target resolution date."},
                    {"id": "5.5", "step": "Update DR Runbooks", "notes": "Incorporate lessons learned. Update procedures, add missing steps, clarify ambiguous instructions."},
                    {"id": "5.6", "step": "Executive Summary Report", "notes": "1-page summary for leadership: test scope, RTO/RPO actuals, pass/fail, critical issues, next steps."},
                ],
            },
        ],
        "success_criteria": [
            {"criterion": "RTO Target Met", "threshold": "Actual RTO ≤ Target RTO + 10% tolerance"},
            {"criterion": "RPO Target Met", "threshold": "Actual RPO ≤ Target RPO (zero data loss for P1 Critical if sync replication)"},
            {"criterion": "All Services Accessible", "threshold": "100% of in-scope services reachable and functional in DR region"},
            {"criterion": "Data Integrity Verified", "threshold": "No data loss, corruption, or missing records detected"},
            {"criterion": "No Critical Errors", "threshold": "No P0/P1 errors in application logs during test execution"},
            {"criterion": "Rollback Successful", "threshold": "If non-disruptive test, production services fully restored with no impact"},
        ],
        "quarterly_schedule": [
            {"quarter": "Q1", "window": "Week of Jan 15 (Wed 6–10 AM)", "scope": "P1 Critical VMs + SQL Databases", "type": "Non-Disruptive Test Failover"},
            {"quarter": "Q2", "window": "Week of Apr 10 (Sat 8AM–12PM)", "scope": "P1 + P2 Workloads (Partial)", "type": "Non-Disruptive Test Failover"},
            {"quarter": "Q3", "window": "Week of Jul 12 (Wed 6–10 AM)", "scope": "P1 Critical VMs Only", "type": "Full Failover Test (Announced)"},
            {"quarter": "Q4", "window": "Week of Oct 15 (Sat 8AM–2PM)", "scope": "All Production Workloads", "type": "Full Failover Test (Announced)"},
        ],
        "roles": [
            {"role": "Test Coordinator", "responsibility": "Overall test execution, timeline tracking, stakeholder coordination"},
            {"role": "Infrastructure Team", "responsibility": "Execute failover procedures, monitor replication, validate connectivity"},
            {"role": "Database Team", "responsibility": "Verify database failover, check data integrity, measure RPO"},
            {"role": "Application Team", "responsibility": "Execute smoke tests, validate business transactions, review application logs"},
            {"role": "Network Team", "responsibility": "Validate DNS changes, verify firewall rules, monitor traffic routing"},
            {"role": "Security Team", "responsibility": "Validate security controls in DR region, check Key Vault access, verify compliance"},
            {"role": "Executive Sponsor", "responsibility": "Approve test window, review post-test report, fund remediation actions"},
        ],
    }


# ---------------------------------------------------------------------------
# Compliance Checklist
# ---------------------------------------------------------------------------
def get_compliance_checklist() -> dict:
    """Return the Qatar BCDR compliance checklist."""
    return {
        "title": "Qatar BCDR Compliance Checklist",
        "note": "Qatar-specific. Qatar PDPPL = Personal Data Protection Privacy Law. NIA = National Information Assurance Authority.",
        "categories": [
            {
                "name": "Data Classification & Residency",
                "items": [
                    {"item": "Data Classification Completed", "status": "pending", "notes": "All resources categorized by data sensitivity: Public, Internal, Confidential, Restricted"},
                    {"item": "Data Inventory Documented", "status": "pending", "notes": "List of all data types, owners, and locations documented"},
                    {"item": "Qatar PDPPL Review Completed", "status": "pending", "notes": "Qatar Personal Data Protection Privacy Law compliance review completed by DPO"},
                    {"item": "Cross-Border Data Transfer Approval", "status": "pending", "notes": "DPO approval obtained for replication of data to secondary regions per Qatar PDPPL"},
                    {"item": "NIA/NCSA Certification Verified", "status": "pending", "notes": "DR region covered by Microsoft Azure Qatar NIA V2.0 certificate (ID: 10018). Certified: Qatar Central, West Europe, North Europe. Sweden Central/Switzerland North are NOT NIA-certified."},
                    {"item": "Data Residency Requirements Confirmed", "status": "pending", "notes": "Customer confirmed which data must remain in Qatar vs. which can be replicated abroad"},
                    {"item": "Encryption in Transit Verified", "status": "pending", "notes": "All cross-region replication uses TLS 1.2+"},
                    {"item": "Encryption at Rest Verified", "status": "pending", "notes": "All data in DR region encrypted at rest (customer-managed keys if required)"},
                ],
            },
            {
                "name": "BCDR Governance & Documentation",
                "items": [
                    {"item": "BCDR Policy Document Approved", "status": "pending", "notes": "Executive-approved BCDR policy defining RPO/RTO targets, DR scope, test frequency"},
                    {"item": "DR Runbooks Documented", "status": "pending", "notes": "Detailed failover and rollback procedures documented and accessible to ops team"},
                    {"item": "Roles & Responsibilities Defined", "status": "pending", "notes": "RACI matrix for DR execution, testing, and maintenance"},
                    {"item": "Escalation Procedures Documented", "status": "pending", "notes": "Contact lists, escalation paths, and communication plan for DR events"},
                    {"item": "SLAs Defined", "status": "pending", "notes": "Internal SLAs for RTO/RPO communicated to business stakeholders"},
                    {"item": "Change Management Integration", "status": "pending", "notes": "DR configuration changes integrated into change management process"},
                ],
            },
            {
                "name": "Technical Readiness",
                "items": [
                    {"item": "IaC Parity Achieved", "status": "pending", "notes": "All infrastructure fully defined in IaC (Bicep/Terraform). Can redeploy to DR region programmatically."},
                    {"item": "Network Connectivity Established", "status": "pending", "notes": "VNet peering, ExpressRoute, or VPN connectivity between primary and DR regions verified"},
                    {"item": "Backup Verification Completed", "status": "pending", "notes": "All critical backups tested for restore. Encryption and retention validated."},
                    {"item": "ASR Test Failover Completed (P1 VMs)", "status": "pending", "notes": "Azure Site Recovery test failover executed for all P1 critical VMs. RTO measured."},
                    {"item": "Database Geo-Replication Configured", "status": "pending", "notes": "SQL, MySQL, PostgreSQL geo-replication enabled for production databases. RPO measured."},
                    {"item": "Storage Replication Validated", "status": "pending", "notes": "Object Replication for Storage Accounts tested (GRS not available in Qatar Central)."},
                    {"item": "Key Vault Sync Configured", "status": "pending", "notes": "Key Vault secrets replicated to DR region (custom Azure Function or backup/restore process)."},
                    {"item": "Monitoring & Alerting in DR Region", "status": "pending", "notes": "Azure Monitor, Log Analytics, and alerts configured for DR region."},
                    {"item": "DNS Failover Mechanism Tested", "status": "pending", "notes": "Traffic Manager or Azure Front Door failover mechanism tested."},
                ],
            },
            {
                "name": "DR Testing & Validation",
                "items": [
                    {"item": "Tabletop Exercise Completed", "status": "pending", "notes": "DR runbook walkthrough with all stakeholders. No actual failover."},
                    {"item": "Non-Disruptive Test Failover Completed", "status": "pending", "notes": "Test failover executed without impacting production. RTO/RPO measured."},
                    {"item": "Full Failover Test Completed", "status": "pending", "notes": "Actual production failover to DR region. All services validated."},
                    {"item": "Rollback Test Completed", "status": "pending", "notes": "Failback from DR to primary tested. Production services restored."},
                    {"item": "RTO Target Met", "status": "pending", "notes": "Actual RTO ≤ Target RTO for all P1 critical workloads"},
                    {"item": "RPO Target Met", "status": "pending", "notes": "Actual RPO ≤ Target RPO for all P1 critical workloads"},
                    {"item": "Post-Test Remediation Completed", "status": "pending", "notes": "All issues from DR test resolved"},
                    {"item": "Quarterly Testing Schedule Established", "status": "pending", "notes": "Recurring DR tests scheduled and communicated"},
                ],
            },
            {
                "name": "Cost & Procurement",
                "items": [
                    {"item": "DR Budget Approved", "status": "pending", "notes": "Executive approval for DR infrastructure costs (compute, storage, bandwidth)"},
                    {"item": "Reserved Instances Purchased", "status": "pending", "notes": "If active-active or warm standby, consider RIs for cost optimization"},
                    {"item": "Azure Hybrid Benefit Applied", "status": "pending", "notes": "Windows Server and SQL Server licenses applied to reduce DR costs"},
                    {"item": "Cost Monitoring Configured", "status": "pending", "notes": "Azure Cost Management alerts configured to monitor DR region spending"},
                ],
            },
            {
                "name": "Security & Access Control",
                "items": [
                    {"item": "RBAC Roles Configured in DR", "status": "pending", "notes": "All RBAC roles and permissions replicated to DR region resources"},
                    {"item": "Service Principals / Managed Identities", "status": "pending", "notes": "DR applications have appropriate managed identities for resource access"},
                    {"item": "Key Vault Access Policies", "status": "pending", "notes": "DR region Key Vault has correct access policies for apps and users"},
                    {"item": "NSGs Configured", "status": "pending", "notes": "NSG rules in DR region match production requirements"},
                    {"item": "Firewall / WAF Rules", "status": "pending", "notes": "Firewall and WAF rules replicated to DR region"},
                    {"item": "Private Endpoints Configured", "status": "pending", "notes": "Private endpoints for PaaS services configured in DR region"},
                    {"item": "Security Baseline Validated", "status": "pending", "notes": "DR region resources meet organizational security baseline"},
                ],
            },
            {
                "name": "Azure Engineering & Preview Features",
                "items": [
                    {"item": "Azure Backup RoC Engaged", "status": "pending", "notes": "Microsoft Engineering engagement initiated for RoC preview (Qatar Central → SDC/SZN)"},
                    {"item": "RoC Enablement Completed", "status": "pending", "notes": "Azure Backup RoC enabled for IaaS VM, SQL, SAP, AFS workloads"},
                    {"item": "RoC Test Restore Completed", "status": "pending", "notes": "Test restore from RoC-enabled vault to verify functionality"},
                ],
            },
        ],
        "signoff": [
            {"role": "Business Owner", "date": "", "name": ""},
            {"role": "IT Director", "date": "", "name": ""},
            {"role": "CISO / Information Security", "date": "", "name": ""},
            {"role": "Data Protection Officer (DPO)", "date": "", "name": ""},
            {"role": "Compliance Officer", "date": "", "name": ""},
            {"role": "Executive Sponsor", "date": "", "name": ""},
        ],
    }


# ---------------------------------------------------------------------------
# BCDR Strategy Reference
# ---------------------------------------------------------------------------
def get_strategy_reference() -> dict:
    """Return the BCDR strategy reference guide."""
    return {
        "title": "Azure BCDR Strategy Reference Guide",
        "patterns": [
            {
                "name": "Active-Active (Multi-Region)",
                "description": "Both regions are LIVE and actively serving production traffic simultaneously. Traffic is load-balanced across regions using global routing (Azure Front Door, Traffic Manager).",
                "when_to_use": "Mission-critical workloads requiring zero downtime and lowest possible RPO/RTO.",
                "rpo": "Near-zero (synchronous or near-synchronous replication)",
                "rto": "Near-zero (traffic automatically re-routes to healthy region)",
                "cost": "High (2x compute cost — both regions fully provisioned and running)",
                "examples": ["Cosmos DB (multi-region writes)", "ACR Premium Geo-Replication", "Static Web Apps (global CDN)", "Traffic Manager / Azure Front Door"],
            },
            {
                "name": "Active-Passive (Warm Standby)",
                "description": "Primary region is LIVE. Secondary region is PRE-PROVISIONED but idle/standby. On DR event, traffic is manually or automatically re-routed to secondary.",
                "when_to_use": "Production workloads with moderate RTO requirements (5–30 minutes). Cost-optimized alternative to Active-Active.",
                "rpo": "Depends on replication method (typically <5 min for databases, near-real-time for VMs via ASR)",
                "rto": "<30 minutes with automation; <5 minutes with health probe-based traffic re-routing",
                "cost": "Medium (secondary infrastructure pre-provisioned but possibly scaled down)",
                "examples": ["VMs (ASR to UAE North)", "Azure SQL (Auto-Failover Groups)", "App Service (standby via Front Door)", "AKS (standby cluster, IaC-provisioned)"],
            },
            {
                "name": "Active-Passive (Cold Standby)",
                "description": "Primary region is LIVE. Secondary region has NO pre-provisioned infrastructure — DR resources are deployed ON-DEMAND during a DR event using IaC.",
                "when_to_use": "Non-production or dev environments with relaxed RTO/RPO. Cost-sensitive scenarios.",
                "rpo": "Dependent on backup frequency (hours to days)",
                "rto": "Hours (infrastructure provisioning + data restore)",
                "cost": "Low (no secondary compute cost — pay only for backup storage)",
                "examples": ["VMs (IaC Redeployment + Backup Restore)", "App Service (IaC Redeployment)", "Logic Apps (Git-based redeployment)"],
            },
            {
                "name": "Backup & Restore",
                "description": "Periodic backups stored in geo-redundant storage. On DR event, data is restored from backup to new/existing instance in DR region.",
                "when_to_use": "Non-critical workloads with longer acceptable RTO/RPO. Data protection (corruption, accidental deletion). Compliance-driven retention.",
                "rpo": "Backup frequency (daily=24h; 4-hourly enhanced=4h)",
                "rto": "Restore time (4–12 hours depending on data size)",
                "cost": "Low (backup storage only)",
                "examples": ["Azure Backup (RSV)", "SQL Geo-Redundant Backup + Geo-Restore", "Blob Soft Delete + Versioning", "ANF Backup"],
            },
            {
                "name": "Geo-Replication (Platform-Managed)",
                "description": "Azure service automatically replicates data to secondary region. Failover is manual or automatic depending on service.",
                "when_to_use": "Leveraging built-in Azure service resilience without custom configuration.",
                "rpo": "Near-zero (<15 seconds)",
                "rto": "Automatic (service-managed) or manual (customer-initiated)",
                "cost": "Typically included in Premium/Standard tiers",
                "examples": ["Azure Storage (GRS/RA-GRS — NOT Qatar Central)", "Cosmos DB Multi-Region", "Event Hubs Premium Geo-Replication", "Service Bus Premium Geo-Replication"],
            },
            {
                "name": "Infrastructure-as-Code (IaC) DR",
                "description": "All infrastructure defined in code (Bicep, Terraform, ARM). DR region provisioned via automated deployment pipelines. Critical for stateless and serverless.",
                "when_to_use": "Cloud-native apps, microservices, serverless (Functions, Logic Apps), container orchestration (AKS).",
                "rpo": "Near-zero (if state externalized to geo-redundant storage)",
                "rto": "Pipeline execution time (15 min – 2 hours)",
                "cost": "Low (no standby infrastructure — deploy only on DR trigger)",
                "examples": ["AKS (GitOps + IaC)", "Functions / App Service (CI/CD redeploy)", "VNets (Bicep redeploy)", "Data Factory (Git-integrated)"],
            },
            {
                "name": "Hybrid/On-Premises Integration",
                "description": "DR strategy extends beyond Azure to include on-premises or multi-cloud. Requires hybrid connectivity (VPN, ExpressRoute) to secondary region.",
                "when_to_use": "Lift-and-shift migrations, hybrid identity (AD DS), on-premises data sources.",
                "rpo": "Dependent on on-premises backup strategy and network bandwidth",
                "rto": "Dependent on on-premises infrastructure and connectivity",
                "cost": "High (on-premises infrastructure + hybrid connectivity)",
                "examples": ["ASR (on-prem VMs to Azure)", "Azure File Sync", "ExpressRoute / VPN Gateway (BGP failover)", "Azure AD Connect"],
            },
        ],
        "qatar_constraints": [
            {"constraint": "No Paired Region", "impact": "No native GRS, CRR, or automatic geo-DR for Qatar Central.", "approach": "Object Replication for Blob Storage. Azure Backup RoC (Preview) to SDC/SZN. Custom cross-region sync for Key Vault. Active-Passive with manual secondary provisioning."},
            {"constraint": "Zone Redundancy Restricted", "impact": "ZRS/GZRS blocked or unavailable for many services. One AZ at full capacity.", "approach": "Do NOT rely solely on zone redundancy. Implement cross-region DR for all Tier-1 and Tier-2 workloads."},
            {"constraint": "Azure Backup RoC Preview", "impact": "Standard CRR does NOT work in Qatar Central. RoC targets: SDC or SZN ONLY.", "approach": "Engage Microsoft for subscription whitelisting. For VM DR, use ASR separately."},
            {"constraint": "IaC Parity is Critical", "impact": "Without IaC, manual DR provisioning can take days.", "approach": "ALL production infra must be in Bicep/ARM/Terraform. CI/CD targeting both regions. Test quarterly."},
            {"constraint": "NIA/NCSA Certification", "impact": "Qatar regulated workloads require NIA-certified regions. Certified: QC, WE, NE.", "approach": "Prefer West Europe or North Europe for DR. SDC/SZN (RoC) are NOT NIA-certified — obtain approval."},
        ],
        "decision_matrix": [
            {"tier": "Tier 1 — Mission-Critical", "rto": "< 5 minutes", "rpo": "< 5 seconds", "strategy": "Active-Active or Active-Passive (Hot Standby)"},
            {"tier": "Tier 2 — Production", "rto": "< 30 minutes", "rpo": "< 1 hour", "strategy": "Active-Passive (Warm Standby)"},
            {"tier": "Tier 3 — Non-Critical", "rto": "< 4 hours", "rpo": "< 24 hours", "strategy": "Backup & Restore or IaC Cold Standby"},
        ],
    }


# ---------------------------------------------------------------------------
# Executive Summary (data-driven from assessments)
# ---------------------------------------------------------------------------
def build_executive_summary(
    assessments: list[ZoneAssessment],
    recommendations: list[BCDRRecommendation],
) -> dict:
    """Build executive summary from real assessment data."""
    total = len(assessments)
    if total == 0:
        return {"total": 0, "message": "No resources assessed"}

    # Priority breakdown
    p1 = sum(1 for r in recommendations if r.sa_priority == "P1")
    p2 = sum(1 for r in recommendations if r.sa_priority == "P2")
    p3 = sum(1 for r in recommendations if r.sa_priority == "P3")
    p4 = sum(1 for r in recommendations if r.sa_priority == "P4")

    # Quick wins
    quick_wins = sum(1 for r in recommendations if r.sa_quick_win == "Yes")

    # Zone status
    non_zonal = sum(1 for a in assessments if a.zone_status in ("NonZonal", "Unknown"))
    zone_redundant = sum(1 for a in assessments if a.zone_status == "ZoneRedundant")
    locally_redundant = sum(1 for a in assessments if a.zone_status == "LocallyRedundant")

    # Qatar Central
    qatar_count = sum(1 for a in assessments if a.is_qatar_central)
    needs_dr = sum(1 for a in assessments if a.needs_dr_action)
    geo_redundant = sum(1 for a in assessments if a.geo_redundant)

    # Risk score
    avg_risk = round(sum(a.zone_risk_score for a in assessments) / total, 1) if total else 0
    high_risk = sum(1 for a in assessments if a.zone_risk_score >= 70)

    # Resource types
    type_counts: dict[str, int] = {}
    for a in assessments:
        short = a.resource_type.split("/")[-1] if "/" in a.resource_type else a.resource_type
        type_counts[short] = type_counts.get(short, 0) + 1

    # Tier breakdown
    tier_counts: dict[str, int] = {}
    for a in assessments:
        tier_counts[a.workload_tier] = tier_counts.get(a.workload_tier, 0) + 1

    # Effort breakdown
    effort_counts: dict[str, int] = {}
    for r in recommendations:
        e = r.sa_implementation_effort or "Unknown"
        effort_counts[e] = effort_counts.get(e, 0) + 1

    return {
        "generated_at": datetime.utcnow().isoformat() + "Z",
        "total_resources": total,
        "priority_breakdown": {"P1": p1, "P2": p2, "P3": p3, "P4": p4},
        "quick_wins_count": quick_wins,
        "zone_status": {
            "zone_redundant": zone_redundant,
            "non_zonal": non_zonal,
            "locally_redundant": locally_redundant,
        },
        "qatar_central_count": qatar_count,
        "needs_dr_action": needs_dr,
        "geo_redundant_count": geo_redundant,
        "average_risk_score": avg_risk,
        "high_risk_count": high_risk,
        "top_resource_types": dict(sorted(type_counts.items(), key=lambda x: -x[1])[:10]),
        "tier_breakdown": tier_counts,
        "effort_breakdown": effort_counts,
        "key_findings": _build_key_findings(assessments, recommendations, qatar_count, non_zonal, high_risk, quick_wins),
    }


def _build_key_findings(
    assessments: list[ZoneAssessment],
    recommendations: list[BCDRRecommendation],
    qatar_count: int,
    non_zonal: int,
    high_risk: int,
    quick_wins: int,
) -> list[dict]:
    """Generate key findings bullets for executive summary."""
    findings: list[dict] = []
    total = len(assessments)
    if qatar_count > 0:
        findings.append({
            "severity": "critical",
            "finding": f"{qatar_count} resources in Qatar Central — NO paired region, zone redundancy restricted. Cross-region DR is mandatory.",
        })
    if non_zonal > 0:
        pct = round(non_zonal / total * 100) if total else 0
        findings.append({
            "severity": "high",
            "finding": f"{non_zonal} resources ({pct}%) are NonZonal — single point of failure within the region.",
        })
    if high_risk > 0:
        findings.append({
            "severity": "high",
            "finding": f"{high_risk} resources have a risk score ≥ 70 — immediate DR action recommended.",
        })
    if quick_wins > 0:
        findings.append({
            "severity": "info",
            "finding": f"{quick_wins} Quick Wins identified — low-effort actions that can improve DR posture immediately.",
        })
    return findings


# ---------------------------------------------------------------------------
# Excel Report Generation
# ---------------------------------------------------------------------------
def generate_excel_report(
    assessments: list[ZoneAssessment],
    recommendations: list[BCDRRecommendation],
) -> io.BytesIO:
    """Generate a multi-sheet BCDR assessment Excel report.

    Returns an in-memory BytesIO object containing the .xlsx file.
    Uses openpyxl for Excel generation.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise RuntimeError("openpyxl is required for Excel report generation. Install with: pip install openpyxl")

    wb = Workbook()

    # Color constants
    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    header_font = Font(bold=True, color="E2E8F0", size=11)
    p1_fill = PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid")
    p2_fill = PatternFill(start_color="F97316", end_color="F97316", fill_type="solid")
    p3_fill = PatternFill(start_color="EAB308", end_color="EAB308", fill_type="solid")
    quick_fill = PatternFill(start_color="22C55E", end_color="22C55E", fill_type="solid")
    section_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    section_font = Font(bold=True, color="60A5FA", size=12)
    thin_border = Border(
        left=Side(style="thin", color="334155"),
        right=Side(style="thin", color="334155"),
        top=Side(style="thin", color="334155"),
        bottom=Side(style="thin", color="334155"),
    )

    # ── Sheet 1: Executive Summary ────────────────────────────────────────
    ws_exec = wb.active
    ws_exec.title = "Executive_Summary"
    ws_exec.sheet_properties.tabColor = "0070C0"
    summary = build_executive_summary(assessments, recommendations)

    ws_exec.append(["BCDR Assessment — Executive Summary"])
    ws_exec.append([f"Generated: {summary.get('generated_at', '')}"])
    ws_exec.append([])
    ws_exec.append(["Metric", "Value"])
    ws_exec.append(["Total Resources Assessed", summary["total_resources"]])
    ws_exec.append(["P1 Critical Actions", summary["priority_breakdown"]["P1"]])
    ws_exec.append(["P2 High Priority", summary["priority_breakdown"]["P2"]])
    ws_exec.append(["P3 Medium Priority", summary["priority_breakdown"]["P3"]])
    ws_exec.append(["Quick Wins Available", summary["quick_wins_count"]])
    ws_exec.append(["Average Risk Score", summary["average_risk_score"]])
    ws_exec.append(["High Risk Resources (≥70)", summary["high_risk_count"]])
    ws_exec.append(["Qatar Central Resources", summary["qatar_central_count"]])
    ws_exec.append(["NonZonal Resources", summary["zone_status"]["non_zonal"]])
    ws_exec.append(["Geo-Redundant Resources", summary["geo_redundant_count"]])
    ws_exec.append([])
    ws_exec.append(["KEY FINDINGS"])
    for f in summary.get("key_findings", []):
        ws_exec.append([f["severity"].upper(), f["finding"]])

    _style_header_row(ws_exec, 1, header_fill, Font(bold=True, color="E2E8F0", size=14))
    _style_header_row(ws_exec, 4, header_fill, header_font)
    ws_exec.column_dimensions["A"].width = 35
    ws_exec.column_dimensions["B"].width = 100

    # ── Sheet 2: SA Recommendations (all resources) ───────────────────────
    ws_sa = wb.create_sheet("SA_Recommendations")
    ws_sa.sheet_properties.tabColor = "4472C4"
    sa_cols = [
        "ResourceName", "ResourceType", "ResourceGroup", "Location", "SubscriptionId",
        "ZoneStatus", "WorkloadTier", "RiskScore", "Priority", "Criticality",
        "BCDRStrategy", "DRMethod", "RPO", "RTO", "ActionRequired",
        "Effort", "CostImpact", "QuickWin", "Dependencies",
        "ComplianceNote", "GapSummary", "ZoneTransitionPath",
    ]
    ws_sa.append(sa_cols)
    for r in recommendations:
        ws_sa.append([
            r.resource_name, r.resource_type, r.resource_group, r.location, r.subscription_id,
            r.zone_status, r.workload_tier, r.zone_risk_score, r.sa_priority, r.sa_criticality,
            r.sa_bcdr_strategy, r.sa_dr_method, r.sa_rpo, r.sa_rto, r.sa_action_required,
            r.sa_implementation_effort, r.sa_cost_impact, r.sa_quick_win, r.sa_dependencies,
            r.sa_compliance_note, r.sa_current_gap_summary, r.sa_zone_transition_path,
        ])
    _style_header_row(ws_sa, 1, header_fill, header_font)
    _auto_width(ws_sa, max_width=50)

    # ── Sheet 3: Quick Wins ───────────────────────────────────────────────
    ws_qw = wb.create_sheet("QuickWins")
    ws_qw.sheet_properties.tabColor = "22C55E"
    qw_cols = ["ResourceName", "ResourceType", "Location", "Priority", "ActionRequired", "Effort", "CostImpact"]
    ws_qw.append(qw_cols)
    for r in sorted(recommendations, key=lambda x: x.sa_priority):
        if r.sa_quick_win == "Yes":
            ws_qw.append([r.resource_name, r.resource_type, r.location, r.sa_priority, r.sa_action_required, r.sa_implementation_effort, r.sa_cost_impact])
    _style_header_row(ws_qw, 1, quick_fill, Font(bold=True, color="000000", size=11))
    _auto_width(ws_qw, max_width=60)

    # ── Sheet 4: P1 Critical Actions ──────────────────────────────────────
    ws_p1 = wb.create_sheet("P1_Critical_Actions")
    ws_p1.sheet_properties.tabColor = "EF4444"
    p_cols = ["ResourceName", "ResourceType", "Location", "Criticality", "BCDRStrategy", "DRMethod", "RPO", "RTO", "ActionRequired", "GapSummary"]
    ws_p1.append(p_cols)
    for r in recommendations:
        if r.sa_priority == "P1":
            ws_p1.append([r.resource_name, r.resource_type, r.location, r.sa_criticality, r.sa_bcdr_strategy, r.sa_dr_method, r.sa_rpo, r.sa_rto, r.sa_action_required, r.sa_current_gap_summary])
    _style_header_row(ws_p1, 1, p1_fill, Font(bold=True, color="FFFFFF", size=11))
    _auto_width(ws_p1, max_width=60)

    # ── Sheet 5: P2 Actions ───────────────────────────────────────────────
    ws_p2 = wb.create_sheet("P2_Actions")
    ws_p2.sheet_properties.tabColor = "F97316"
    ws_p2.append(p_cols)
    for r in recommendations:
        if r.sa_priority == "P2":
            ws_p2.append([r.resource_name, r.resource_type, r.location, r.sa_criticality, r.sa_bcdr_strategy, r.sa_dr_method, r.sa_rpo, r.sa_rto, r.sa_action_required, r.sa_current_gap_summary])
    _style_header_row(ws_p2, 1, p2_fill, Font(bold=True, color="000000", size=11))
    _auto_width(ws_p2, max_width=60)

    # ── Sheet 6: P3 Actions ───────────────────────────────────────────────
    ws_p3 = wb.create_sheet("P3_Actions")
    ws_p3.sheet_properties.tabColor = "EAB308"
    ws_p3.append(p_cols)
    for r in recommendations:
        if r.sa_priority == "P3":
            ws_p3.append([r.resource_name, r.resource_type, r.location, r.sa_criticality, r.sa_bcdr_strategy, r.sa_dr_method, r.sa_rpo, r.sa_rto, r.sa_action_required, r.sa_current_gap_summary])
    _style_header_row(ws_p3, 1, p3_fill, Font(bold=True, color="000000", size=11))
    _auto_width(ws_p3, max_width=60)

    # ── Sheet 7: Summary by Resource Type ─────────────────────────────────
    ws_rt = wb.create_sheet("Summary_ByResourceType")
    ws_rt.sheet_properties.tabColor = "8B5CF6"
    type_summary: dict[str, dict] = {}
    for r in recommendations:
        rt = r.resource_type
        if rt not in type_summary:
            type_summary[rt] = {"count": 0, "p1": 0, "p2": 0, "quick_wins": 0}
        type_summary[rt]["count"] += 1
        if r.sa_priority == "P1":
            type_summary[rt]["p1"] += 1
        elif r.sa_priority == "P2":
            type_summary[rt]["p2"] += 1
        if r.sa_quick_win == "Yes":
            type_summary[rt]["quick_wins"] += 1
    ws_rt.append(["ResourceType", "Count", "P1_Critical", "P2_High", "QuickWins"])
    for rt, data in sorted(type_summary.items(), key=lambda x: -x[1]["count"]):
        ws_rt.append([rt, data["count"], data["p1"], data["p2"], data["quick_wins"]])
    _style_header_row(ws_rt, 1, header_fill, header_font)
    _auto_width(ws_rt, max_width=60)

    # ── Sheet 8: Timeline Action Plan ─────────────────────────────────────
    ws_tl = wb.create_sheet("Timeline_ActionPlan")
    ws_tl.sheet_properties.tabColor = "FFC000"
    timeline = get_timeline_action_plan()
    ws_tl.append(["BCDR IMPLEMENTATION TIMELINE — TENTATIVE"])
    ws_tl.append([])
    ws_tl.append(["DISCLAIMER"])
    for k, v in timeline["disclaimer"].items():
        ws_tl.append([k.replace("_", " ").title(), v])
    ws_tl.append([])
    for phase in timeline["phases"]:
        ws_tl.append([f"{phase['name']} (Weeks {phase['weeks']})"])
        ws_tl.append(["Action", "Deliverable & Notes"])
        for a in phase["actions"]:
            ws_tl.append([f"{a['id']} {a['action']}", a["notes"]])
        ws_tl.append([])
    ws_tl.append(["ONGOING: SUSTAINING OPERATIONS"])
    for a in timeline["ongoing"]:
        ws_tl.append([a["action"], a["notes"]])
    _style_header_row(ws_tl, 1, PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid"), Font(bold=True, size=14))
    ws_tl.column_dimensions["A"].width = 50
    ws_tl.column_dimensions["B"].width = 100

    # ── Sheet 9: DR Testing Plan ──────────────────────────────────────────
    ws_tp = wb.create_sheet("DR_Testing_Plan")
    ws_tp.sheet_properties.tabColor = "92D050"
    plan = get_dr_testing_plan()
    ws_tp.append(["BCDR TESTING PLAN TEMPLATE"])
    ws_tp.append([])
    ws_tp.append(["Objective", plan["purpose"]["objective"]])
    ws_tp.append(["Frequency", plan["purpose"]["frequency"]])
    ws_tp.append([])
    for section in plan["checklist"]:
        ws_tp.append([section["phase"]])
        ws_tp.append(["Step", "Action & Notes"])
        for s in section["steps"]:
            ws_tp.append([f"{s['id']} {s['step']}", s["notes"]])
        ws_tp.append([])
    ws_tp.append(["TEST SUCCESS CRITERIA"])
    ws_tp.append(["Criterion", "Pass/Fail Threshold"])
    for c in plan["success_criteria"]:
        ws_tp.append([c["criterion"], c["threshold"]])
    ws_tp.append([])
    ws_tp.append(["QUARTERLY TEST SCHEDULE"])
    ws_tp.append(["Quarter", "Window", "Scope", "Type"])
    for q in plan["quarterly_schedule"]:
        ws_tp.append([q["quarter"], q["window"], q["scope"], q["type"]])
    ws_tp.append([])
    ws_tp.append(["ROLES & RESPONSIBILITIES"])
    for r in plan["roles"]:
        ws_tp.append([r["role"], r["responsibility"]])
    _style_header_row(ws_tp, 1, PatternFill(start_color="92D050", end_color="92D050", fill_type="solid"), Font(bold=True, size=14))
    ws_tp.column_dimensions["A"].width = 50
    ws_tp.column_dimensions["B"].width = 100

    # ── Sheet 10: Compliance Checklist ────────────────────────────────────
    ws_cc = wb.create_sheet("Compliance_Checklist")
    ws_cc.sheet_properties.tabColor = "C00000"
    cl = get_compliance_checklist()
    ws_cc.append(["QATAR BCDR COMPLIANCE CHECKLIST"])
    ws_cc.append([cl["note"]])
    ws_cc.append([])
    for cat in cl["categories"]:
        ws_cc.append([cat["name"]])
        ws_cc.append(["Item", "Status", "Notes / Evidence"])
        for item in cat["items"]:
            ws_cc.append([item["item"], "☐ Pending", item["notes"]])
        ws_cc.append([])
    ws_cc.append(["SIGN-OFF & APPROVAL"])
    ws_cc.append(["Role", "Sign-Off Date", "Name"])
    for s in cl["signoff"]:
        ws_cc.append([s["role"], "", ""])
    _style_header_row(ws_cc, 1, PatternFill(start_color="C00000", end_color="C00000", fill_type="solid"), Font(bold=True, color="FFFFFF", size=14))
    ws_cc.column_dimensions["A"].width = 50
    ws_cc.column_dimensions["B"].width = 20
    ws_cc.column_dimensions["C"].width = 80

    # ── Sheet 11: BCDR Strategy Reference ─────────────────────────────────
    ws_ref = wb.create_sheet("BCDR_Strategy_Reference")
    ws_ref.sheet_properties.tabColor = "0078D4"
    ref = get_strategy_reference()
    ws_ref.append(["AZURE BCDR STRATEGY REFERENCE GUIDE"])
    ws_ref.append([])
    ws_ref.append(["Strategy Pattern", "Description", "RPO", "RTO", "Cost", "Examples"])
    for p in ref["patterns"]:
        ws_ref.append([p["name"], p["description"], p["rpo"], p["rto"], p["cost"], "; ".join(p["examples"])])
    ws_ref.append([])
    ws_ref.append(["QATAR CENTRAL SPECIFIC CONSTRAINTS"])
    ws_ref.append(["Constraint", "Impact", "Recommended Approach"])
    for c in ref["qatar_constraints"]:
        ws_ref.append([c["constraint"], c["impact"], c["approach"]])
    ws_ref.append([])
    ws_ref.append(["DR STRATEGY DECISION MATRIX"])
    ws_ref.append(["Tier", "Acceptable RTO", "Acceptable RPO", "Recommended Strategy"])
    for d in ref["decision_matrix"]:
        ws_ref.append([d["tier"], d["rto"], d["rpo"], d["strategy"]])
    _style_header_row(ws_ref, 1, PatternFill(start_color="0078D4", end_color="0078D4", fill_type="solid"), Font(bold=True, color="FFFFFF", size=14))
    _auto_width(ws_ref, max_width=80)

    # ── Sheet 12: Risk Heatmap ────────────────────────────────────────────
    ws_rh = wb.create_sheet("Risk_Heatmap")
    ws_rh.sheet_properties.tabColor = "FF6384"
    ws_rh.append(["ResourceName", "ResourceType", "Location", "ZoneStatus", "RiskScore", "Priority", "WorkloadTier", "GapSummary"])
    for r in sorted(recommendations, key=lambda x: x.zone_risk_score, reverse=True):
        if r.zone_risk_score >= 50:
            ws_rh.append([r.resource_name, r.resource_type, r.location, r.zone_status, r.zone_risk_score, r.sa_priority, r.workload_tier, r.sa_current_gap_summary])
    _style_header_row(ws_rh, 1, header_fill, header_font)
    _auto_width(ws_rh, max_width=60)

    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ---------------------------------------------------------------------------
# Helper functions for Excel styling
# ---------------------------------------------------------------------------
def _style_header_row(ws, row_num: int, fill, font):
    """Apply fill and font to all cells in a row."""
    for cell in ws[row_num]:
        cell.fill = fill
        cell.font = font


def _auto_width(ws, max_width: int = 50):
    """Auto-size columns based on content, capped at max_width."""
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val = str(cell.value) if cell.value else ""
            max_len = max(max_len, min(len(val), max_width))
        ws.column_dimensions[col_letter].width = max(max_len + 2, 12)
