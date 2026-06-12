"""
Azure Cost Optimization Tool — FastAPI Backend v3
Features: Cost Management · Monitor · Advisor · AI · Activity Log ·
          Carbon · Right-Sizing · Tag Compliance · SSE Streaming ·
          Settings API · Demo Mode · Multi-Subscription · Resource Group Filter
"""
from __future__ import annotations

import asyncio
import json
import logging
import os
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime, timedelta, timezone
from functools import partial
from typing import Any, AsyncGenerator, Dict, List, Optional

import anthropic
from dotenv import load_dotenv
import pathlib
from pathlib import Path

from fastapi import FastAPI, HTTPException, Request, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from starlette.middleware.base import BaseHTTPMiddleware
from fastapi.responses import FileResponse, StreamingResponse, JSONResponse, Response
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

from models.schemas import (
    AdvisorRecommendation, AppSettings, CacheStatus, CostAnomaly, DashboardData,
    KPIData, OrphanResource, ResourceMetrics, ResourceTypeSummary,
    RightSizeOpportunity, SavingsRecommendation, ScoreDistribution,
    ScoreLabel, SubscriptionSummary, TrendDirection,
)
from services.cost_service     import get_two_month_costs, get_daily_costs, get_monthly_cost_history, get_total_daily_costs, get_reservation_covered_resource_ids
from services.metrics_service  import get_resource_metrics
from services.resource_service import list_all_resources, find_orphans, get_app_service_plan_links, get_vm_power_states, get_resource_locks, get_app_insights_links, get_vm_attachments, get_rbac_signals, get_reservation_coverage, get_reservation_recommendations, get_private_endpoint_targets, get_sql_replica_ids, get_app_service_details, get_backup_protected_ids, get_openai_deployments
from services.storage_access_service  import get_storage_access_signals
from services.keyvault_access_service import get_keyvault_signals
from services.advisor_service  import get_advisor_recommendations
from services.ai_service       import get_ai_verdicts, get_active_provider, get_ai_narrative
from services.scoring_service       import score_resource, estimate_savings, is_infrastructure_resource, get_safe_action_steps
from services.observability_service import get_data_confidence, should_suppress_idle_penalty
from services.activity_service import get_subscription_activity
from services.carbon_service   import estimate_carbon, carbon_equivalents
from services.rightsize_service import get_rightsize_recommendations, RightSizeRec
from services.security_service      import identify_security_gaps
from services.defender_service      import get_full_security_posture
from services.modernization_service import detect_modernization_opportunities, build_migration_assessment
from services.waf_service           import compute_waf_scorecard
from services.innovation_service    import detect_innovation_gaps
from services.maturity_service      import compute_cloud_maturity
from services.licensing_service     import detect_licensing_opportunities, build_reservation_analysis
from services.backup_service        import analyze_backup_coverage
from services.backup_enhanced_service import get_enhanced_backup_analysis, build_rpo_rto_matrix, build_ransomware_readiness
from services.security_enhanced_service import build_zero_trust_scorecard, build_attack_surface_analysis
from services.security_findings_service import (
    persist_findings as persist_security_findings,
    query_findings as query_security_findings,
    export_findings_csv as export_security_csv,
    get_scan_history as get_security_scan_history,
    get_findings_summary as get_security_summary,
)
from services.acr_service           import analyze_acr_opportunities
from services.dependency_service    import (
    build_dependency_graph,
    get_blast_radius,
    get_resource_dependencies,
    get_graph_summary,
)
from models.dependency_models import DependencyGraph, DependencyGraphSummary
import services.project_service as project_svc
import services.settings_service as settings_svc
import services.persistence_service as persistence_svc
import services.cache_service as cache_svc
import services.auth_service as auth_svc

try:
    import services.tagging_service as tagging_svc
except Exception as _e:
    logger.warning("tagging_service unavailable: %s", _e)
    tagging_svc = None  # type: ignore

try:
    import services.ai_infra_service as ai_infra_svc
except Exception as _e:
    logger.warning("ai_infra_service unavailable: %s", _e)
    ai_infra_svc = None  # type: ignore

try:
    import services.bcdr_metadata_service as bcdr_meta_svc
except Exception as _e:
    logger.warning("bcdr_metadata_service unavailable: %s", _e)
    bcdr_meta_svc = None  # type: ignore

try:
    import services.arc_service as arc_svc
except Exception as _e:
    logger.warning("arc_service unavailable: %s", _e)
    arc_svc = None  # type: ignore

from services.bcdr_assessment_service import (
    assess_all_resources,
    build_bcdr_dashboard_summary,
    ZoneAssessment,
)
from services.bcdr_recommendation_service import (
    generate_all_recommendations,
    build_quick_wins,
    build_priority_summary,
    BCDRRecommendation,
)
from services.bcdr_deliverables_service import (
    get_timeline_action_plan,
    get_dr_testing_plan,
    get_compliance_checklist,
    get_strategy_reference,
    build_executive_summary,
    generate_excel_report,
)
from services.bcdr_enhanced_service import (
    build_business_impact_analysis,
    build_recovery_sequence_plan,
)

load_dotenv()
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(name)s: %(message)s")
logger = logging.getLogger(__name__)

# Suppress extremely verbose Azure SDK HTTP-level logging (headers, bodies, tokens)
# These flood the console with 100+ lines per query and can crash terminals.
logging.getLogger("azure.core.pipeline.policies.http_logging_policy").setLevel(logging.WARNING)
logging.getLogger("azure.identity").setLevel(logging.WARNING)

app = FastAPI(title="Azure Infra IQ API", version="3.0.0")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173", "http://localhost:3000", "http://localhost:8000", "http://localhost:80"],
    allow_credentials=True, allow_methods=["GET", "POST", "OPTIONS"], allow_headers=["Content-Type", "Authorization"],
)

class SecurityHeadersMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request: Request, call_next):
        response = await call_next(request)
        response.headers["X-Content-Type-Options"] = "nosniff"
        # The embedded Architecture Map renders in a same-origin iframe under
        # /zuremap/*, so that prefix needs SAMEORIGIN; everything else stays DENY.
        _p = request.url.path
        response.headers["X-Frame-Options"] = "SAMEORIGIN" if (_p == "/zuremap" or _p.startswith("/zuremap/")) else "DENY"
        response.headers["X-XSS-Protection"] = "1; mode=block"
        return response

app.add_middleware(SecurityHeadersMiddleware)

# ── Entra ID sign-in gate ───────────────────────────────────────────────────
# When ENTRA_CLIENT_ID/ENTRA_TENANT_ID are configured (locally via backend/.env or
# by the deployment), every /api call must carry a valid Entra token. The SPA gets
# it via MSAL and sends it as a Bearer header — or as ?access_token= on the SSE
# stream, since the browser EventSource API cannot set headers. With no config the
# gate is inert, so open/local dev and the in-app setup wizard are unaffected.
_AUTH_PUBLIC_PATHS = {"/api/auth/config", "/api/version", "/health"}


@app.get("/api/auth/config", include_in_schema=False)
async def auth_config():
    """Public: runtime auth config the SPA reads at boot to initialise MSAL."""
    return auth_svc.public_config()


@app.middleware("http")
async def _entra_auth_gate(request: Request, call_next):
    if not auth_svc.is_enabled() or request.method == "OPTIONS":
        return await call_next(request)
    path = request.url.path
    # The embedded Architecture Map engine is served same-origin under /zuremap/*
    # (reverse-proxied to the in-container engine). The iframe cannot attach our
    # Entra Bearer header, so these routes are gated by a short-lived signed cookie
    # issued by POST /api/zuremap/session (which DOES pass the Bearer gate). A valid
    # Bearer is also accepted. Without this gate, /zuremap/api/az/* would be an
    # unauthenticated Azure-scanning endpoint on the public ingress.
    if path == "/zuremap" or path.startswith("/zuremap/"):
        if auth_svc.check_zuremap_cookie(request.cookies.get("zm_sess")):
            return await call_next(request)
        authz = request.headers.get("authorization")
        tok = authz[7:].strip() if authz and authz[:7].lower() == "bearer " else None
        if auth_svc.validate_bearer(tok):
            return await call_next(request)
        return JSONResponse({"detail": "Authentication required. Please sign in."}, status_code=401)
    # Only API routes are gated. Static assets and the SPA shell stay public so the
    # login page itself can load before the user has a token.
    if not path.startswith("/api/") or path in _AUTH_PUBLIC_PATHS:
        return await call_next(request)
    token = None
    authz = request.headers.get("authorization")
    if authz and authz[:7].lower() == "bearer ":
        token = authz[7:].strip()
    if not token:
        token = request.query_params.get("access_token")
    if not auth_svc.validate_bearer(token):
        return JSONResponse({"detail": "Authentication required. Please sign in."}, status_code=401)
    return await call_next(request)

_cache: dict = {}
_auto_refresh_task: Optional[asyncio.Task] = None   # background scheduler handle
_pool = ThreadPoolExecutor(max_workers=12)  # shared executor for blocking I/O (data) calls
# Dedicated pool for AI calls so long (~15-20s) Azure OpenAI requests can never
# starve the data pool and make every FinOps tab hang on a spinner.
_ai_pool = ThreadPoolExecutor(max_workers=4)
# ── Arc security findings cache (15 min TTL) ──────────────────────────────────
import time as _time_mod
_arc_security_cache: dict = {}  # {"data": [...], "ts": float}
_ARC_SECURITY_TTL = 900  # 15 minutes

REQUIRED_TAGS = ["owner", "environment", "project", "cost-center"]

SCORE_COLORS = {
    ScoreLabel.NOT_USED:      "#ef4444",
    ScoreLabel.RARELY_USED:   "#f97316",
    ScoreLabel.ACTIVELY_USED: "#eab308",
    ScoreLabel.FULLY_USED:    "#22c55e",
    ScoreLabel.UNKNOWN:       "#6b7280",
}

RESOURCE_TYPE_DISPLAY: dict[str, str] = {
    "microsoft.compute/virtualmachines":              "Virtual Machines",
    "microsoft.compute/virtualmachinescalesets":      "VM Scale Sets",
    "microsoft.compute/disks":                        "Managed Disks",
    "microsoft.storage/storageaccounts":              "Storage Accounts",
    "microsoft.sql/servers/databases":                "SQL Databases",
    "microsoft.sql/servers/elasticpools":             "SQL Elastic Pools",
    "microsoft.dbformysql/flexibleservers":           "MySQL Flexible",
    "microsoft.dbformysql/servers":                   "MySQL",
    "microsoft.dbforpostgresql/flexibleservers":      "PostgreSQL Flexible",
    "microsoft.dbforpostgresql/servers":              "PostgreSQL",
    "microsoft.web/sites":                            "App Services / Functions",
    "microsoft.web/serverfarms":                      "App Service Plans",
    "microsoft.logic/workflows":                      "Logic Apps",
    "microsoft.cache/redis":                          "Redis Cache",
    "microsoft.documentdb/databaseaccounts":          "Cosmos DB",
    "microsoft.eventhub/namespaces":                  "Event Hubs",
    "microsoft.servicebus/namespaces":                "Service Bus",
    "microsoft.network/applicationgateways":          "App Gateways",
    "microsoft.network/loadbalancers":                "Load Balancers",
    "microsoft.network/publicipaddresses":            "Public IPs",
    "microsoft.network/virtualnetworkgateways":       "VPN Gateways",
    "microsoft.network/expressroutecircuits":         "ExpressRoute",
    "microsoft.network/frontdoors":                   "Front Door",
    "microsoft.cdn/profiles":                         "CDN",
    "microsoft.keyvault/vaults":                      "Key Vaults",
    "microsoft.containerservice/managedclusters":     "AKS Clusters",
    "microsoft.containerinstance/containergroups":    "Container Instances",
    "microsoft.containerregistry/registries":         "Container Registry",
    "microsoft.apimanagement/service":                "API Management",
    "microsoft.datafactory/factories":                "Data Factory",
    "microsoft.cognitiveservices/accounts":           "Cognitive Services",
    "microsoft.search/searchservices":                "AI Search",
    "microsoft.machinelearningservices/workspaces":   "ML Workspaces",
    "microsoft.synapse/workspaces":                   "Synapse Analytics",
    "microsoft.hdinsight/clusters":                   "HDInsight",
    "microsoft.databricks/workspaces":                "Databricks",
    "microsoft.devices/iothubs":                      "IoT Hubs",
    "microsoft.network/networksecuritygroups":        "NSGs",
    "microsoft.network/networkinterfaces":            "Network Interfaces",
    "microsoft.operationalinsights/workspaces":       "Log Analytics",
    "microsoft.signalrservice/signalr":               "SignalR",
}


RESOURCE_CATEGORIES: dict[str, str] = {
    # Compute
    "microsoft.compute/virtualmachines":           "compute",
    "microsoft.compute/virtualmachinescalesets":   "compute",
    "microsoft.compute/disks":                     "compute",
    "microsoft.web/sites":                         "compute",
    "microsoft.web/serverfarms":                   "compute",
    "microsoft.containerservice/managedclusters":  "compute",
    "microsoft.containerinstance/containergroups": "compute",
    # Storage
    "microsoft.storage/storageaccounts":           "storage",
    "microsoft.compute/snapshots":                 "storage",
    "microsoft.containerregistry/registries":      "storage",
    # Data / Databases
    "microsoft.sql/servers/databases":             "data",
    "microsoft.sql/servers/elasticpools":          "data",
    "microsoft.dbformysql/flexibleservers":        "data",
    "microsoft.dbforpostgresql/flexibleservers":   "data",
    "microsoft.documentdb/databaseaccounts":       "data",
    "microsoft.cache/redis":                       "data",
    "microsoft.synapse/workspaces":                "data",
    "microsoft.databricks/workspaces":             "data",
    # AI / ML
    "microsoft.cognitiveservices/accounts":        "ai",
    "microsoft.machinelearningservices/workspaces":"ai",
    "microsoft.search/searchservices":             "ai",
    "microsoft.openai":                            "ai",
    # Networking infrastructure
    "microsoft.network/virtualnetworks":           "infrastructure",
    "microsoft.network/networksecuritygroups":     "infrastructure",
    "microsoft.network/privateendpoints":          "infrastructure",
    "microsoft.network/privatednszones":           "infrastructure",
    "microsoft.network/dnszones":                  "infrastructure",
    "microsoft.network/routetables":               "infrastructure",
    "microsoft.network/networkwatchers":           "infrastructure",
    "microsoft.network/natgateways":               "infrastructure",
}


def _resource_category(resource_type: str) -> str:
    t = resource_type.lower()
    for prefix, cat in RESOURCE_CATEGORIES.items():
        if t.startswith(prefix):
            return cat
    return "other"


def _check_tag_compliance(tags: dict) -> list[str]:
    tag_keys = {k.lower() for k in tags}
    missing = []
    for req in REQUIRED_TAGS:
        if not any(req.replace("-", "") in tk.replace("-", "") for tk in tag_keys):
            missing.append(req)
    return missing


def _portal_url(resource_id: str) -> str:
    tenant = os.getenv("AZURE_TENANT_ID", "")
    return f"https://portal.azure.com/#@{tenant}/resource{resource_id}"


def _month_daily_arrays(
    daily_data: list[tuple[str, float]],
) -> tuple[list[float], list[float]]:
    """
    Returns (curr_month_daily, prev_month_daily) where:
      curr_month_daily — one cost per day from day 1 to today (current month)
      prev_month_daily — one cost per day for the full previous calendar month
    Both arrays are indexed from 0 = day 1 of that month.
    """
    now   = datetime.now(tz=timezone.utc)
    today = datetime(now.year, now.month, now.day, tzinfo=timezone.utc)
    by_date = {d: c for d, c in daily_data}

    # Current month: day 1 … today
    curr = [
        round(by_date.get(f"{now.year}-{now.month:02d}-{day:02d}", 0.0), 4)
        for day in range(1, today.day + 1)
    ]

    # Previous month: full month
    first_of_this = datetime(now.year, now.month, 1, tzinfo=timezone.utc)
    last_of_prev  = first_of_this - timedelta(days=1)
    prev_year, prev_month, days_in_prev = last_of_prev.year, last_of_prev.month, last_of_prev.day
    prev = [
        round(by_date.get(f"{prev_year}-{prev_month:02d}-{day:02d}", 0.0), 4)
        for day in range(1, days_in_prev + 1)
    ]

    return curr, prev


def _sparkline_array(daily_data: list[tuple[str, float]], days: int = 30) -> list[float]:
    if not daily_data:
        return [0.0] * days
    by_date = {d: c for d, c in daily_data}
    # Anchor to today's UTC midnight so subtracting whole-day deltas never
    # lands on the wrong calendar date near end-of-day.
    now = datetime.now(tz=timezone.utc)
    today = datetime(now.year, now.month, now.day, tzinfo=timezone.utc)
    return [
        round(by_date.get((today - timedelta(days=i)).strftime("%Y-%m-%d"), 0.0), 4)
        for i in range(days - 1, -1, -1)
    ]


def _7d_trend(daily_vals: list[float]) -> Optional[float]:
    if len(daily_vals) < 14:
        return None
    recent = sum(daily_vals[-7:]) / 7
    prior  = sum(daily_vals[-14:-7]) / 7
    if prior <= 0:
        return None
    return round((recent - prior) / prior * 100, 1)


# ── Cost Score KPI ────────────────────────────────────────────────────────────

def _compute_cost_score(
    orphan_cost:   float,
    not_used_cost: float,
    total_curr:    float,
    health_pct:    float,
    resources:     list,
) -> tuple:
    """Composite 0–100 cost-efficiency score across 5 weighted dimensions:
    orphan waste (25%), confirmed waste (25%), Azure Advisor (20%),
    reservation coverage (20%), resource health (10%).
    """
    base = max(total_curr, 0.01)

    # 1. Orphan efficiency (25%) — orphaned spend as % of total bill
    orphan_score  = max(0.0, 100.0 - (orphan_cost  / base) * 500.0)

    # 2. Waste efficiency (25%) — confirmed-waste spend as % of total bill
    waste_score   = max(0.0, 100.0 - (not_used_cost / base) * 400.0)

    # 3. Advisor compliance (20%) — weighted severity density per resource
    adv_h = sum(sum(1 for a in r.advisor_recommendations if a.impact == "High")   for r in resources)
    adv_m = sum(sum(1 for a in r.advisor_recommendations if a.impact == "Medium") for r in resources)
    adv_l = sum(sum(1 for a in r.advisor_recommendations if a.impact == "Low")    for r in resources)
    density       = (adv_h * 3 + adv_m * 2 + adv_l) / max(1, len(resources))
    advisor_score = max(0.0, 100.0 - density * 40.0)

    # 4. Reservation coverage (20%) — cost covered by RIs / total RI-eligible cost
    eligible_res = [r for r in resources if r.ri_eligible]
    if eligible_res:
        elig_cost = sum(r.cost_current_month for r in eligible_res)
        cov_cost  = sum(r.cost_current_month for r in eligible_res if r.ri_covered)
        ri_score  = (cov_cost / max(elig_cost, 0.01)) * 100.0
    else:
        ri_score  = 75.0   # neutral — nothing is RI-eligible

    # 5. Resource health (10%) — % actively/fully used
    health_score = health_pct

    composite = (
        orphan_score  * 0.25 +
        waste_score   * 0.25 +
        advisor_score * 0.20 +
        ri_score      * 0.20 +
        health_score  * 0.10
    )
    composite = round(min(max(composite, 0.0), 100.0), 1)

    if   composite >= 85: grade, label = "A", "Excellent"
    elif composite >= 70: grade, label = "B", "Good"
    elif composite >= 55: grade, label = "C", "Fair"
    elif composite >= 40: grade, label = "D", "Poor"
    else:                 grade, label = "F", "Critical"

    return composite, grade, label, {
        "orphans":      round(orphan_score,  1),
        "waste":        round(waste_score,   1),
        "advisor":      round(advisor_score, 1),
        "reservations": round(ri_score,      1),
        "health":       round(health_score,  1),
    }


def _cost_series_empty(*series) -> bool:
    """True when every provided daily-cost series is missing or entirely zero.

    A list of zeros is truthy (non-empty), so a plain ``not (cm or pm)`` test
    treats an all-zero array as "present" and skips the snapshot backfill,
    leaving the Spend Trend rendered as a flat $0 line even though the durable
    cost snapshot holds real values. Test the actual numbers instead so all-zero
    series correctly trigger backfill.
    """
    for s in series:
        try:
            if any(float(x or 0) for x in (s or [])):
                return False
        except (TypeError, ValueError):
            # Unexpected non-numeric payload — treat as having data (don't clobber).
            return False
    return True


# ── Core build function ────────────────────────────────────────────────────────

async def _build_dashboard(
    refresh: bool,
    progress_cb=None,
    resource_group_filter: Optional[str] = None,
    skip_metrics: bool = False,
    subscription_id: Optional[str] = None,
) -> DashboardData:

    async def report(step: str, msg: str, pct: int):
        if progress_cb:
            await progress_cb({"type": "progress", "step": step, "message": msg, "pct": pct})

    loop     = asyncio.get_event_loop()
    executor = ThreadPoolExecutor(max_workers=10)
    cfg      = settings_svc.get()
    sub_ids  = settings_svc.get_subscription_ids()
    # Full set the identity can access, captured BEFORE any single-subscription
    # scoping below, so the scope selector can list every switchable subscription
    # (not only the ones in the current scan).
    all_accessible_sub_ids = list(sub_ids)

    def _fetch_subscription_names(ids: list[str]) -> dict[str, str]:
        try:
            from azure.mgmt.subscription import SubscriptionClient
            from services.azure_auth import get_credential
            client = SubscriptionClient(get_credential())
            return {s.subscription_id: s.display_name for s in client.subscriptions.list()
                    if s.subscription_id in ids}
        except Exception:
            return {}

    # Apply scan scope defaults (for testing/validation) when no explicit filter passed.
    # None  = caller did not specify a filter → apply scope defaults if configured.
    # ""    = caller explicitly selected "All" → clear any scope override.
    scope_sub = settings_svc.get_value("SCAN_SCOPE_SUBSCRIPTION_ID", "").strip()
    scope_rg  = settings_svc.get_value("SCAN_SCOPE_RESOURCE_GROUP",  "").strip()
    if scope_sub and resource_group_filter is None:
        sub_ids = [scope_sub] if scope_sub in sub_ids or not sub_ids else [scope_sub]
    if scope_rg and resource_group_filter is None:
        resource_group_filter = scope_rg

    scope_active = bool(scope_sub or scope_rg)

    # ── Live subscription scoping ────────────────────────────────────────────
    # When a subscription (or a comma-separated set — e.g. every subscription under a
    # selected management group) is chosen in the UI, scope the entire build to just
    # those subscriptions so the panel loads quickly with live data.
    if subscription_id:
        _scoped = [s.strip() for s in str(subscription_id).split(",") if s.strip()]
        if _scoped:
            sub_ids = _scoped
            scope_active = True
    if settings_svc.check_and_wipe_if_expired():
        raise EnvironmentError(
            "Credentials have been automatically cleared after the configured timeout. "
            "Please re-enter your service principal credentials in Settings."
        )

    await report("resources", f"Listing resources across {len(sub_ids)} subscription(s)…", 5)

    # On the fast initial open (skip_metrics), only the headline 2-month cost
    # figures are needed. The cost-trend queries (60-day daily, 6-month history,
    # tenant total-daily) feed charts/sparklines that are not required for first
    # paint, so we defer them — this cuts concurrent Cost Management calls from
    # four groups to one and largely eliminates tenant-shared 429 throttling on
    # open. They populate on a manual refresh (skip_metrics=False).
    fetch_cost_trends = not skip_metrics

    async def _resolved(value):
        return value

    resources_task      = loop.run_in_executor(executor, partial(list_all_resources, sub_ids))
    costs_task          = loop.run_in_executor(executor, partial(get_two_month_costs, sub_ids))
    advisor_task        = loop.run_in_executor(executor, partial(get_advisor_recommendations, sub_ids))
    daily_task          = loop.run_in_executor(executor, partial(get_daily_costs, 60, sub_ids))          if fetch_cost_trends else _resolved({})
    monthly_hist_task   = loop.run_in_executor(executor, partial(get_monthly_cost_history, 6, sub_ids))  if fetch_cost_trends else _resolved({})
    total_daily_task    = loop.run_in_executor(executor, partial(get_total_daily_costs, sub_ids))        if fetch_cost_trends else _resolved(([], []))
    sub_names_task      = loop.run_in_executor(executor, partial(_fetch_subscription_names, sub_ids))

    await report("costs", f"Fetching 2 months of cost data across {len(sub_ids)} subscription(s)…", 15)

    # Wrap each task with a timeout so a slow/rate-limited API call
    # cannot block the entire dashboard. Cost tasks get a slightly longer
    # timeout to absorb the bounded 429 retries (2 retries × ~20 s cap each).
    async def _guarded(coro, default, timeout=90):
        try:
            return await asyncio.wait_for(asyncio.shield(coro), timeout=timeout)
        except Exception:
            return default

    resources, (curr_costs, prev_costs, cost_fetch_error), advisor_map, daily_costs_raw, monthly_hist_raw, (total_daily_cm, total_daily_pm), sub_names = await asyncio.gather(
        _guarded(resources_task,     []),
        # If the cost task times out it's almost always tenant-shared 429 throttling,
        # not a permissions problem — carry a throttle-flagged error so the banner
        # renders as a transient (amber) notice rather than a red RBAC failure.
        _guarded(costs_task,         ({}, {}, "Cost Management timed out — likely rate-limited (429). Showing $0 for this scan; refresh in a few minutes."), timeout=120),  # bounded 429 retries → degrades to cached/partial
        _guarded(advisor_task,       {}),
        _guarded(daily_task,         {}, timeout=120),
        _guarded(monthly_hist_task,  {}, timeout=120),
        _guarded(total_daily_task,   ([], []), timeout=120),
        _guarded(sub_names_task,     {}),
    )

    # Credentials successfully used — reset the inactivity timer (SEC1)
    settings_svc.touch_credential_use()

    # ── Hydrate cost trend from the latest persisted snapshot ─────────────────
    # On the fast-open path (skip_metrics) the total-daily series is deferred, and
    # under 429 throttling it degrades to empty. Backfill from the cost_snapshots
    # table so the home SpendTrend never renders $0.00 / empty.
    if _cost_series_empty(total_daily_cm, total_daily_pm):
        try:
            # L2 (Redis) cache-aside in front of the durable SQL snapshot so a
            # restarted process / second replica gets the cost series instantly.
            _cost_snap = cache_svc.get_json("snap:cost:latest")
            if not _cost_snap:
                _cost_snap = persistence_svc.load_latest_cost_snapshot()
                if _cost_snap:
                    cache_svc.set_json("snap:cost:latest", _cost_snap, ttl_seconds=3600)
            if _cost_snap:
                total_daily_cm = _cost_snap.get("total_daily_cm") or total_daily_cm
                total_daily_pm = _cost_snap.get("total_daily_pm") or total_daily_pm
                if total_daily_cm or total_daily_pm:
                    logger.info(
                        "Dashboard: hydrated total-daily cost from snapshot (%s, %d cm / %d pm days)",
                        _cost_snap.get("captured_at"), len(total_daily_cm or []), len(total_daily_pm or []),
                    )
        except Exception as _che:
            logger.warning("Dashboard: cost snapshot hydrate failed: %s", _che)


    if resource_group_filter:
        resources = [r for r in resources if r["resource_group"].lower() == resource_group_filter.lower()]

    await report("activity", "Querying activity logs…", 28)
    activity_map, asp_links, ai_links, app_detail_map = await asyncio.gather(
        _guarded(loop.run_in_executor(executor, partial(get_subscription_activity,   sub_ids)),  {}),
        _guarded(loop.run_in_executor(executor, partial(get_app_service_plan_links,  sub_ids)),  {}),
        _guarded(loop.run_in_executor(executor, partial(get_app_insights_links,      resources, sub_ids)), {}),
        _guarded(loop.run_in_executor(executor, partial(get_app_service_details,     resources, sub_ids)), {}),
    )

    # Stamp server_farm_id onto each web app resource so the frontend can group apps under plans
    for r in resources:
        if r["type"] == "microsoft.web/sites":
            r["server_farm_id"] = asp_links.get(r["id"].lower())
            # Stamp A1–A8 app service detail fields
            detail = app_detail_map.get(r["id"].lower(), {})
            for key in ("app_kind", "runtime_stack", "last_modified", "custom_domain_count",
                        "health_check_enabled", "health_check_path", "ssl_expiry_date",
                        "slot_count", "has_linked_storage", "app_state"):
                r[key] = detail.get(key)

    await report("orphans", f"Checking {len(resources)} resources for orphans…", 35)
    orphan_results, (vm_power_map, vm_size_map), lock_set, vm_attachments, backup_ids = await asyncio.gather(
        _guarded(loop.run_in_executor(executor, partial(find_orphans,                resources, sub_ids)), []),
        _guarded(loop.run_in_executor(executor, partial(get_vm_power_states,         resources, sub_ids)), ({}, {})),
        _guarded(loop.run_in_executor(executor, partial(get_resource_locks,          sub_ids)),            set()),
        _guarded(loop.run_in_executor(executor, partial(get_vm_attachments,          resources, sub_ids)), {}),
        _guarded(loop.run_in_executor(executor, partial(get_backup_protected_ids,    resources, sub_ids)), set()),
    )
    orphan_map: dict[str, str] = dict(orphan_results)
    (storage_signals, kv_signals, rbac_map,
     ri_coverage_result, pe_targets, sql_replicas,
     reservation_recommendations,
     billing_covered_ids) = await asyncio.gather(
        _guarded(loop.run_in_executor(executor, partial(get_storage_access_signals,           resources, sub_ids)), {}),
        _guarded(loop.run_in_executor(executor, partial(get_keyvault_signals,                 resources, sub_ids)), {}),
        _guarded(loop.run_in_executor(executor, partial(get_rbac_signals,                     sub_ids)),            {}),
        _guarded(loop.run_in_executor(executor, partial(get_reservation_coverage,             sub_ids)),            (set(), [])),
        _guarded(loop.run_in_executor(executor, partial(get_private_endpoint_targets,         sub_ids)),            set()),
        _guarded(loop.run_in_executor(executor, partial(get_sql_replica_ids,      resources,  sub_ids)),            set()),
        _guarded(loop.run_in_executor(executor, partial(get_reservation_recommendations,      sub_ids)),            []),
        _guarded(loop.run_in_executor(executor, partial(get_reservation_covered_resource_ids, sub_ids)),            set()),
    )
    ri_covered_set, active_reservations = ri_coverage_result
    # Stamp power_state and VM size onto VM resources for display
    for r in resources:
        if r["type"] == "microsoft.compute/virtualmachines":
            vid = r["id"].lower()
            r["power_state"] = vm_power_map.get(vid, "unknown")
            vm_size = vm_size_map.get(vid)
            if vm_size:
                r["sku"] = vm_size
    # Stamp has_lock onto all resources.
    # has_lock = True ONLY when the resource itself has a direct lock.
    # RG/subscription-level locks are captured as has_inherited_lock which
    # contributes to is_protected (prevents "Not Used") but does NOT floor
    # the score at 60 — that floor is reserved for explicitly locked resources.
    for r in resources:
        rid_lower  = r["id"].lower()
        rg_prefix  = "/".join(rid_lower.split("/")[:5])
        sub_prefix = "/".join(rid_lower.split("/")[:3])
        r["has_lock"]          = rid_lower  in lock_set
        r["has_inherited_lock"] = (rg_prefix in lock_set or sub_prefix in lock_set)

    if skip_metrics:
        await report("metrics", "Metrics load on demand per resource — skipping bulk pull…", 42)
    else:
        await report("metrics", f"Pulling 30-day metrics for {len(resources)} resources…", 42)
    BATCH = 20
    all_metrics: dict[str, Any] = {}

    # ── Delta metrics cache ────────────────────────────────────────────────────
    # Load per-resource metrics that were persisted during the previous scan
    # (delta-cache TTL: 6 hours). Resources whose entry is still fresh are
    # skipped so we avoid hundreds of Azure Monitor API calls on every re-scan.
    #
    # On the fast-open path (skip_metrics) we ALSO load persisted metrics — but
    # with a much more generous "display" TTL — to hydrate utilisation so the
    # Waste Quadrant and scoring have real values on instant open instead of
    # all-zero / all-Unknown. A dedicated background job keeps these fresh.
    _metrics_ttl_hours = float(settings_svc.get_value("metrics_cache_ttl_hours", 6.0))
    _metrics_display_ttl_hours = float(settings_svc.get_value("metrics_display_ttl_hours", 24.0))
    _cached_metrics_raw: dict[str, dict] = {}
    if not refresh:
        # Use the cache whenever this is NOT a forced refresh. On the fast path
        # use the longer display TTL so the quadrant is never empty.
        _load_ttl = _metrics_display_ttl_hours if skip_metrics else _metrics_ttl_hours
        # L2 (Redis) cache-aside: short TTL so repeated dashboard opens skip the
        # SQL round-trip while staying fresh relative to the snapshot job.
        _mkey = f"snap:metrics:{_load_ttl}"
        _cached_metrics_raw = cache_svc.get_json(_mkey) or {}
        if not _cached_metrics_raw:
            _cached_metrics_raw = persistence_svc.load_resource_metrics(ttl_hours=_load_ttl)
            if _cached_metrics_raw:
                cache_svc.set_json(_mkey, _cached_metrics_raw, ttl_seconds=300)
    _cached_metrics_hits = 0

    async def fetch_batch(batch):
        nonlocal _cached_metrics_hits
        # Separate resources into cache-hit and cache-miss
        fresh_batch = []
        for r in batch:
            rid = r["id"].lower()
            if rid in _cached_metrics_raw:
                # Reconstruct a lightweight namespace object from the cached dict
                # so downstream code can access .primary_utilization etc.
                from types import SimpleNamespace
                all_metrics[rid] = SimpleNamespace(**_cached_metrics_raw[rid])
                _cached_metrics_hits += 1
            else:
                fresh_batch.append(r)
        if not fresh_batch:
            return
        tasks = [loop.run_in_executor(executor, partial(
            get_resource_metrics, r["id"], r["type"], r.get("subscription_id", "")
        )) for r in fresh_batch]
        for r, res in zip(fresh_batch, await asyncio.gather(*tasks, return_exceptions=True)):
            if not isinstance(res, Exception):
                all_metrics[r["id"].lower()] = res

    if not skip_metrics:
        for i in range(0, len(resources), BATCH):
            await fetch_batch(resources[i: i + BATCH])
    elif _cached_metrics_raw:
        # Fast-open path: no live pull, but hydrate utilisation from persisted
        # metrics so the Waste Quadrant / scoring are populated rather than 0%.
        from types import SimpleNamespace
        for r in resources:
            rid = r["id"].lower()
            if rid in _cached_metrics_raw and rid not in all_metrics:
                all_metrics[rid] = SimpleNamespace(**_cached_metrics_raw[rid])
                _cached_metrics_hits += 1

    if _cached_metrics_hits:
        logger.info(
            "Delta cache: reused metrics for %d/%d resources (skipped Monitor API calls)",
            _cached_metrics_hits, len(resources),
        )

    # Sanity-check: if zero resources got any metrics data, the Monitor API is
    # almost certainly blocked by a missing 'Monitoring Reader' role assignment.
    # Surface this as a loud warning rather than silently producing all-Unknown scores.
    from services.observability_service import NATIVE_METRICS_TYPES
    native_rids = {
        r["id"].lower() for r in resources
        if r["type"].lower() in NATIVE_METRICS_TYPES
    }
    native_with_metrics = sum(
        1 for rid in native_rids
        if rid in all_metrics and all_metrics[rid].primary_utilization is not None
    )
    if native_rids and native_with_metrics == 0 and not skip_metrics:
        logger.warning(
            "METRICS UNAVAILABLE: 0 of %d native-metrics resources returned any data. "
            "All resources will score as 'Unknown'. "
            "Most likely cause: the service principal is missing the 'Monitoring Reader' "
            "role on subscription %s. "
            "Assign it in Azure Portal > Subscriptions > Access Control (IAM).",
            len(native_rids),
            sub_ids[0] if sub_ids else "unknown",
        )

    # Merge App Insights metrics into linked web app metrics.
    # If a web app has low/no CPU metrics but its App Insights component shows
    # requests or active users, that is strong evidence the app is in use.
    for app_rid, ai_rid in ai_links.items():
        ai_metrics = all_metrics.get(ai_rid)
        if not ai_metrics:
            continue
        app_metrics = all_metrics.get(app_rid)
        if app_metrics is None:
            # Web app had no metrics at all — use App Insights as the primary source
            all_metrics[app_rid] = ai_metrics
        else:
            # Web app has some metrics — boost activity signal if App Insights shows traffic
            if ai_metrics.has_any_activity:
                app_metrics.has_any_activity = True
            # Use App Insights request count as primary utilization if it's stronger
            if ai_metrics.primary_utilization is not None:
                if app_metrics.primary_utilization is None or ai_metrics.primary_utilization > app_metrics.primary_utilization:
                    app_metrics.primary_utilization = ai_metrics.primary_utilization

    await report("scoring", "Scoring all resources…", 62)

    cost_floor = cfg.get("cost_floor_usd", 1.0)
    resource_dicts: list[dict] = []

    # ── S10: Auto-shutdown schedule detection ────────────────────────────────
    # Azure DevTest Labs auto-shutdown creates a free microsoft.devtestlab/schedules
    # resource named "shutdown-computevm-{vm-name}" in the same resource group.
    # VMs with a schedule are intentionally managed — never flag as Not Used.
    auto_shutdown_vms: set[str] = set()
    for r in resources:
        if r["type"] == "microsoft.devtestlab/schedules":
            sched_name = r["name"].lower()
            if sched_name.startswith("shutdown-computevm-"):
                vm_name = sched_name[len("shutdown-computevm-"):]
                auto_shutdown_vms.add(f"{vm_name}|{r['resource_group'].lower()}")

    # ── Partial month detection (B3) ────────────────────────────────────────
    # When today is the 1st–6th of the month, cost_current_month contains only
    # a fraction of normal monthly spend. Savings estimates derived from that
    # partial spend would be misleadingly small.
    # Use cost_previous_month as the savings baseline in this window.
    now = datetime.now(tz=timezone.utc)
    is_partial_month = now.day <= 6

    # Pre-compute last month's year/month for MTD delta calculation (used inside loop)
    _prev_year  = now.year if now.month > 1 else now.year - 1
    _prev_month = now.month - 1 if now.month > 1 else 12
    # Build the set of date strings covering last month days 1..now.day
    # e.g. today = April 6 → {"2026-03-01","2026-03-02",...,"2026-03-06"}
    _prev_mtd_dates = {
        f"{_prev_year}-{_prev_month:02d}-{day:02d}"
        for day in range(1, now.day + 1)
    }

    for r in resources:
        rid_lower = r["id"].lower()
        cost_curr = curr_costs.get(rid_lower, 0.0)
        cost_prev = prev_costs.get(rid_lower, 0.0)
        if cost_curr < cost_floor and cost_prev < cost_floor:
            continue

        # ── MTD-to-MTD delta (the only fair comparison during a live month) ───
        # Comparing April 1-6 ($33) to full March ($195) always shows a fake
        # ~83% drop. Instead compare April 1-6 to March 1-6 — same elapsed days.
        # cost_prev (full month) is kept for savings estimates and B3 logic.
        daily_data   = daily_costs_raw.get(rid_lower, [])
        cost_prev_mtd = sum(cost for date_str, cost in daily_data if date_str in _prev_mtd_dates)

        # Use MTD comparison when we have daily data for last month; fall back to
        # full-month delta if no daily data exists for last month's early days.
        if cost_prev_mtd > 0:
            cost_delta_pct = ((cost_curr - cost_prev_mtd) / cost_prev_mtd * 100)
            delta_is_mtd   = True
        else:
            cost_delta_pct = ((cost_curr - cost_prev) / cost_prev * 100) if cost_prev > 0 else 0.0
            delta_is_mtd   = False
        metrics        = all_metrics.get(rid_lower)
        util_pct       = metrics.primary_utilization if metrics else None
        has_activity   = metrics.has_any_activity    if metrics else False

        is_orphan     = rid_lower in orphan_map
        orphan_reason = orphan_map.get(rid_lower)

        activity          = activity_map.get(rid_lower)
        days_since        = activity.days_since_active if activity else None
        log_count         = activity.event_count       if activity else 0
        last_active       = activity.last_active_date  if activity else None
        recently_deployed = activity.recently_deployed if activity else False

        adv_recs    = advisor_map.get(rid_lower, [])
        adv_delta   = sum(rec.score_impact for rec in adv_recs if rec.category == "cost")
        adv_savings = sum(rec.potential_savings for rec in adv_recs if rec.potential_savings > 0)

        sparkline   = _sparkline_array(daily_data, 30)
        trend_7d    = _7d_trend(sparkline)
        daily_cm, daily_pm = _month_daily_arrays(daily_data)

        # Anomaly: last 7-day avg > 2× prior 23-day avg
        is_anomaly = False
        if len(sparkline) >= 14:
            recent_avg = sum(sparkline[-7:]) / 7
            older_avg  = sum(sparkline[-30:-7]) / max(len(sparkline[-30:-7]), 1)
            is_anomaly = older_avg > 0 and recent_avg > older_avg * 2.0

        resource_cat  = _resource_category(r["type"])
        is_infra      = is_infrastructure_resource(r["type"])

        data_conf, telem_src = get_data_confidence(
            r["type"], util_pct, has_activity, cost_curr,
        )

        # A deallocated VM has 0% CPU by design — do not treat it as idle/unused.
        # Its zero utilization is expected and intentional, not a signal of waste.
        vm_is_deallocated = r.get("power_state") in ("deallocated", "stopped")

        # idle_confirmed requires HIGH-confidence monitoring data.
        # Absence of metrics ≠ proof of idleness — only confirm idle when
        # Monitor data was actually fetched and shows no activity.
        idle_confirmed = (
            not has_activity
            and log_count == 0
            and not is_orphan
            and not vm_is_deallocated
            and data_conf == "high"
        )

        # ── S8: Recent deployment history ────────────────────────────────────
        # A resource deployed or updated via ARM/Bicep/Terraform in the last 30
        # days is actively maintained. Override idle signals — never flag as Not Used.
        if recently_deployed:
            has_activity   = True
            idle_confirmed = False

        # ── S10: Auto-shutdown schedule ───────────────────────────────────────
        # A VM with a DevTest Labs auto-shutdown schedule is intentionally managed
        # (dev/test VMs stopped nightly). Treat as actively used — never flag as waste.
        vm_has_auto_shutdown = (
            r["type"] == "microsoft.compute/virtualmachines"
            and f"{r['name'].lower()}|{r['resource_group'].lower()}" in auto_shutdown_vms
        )
        if vm_has_auto_shutdown:
            has_activity   = True
            idle_confirmed = False

        # S17: Intent/protection signals — track separately from usage signals.
        # These block deletion recommendations but must NOT set has_activity=True,
        # which would artificially inflate the utilization score. A VM with RBAC
        # but CPU=0% for 30 days should score low — it is protected, not active.
        is_protected      = False
        protection_reasons: list[str] = []

        # Direct lock on the resource itself → is_protected + score floor at 60 (handled in scorer)
        # Inherited RG/subscription lock → is_protected only (floor at 26, not 60)
        if r.get("has_lock", False):
            is_protected = True
            protection_reasons.append("resource lock")
        if r.get("has_inherited_lock", False):
            is_protected = True
            protection_reasons.append("resource group lock")

        # ── S7: Direct RBAC assignments ───────────────────────────────────────
        # RBAC = intent signal: someone explicitly granted access.
        # Prevents idle_confirmed penalty but does NOT boost utilization score.
        rbac_count = rbac_map.get(rid_lower, 0)
        if rbac_count > 0:
            idle_confirmed = False
            is_protected   = True
            protection_reasons.append(f"{rbac_count} RBAC assignment{'s' if rbac_count > 1 else ''}")

        # ── S16: Private endpoint target ──────────────────────────────────────
        # Private endpoint = intent signal: another resource consciously targets this one.
        # Prevents idle_confirmed penalty but does NOT boost utilization score.
        # For child resources (e.g. SQL databases, KV certs), check parent resource too
        has_private_endpoint = rid_lower in pe_targets
        if not has_private_endpoint:
            # Check parent resource — PE targets server not individual databases/certs
            parts = rid_lower.split("/")
            # SQL: .../servers/name/databases/name → check .../servers/name
            # KV:  .../vaults/name (already direct, no child issue)
            if len(parts) > 2:
                parent_id = "/".join(parts[:-2])
                if parent_id in pe_targets:
                    has_private_endpoint = True
        if has_private_endpoint:
            idle_confirmed = False
            is_protected   = True
            protection_reasons.append("private endpoint")

        # ── S15: Azure Backup coverage ────────────────────────────────────────
        # Backup policy = intent signal: someone set up protection for this resource.
        has_backup = rid_lower in backup_ids
        if has_backup:
            is_protected = True
            protection_reasons.append("backup policy")

        # ── S11: SQL geo-replica ───────────────────────────────────────────────
        # Replica = structural usage signal: it IS actively serving the primary.
        # This is real usage (replication traffic), not just intent — keep has_activity.
        is_sql_replica = rid_lower in sql_replicas
        if is_sql_replica:
            has_activity   = True
            idle_confirmed = False
            is_orphan      = False
            orphan_reason  = None

        # ── S3: Reservation / Savings Plan coverage ───────────────────────────
        # RI coverage = intent signal: customer has committed spend to this resource.
        # Prevents idle_confirmed penalty but does NOT boost utilization score.
        ri_coverage_key = f"{r['type']}|{(r.get('location') or '').lower().replace(' ', '')}"
        ri_covered = (
            ri_coverage_key in ri_covered_set
            or f"{r['type']}|*" in ri_covered_set
            or r["id"].lower() in billing_covered_ids
        )
        if ri_covered:
            idle_confirmed = False
            is_protected   = True
            protection_reasons.append("Reserved Instance")

        # ── Storage: last access time confirmation ────────────────────────────
        # If last access time tracking is enabled AND no transactions → confirmed unused.
        # If a lifecycle policy exists → actively managed, protect from Not Used.
        storage_signal = storage_signals.get(rid_lower)
        if storage_signal:
            if storage_signal.last_access_tracking_enabled and not has_activity:
                idle_confirmed = True
            if storage_signal.has_lifecycle_policy:
                # Lifecycle policy = someone actively manages this account
                has_activity = True

        # ── Key Vault: protection signals ─────────────────────────────────────
        # Purge-protected or infra-linked vaults are intentionally maintained.
        # Treat them like tagged production resources — floor at Actively Used.
        kv_signal = kv_signals.get(rid_lower)
        kv_is_protected = kv_signal.is_protected if kv_signal else False

        # Calculate actual resource age from creation date if available
        resource_age_days = 30  # fallback
        created_at_str = r.get("created_at", "")
        if created_at_str:
            try:
                created_dt = datetime.fromisoformat(created_at_str.replace("Z", "+00:00"))
                resource_age_days = max(0, (datetime.now(tz=timezone.utc) - created_dt).days)
            except Exception:
                pass

        has_lock = r.get("has_lock", False)

        # Merge Key Vault protection into tags so the existing tag-guard in
        # score_resource() floors the score at Rarely Used (40) at minimum.
        effective_tags = dict(r.get("tags", {}))
        if kv_is_protected:
            effective_tags.setdefault("criticality", "high")

        peak_util_pct = metrics.peak_utilization if metrics else None

        # Use MTD-to-MTD cost comparison for trend detection — prevents IDLE/FALLING
        # misclassification from comparing partial current month to full prior month.
        _cost_prev_for_trend = cost_prev_mtd if cost_prev_mtd > 0 else cost_prev

        base, final, trend_mod, trend, label = score_resource(
            util_pct=util_pct, cost_current=cost_curr, cost_previous=_cost_prev_for_trend,
            is_orphan=is_orphan, advisor_score_delta=adv_delta,
            has_any_activity=has_activity, resource_age_days=resource_age_days,
            days_since_active=days_since, activity_log_count=log_count,
            idle_confirmed=idle_confirmed,
            is_infrastructure=is_infra,
            data_confidence=data_conf,
            tags=effective_tags,
            vm_is_deallocated=vm_is_deallocated,
            has_lock=has_lock,
            has_inherited_lock=r.get("has_inherited_lock", False),
            is_protected=is_protected,
            peak_util_pct=peak_util_pct,
        )
        # Partial month: use previous month as savings baseline so estimates
        # reflect a realistic full-month figure rather than 2–6 days of spend.
        if is_partial_month and cost_prev > 0:
            savings_basis = cost_prev
        elif is_partial_month and cost_curr > 0 and now.day > 0:
            savings_basis = round(cost_curr * (30.0 / now.day), 4)
        else:
            savings_basis = cost_curr

        savings, recommendation = estimate_savings(
            savings_basis, final, is_orphan, adv_savings,
            has_metrics=(util_pct is not None),
        )

        missing_tags = _check_tag_compliance(r.get("tags", {}))
        carbon       = estimate_carbon(r["type"], r.get("location", ""), r.get("sku"))

        # ── D2: Waste Age ─────────────────────────────────────────────────────
        # How long has this resource been idle and how much has it cost?
        # "Idle 47 days · $382 wasted since Jan 15" creates urgency that a score never does.
        idle_since_date      = None
        days_idle            = None
        cumulative_waste_usd = None

        is_waste_candidate = label in (ScoreLabel.NOT_USED, ScoreLabel.RARELY_USED) and not is_infra
        if is_waste_candidate:
            ref_date_str = last_active or created_at_str
            if ref_date_str:
                try:
                    ref_dt   = datetime.fromisoformat(ref_date_str.replace("Z", "+00:00"))
                    days_idle = max(0, (now - ref_dt).days)
                    idle_since_date = ref_dt.date().isoformat()
                    if days_idle > 0 and cost_curr > 0:
                        daily_rate = cost_curr / 30.0
                        cumulative_waste_usd = round(daily_rate * days_idle, 2)
                except Exception:
                    pass

        # ── S19: Workload pattern classification ──────────────────────────────
        # Classifies the resource usage pattern for display and AI context.
        workload_pattern: Optional[str] = None
        if is_orphan:
            workload_pattern = "inactive"
        elif peak_util_pct is not None and util_pct is not None and util_pct > 0 and peak_util_pct > max(3 * util_pct, 40.0):
            workload_pattern = "bursty"      # big spikes vs average → scheduled job / event-driven
        elif trend == TrendDirection.FALLING and util_pct is not None and util_pct < 20:
            workload_pattern = "declining"   # usage trending down → optimization candidate
        elif (util_pct is None or util_pct < 3) and not has_activity:
            workload_pattern = "inactive"    # nothing running
        elif util_pct is not None and util_pct < 20:
            workload_pattern = "steady_low"  # consistently low but something running
        elif util_pct is not None:
            workload_pattern = "normal"

        # ── S22: "Why NOT waste" explanation ─────────────────────────────────
        # Surfaces the highest-confidence reason a resource was kept, so users
        # trust clean scans as much as flagged ones.
        protection_reason: Optional[str] = None
        if label not in (ScoreLabel.NOT_USED,) or is_protected:
            if has_lock:
                protection_reason = "Protected — has resource lock"
            elif ri_covered:
                protection_reason = "Covered — Reserved Instance active"
            elif peak_util_pct is not None and peak_util_pct > 60:
                protection_reason = f"Active — peak utilization {peak_util_pct:.0f}% in last 30 days"
            elif recently_deployed:
                protection_reason = "Active — deployed or updated in last 30 days"
            elif metrics and metrics.raw_absolute:
                calls = (metrics.raw_absolute.get("TotalCalls") or
                         metrics.raw_absolute.get("Requests") or
                         metrics.raw_absolute.get("requests/count"))
                if calls and calls > 0:
                    protection_reason = f"Active — {int(calls):,} requests in last 30 days"
            if protection_reason is None and has_private_endpoint:
                protection_reason = "Protected — has private endpoint"
            if protection_reason is None and rbac_count > 0:
                protection_reason = f"Protected — {rbac_count} role assignment{'s' if rbac_count > 1 else ''} on this resource"
            if protection_reason is None and has_backup:
                protection_reason = "Protected — has backup policy"

        # RI / Savings Plan opportunity — rates by resource type (1yr, 3yr discount vs on-demand)
        # NOTE: azure ml *workspaces* are NOT reservable — only the compute instances/clusters
        # inside them are. Workspaces are excluded; users are directed to portal to reserve compute.
        _RI_RATES = {
            # Confirmed available in Azure Portal → Purchase Reservations
            "microsoft.compute/virtualmachines":             (0.37, 0.57),
            "microsoft.sql/servers/databases":               (0.33, 0.44),
            "microsoft.sql/managedinstances":                (0.33, 0.55),
            "microsoft.sql/servers/elasticpools":            (0.33, 0.44),
            "microsoft.web/serverfarms":                     (0.35, 0.55),
            "microsoft.cache/redis":                         (0.37, 0.55),
            "microsoft.cache/redisenterprise":               (0.37, 0.55),
            "microsoft.documentdb/databaseaccounts":         (0.24, 0.48),
            "microsoft.dbforpostgresql/flexibleservers":     (0.33, 0.50),
            "microsoft.dbformysql/flexibleservers":          (0.33, 0.50),
            "microsoft.databricks/workspaces":               (0.40, 0.60),
            "microsoft.synapse/workspaces":                  (0.40, 0.60),
            "microsoft.compute/disks":                       (0.20, 0.38),
            "microsoft.kusto/clusters":                      (0.22, 0.42),
            "microsoft.compute/dedicatedhosts":              (0.30, 0.45),
            "microsoft.avs/privateclouds":                   (0.28, 0.46),
            "microsoft.netapp/netappaccounts/capacitypools": (0.17, 0.31),
        }
        # Eligibility: any resource of an RI-eligible type that costs money and isn't
        # already covered. Score is NOT used as a gate — a resource incurring $200/mo
        # is a valid RI candidate regardless of utilisation data availability.
        # Orphans are excluded (structurally confirmed dead resources).
        # Score and label are still passed to the UI to drive the term recommendation
        # (3yr / 1yr / Verify) so the user can make an informed commitment decision.
        _ri_eligible = cost_curr > 50 and not ri_covered and not is_orphan
        _ri_rates    = _RI_RATES.get(r["type"]) if _ri_eligible else None
        ri_1yr = round(cost_curr * _ri_rates[0], 2) if _ri_rates else 0.0
        ri_3yr = round(cost_curr * _ri_rates[1], 2) if _ri_rates else 0.0

        portal_url = _portal_url(r["id"])
        cli_delete = f'az resource delete --ids "{r["id"]}" --verbose'
        cli_resize = ""

        resource_dicts.append({
            "resource_id": r["id"], "resource_name": r["name"],
            "resource_type": r["type"], "resource_group": r["resource_group"],
            "location": r.get("location",""), "sku": r.get("sku"),
            "cost_current_month": round(cost_curr, 4), "cost_previous_month": round(cost_prev, 4),
            "cost_previous_month_mtd": round(cost_prev_mtd, 4),
            "cost_delta_is_mtd": delta_is_mtd,
            "cost_delta_pct": round(cost_delta_pct, 2),
            "avg_cpu_pct":    round(metrics.cpu,    2) if metrics and metrics.cpu    is not None else None,
            "avg_memory_pct": round(metrics.memory, 2) if metrics and metrics.memory is not None else None,
            "avg_disk_pct":   round(metrics.disk,   2) if metrics and metrics.disk   is not None else None,
            "avg_network_pct":round(metrics.network,2) if metrics and metrics.network is not None else None,
            "primary_utilization_pct": round(util_pct, 2) if util_pct is not None else None,
            "has_any_activity": has_activity,
            "has_lock": has_lock,
            "has_inherited_lock": r.get("has_inherited_lock", False),
            "base_score": round(base, 2), "advisor_score_delta": adv_delta,
            "trend_modifier": trend_mod, "ai_score_adjustment": 0,
            "final_score": round(final, 2), "score_label": label, "trend": trend,
            "advisor_recommendations": [
                {"category":rec.category,"impact":rec.impact,
                 "short_description":rec.short_description,
                 "score_impact":rec.score_impact,"potential_savings":rec.potential_savings}
                for rec in adv_recs
            ],
            "ai_confidence": None, "ai_action": None, "ai_explanation": None,
            "last_active_date": last_active, "days_since_active": days_since,
            "activity_log_count": log_count, "idle_confirmed": idle_confirmed,
            "rightsize_sku": None, "rightsize_savings_pct": 0.0,
            "ri_1yr_monthly_savings": ri_1yr,
            "ri_3yr_monthly_savings": ri_3yr,
            "ri_eligible": bool(_ri_eligible and _ri_rates),
            "missing_tags": missing_tags,
            "carbon_kg_per_month": carbon,
            "portal_url": portal_url, "cli_delete_cmd": cli_delete, "cli_resize_cmd": cli_resize,
            "is_anomaly": is_anomaly,
            "daily_costs": sparkline, "cost_7d_trend_pct": trend_7d,
            "daily_costs_cm": daily_cm, "daily_costs_pm": daily_pm,
            "monthly_cost_history": monthly_hist_raw.get(rid_lower, []),
            "estimated_monthly_savings": savings, "recommendation": recommendation,
            "savings_basis": savings_basis,   # may differ from cost_current_month in partial-month window
            "is_orphan": is_orphan, "orphan_reason": orphan_reason,
            "subscription_id": r.get("subscription_id", sub_ids[0] if sub_ids else ""),
            "resource_category": resource_cat,
            "is_infrastructure": is_infra,
            "data_confidence":  data_conf,
            "telemetry_source": telem_src,
            "tags": r.get("tags", {}),
            "instance_count": r.get("instance_count"),
            "server_farm_id": r.get("server_farm_id"),
            "power_state": r.get("power_state"),  # VMs only: running/deallocated/stopped/unknown
            # A1–A8: App Service detail fields (web apps only)
            "app_kind":             r.get("app_kind"),
            "runtime_stack":        r.get("runtime_stack"),
            "last_modified":        r.get("last_modified"),
            "custom_domain_count":  r.get("custom_domain_count", 0),
            "health_check_enabled": r.get("health_check_enabled", False),
            "health_check_path":    r.get("health_check_path"),
            "ssl_expiry_date":      r.get("ssl_expiry_date"),
            "slot_count":           r.get("slot_count", 0),
            "has_linked_storage":   r.get("has_linked_storage", False),
            "app_state":            r.get("app_state"),
            "storage_last_access_tracking": storage_signal.last_access_tracking_enabled if storage_signal else False,
            "storage_has_lifecycle_policy":  storage_signal.has_lifecycle_policy          if storage_signal else False,
            # AI1–AI7: Cognitive Services / OpenAI token + billing signals
            # AI Foundry uses InputTokens/OutputTokens/TotalTokens/ModelRequests;
            # classic Azure OpenAI uses ProcessedPromptTokens/ProcessedCompletionTokens/TotalCalls.
            "prompt_tokens":     (metrics.raw_absolute.get("InputTokens") or
                                  metrics.raw_absolute.get("ProcessedPromptTokens"))     if metrics else None,
            "completion_tokens": (metrics.raw_absolute.get("OutputTokens") or
                                  metrics.raw_absolute.get("ProcessedCompletionTokens")) if metrics else None,
            "total_tokens":      (metrics.raw_absolute.get("TotalTokens") or (
                (metrics.raw_absolute.get("InputTokens", 0) or 0) +
                (metrics.raw_absolute.get("OutputTokens", 0) or 0)) or (
                (metrics.raw_absolute.get("ProcessedPromptTokens", 0) or 0) +
                (metrics.raw_absolute.get("ProcessedCompletionTokens", 0) or 0))
            ) or None if metrics else None,
            "total_calls":    (metrics.raw_absolute.get("ModelRequests") or
                               metrics.raw_absolute.get("TotalCalls"))                   if metrics else None,
            "blocked_calls":  metrics.raw_absolute.get("BlockedCalls")  if metrics else None,
            "billing_type":   (
                "ptu" if "provisioned" in (r.get("sku") or "").lower() else "consumption"
            ) if r["type"].lower().startswith("microsoft.cognitiveservices") else None,
            "auto_shutdown": vm_has_auto_shutdown,
            "rbac_assignment_count": rbac_count,
            "ri_covered": ri_covered,
            "has_private_endpoint": has_private_endpoint,
            "is_sql_replica": is_sql_replica,
            "has_backup": has_backup,
            # S17: Intent vs Usage separation
            "is_protected": is_protected,
            "protection_reasons": protection_reasons,
            # S18: Peak utilization
            "peak_utilization_pct": round(peak_util_pct, 2) if peak_util_pct is not None else None,
            # D2: Waste Age
            "idle_since_date": idle_since_date,
            "days_idle": days_idle,
            "cumulative_waste_usd": cumulative_waste_usd,
            # S19: Workload pattern
            "workload_pattern": workload_pattern,
            # S22: Protection reason
            "protection_reason": protection_reason,
            "safe_action_steps": [],  # populated after AI scoring
        })

    # ── AI scoring ─────────────────────────────────────────────────────────
    active_ai   = get_active_provider()
    ai_enabled  = active_ai != "none"
    ai_reviewed = 0

    if ai_enabled:
        provider_label = {"claude": "Claude", "azure_openai": "Azure OpenAI"}.get(active_ai, "AI")
        await report("ai", f"Running {provider_label} analysis…", 72)
        verdicts = await loop.run_in_executor(executor, partial(get_ai_verdicts, resource_dicts))
        ai_map   = {v.resource_id: v for v in verdicts if not v.error}
        ai_reviewed = len(ai_map)

        for rd in resource_dicts:
            v = ai_map.get(rd["resource_id"].lower())
            if not v:
                continue
            rd["ai_score_adjustment"] = v.score_adjustment
            rd["ai_confidence"]       = v.confidence
            rd["ai_action"]           = v.action
            rd["ai_explanation"]      = v.explanation

            # Use the same MTD-adjusted cost baseline as the first pass so that
            # trend direction is consistent between the two scoring rounds.
            _ai_cost_prev = rd.get("cost_previous_month_mtd") or rd["cost_previous_month"]
            base, final, trend_mod, trend, label = score_resource(
                util_pct=rd["primary_utilization_pct"],
                cost_current=rd["cost_current_month"],
                cost_previous=_ai_cost_prev,
                is_orphan=rd["is_orphan"],
                advisor_score_delta=rd["advisor_score_delta"],
                ai_score_adjustment=v.score_adjustment,
                has_any_activity=rd["has_any_activity"],
                days_since_active=rd["days_since_active"],
                activity_log_count=rd["activity_log_count"],
                idle_confirmed=rd["idle_confirmed"],
                data_confidence=rd.get("data_confidence", "high"),
                vm_is_deallocated=rd.get("power_state") in ("deallocated", "stopped"),
                has_lock=rd.get("has_lock", False),
                has_inherited_lock=rd.get("has_inherited_lock", False),
                is_protected=rd.get("is_protected", False),
                peak_util_pct=rd.get("peak_utilization_pct"),
            )
            sav, rec = estimate_savings(
                rd.get("savings_basis", rd["cost_current_month"]), final, rd["is_orphan"],
                sum(a["potential_savings"] for a in rd["advisor_recommendations"])
            )
            rd.update({"final_score": round(final,2), "score_label": label, "trend": trend,
                        "estimated_monthly_savings": sav,
                        "recommendation": v.explanation or rec,
                        "safe_action_steps": v.action_steps or []})

    # ── Right-sizing ────────────────────────────────────────────────────────
    await report("rightsize", "Computing right-sizing recommendations…", 82)
    rs_recs: List[RightSizeRec] = await loop.run_in_executor(
        executor, partial(get_rightsize_recommendations, resource_dicts)
    )
    rs_map = {r.resource_id.lower(): r for r in rs_recs}
    for rd in resource_dicts:
        rs = rs_map.get(rd["resource_id"].lower())
        if rs:
            rd["rightsize_sku"]          = rs.suggested_sku
            rd["rightsize_savings_pct"]  = rs.savings_pct
            rd["cli_resize_cmd"]         = (
                f'az vm resize --resource-group "{rd["resource_group"]}" '
                f'--name "{rd["resource_name"]}" --size {rs.suggested_sku}'
                if "virtualmachines" in rd["resource_type"].lower() else
                f'az sql db update --resource-group "{rd["resource_group"]}" '
                f'--name "{rd["resource_name"]}" --service-objective {rs.suggested_sku}'
            )

    # ── S9: Dependency score propagation ────────────────────────────────────
    # Attached NICs, disks, and public IPs should not score worse than the VM
    # they serve. If a resource is attached to an active VM, floor its score
    # at the VM's level and clear any orphan/waste flags.
    if vm_attachments:
        vm_score_map = {
            rd["resource_id"].lower(): rd
            for rd in resource_dicts
            if rd["resource_type"] == "microsoft.compute/virtualmachines"
        }
        for rd in resource_dicts:
            rid_lower = rd["resource_id"].lower()
            vm_id = vm_attachments.get(rid_lower)
            if not vm_id:
                continue
            vm_rd = vm_score_map.get(vm_id)
            if not vm_rd:
                continue
            vm_score = vm_rd["final_score"]
            if vm_score <= rd["final_score"]:
                continue  # already scoring at least as well as the parent VM
            # Propagate VM score — the attached resource is needed while the VM runs
            rd["final_score"]              = vm_rd["final_score"]
            rd["score_label"]              = vm_rd["score_label"]
            rd["has_any_activity"]         = True
            rd["is_orphan"]                = False
            rd["orphan_reason"]            = None
            rd["estimated_monthly_savings"] = 0.0
            rd["recommendation"] = (
                f"Attached to VM '{vm_rd['resource_name']}' ({vm_rd['score_label']}). "
                f"This resource is in use and should not be removed independently."
            )
            rd["ai_action"] = "keep"
            rd["safe_action_steps"] = []  # will be regenerated below

    # ── Safe action steps ───────────────────────────────────────────────────
    # If AI provided specific steps, keep them. Otherwise fall back to
    # rule-based generic steps that at least know the resource type.
    for rd in resource_dicts:
        if not rd.get("safe_action_steps"):
            rd["safe_action_steps"] = get_safe_action_steps(
                resource_type=rd["resource_type"],
                score_label=rd.get("score_label", ""),
                is_orphan=rd.get("is_orphan", False),
                orphan_reason=rd.get("orphan_reason", "") or "",
                ai_action=rd.get("ai_action", "") or "",
            )
            rd["steps_source"] = "rules"
        else:
            rd["steps_source"] = "ai"

        # ── Substitute real resource values into CLI template placeholders ──
        # Steps use {name}, {rg}, {id}, {nsg} as placeholders — fill them in
        # so the Copy button gives the user a command that actually works.
        name = rd.get("resource_name", "")
        rg   = rd.get("resource_group", "")
        rid  = rd.get("resource_id", "")
        for step in rd["safe_action_steps"]:
            if step.get("az_cli"):
                step["az_cli"] = (
                    step["az_cli"]
                    .replace("{name}", name)
                    .replace("{rg}",   rg)
                    .replace("{id}",   rid)
                    .replace("{nsg}",  name)   # NSG name matches resource name
                )

    await report("assembling", "Assembling dashboard…", 92)

    # ── Build ResourceMetrics list ──────────────────────────────────────────
    _INTERNAL_KEYS = {"advisor_recommendations", "savings_basis"}
    resource_metrics_list: List[ResourceMetrics] = [
        ResourceMetrics(
            **{k: v for k, v in rd.items() if k not in _INTERNAL_KEYS},
            advisor_recommendations=[AdvisorRecommendation(**a) for a in rd["advisor_recommendations"]],
        )
        for rd in resource_dicts
    ]

    # ── KPI ─────────────────────────────────────────────────────────────────
    total_curr  = sum(r.cost_current_month  for r in resource_metrics_list)
    total_prev  = sum(r.cost_previous_month for r in resource_metrics_list)
    orphan_list = [r for r in resource_metrics_list if r.is_orphan]
    orphan_cost = sum(r.cost_current_month for r in orphan_list)
    scores      = [r.final_score for r in resource_metrics_list]
    avg_score   = sum(scores) / len(scores) if scores else 0.0
    total_save  = sum(r.estimated_monthly_savings for r in resource_metrics_list)
    total_adv   = sum(len(r.advisor_recommendations) for r in resource_metrics_list)
    total_carbon = sum(r.carbon_kg_per_month for r in resource_metrics_list)
    untagged     = sum(1 for r in resource_metrics_list if r.missing_tags)
    tag_pct      = round((len(resource_metrics_list) - untagged) / len(resource_metrics_list) * 100, 1) if resource_metrics_list else 100.0
    mom_delta    = total_curr - total_prev
    mom_pct      = (mom_delta / total_prev * 100) if total_prev > 0 else 0.0

    # Health metrics — exclude infrastructure from utilisation-based scoring
    scorable       = [r for r in resource_metrics_list if not r.is_infrastructure and not r.is_orphan]
    not_used_list  = [r for r in scorable if r.score_label == ScoreLabel.NOT_USED]
    healthy_list   = [r for r in scorable if r.score_label in (ScoreLabel.ACTIVELY_USED, ScoreLabel.FULLY_USED)]
    health_pct     = round(len(healthy_list) / len(scorable) * 100, 1) if scorable else 100.0
    infra_list     = [r for r in resource_metrics_list if r.is_infrastructure]

    _not_used_cost = round(sum(r.cost_current_month for r in not_used_list), 2)
    _cs, _cg, _cl, _cc = _compute_cost_score(
        orphan_cost   = orphan_cost,
        not_used_cost = _not_used_cost,
        total_curr    = total_curr,
        health_pct    = health_pct,
        resources     = resource_metrics_list,
    )

    kpi = KPIData(
        total_cost_current_month=round(total_curr, 2),
        total_cost_previous_month=round(total_prev, 2),
        mom_cost_delta=round(mom_delta, 2),
        mom_cost_delta_pct=round(mom_pct, 2),
        total_resources=len(resource_metrics_list),
        avg_optimization_score=round(avg_score, 1),
        total_potential_savings=round(total_save, 2),
        orphan_count=len(orphan_list),
        orphan_cost=round(orphan_cost, 2),
        advisor_total_recs=total_adv,
        ai_reviewed_count=ai_reviewed,
        not_used_count=len(not_used_list),
        not_used_cost=_not_used_cost,
        infrastructure_count=len(infra_list),
        health_score_pct=health_pct,
        subscription_count=len(sub_ids),
        billing_basis="previous_month" if is_partial_month else "current_month",
        billing_days_current=now.day,
        cost_score=_cs,
        cost_grade=_cg,
        cost_score_label=_cl,
        cost_score_components=_cc,
    )

    # ── Score distribution ──────────────────────────────────────────────────
    dist = {l: {"count": 0, "total_cost": 0.0} for l in ScoreLabel}
    for r in resource_metrics_list:
        dist[r.score_label]["count"]      += 1
        dist[r.score_label]["total_cost"] += r.cost_current_month
    score_distribution = [
        ScoreDistribution(label=l.value, count=d["count"], total_cost=round(d["total_cost"],2), color=SCORE_COLORS[l])
        for l, d in dist.items() if d["count"] > 0
    ]

    # ── Type summary ────────────────────────────────────────────────────────
    tmap: Dict[str, dict] = {}
    for r in resource_metrics_list:
        t = r.resource_type
        if t not in tmap:
            tmap[t] = {"count":0,"cost_curr":0.0,"cost_prev":0.0,"scores":[],"adv":0}
        tmap[t]["count"] += 1; tmap[t]["cost_curr"] += r.cost_current_month
        tmap[t]["cost_prev"] += r.cost_previous_month
        tmap[t]["scores"].append(r.final_score); tmap[t]["adv"] += len(r.advisor_recommendations)
    resource_type_summary = [
        ResourceTypeSummary(
            resource_type=t, display_name=RESOURCE_TYPE_DISPLAY.get(t, t.split("/")[-1].title()),
            count=d["count"], cost_current_month=round(d["cost_curr"],2),
            cost_previous_month=round(d["cost_prev"],2),
            avg_score=round(sum(d["scores"])/len(d["scores"]),1), advisor_rec_count=d["adv"],
        )
        for t, d in sorted(tmap.items(), key=lambda x: -x[1]["cost_curr"])
    ]

    # ── Orphan panel ────────────────────────────────────────────────────────
    orphans_panel = [
        OrphanResource(resource_id=r.resource_id, resource_name=r.resource_name,
                        resource_type=r.resource_type, resource_group=r.resource_group,
                        orphan_reason=r.orphan_reason or "Orphaned", monthly_cost=r.cost_current_month,
                        estimated_savings=r.estimated_monthly_savings)
        for r in orphan_list
    ]

    # ── Savings recs ────────────────────────────────────────────────────────
    savings_recs = sorted(
        [SavingsRecommendation(
            resource_id=r.resource_id, resource_name=r.resource_name,
            resource_type=r.resource_type, resource_group=r.resource_group,
            current_monthly_cost=r.cost_current_month,
            estimated_monthly_savings=r.estimated_monthly_savings,
            savings_pct=round(r.estimated_monthly_savings/r.cost_current_month*100 if r.cost_current_month>0 else 0, 1),
            recommendation=r.recommendation or "",
            ai_explanation=r.ai_explanation, ai_action=r.ai_action,
            priority="High" if r.final_score<=25 else "Medium" if r.final_score<=50 else "Low",
            score=r.final_score, advisor_count=len(r.advisor_recommendations),
        ) for r in resource_metrics_list if r.estimated_monthly_savings > 0],
        key=lambda x: -x.estimated_monthly_savings,
    )[:50]

    # ── Anomalies ──────────────────────────────────────────────────────────
    cost_anomalies = [
        CostAnomaly(
            resource_id=r.resource_id, resource_name=r.resource_name,
            resource_type=r.resource_type, resource_group=r.resource_group,
            avg_daily_cost_30d=round(r.cost_current_month/30, 4),
            latest_daily_cost=round(r.daily_costs[-1] if r.daily_costs else 0, 4),
            anomaly_factor=round((r.daily_costs[-1]/(r.cost_current_month/30)) if r.daily_costs and r.cost_current_month>0 else 1, 2),
        )
        for r in resource_metrics_list if r.is_anomaly
    ]

    # ── Right-size opportunities ────────────────────────────────────────────
    rightsize_opps = [
        RightSizeOpportunity(
            resource_id=r.resource_id, resource_name=r.resource_name,
            resource_type=r.resource_type, resource_group=r.resource_group,
            current_sku=r.sku or "", suggested_sku=r.rightsize_sku or "",
            current_cost=r.cost_current_month,
            estimated_savings=round(r.cost_current_month * r.rightsize_savings_pct / 100, 2),
            savings_pct=r.rightsize_savings_pct,
            reason=r.recommendation or "", cpu_pct=r.avg_cpu_pct,
        )
        for r in resource_metrics_list if r.rightsize_sku
    ]

    active_provider = get_active_provider()

    # ── Per-subscription summary ────────────────────────────────────────────
    sub_summaries: dict[str, dict] = {}
    for r in resource_metrics_list:
        sid = r.subscription_id
        if sid not in sub_summaries:
            sub_summaries[sid] = {"resource_count":0,"cost_current":0.0,"cost_previous":0.0,"orphan_count":0,"advisor_rec_count":0}
        sub_summaries[sid]["resource_count"] += 1
        sub_summaries[sid]["cost_current"]   += r.cost_current_month
        sub_summaries[sid]["cost_previous"]  += r.cost_previous_month
        if r.is_orphan:
            sub_summaries[sid]["orphan_count"] += 1
        sub_summaries[sid]["advisor_rec_count"] += len(r.advisor_recommendations)

    # Surface EVERY accessible subscription in the scope selector — scanned subs
    # carry real counts; the rest appear with zeros so the user can switch to them
    # even when the current scan is scoped to a single subscription.
    for _sid in all_accessible_sub_ids:
        if _sid and _sid not in sub_summaries:
            sub_summaries[_sid] = {"resource_count": 0, "cost_current": 0.0, "cost_previous": 0.0, "orphan_count": 0, "advisor_rec_count": 0}
    _missing_name_ids = [sid for sid in sub_summaries if sid not in sub_names]
    if _missing_name_ids:
        try:
            sub_names = {**_fetch_subscription_names(_missing_name_ids), **sub_names}
        except Exception:
            pass

    subscription_list = [
        SubscriptionSummary(
            subscription_id=sid,
            subscription_name=sub_names.get(sid, ""),
            resource_count=d["resource_count"],
            cost_current=round(d["cost_current"], 2),
            cost_previous=round(d["cost_previous"], 2),
            orphan_count=d["orphan_count"],
            advisor_rec_count=d["advisor_rec_count"],
        )
        for sid, d in sub_summaries.items()
    ]

    # ── Distinct resource groups for filter ────────────────────────────────
    # Use pre-cost-floor resources so RGs with only cheap resources still appear in the dropdown
    rg_list = sorted({r["resource_group"] for r in resources if r.get("resource_group")})

    # ── Resources-first partial render (additive, best-effort) ───────────────
    # The full CORE dashboard is now computed — resources, cost, scores, KPIs,
    # charts, orphans, savings, anomalies and spend trends. Only the AI narrative
    # (an LLM call that can be slow or even hang) and the CPU-only strategic panels
    # remain. Stream this complete-core snapshot NOW so the portal paints the whole
    # home view immediately instead of blocking on AI; the final "done" event below
    # re-sends the same data plus the AI narrative + strategic panels. Purely
    # additive and best-effort: any failure here is swallowed and never affects the
    # full DashboardData returned at the end of this function.
    if progress_cb is not None:
        try:
            _core_partial = DashboardData(
                kpi=kpi, score_distribution=score_distribution,
                resource_type_summary=resource_type_summary,
                resources=resource_metrics_list, orphans=orphans_panel,
                savings_recommendations=savings_recs,
                last_refreshed=datetime.now(tz=timezone.utc).isoformat(),
                ai_enabled=ai_enabled, ai_provider=active_provider,
                ai_narrative=None,
                demo_mode=False,
                total_carbon_kg=round(total_carbon, 1),
                tag_compliance_pct=tag_pct, total_untagged=untagged,
                cost_anomalies=cost_anomalies, rightsize_opportunities=rightsize_opps,
                subscriptions=subscription_list, resource_groups=rg_list,
                active_resource_group=resource_group_filter or "",
                active_subscription_id=sub_ids[0] if scope_sub and len(sub_ids) == 1 else "",
                scan_scope_active=scope_active,
                active_reservations=active_reservations,
                total_daily_cm=total_daily_cm, total_daily_pm=total_daily_pm,
                cost_data_warning=None,
            )
            await progress_cb({"type": "partial", "pct": 95, "data": _core_partial.model_dump()})
        except Exception as _cpe:
            logger.debug("Resources-first partial render skipped: %s", _cpe)

    # ── AI Narrative summary ────────────────────────────────────────────────
    ai_narrative: Optional[str] = None
    if ai_enabled:
        await report("narrative", "Generating AI subscription summary…", 96)
        ai_narrative = await loop.run_in_executor(
            executor,
            partial(get_ai_narrative, resource_metrics_list, kpi),
        )

    # Detect when cost data returned nothing — surface a visible warning
    cost_data_warning: Optional[str] = None
    if not curr_costs and not prev_costs and resource_metrics_list:
        if cost_fetch_error:
            err_lower = str(cost_fetch_error).lower()
            is_throttle = any(t in err_lower for t in (
                "429", "too many requests", "rate limit", "rate-limited",
                "timed out", "timeout", "temporarily unavailable",
            ))
            if is_throttle:
                cost_data_warning = (
                    "Cost Management API rate limit (429). "
                    "Cost figures show $0 for this scan. "
                    "Re-run the scan in a few minutes to get current cost data."
                )
            else:
                cost_data_warning = (
                    f"Azure Cost Management error: {cost_fetch_error}. "
                    "Cost, savings, and trend figures will show $0. "
                    "Check the backend logs for details."
                )
        else:
            cost_data_warning = (
                "Azure Cost Management returned no billing data. "
                "Cost, savings, and trend figures will show $0. "
                "Ensure the service principal has the Cost Management Reader role "
                "at the subscription scope, then refresh."
            )

    # ── Reservation over-commitment (F10) ──────────────────────────────────────
    # Augment each reservation dict with covered_cost, over_commitment_usd,
    # and days_to_expiry so the frontend can render F10 and Expiring Soon.
    covered_cost_by_key: dict[str, float] = {}
    for r in resource_metrics_list:
        if r.ri_covered:
            key = f"{r.resource_type}|{(r.location or '').lower().replace(' ', '')}"
            covered_cost_by_key[key] = covered_cost_by_key.get(key, 0.0) + r.cost_current_month

    # ── Synthetic reservations from billing when Reservations API is unavailable ──
    # If the Reservations API returned nothing (e.g. 403) but billing data shows
    # reservation-covered resources, build one synthetic entry per individual resource
    # so the "Already Reserved" section shows each resource as its own row.
    if not active_reservations and billing_covered_ids:
        synthetic: list[dict] = []
        for r in resource_metrics_list:
            if r.resource_id.lower() in billing_covered_ids:
                t = r.resource_type
                synthetic.append({
                    "reservation_id":     f"billing-{r.resource_id}",
                    "name":               r.resource_name or r.resource_id,
                    "display_name":       r.resource_name or r.resource_id,
                    "resource_type":      t,
                    "type_label":         RESOURCE_TYPE_DISPLAY.get(t, t.split("/")[-1]),
                    "sku":                r.sku or "",
                    "location":           r.location or "",
                    "term":               "",
                    "quantity":           1,
                    "expiry_date":        "",
                    "effective_date":     "",
                    "utilization_pct":    None,
                    "provisioning_state": "covered",
                    "covered_cost":       round(r.cost_current_month, 2),
                    "over_commitment_usd": 0.0,
                    "days_to_expiry":     None,
                    "from_billing":       True,
                    "resources":          [r.resource_name or r.resource_id],
                })
        active_reservations = synthetic
        logger.info("Generated %d synthetic reservation entries from billing coverage", len(active_reservations))

    ri_over_commitment_total = 0.0
    for res in active_reservations:
        # Days to expiry (skip for billing-synthetic entries which have no expiry date)
        if not res.get("from_billing"):
            try:
                from dateutil.parser import parse as _parse_dt
                exp_dt = _parse_dt(res["expiry_date"])
                res["days_to_expiry"] = max(0, (exp_dt.replace(tzinfo=None) - datetime.now()).days)
            except Exception:
                res["days_to_expiry"] = None

        # Over-commitment estimate (skip for billing-synthetic entries — covered_cost already set)
        if not res.get("from_billing"):
            key = f"{res.get('resource_type', '')}|{res.get('location', '')}"
            covered = covered_cost_by_key.get(key, 0.0)
            res["covered_cost"] = round(covered, 2)
            util = res.get("utilization_pct")
            if util is not None and 0 < util < 100 and covered > 0:
                wasted = covered * (100.0 - util) / util
                res["over_commitment_usd"] = round(wasted, 2)
                ri_over_commitment_total += wasted
            else:
                res["over_commitment_usd"] = 0.0

    ri_over_commitment_total = round(ri_over_commitment_total, 2)

    # ── Persist resource-level metrics for delta scanning on next run ─────────
    if all_metrics and not resource_group_filter:
        try:
            persistence_svc.save_resource_metrics(all_metrics)
        except Exception as _me:
            logger.warning("Could not persist resource metrics: %s", _me)

    # ── Strategic feature computation (no extra Azure API calls) ──────────────
    try:
        security_gaps_list = identify_security_gaps(resource_metrics_list)
    except Exception as _se:
        logger.warning("Security gap analysis failed: %s", _se)
        security_gaps_list = []

    try:
        modernization_list = detect_modernization_opportunities(resource_metrics_list)
    except Exception as _me:
        logger.warning("Modernization detection failed: %s", _me)
        modernization_list = []

    try:
        waf = compute_waf_scorecard(
            resource_metrics_list, kpi, orphans_panel,
            rightsize_opps, security_gaps_list,
        )
    except Exception as _we:
        logger.warning("WAF scorecard computation failed: %s", _we)
        waf = None

    try:
        innovation_gaps_list, service_adoption_list = detect_innovation_gaps(resource_metrics_list)
    except Exception as _ie:
        logger.warning("Innovation gap analysis failed: %s", _ie)
        innovation_gaps_list, service_adoption_list = [], []

    try:
        cloud_maturity = compute_cloud_maturity(resource_metrics_list, security_gaps_list, waf)
    except Exception as _ce:
        logger.warning("Cloud maturity computation failed: %s", _ce)
        cloud_maturity = None

    try:
        licensing_opps_list = detect_licensing_opportunities(resource_metrics_list)
    except Exception as _le:
        logger.warning("Licensing opportunity detection failed: %s", _le)
        licensing_opps_list = []

    try:
        backup_coverage_obj = analyze_backup_coverage(resource_metrics_list)
    except Exception as _be:
        logger.warning("Backup coverage analysis failed: %s", _be)
        backup_coverage_obj = None

    try:
        acr_opps_obj = analyze_acr_opportunities(resource_metrics_list)
    except Exception as _ae:
        logger.warning("ACR opportunity analysis failed: %s", _ae)
        acr_opps_obj = None

    return DashboardData(
        kpi=kpi, score_distribution=score_distribution,
        resource_type_summary=resource_type_summary,
        resources=resource_metrics_list, orphans=orphans_panel,
        savings_recommendations=savings_recs,
        last_refreshed=datetime.now(tz=timezone.utc).isoformat(),
        ai_enabled=ai_enabled, ai_provider=active_provider,
        ai_narrative=ai_narrative,
        demo_mode=False,
        total_carbon_kg=round(total_carbon, 1),
        tag_compliance_pct=tag_pct, total_untagged=untagged,
        cost_anomalies=cost_anomalies, rightsize_opportunities=rightsize_opps,
        subscriptions=subscription_list,
        resource_groups=rg_list,
        active_resource_group=resource_group_filter or "",
        active_subscription_id=sub_ids[0] if scope_sub and len(sub_ids) == 1 else "",
        scan_scope_active=scope_active,
        active_reservations=active_reservations,
        reservation_over_commitment_usd=ri_over_commitment_total,
        reservation_recommendations=reservation_recommendations,
        cost_data_warning=cost_data_warning,
        total_daily_cm=total_daily_cm,
        total_daily_pm=total_daily_pm,
        waf_scorecard=waf,
        security_gaps=security_gaps_list,
        modernization_opportunities=modernization_list,
        innovation_gaps=innovation_gaps_list,
        service_adoption_scores=service_adoption_list,
        cloud_maturity=cloud_maturity,
        licensing_opportunities=licensing_opps_list,
        backup_coverage=backup_coverage_obj,
        acr_opportunities=acr_opps_obj,
    )


# ── Health check endpoint ──────────────────────────────────────────────────────

@app.get("/health")
async def health_check():
    """Health check endpoint for monitoring and load balancers."""
    return {
        "status": "healthy",
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "version": "3.0.0",
    }


@app.get("/api/health-score")
async def get_health_score():
    """Composite infrastructure health score from cached dashboard data."""
    data = _cache.get("data:default") or _cache.get("data")
    if not data:
        raise HTTPException(status_code=404, detail="No scan data available. Run a scan first.")
    score = getattr(data, "composite_score", None) or data.composite_score if hasattr(data, "composite_score") else None
    grade = getattr(data, "composite_grade", None) or ""
    label = getattr(data, "composite_label", None) or ""
    breakdown = getattr(data, "composite_breakdown", None) or {}
    total_resources = getattr(data, "total_resources", 0)
    total_cost = getattr(data, "total_monthly_cost", 0)
    total_savings = getattr(data, "total_savings_potential", 0)
    return {
        "score": score,
        "grade": grade,
        "label": label,
        "breakdown": breakdown,
        "total_resources": total_resources,
        "total_monthly_cost": round(total_cost, 2) if total_cost else 0,
        "total_savings_potential": round(total_savings, 2) if total_savings else 0,
        "scanned_at": getattr(data, "scanned_at", None),
    }


# ── SSE streaming endpoint ─────────────────────────────────────────────────────

@app.get("/api/dashboard/stream")
async def stream_dashboard(
    refresh: bool = False,
    resource_group: Optional[str] = None,
    subscription: Optional[str] = None,
    fast: bool = True,
):
    """SSE endpoint — streams progress then live data.

    Live-data model:
      - `subscription` scopes the build to a single subscription (fast open).
      - `fast=True` (default) skips the bulk per-resource metrics pull; metrics
        are loaded on demand when a resource is expanded. This is what keeps the
        portal responsive instead of blocking for many minutes.
    """

    if settings_svc.get_value("demo_mode", False):
        async def demo_gen():
            from demo_data import build_demo_dashboard
            for step in [
                ("resources","Loading demo resources…",10),
                ("costs","Applying demo cost data…",30),
                ("metrics","Generating demo metrics…",55),
                ("scoring","Scoring demo resources…",75),
                ("ai","Applying demo AI insights…",90),
            ]:
                yield f"data: {json.dumps({'type':'progress','step':step[0],'message':step[1],'pct':step[2]})}\n\n"
                await asyncio.sleep(0.3)
            data = build_demo_dashboard()
            yield f"data: {json.dumps({'type':'done','pct':100,'data':data})}\n\n"
        return StreamingResponse(demo_gen(), media_type="text/event-stream",
                                  headers={"Cache-Control":"no-cache","X-Accel-Buffering":"no"})

    # resource_group=None  → no filter specified (scope defaults may apply)
    # resource_group=""    → user explicitly selected "All" (clear scope override)
    # resource_group="rg"  → filter to specific group
    # Normalise empty string to None so both map to the same "data:*" cache key.
    # Without this, resource_group="" produces cache key "data:" which is never
    # populated on startup, causing a spurious cache miss → full scan.
    _eff_rg = resource_group if resource_group else None
    cache_key = f"data:{_eff_rg if _eff_rg is not None else '*'}"
    now = datetime.now(tz=timezone.utc).timestamp()
    if not refresh and cache_key in _cache and now - _cache.get(f"{cache_key}:ts", 0) < settings_svc.get_value("cache_ttl_seconds", 1800):
        async def cached_gen():
            yield f"data: {json.dumps({'type':'progress','step':'cache','message':'Returning cached data…','pct':90})}\n\n"
            await asyncio.sleep(0.1)
            yield f"data: {json.dumps({'type':'done','pct':100,'data':_cache[cache_key].model_dump()})}\n\n"
        return StreamingResponse(cached_gen(), media_type="text/event-stream",
                                  headers={"Cache-Control":"no-cache","X-Accel-Buffering":"no"})

    progress_q: asyncio.Queue = asyncio.Queue()

    async def progress_cb(event: dict):
        await progress_q.put(event)

    async def build_task():
        try:
            data = await _build_dashboard(
                refresh, progress_cb,
                resource_group_filter=resource_group,
                skip_metrics=fast,
                subscription_id=subscription,
            )
            # A scan that comes back EMPTY (0 resources) is almost always transient — a
            # Cost/ARM throttle, a per-subscription API hiccup, or the managed identity's
            # Reader role still propagating right after a fresh deploy. NEVER let that blank
            # a dashboard the user is already looking at: if we have a prior non-empty
            # result (in memory or the durable snapshot), keep serving it instead of "0".
            if not (data.resources or []):
                prior = _cache.get(cache_key)
                if not (prior is not None and getattr(prior, "resources", None)):
                    try:
                        _snap = persistence_svc.load_latest_dashboard()
                        if _snap and (_snap.get("resources") or []):
                            prior = DashboardData(**{k: v for k, v in _snap.items() if k in DashboardData.model_fields})
                    except Exception:
                        prior = None
                if prior is not None and getattr(prior, "resources", None):
                    logger.warning(
                        "Scan returned 0 resources — serving last-good dashboard (%d resources) "
                        "instead of blanking the view", len(prior.resources or [])
                    )
                    await progress_q.put({"type": "done", "pct": 100, "data": prior.model_dump()})
                    return
                # else: genuinely no good data yet (fresh deploy, pre-RBAC) — fall through
                # and surface the empty result; the startup catch-up will self-heal.
            _cache[cache_key] = data
            _cache[f"{cache_key}:ts"] = datetime.now(tz=timezone.utc).timestamp()
            # Also populate the unkeyed fallback used by /api/dashboard (legacy)
            _cache["data"] = data
            _cache["cached_at"] = _cache[f"{cache_key}:ts"]
            # Durable snapshot — persisted to the DB (and mirrored to Redis L2) so the
            # portal opens instantly with last-good data after a restart
            # (stale-while-revalidate). Skip persisting an EMPTY (0-resource) scan: on a
            # fresh deploy the first scan can run before the managed identity's Reader
            # role has finished propagating and read 0 resources — saving that would make
            # the empty result stick until a manual refresh. Keeping the prior good
            # snapshot (or none) lets the startup catch-up self-heal once RBAC is ready.
            if data.resources:
                try:
                    _dash_json = json.loads(data.model_dump_json())
                    persistence_svc.save_dashboard(_dash_json)
                    if cache_svc.is_enabled():
                        cache_svc.set_json("dash:latest", _dash_json, ttl_seconds=24 * 3600)
                except Exception as _pe:
                    logger.warning("Could not persist dashboard snapshot: %s", _pe)
            else:
                logger.warning(
                    "Skipping dashboard snapshot persist: scan returned 0 resources "
                    "(managed-identity Reader role may still be propagating)"
                )
            await progress_q.put({"type": "done", "pct": 100, "data": data.model_dump()})
        except Exception as exc:
            logger.exception("Dashboard build failed")
            await progress_q.put({"type": "error", "message": str(exc)})

    asyncio.create_task(build_task())

    async def event_gen():
        while True:
            try:
                event = await asyncio.wait_for(progress_q.get(), timeout=120)
                yield f"data: {json.dumps(event, default=str)}\n\n"
                if event.get("type") in ("done", "error"):
                    break
            except asyncio.TimeoutError:
                yield f"data: {json.dumps({'type':'progress','step':'waiting','message':'Still working…','pct':50})}\n\n"

    return StreamingResponse(event_gen(), media_type="text/event-stream",
                              headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


# ── Legacy non-streaming endpoint (kept for compatibility) ─────────────────────

@app.get("/api/dashboard", response_model=DashboardData)
async def get_dashboard(refresh: bool = False):
    if settings_svc.get_value("demo_mode", False):
        from demo_data import build_demo_dashboard
        raw = build_demo_dashboard()
        return DashboardData(**{k: v for k, v in raw.items() if k in DashboardData.model_fields})

    # If refresh is explicitly requested, bypass cache
    if not refresh and "data" in _cache:
        # Always serve cached/persisted data immediately — never auto-scan
        # on a GET. Staleness is shown via data.last_refreshed; the user or
        # auto-refresh scheduler can trigger a rescan explicitly.
        return _cache["data"]
    try:
        # Live model: metrics are loaded on demand per resource, so the legacy
        # REST fallback also skips the bulk metrics pull to stay responsive.
        data = await _build_dashboard(refresh, skip_metrics=True)
        _cache.update({"data": data, "cached_at": datetime.now(tz=timezone.utc).timestamp()})
        return data
    except EnvironmentError as exc:
        raise HTTPException(status_code=500, detail=str(exc))
    except Exception as exc:
        logger.exception("Dashboard build failed")
        raise HTTPException(status_code=500, detail=str(exc))


@app.get("/api/resources", response_model=List[ResourceMetrics])
async def get_resources(resource_group: str | None = None, resource_type: str | None = None,
                         score_label: str | None = None, orphans_only: bool = False):
    if "data" not in _cache:
        await get_dashboard()
    resources = _cache["data"].resources
    if resource_group: resources = [r for r in resources if r.resource_group.lower() == resource_group.lower()]
    if resource_type:  resources = [r for r in resources if resource_type.lower() in r.resource_type.lower()]
    if score_label:    resources = [r for r in resources if r.score_label.value.lower() == score_label.lower()]
    if orphans_only:   resources = [r for r in resources if r.is_orphan]
    return resources


def _advisor_arg_index() -> dict:
    """Cached {resource_id_lower: [advisor ARG item, ...]} with problem+solution
    text — gives the resource drawer a full Advisor drill-down (30-min cache)."""
    import time
    idx = _cache.get("advisor_arg_index")
    ts = _cache.get("advisor_arg_index_ts", 0)
    if idx is not None and (time.time() - ts) < 1800:
        return idx
    try:
        from services.advisor_service import get_advisor_overview
        ov = get_advisor_overview()
        new_idx: dict = {}
        for it in ov.get("items", []):
            rid = (it.get("resource_id") or "").lower()
            if rid:
                new_idx.setdefault(rid, []).append(it)
        _cache["advisor_arg_index"] = new_idx
        _cache["advisor_arg_index_ts"] = time.time()
        return new_idx
    except Exception as exc:
        logger.debug("advisor ARG index build failed: %s", exc)
        return _cache.get("advisor_arg_index") or {}


@app.get("/api/resources/{resource_id:path}/detail")
async def get_resource_detail(resource_id: str):
    """
    Universal resource detail aggregator.
    Combines ResourceMetrics, security gaps, advisor recommendations, BCDR assessment,
    backup status, and blast-radius context — all from in-memory/cached data.
    No new Azure API calls are made.
    """
    if "data" not in _cache:
        await get_dashboard()

    resources = _cache["data"].resources
    # Find by exact resource_id or case-insensitive suffix match
    resource = None
    rid_lower = resource_id.lower()
    for r in resources:
        if r.resource_id.lower() == rid_lower or r.resource_id.lower().endswith(rid_lower):
            resource = r
            break

    if not resource:
        from fastapi import HTTPException
        raise HTTPException(status_code=404, detail=f"Resource '{resource_id}' not found in cache")

    r_dict = resource.model_dump()

    # ── Security gaps ──────────────────────────────────────────────────────────
    security_gaps = []
    try:
        from services.security_service import identify_security_gaps
        all_gaps = identify_security_gaps([r_dict])
        security_gaps = [g if isinstance(g, dict) else g.__dict__ for g in all_gaps]
    except Exception:
        pass

    # ── BCDR assessment ────────────────────────────────────────────────────────
    bcdr_assessment = None
    try:
        assessments = _get_bcdr_assessments()
        for a in assessments:
            a_id = a.get("resource_id", "") if isinstance(a, dict) else getattr(a, "resource_id", "")
            if a_id and (a_id.lower() == rid_lower or a_id.lower().endswith(rid_lower)):
                bcdr_assessment = a if isinstance(a, dict) else a.__dict__
                break
    except Exception:
        pass

    # ── Dependency / blast radius ──────────────────────────────────────────────
    blast_radius = {}
    try:
        graph = _cache.get("dependency_graph")
        if graph:
            blast_radius_obj = get_blast_radius(resource_id, graph)
            if blast_radius_obj:
                blast_radius = blast_radius_obj.model_dump() if hasattr(blast_radius_obj, "model_dump") else (blast_radius_obj if isinstance(blast_radius_obj, dict) else {})
    except Exception:
        pass

    # ── Cached AI findings that mention this resource ──────────────────────────
    ai_findings = []
    try:
        from services.tagging_service import get_latest_ai_analysis
        for analysis_type in ("ai_security_posture", "ai_cloud_maturity", "ai_resilience", "ai_backup"):
            cached = get_latest_ai_analysis(analysis_type, None, max_age_hours=48)
            if not cached:
                continue
            result = cached.get("result", cached)
            # Walk all nested findings for ones that reference this resource
            def _scan(obj):
                if isinstance(obj, dict):
                    ar = obj.get("affected_resources", [])
                    for item in ar:
                        if isinstance(item, dict):
                            iid = item.get("resource_id", "").lower()
                            iname = item.get("resource_name", "").lower()
                            if (iid and (iid == rid_lower or iid.endswith(rid_lower))) or \
                               (iname and iname == resource.resource_name.lower()):
                                ai_findings.append({
                                    "analysis_type": analysis_type,
                                    "title": obj.get("title", obj.get("name", "")),
                                    "severity": obj.get("severity", ""),
                                    "detail": obj.get("detail", obj.get("description", "")),
                                    "recommendation": obj.get("recommendation", obj.get("remediation", "")),
                                })
                    for v in obj.values():
                        _scan(v)
                elif isinstance(obj, list):
                    for item in obj:
                        _scan(item)
            _scan(result)
    except Exception:
        pass

    # ── Advisor recommendations (enriched: problem + solution + category + savings + portal) ──
    _CAT_LABEL = {"cost": "Cost", "security": "Security", "performance": "Performance",
                  "highavailability": "Reliability", "operationalexcellence": "Operational Excellence"}
    portal_base = f"https://portal.azure.com/#@/resource{resource.resource_id}"
    advisor_out: list = []
    seen_adv: set = set()
    try:
        for it in _advisor_arg_index().get(resource.resource_id.lower(), []):
            problem = (it.get("problem") or "").strip()
            solution = (it.get("solution") or "").strip()
            key = (problem + "|" + solution)[:120]
            if not (problem or solution) or key in seen_adv:
                continue
            seen_adv.add(key)
            cat = (it.get("category") or "").lower()
            advisor_out.append({
                "category": cat,
                "category_label": it.get("category_label") or _CAT_LABEL.get(cat, it.get("category", "")),
                "impact": it.get("impact", ""),
                "problem": problem,
                "solution": solution,
                "potential_savings": 0,
                "portal_url": portal_base,
            })
    except Exception:
        pass
    for a in (resource.advisor_recommendations or []):
        ad = a.model_dump() if hasattr(a, "model_dump") else (a if isinstance(a, dict) else a.__dict__)
        txt = (ad.get("short_description") or "").strip()
        key = (txt + "|")[:120]
        if not txt or key in seen_adv:
            continue
        seen_adv.add(key)
        cat = (ad.get("category") or "").lower()
        advisor_out.append({
            "category": cat,
            "category_label": _CAT_LABEL.get(cat, (ad.get("category") or "").title()),
            "impact": ad.get("impact", ""),
            "problem": txt,
            "solution": "",
            "potential_savings": ad.get("potential_savings", 0) or 0,
            "portal_url": portal_base,
        })

    return {
        "resource": r_dict,
        "security_gaps": security_gaps,
        "bcdr_assessment": bcdr_assessment,
        "blast_radius": blast_radius,
        "ai_findings": ai_findings[:20],  # cap at 20 cross-module findings
        "portal_url": portal_base,
        "advisor_recommendations": advisor_out,
    }


@app.get("/api/resources/by-name/{resource_name:path}", tags=["Resources"])
async def get_resource_id_by_name(resource_name: str):
    """
    Resolve a resource name → resource_id.
    Used by the frontend to open ResourceDetailDrawer from AI finding chips
    that only have resource names, not full ARM IDs.
    Returns the first matching resource (case-insensitive, partial match allowed).
    """
    if "data" not in _cache:
        await get_dashboard()

    resources = _cache["data"].resources
    name_lower = resource_name.strip().lower()

    # First: exact name match
    for r in resources:
        if (r.resource_name or "").lower() == name_lower:
            return {"resource_id": r.resource_id, "resource_name": r.resource_name, "resource_type": r.resource_type}

    # Second: case-insensitive suffix/contains match
    for r in resources:
        rname = (r.resource_name or "").lower()
        if rname and (name_lower in rname or rname in name_lower):
            return {"resource_id": r.resource_id, "resource_name": r.resource_name, "resource_type": r.resource_type}

    from fastapi import HTTPException
    raise HTTPException(status_code=404, detail=f"Resource '{resource_name}' not found")


# ── Settings endpoints ─────────────────────────────────────────────────────────

@app.get("/api/settings", response_model=AppSettings)
async def get_settings_endpoint():
    s = settings_svc.safe_export()
    # The app can scan Azure when it has SP creds OR a managed identity is available
    # (Azure Container Apps / App Service inject IDENTITY_ENDPOINT; AKS workload identity
    # sets AZURE_FEDERATED_TOKEN_FILE). This lets managed-identity deployments auto-load the
    # dashboard exactly like a local run that has a service principal in .env (instead of
    # falling back to the manual "Ready to scan" wizard).
    _mi_ready = bool(os.getenv("IDENTITY_ENDPOINT") or os.getenv("MSI_ENDPOINT") or os.getenv("AZURE_FEDERATED_TOKEN_FILE"))
    _sp_ready = bool(settings_svc.get_value("AZURE_CLIENT_SECRET","")) or bool(s.get("AZURE_CLIENT_ID",""))
    return AppSettings(
        azure_client_id          = s.get("AZURE_CLIENT_ID",""),
        azure_client_secret      = s.get("AZURE_CLIENT_SECRET",""),
        azure_tenant_id          = s.get("AZURE_TENANT_ID",""),
        azure_subscription_id    = s.get("AZURE_SUBSCRIPTION_ID",""),
        azure_subscription_ids   = s.get("AZURE_SUBSCRIPTION_IDS",""),
        has_azure_secret         = bool(settings_svc.get_value("AZURE_CLIENT_SECRET","")),
        auth_ready               = bool(_mi_ready or _sp_ready),
        ai_provider              = s.get("ai_provider", "none"),
        has_anthropic_key        = bool(settings_svc.get_value("ANTHROPIC_API_KEY","")),
        anthropic_api_key        = s.get("ANTHROPIC_API_KEY",""),
        azure_openai_endpoint    = s.get("AZURE_OPENAI_ENDPOINT",""),
        azure_openai_key         = s.get("AZURE_OPENAI_KEY",""),
        azure_openai_deployment  = s.get("AZURE_OPENAI_DEPLOYMENT","gpt-4o-mini"),
        has_azure_openai_key     = bool(settings_svc.get_value("AZURE_OPENAI_KEY","")),
        azure_ai_endpoint        = s.get("AZURE_AI_ENDPOINT",""),
        has_azure_ai_key         = bool(settings_svc.get_value("AZURE_AI_KEY","")),
        idle_threshold_pct       = s.get("idle_threshold_pct", 3.0),
        no_metrics_age_days      = s.get("no_metrics_age_days", 7),
        cost_floor_usd           = s.get("cost_floor_usd", 1.0),
        ai_cost_threshold_usd    = s.get("ai_cost_threshold_usd", 20.0),
        cache_ttl_seconds        = s.get("cache_ttl_seconds", 1800),
        demo_mode                = s.get("demo_mode", False),
        auto_refresh_interval_hours = int(s.get("auto_refresh_interval_hours", 0)),
        scan_scope_subscription_id = s.get("SCAN_SCOPE_SUBSCRIPTION_ID", ""),
        scan_scope_resource_group  = s.get("SCAN_SCOPE_RESOURCE_GROUP",  ""),
    )


@app.get("/api/settings/preflight")
async def preflight_check():
    """
    Check prerequisites for az login auth method:
    - Is Azure CLI installed?
    - Is the user already logged in?
    - Which accounts are accessible?
    """
    import shutil, subprocess, json as _json
    result = {"az_installed": False, "az_logged_in": False, "accounts": []}

    if not shutil.which("az"):
        return result

    result["az_installed"] = True
    try:
        proc = subprocess.run(
            ["az", "account", "list", "--output", "json"],
            capture_output=True, text=True, timeout=15,
        )
        if proc.returncode == 0:
            accounts = _json.loads(proc.stdout or "[]")
            enabled  = [a for a in accounts if str(a.get("state", "")).lower() == "enabled"]
            result["az_logged_in"] = len(enabled) > 0
            result["accounts"]     = [
                {"name": a.get("name"), "id": a.get("id")}
                for a in enabled
            ]
    except Exception:
        pass
    return result


@app.get("/api/settings/discover-subscriptions")
async def discover_subscriptions_endpoint(auth_method: str = ""):
    """
    Auto-discover subscriptions accessible to the current credential.
    Works with service principal AND `az login` — no setup required for CLI users.
    Pass auth_method=az_login to use AzureCliCredential directly (avoids stale env vars).
    """
    from services.azure_auth import discover_subscriptions
    try:
        subs = discover_subscriptions(auth_method=auth_method or None)
        return {"subscriptions": subs}
    except Exception as exc:
        raise HTTPException(status_code=400, detail=str(exc))


@app.get("/api/settings/auth-method")
async def get_auth_method_endpoint():
    """Returns which auth method is currently active."""
    from services.azure_auth import get_auth_method
    return {"method": get_auth_method()}


@app.get("/api/settings/resource-groups")
async def list_resource_groups_endpoint(subscription_id: str = ""):
    """List resource groups for a given subscription (or the configured one)."""
    from services.azure_auth import get_credential
    from azure.mgmt.resource import ResourceManagementClient
    sub_id = subscription_id or settings_svc.get_value("AZURE_SUBSCRIPTION_ID", "")
    if not sub_id:
        return {"resource_groups": []}
    try:
        client = ResourceManagementClient(get_credential(), sub_id)
        rgs = sorted([rg.name for rg in client.resource_groups.list()])
        return {"resource_groups": rgs}
    except Exception as exc:
        raise HTTPException(status_code=400, detail=str(exc))


@app.post("/api/settings")
async def update_settings_endpoint(body: dict):
    persist = body.pop("persist_to_env", False)
    # If the user is saving real SP credentials, auto-disable demo mode —
    # providing credentials is an explicit signal they want real data.
    has_sp_creds = all([
        body.get("AZURE_TENANT_ID",     "").strip(),
        body.get("AZURE_CLIENT_ID",     "").strip(),
        body.get("AZURE_CLIENT_SECRET", "").strip(),
    ])
    if has_sp_creds:
        body["demo_mode"] = False
    settings_svc.update(body, persist=persist)
    # If auto_refresh_interval changed, reschedule immediately
    if "auto_refresh_interval_hours" in body:
        _schedule_next_refresh(int(body["auto_refresh_interval_hours"]))
    _cache.clear()  # force fresh data with new settings
    # Wipe the per-resource metrics cache so delta scanning picks up fresh data
    persistence_svc.clear_resource_metrics()
    return {"ok": True, "message": "Settings updated. Refresh dashboard to apply."}


@app.post("/api/settings/test-azure")
async def test_azure(body: dict):
    import os as _os
    tid = body.get("AZURE_TENANT_ID") or settings_svc.get_value("AZURE_TENANT_ID","")
    cid = body.get("AZURE_CLIENT_ID") or settings_svc.get_value("AZURE_CLIENT_ID","")
    sec = body.get("AZURE_CLIENT_SECRET") or settings_svc.get_value("AZURE_CLIENT_SECRET","")
    sub = body.get("AZURE_SUBSCRIPTION_ID") or settings_svc.get_value("AZURE_SUBSCRIPTION_ID","")
    if not all([tid, cid, sec, sub]):
        raise HTTPException(status_code=400, detail="All four Azure fields are required.")
    try:
        from azure.identity import ClientSecretCredential
        from azure.mgmt.resource import ResourceManagementClient
        cred   = ClientSecretCredential(tenant_id=tid, client_id=cid, client_secret=sec)
        client = ResourceManagementClient(cred, sub)
        list(client.resource_groups.list())
        return {"ok": True, "message": "Connected successfully."}
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Connection failed: {exc}")


@app.post("/api/settings/test-ai")
async def test_ai(body: dict):
    provider = body.get("ai_provider") or settings_svc.get_value("ai_provider", "claude")

    if provider == "claude":
        key = body.get("ANTHROPIC_API_KEY") or settings_svc.get_value("ANTHROPIC_API_KEY", "")
        if not key:
            raise HTTPException(status_code=400, detail="Anthropic API key is required.")
        try:
            client = anthropic.Anthropic(api_key=key)
            client.messages.create(model="claude-haiku-4-5-20251001", max_tokens=10,
                                    messages=[{"role": "user", "content": "Hi"}])
            return {"ok": True, "message": "Claude (Anthropic) API key is valid."}
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"Claude key validation failed: {exc}")

    elif provider == "azure_openai":
        endpoint   = body.get("AZURE_OPENAI_ENDPOINT")   or settings_svc.get_value("AZURE_OPENAI_ENDPOINT", "")
        api_key    = body.get("AZURE_OPENAI_KEY")        or settings_svc.get_value("AZURE_OPENAI_KEY", "")
        deployment = body.get("AZURE_OPENAI_DEPLOYMENT") or settings_svc.get_value("AZURE_OPENAI_DEPLOYMENT", "gpt-4o-mini")
        if not endpoint or not api_key:
            raise HTTPException(status_code=400, detail="Azure OpenAI endpoint and key are required.")
        # Normalise endpoint — strip /openai/... paths, the SDK adds them
        import re as _re
        endpoint = _re.sub(r"/openai/.*$", "", endpoint.rstrip("/")) + "/"
        try:
            from openai import AzureOpenAI, NotFoundError, AuthenticationError
            client = AzureOpenAI(azure_endpoint=endpoint, api_key=api_key, api_version="2024-10-21")
            try:
                client.chat.completions.create(
                    model=deployment, max_completion_tokens=5,
                    messages=[{"role": "user", "content": "Hi"}],
                )
            except Exception as _e:
                if "max_completion_tokens" in str(_e) or "unsupported_parameter" in str(_e):
                    client.chat.completions.create(
                        model=deployment, max_tokens=5, temperature=0.1,
                        messages=[{"role": "user", "content": "Hi"}],
                    )
                else:
                    raise
            return {"ok": True, "message": f"Azure OpenAI connected successfully (deployment: {deployment})."}
        except Exception as exc:
            err = str(exc)
            if "404" in err or "not found" in err.lower():
                raise HTTPException(status_code=400, detail=(
                    f"Deployment '{deployment}' not found. "
                    "Go to Azure Portal → Azure OpenAI → Model deployments and copy the exact deployment name."
                ))
            if "401" in err or "authentication" in err.lower() or "unauthorized" in err.lower():
                raise HTTPException(status_code=400, detail="Invalid API key. Check Azure Portal → Azure OpenAI → Keys and Endpoint.")
            if "name or service not known" in err.lower() or "nodename" in err.lower():
                raise HTTPException(status_code=400, detail=f"Endpoint URL not reachable: {endpoint}. Check the URL in Azure Portal → Azure OpenAI → Keys and Endpoint.")
            raise HTTPException(status_code=400, detail=f"Azure OpenAI error: {err}")

    raise HTTPException(status_code=400, detail="Select a provider (claude or azure_openai) to test.")


# ── Auto-refresh scheduler ─────────────────────────────────────────────────────

_is_refreshing: bool = False   # True while a background scan is running
_refresh_started_ts: Optional[float] = None  # when the current scan started (stale-guard)
_next_refresh_ts: Optional[float] = None   # Unix timestamp of next scheduled scan


def _schedule_next_refresh(interval_hours: int) -> None:
    """Update the _next_refresh_ts based on the configured interval."""
    global _next_refresh_ts
    if interval_hours > 0:
        _next_refresh_ts = datetime.now(tz=timezone.utc).timestamp() + interval_hours * 3600
    else:
        _next_refresh_ts = None


async def _regenerate_ai_narrative() -> None:
    """Regenerate AI narrative for the cached data without a full re-scan."""
    try:
        loop     = asyncio.get_event_loop()
        executor = ThreadPoolExecutor(max_workers=2)
        for slot in ("data:*", "data"):
            cached: Optional[DashboardData] = _cache.get(slot)
            if cached and cached.resources and cached.kpi:
                narrative = await loop.run_in_executor(
                    executor,
                    partial(get_ai_narrative, cached.resources, cached.kpi),
                )
                if narrative:
                    cached.ai_narrative = narrative
                    cached.ai_enabled   = True
                    cached.ai_provider  = get_active_provider()
                    logger.info("AI narrative regenerated (%d chars)", len(narrative))
                    # Persist updated data
                    try:
                        persistence_svc.save_dashboard(json.loads(cached.model_dump_json()))
                    except Exception:
                        pass
                break
    except Exception as exc:
        logger.warning("AI narrative regeneration failed: %s", exc)


# ── FinOps Pre-Warm Cache ─────────────────────────────────────────────────────
# Background job that proactively warms FinOps cost data every 30 minutes so
# dashboard endpoints respond instantly instead of calling Azure live.

_finops_warm_cache: dict = {}                      # pre-warmed FinOps data
_FINOPS_WARMUP_INTERVAL_SECONDS = 30 * 60          # 30 minutes


async def _finops_cache_warmup() -> None:
    """Pre-fetch frequently-requested FinOps data and store in _finops_warm_cache."""
    if not _FINOPS_AVAILABLE:
        return
    logger.info("FinOps cache warmup: starting…")
    loop = asyncio.get_event_loop()
    sub_ids = finops_data_svc.get_subscription_ids()
    if not sub_ids:
        logger.warning("FinOps cache warmup: no subscription IDs configured — skipping")
        return

    ts = datetime.now(tz=timezone.utc).timestamp()

    # ── 1. Summary KPIs (from dashboard cache — no API call needed) ───────
    try:
        dash_for_warmup = _cache.get("data:*") or _cache.get("data")
        if not dash_for_warmup:
            # Proactively rehydrate from the durable persistence snapshot
            dash_for_warmup = _ensure_dashboard_in_cache()
        if dash_for_warmup and dash_for_warmup.kpi:
            # Use the same fast-path logic as the endpoint
            k = dash_for_warmup.kpi
            from datetime import date as _date_cls
            today_w = datetime.now(tz=timezone.utc).date()
            days_elapsed = today_w.day
            eom_fc = (k.total_cost_current_month / days_elapsed * 30) if days_elapsed > 0 else 0.0
            daily_cm = list(dash_for_warmup.total_daily_cm or [])
            daily_pm = list(dash_for_warmup.total_daily_pm or [])
            # All-zero arrays render as a flat $0 trend — backfill from the durable
            # cost snapshot so the warmed FinOps summary matches the live endpoint.
            if _cost_series_empty(daily_cm, daily_pm):
                _wc_snap = cache_svc.get_json("snap:cost:latest") or persistence_svc.load_latest_cost_snapshot()
                if _wc_snap:
                    daily_cm = list(_wc_snap.get("total_daily_cm") or daily_cm)
                    daily_pm = list(_wc_snap.get("total_daily_pm") or daily_pm)
            combined = daily_pm + daily_cm
            trend_30d = combined[-30:] if len(combined) >= 30 else ([0.0] * (30 - len(combined)) + combined)
            trend_dates = [str(today_w - timedelta(days=29 - i)) for i in range(30)]
            subs_sorted = sorted(dash_for_warmup.subscriptions or [], key=lambda s: -(s.cost_current or 0))
            by_sub = [{"id": s.subscription_id, "name": s.subscription_name or (s.subscription_id[:8] + "…"), "cost": round(s.cost_current or 0.0, 2)} for s in subs_sorted[:5]]
            extras = _derive_finops_kpi_extras(dash_for_warmup)
            from models.schemas import FinOpsKPI as _FinOpsKPI
            summary = _FinOpsKPI(
                total_spend_mtd=round(k.total_cost_current_month, 2),
                total_spend_last_month=round(k.total_cost_previous_month, 2),
                mom_delta_usd=round(k.mom_cost_delta, 2),
                mom_delta_pct=round(k.mom_cost_delta_pct, 1),
                forecast_eom_usd=round(eom_fc, 2),
                savings_identified_usd=round(getattr(dash_for_warmup, "total_potential_savings", None) or k.total_potential_savings, 2),
                budget_utilization_pct=extras["budget_utilization_pct"],
                budgets_exceeded=extras["budgets_exceeded"],
                budgets_at_risk=extras["budgets_at_risk"],
                ri_coverage_pct=extras["ri_coverage_pct"],
                ri_utilization_pct=extras["ri_utilization_pct"],
                has_reservations=extras["has_reservations"],
                has_budgets=extras["has_budgets"],
                tagging_compliance_pct=round(getattr(dash_for_warmup, "tag_compliance_pct", None) or 0.0, 1),
                total_untagged=int(getattr(dash_for_warmup, "total_untagged", 0) or 0),
                tag_required_keys=list(REQUIRED_TAGS),
                subscription_count=k.subscription_count,
                total_resource_count=len(dash_for_warmup.resources or []),
                cost_trend_30d=[round(v, 2) for v in trend_30d],
                cost_trend_dates=trend_dates,
                by_subscription=by_sub,
                data_source="dashboard_cache",
                generated_at=datetime.now(tz=timezone.utc).isoformat(),
            )
            _finops_warm_cache["summary"]    = summary
            _finops_warm_cache["summary_ts"] = ts
            logger.info("FinOps warmup: summary KPIs cached from dashboard cache (instant)")
        else:
            summary = await loop.run_in_executor(_pool, lambda: finops_svc.get_finops_kpi(sub_ids))
            _finops_warm_cache["summary"]    = summary
            _finops_warm_cache["summary_ts"] = ts
            logger.info("FinOps warmup: summary KPIs cached from live API")
    except Exception as e:
        logger.warning("FinOps warmup: summary failed: %s", e)

    # ── 2. Savings (dashboard-cache-backed, fast) ─────────────────────────
    try:
        dash = dash_for_warmup or _cache.get("data:*") or _cache.get("data")
        sv = await loop.run_in_executor(_pool, lambda: finops_svc.get_savings_summary(dash))
        _finops_warm_cache["savings"]    = sv
        _finops_warm_cache["savings_ts"] = ts
        logger.info("FinOps warmup: savings cached")
    except Exception as e:
        logger.warning("FinOps warmup: savings failed: %s", e)

    # ── 3. Commitments ────────────────────────────────────────────────────
    try:
        cm = await asyncio.wait_for(
            loop.run_in_executor(_pool, commitment_svc.get_commitment_summary),
            timeout=30.0,
        )
        _finops_warm_cache["commitments"]    = cm
        _finops_warm_cache["commitments_ts"] = ts
        logger.info("FinOps warmup: commitments cached")
    except (asyncio.TimeoutError, Exception) as e:
        logger.warning("FinOps warmup: commitments failed/timeout: %s", e)

    # ── 4. Allocation breakdowns (3 key dimensions) ───────────────────────
    for dim in ("SubscriptionId", "ServiceFamily", "ResourceType"):
        try:
            alloc = await loop.run_in_executor(
                _pool, lambda d=dim: finops_svc.get_cost_allocation(d, "last_30d", sub_ids))
            _finops_warm_cache[f"alloc_{dim}"]    = alloc
            _finops_warm_cache[f"alloc_{dim}_ts"] = ts
            logger.info("FinOps warmup: allocation by %s cached", dim)
        except Exception as e:
            logger.warning("FinOps warmup: allocation(%s) failed: %s", dim, e)

    # ── 5. Chargeback ─────────────────────────────────────────────────────
    try:
        cb = await loop.run_in_executor(_pool, lambda: finops_svc.get_chargeback_report("last_30d", sub_ids))
        _finops_warm_cache["chargeback"]    = cb
        _finops_warm_cache["chargeback_ts"] = ts
        logger.info("FinOps warmup: chargeback cached")
    except Exception as e:
        logger.warning("FinOps warmup: chargeback failed: %s", e)

    # ── 6. Forecast — runs last; capped at 90s so it doesn't block others ─
    try:
        fc = await asyncio.wait_for(
            loop.run_in_executor(_pool, lambda: forecast_finops_svc.get_forecast(horizon_days=30, subscription_ids=sub_ids)),
            timeout=90.0,
        )
        _finops_warm_cache["forecast"]    = fc
        _finops_warm_cache["forecast_ts"] = ts
        logger.info("FinOps warmup: forecast cached")
    except (asyncio.TimeoutError, Exception) as e:
        logger.warning("FinOps warmup: forecast failed/timeout: %s", e)

    _finops_warm_cache["last_warmup"]    = datetime.now(tz=timezone.utc).isoformat()
    _finops_warm_cache["last_warmup_ts"] = ts
    elapsed = datetime.now(tz=timezone.utc).timestamp() - ts
    logger.info("FinOps cache warmup: completed in %.1fs", elapsed)


async def _finops_cache_warmup_loop() -> None:
    """Background loop: run FinOps cache warmup every 30 minutes."""
    await asyncio.sleep(5)  # wait for app to fully initialize
    while True:
        try:
            await _finops_cache_warmup()
        except Exception as e:
            logger.error("FinOps cache warmup loop error: %s", e)
        await asyncio.sleep(_FINOPS_WARMUP_INTERVAL_SECONDS)


async def _background_refresh_loop() -> None:
    """Background task that periodically refreshes the full-scan cache."""
    global _is_refreshing, _refresh_started_ts, _next_refresh_ts
    while True:
        interval_hours = int(settings_svc.get_value("auto_refresh_interval_hours", 0))
        if interval_hours <= 0:
            # Scheduler disabled — sleep and re-check every minute
            _next_refresh_ts = None
            await asyncio.sleep(60)
            continue

        _schedule_next_refresh(interval_hours)
        wait_secs = _next_refresh_ts - datetime.now(tz=timezone.utc).timestamp()
        if wait_secs > 0:
            await asyncio.sleep(wait_secs)

        # Double-check interval hasn't been disabled while we were sleeping
        interval_hours = int(settings_svc.get_value("auto_refresh_interval_hours", 0))
        if interval_hours <= 0:
            continue

        if _is_refreshing:
            await asyncio.sleep(30)
            continue

        logger.info("Auto-refresh: starting scheduled background scan")
        _is_refreshing = True
        _refresh_started_ts = datetime.now(tz=timezone.utc).timestamp()
        try:
            # Light scheduled refresh: skip the bulk Azure Monitor metrics pull. The
            # dedicated, lock-protected metrics-snapshot job keeps utilisation fresh.
            # A FULL metrics scan here ran on EVERY replica every interval and could
            # overload the app — making /api/settings time out, which made the SPA
            # fall back to the manual setup wizard. skip_metrics=True also defers the
            # heavy cost-trend queries, so the periodic refresh stays cheap.
            data = await _build_dashboard(True, None, resource_group_filter=None, skip_metrics=True)
            # Persist result in the unfiltered cache slot
            cache_key = "data:*"
            now_ts = datetime.now(tz=timezone.utc).timestamp()
            _cache[cache_key] = data
            _cache[f"{cache_key}:ts"] = now_ts
            _cache["data"] = data
            _cache["cached_at"] = now_ts
            # Persist to SQLite so the next restart loads fresh data
            try:
                persistence_svc.save_dashboard(json.loads(data.model_dump_json()))
            except Exception as _pe:
                logger.warning("Auto-refresh: could not persist dashboard: %s", _pe)
            logger.info("Auto-refresh: scan complete")
        except Exception as exc:
            logger.error("Auto-refresh scan failed: %s", exc)
        finally:
            _is_refreshing = False
            _refresh_started_ts = None
        _schedule_next_refresh(interval_hours)


# ── Utilisation metrics snapshot job ──────────────────────────────────────────
# Periodically pulls 30-day Azure Monitor metrics for every resource and persists
# them to the resource_metrics table. The dashboard fast-open path then hydrates
# utilisation from this store (see _build_dashboard), so the Waste Quadrant and
# scoring are populated instantly instead of showing 0% on every open.
#
# This job hits Azure Monitor only (NOT Cost Management), so it does not
# reintroduce the tenant-shared cost 429 throttling that the live-data model
# was designed to avoid.

_metrics_snapshot_state: dict = {
    "last_run": None,        # ISO timestamp of last completed run
    "last_run_ts": None,     # epoch seconds
    "resource_count": 0,     # resources persisted in last run
    "running": False,
    "last_error": None,
}


async def _metrics_snapshot_run() -> dict:
    """Fetch 30-day metrics for all resources and persist to resource_metrics."""
    if _metrics_snapshot_state.get("running"):
        return {"skipped": "already running"}
    # Cross-replica lock: when Redis is configured, only one process/replica runs
    # the (Monitor-heavy) pull at a time. Returns "LOCAL" when Redis is absent, so
    # the in-process flag above governs single-process behavior unchanged.
    _lock_token = cache_svc.acquire_lock("lock:metrics_snapshot", ttl_seconds=3600)
    if _lock_token is None:
        return {"skipped": "locked by another replica"}
    _metrics_snapshot_state["running"] = True
    started = datetime.now(tz=timezone.utc)
    loop = asyncio.get_event_loop()
    try:
        sub_ids = settings_svc.get_subscription_ids()
        if not sub_ids:
            logger.warning("Metrics snapshot: no subscription IDs configured — skipping")
            _metrics_snapshot_state["last_error"] = "no subscriptions configured"
            return {"error": "no subscriptions configured"}

        logger.info("Metrics snapshot: starting full metrics pull…")
        resources = await loop.run_in_executor(_pool, partial(list_all_resources, sub_ids))
        all_metrics: dict[str, Any] = {}
        BATCH = 20
        for i in range(0, len(resources), BATCH):
            batch = resources[i: i + BATCH]
            tasks = [
                loop.run_in_executor(
                    _pool,
                    partial(get_resource_metrics, r["id"], r["type"], r.get("subscription_id", "")),
                )
                for r in batch
            ]
            for r, res in zip(batch, await asyncio.gather(*tasks, return_exceptions=True)):
                if not isinstance(res, Exception):
                    all_metrics[r["id"].lower()] = res

        if all_metrics:
            await loop.run_in_executor(_pool, partial(persistence_svc.save_resource_metrics, all_metrics))
            # Invalidate the short-lived Redis metrics cache so the rebuild below
            # and the next dashboard open pick up the freshly-persisted metrics.
            try:
                _disp = float(settings_svc.get_value("metrics_display_ttl_hours", 24.0))
                _cttl = float(settings_svc.get_value("metrics_cache_ttl_hours", 6.0))
                cache_svc.delete(f"snap:metrics:{_disp}")
                cache_svc.delete(f"snap:metrics:{_cttl}")
            except Exception:
                pass

        # Rebuild and persist the dashboard snapshot so utilisation-derived
        # fields (primary_utilization_pct → Waste Quadrant) are populated on the
        # snapshot-first fast path. The fast build (refresh=False, skip_metrics=
        # True) hydrates utilisation from the metrics we just persisted and
        # backfills cost trend from the durable cost snapshot — no extra live
        # Monitor calls.
        try:
            data = await _build_dashboard(False, None, resource_group_filter=None, skip_metrics=True)
            # Don't let an empty rebuild overwrite a good snapshot or persist an empty one:
            # on a fresh deploy the managed identity's Reader role may not have propagated
            # yet, so this scan can return 0 resources. Persisting that would defeat the
            # dashboard self-heal (the portal would serve an empty snapshot). Only cache /
            # persist when the rebuild actually returned resources.
            if data.resources:
                _now_ts = datetime.now(tz=timezone.utc).timestamp()
                _cache["data:*"] = data
                _cache["data:*:ts"] = _now_ts
                _cache["data"] = data
                _cache["cached_at"] = _now_ts
                _dash_json = json.loads(data.model_dump_json())
                await loop.run_in_executor(
                    _pool, partial(persistence_svc.save_dashboard, _dash_json)
                )
                # Mirror the rebuilt dashboard into Redis so a restarted process or a
                # second replica serves a warm Waste-Quadrant-populated dashboard
                # instantly instead of rebuilding under throttle.
                try:
                    _disp_ttl = float(settings_svc.get_value("metrics_display_ttl_hours", 24.0))
                    cache_svc.set_json("dash:latest", _dash_json, ttl_seconds=int(_disp_ttl * 3600))
                except Exception:
                    pass
                logger.info("Metrics snapshot: dashboard snapshot rebuilt with fresh utilisation")
            else:
                logger.warning(
                    "Metrics snapshot: dashboard rebuild returned 0 resources — not persisting "
                    "(managed-identity Reader role may still be propagating)"
                )
        except Exception as _de:
            logger.warning("Metrics snapshot: dashboard rebuild failed: %s", _de)

        now = datetime.now(tz=timezone.utc)
        _metrics_snapshot_state.update({
            "last_run": now.isoformat(),
            "last_run_ts": now.timestamp(),
            "resource_count": len(all_metrics),
            "last_error": None,
        })
        elapsed = (now - started).total_seconds()
        logger.info(
            "Metrics snapshot: persisted %d/%d resources in %.1fs",
            len(all_metrics), len(resources), elapsed,
        )
        return {"resource_count": len(all_metrics), "elapsed_seconds": round(elapsed, 1)}
    except Exception as exc:
        logger.error("Metrics snapshot run failed: %s", exc)
        _metrics_snapshot_state["last_error"] = str(exc)
        return {"error": str(exc)}
    finally:
        _metrics_snapshot_state["running"] = False
        cache_svc.release_lock("lock:metrics_snapshot", _lock_token)


async def _metrics_snapshot_loop() -> None:
    """Background loop: refresh utilisation metrics every N hours.

    Performs a startup catch-up run when no fresh metrics exist, then loops on
    METRICS_SNAPSHOT_INTERVAL_HOURS (default 6). Set the interval to 0 to disable.
    """
    await asyncio.sleep(20)  # let the app finish initialising first

    # Startup catch-up: if the persisted metrics are missing/stale, pull now so
    # the Waste Quadrant is populated on the first dashboard open.
    try:
        interval_hours = float(settings_svc.get_value("METRICS_SNAPSHOT_INTERVAL_HOURS", 6.0))
        if interval_hours > 0:
            existing = persistence_svc.load_resource_metrics(ttl_hours=interval_hours)
            if not existing:
                logger.info("Metrics snapshot: no fresh metrics on startup — running catch-up pull")
                await _metrics_snapshot_run()
    except Exception as exc:
        logger.warning("Metrics snapshot startup catch-up failed: %s", exc)

    while True:
        interval_hours = float(settings_svc.get_value("METRICS_SNAPSHOT_INTERVAL_HOURS", 6.0))
        if interval_hours <= 0:
            await asyncio.sleep(300)  # disabled — re-check every 5 min
            continue
        await asyncio.sleep(interval_hours * 3600)
        try:
            await _metrics_snapshot_run()
        except Exception as exc:
            logger.error("Metrics snapshot loop error: %s", exc)


# ── Cost snapshot job ─────────────────────────────────────────────────────────
# Periodically downloads the full cost bundle (tenant total-daily series + FinOps
# KPIs) and persists it to the cost_snapshots table. The dashboard / FinOps read
# the latest snapshot so the home SpendTrend never shows $0.00 even when live Cost
# Management calls are 429-throttled.

_cost_snapshot_state: dict = {
    "last_run": None,
    "last_run_ts": None,
    "running": False,
    "last_error": None,
    "last_summary": None,
}


async def _cost_snapshot_run() -> dict:
    """Build and persist a cost-bundle snapshot."""
    if _cost_snapshot_state.get("running"):
        return {"skipped": "already running"}
    # Cross-replica lock so only one process captures the cost bundle (which hits
    # the heavily-throttled Cost Management API) at a time.
    _lock_token = cache_svc.acquire_lock("lock:cost_snapshot", ttl_seconds=3600)
    if _lock_token is None:
        return {"skipped": "locked by another replica"}
    _cost_snapshot_state["running"] = True
    loop = asyncio.get_event_loop()
    try:
        import services.cost_snapshot_service as cost_snapshot_svc
        summary = await loop.run_in_executor(_pool, cost_snapshot_svc.capture_and_save)
        now = datetime.now(tz=timezone.utc)
        _cost_snapshot_state.update({
            "last_run": now.isoformat(),
            "last_run_ts": now.timestamp(),
            "last_error": None if summary.get("ok") else summary.get("error"),
            "last_summary": summary,
        })
        # Refresh the Redis cost cache-aside slot so other replicas / a restarted
        # process see the new series without a SQL round-trip.
        try:
            _fresh = persistence_svc.load_latest_cost_snapshot()
            if _fresh:
                cache_svc.set_json("snap:cost:latest", _fresh, ttl_seconds=3600)
        except Exception:
            pass
        return summary
    except Exception as exc:
        logger.error("Cost snapshot run failed: %s", exc)
        _cost_snapshot_state["last_error"] = str(exc)
        return {"error": str(exc)}
    finally:
        _cost_snapshot_state["running"] = False
        cache_svc.release_lock("lock:cost_snapshot", _lock_token)


async def _cost_snapshot_loop() -> None:
    """Background loop: refresh the cost snapshot every N hours (default 24).

    Runs a startup catch-up when no snapshot exists yet, then loops on
    COST_SNAPSHOT_INTERVAL_HOURS. Set the interval to 0 to disable.
    """
    await asyncio.sleep(25)  # let the app finish initialising

    try:
        interval_hours = float(settings_svc.get_value("COST_SNAPSHOT_INTERVAL_HOURS", 12.0))
        # Re-capture on startup when there is NO snapshot OR the latest one is all-zero
        # (a transient cost-API 429 / managed-identity permission lag at a previous boot
        # can persist a $0 bundle). Retry with backoff so the home Spend Trend self-heals
        # instead of staying flat until the next 12h cycle.
        def _snap_is_empty(s) -> bool:
            if not s:
                return True
            tot = (sum(float(x or 0) for x in (s.get("total_daily_cm") or []))
                   + sum(float(x or 0) for x in (s.get("total_daily_pm") or [])))
            return tot <= 0
        if interval_hours > 0 and _snap_is_empty(persistence_svc.load_latest_cost_snapshot()):
            for _attempt in range(1, 7):
                logger.info("Cost snapshot: missing/empty on startup — capture attempt %d", _attempt)
                _res = await _cost_snapshot_run()
                if _res and _res.get("ok"):
                    logger.info("Cost snapshot: startup capture succeeded")
                    break
                await asyncio.sleep(min(120 * _attempt, 600))
    except Exception as exc:
        logger.warning("Cost snapshot startup catch-up failed: %s", exc)

    while True:
        interval_hours = float(settings_svc.get_value("COST_SNAPSHOT_INTERVAL_HOURS", 12.0))
        if interval_hours <= 0:
            await asyncio.sleep(300)  # disabled — re-check every 5 min
            continue
        await asyncio.sleep(interval_hours * 3600)
        try:
            await _cost_snapshot_run()
        except Exception as exc:
            logger.error("Cost snapshot loop error: %s", exc)


async def _dashboard_snapshot_catchup() -> None:
    """Startup self-heal for the durable dashboard (resource) snapshot.

    A fresh deployment gets its own managed identity, and the very first dashboard
    scan can run before that identity's Reader role has propagated — so it reads 0
    resources. Unlike the cost / warehouse snapshots (which already self-heal), the
    resource scan had no catch-up, so an empty snapshot would stick and the portal
    would show "0 resources" until a manual refresh or the next scheduled refresh.

    This runs once on startup: if the persisted dashboard snapshot is missing or has
    0 resources, it re-scans with backoff until resources land, then populates the
    in-memory cache + durable snapshot (DB + Redis L2). It is a no-op on a healthy
    restart (a non-empty snapshot already exists) and exits early if another replica
    or a user refresh populates the snapshot first.
    """
    await asyncio.sleep(35)  # let the app initialise and give RBAC a moment to settle
    try:
        snap = persistence_svc.load_latest_dashboard()
        if snap and (snap.get("resources") or []):
            return  # already have a good snapshot — nothing to heal

        for _attempt in range(1, 7):
            # Re-check each round so we stop as soon as another worker / a user
            # refresh has populated the snapshot.
            snap = persistence_svc.load_latest_dashboard()
            if snap and (snap.get("resources") or []):
                logger.info("Dashboard snapshot catch-up: snapshot already populated — done")
                return

            logger.info("Dashboard snapshot: missing/empty on startup — scan attempt %d", _attempt)
            try:
                data = await _build_dashboard(True, None, skip_metrics=True)
                if data.resources:
                    _ts = datetime.now(tz=timezone.utc).timestamp()
                    _cache["data:*"]    = data
                    _cache["data:*:ts"] = _ts
                    _cache["data"]      = data
                    _cache["cached_at"] = _ts
                    try:
                        _dash_json = json.loads(data.model_dump_json())
                        persistence_svc.save_dashboard(_dash_json)
                        if cache_svc.is_enabled():
                            cache_svc.set_json("dash:latest", _dash_json, ttl_seconds=24 * 3600)
                    except Exception as _pe:
                        logger.warning("Dashboard snapshot catch-up persist failed: %s", _pe)
                    logger.info(
                        "Dashboard snapshot: startup scan populated %d resources", len(data.resources)
                    )
                    return
                logger.warning(
                    "Dashboard snapshot: scan returned 0 resources (attempt %d) — "
                    "managed-identity Reader role may still be propagating", _attempt
                )
            except Exception as _se:
                logger.warning("Dashboard snapshot: startup scan attempt %d failed: %s", _attempt, _se)
            await asyncio.sleep(min(120 * _attempt, 600))
    except Exception as exc:
        logger.warning("Dashboard snapshot startup catch-up failed: %s", exc)


async def _finops_warehouse_scheduler() -> None:
    """
    Background scheduler: run the FinOps Warehouse ETL at midnight UTC every night.

    On first startup, triggers an immediate run if the warehouse has no data.
    After that, calculates seconds until next midnight UTC and sleeps until then.
    """
    await asyncio.sleep(30)  # let the app finish starting up

    # First-run: if the warehouse has no cost ROWS, collect now. The first ETL after
    # a fresh deploy can come back empty (Cost Management 429, or the managed identity's
    # Cost Management Reader role not yet propagated at boot), so retry with backoff
    # until real data lands instead of leaving the Cost Warehouse dashboards empty.
    try:
        if _FINOPS_WAREHOUSE_AVAILABLE and not finops_warehouse_svc.has_warehouse_data():
            for _attempt in range(1, 7):
                logger.info("FinOps Warehouse: no data — initial collection attempt %d", _attempt)
                await _run_warehouse_etl_async("startup_initial")
                if finops_warehouse_svc.has_warehouse_data():
                    logger.info("FinOps Warehouse: initial collection populated data")
                    break
                logger.warning("FinOps Warehouse: still empty after attempt %d — retrying", _attempt)
                await asyncio.sleep(min(120 * _attempt, 600))
    except Exception as exc:
        logger.warning("FinOps Warehouse: startup initial collection failed: %s", exc)

    while True:
        sleep_secs = 12 * 3600  # run every 12 hours
        logger.info("FinOps Warehouse scheduler: next run in %.0f seconds (12h interval)", sleep_secs)
        await asyncio.sleep(sleep_secs)
        try:
            logger.info("FinOps Warehouse: 12h ETL starting")
            await _run_warehouse_etl_async("scheduler")
        except Exception as exc:
            logger.error("FinOps Warehouse scheduler error: %s", exc)


# ═══════════════════════════════════════════════════════════════════════════════
# ENTERPRISE FINOPS MODULE — API ROUTES
# All data sourced live from Azure Cost Management APIs.
# Numbers are byte-for-byte identical to Azure Portal Cost Analysis.
# ═══════════════════════════════════════════════════════════════════════════════

try:
    import services.finops_service    as finops_svc
    import services.finops_data_service as finops_data_svc
    import services.budget_service    as budget_svc
    import services.forecast_service  as forecast_finops_svc
    import services.commitment_service as commitment_svc
    import services.tag_analytics_service as tag_analytics_svc
    _FINOPS_AVAILABLE = True
except Exception as _fe:
    logger.warning("FinOps module unavailable: %s", _fe)
    _FINOPS_AVAILABLE = False

try:
    import services.finops_warehouse_service as finops_warehouse_svc
    _FINOPS_WAREHOUSE_AVAILABLE = True
except Exception as _fwe:
    logger.warning("FinOps Warehouse module unavailable: %s", _fwe)
    _FINOPS_WAREHOUSE_AVAILABLE = False

# Track in-progress warehouse ETL to prevent concurrent runs
_warehouse_etl_task: Optional[asyncio.Task] = None

from models.schemas import (
    FinOpsKPI, FinOpsCostExplorerQuery, FinOpsCostExplorerResult,
    FinOpsAllocationReport, FinOpsChargebackReport,
    FinOpsForecastResult, FinOpsCommitmentSummary,
    FinOpsSavingsSummary, FinOpsTagAnalyticsResult, FinOpsTagCostMatrix,
    FinOpsBudgetDefinition, FinOpsBudgetVariance, FinOpsBudgetAlert,
    FinOpsTopMover,
)


def _require_finops():
    if not _FINOPS_AVAILABLE:
        raise HTTPException(status_code=503, detail="FinOps module is not available — check backend logs")


# ── FinOps KPI Summary ────────────────────────────────────────────────────────

# RI-eligible resource types — used to derive RI coverage from the cached dashboard
# so the FinOps summary reports a real number even when the Reservations API 403s.
_RI_ELIGIBLE_TYPES = {
    "microsoft.compute/virtualmachines",
    "microsoft.compute/virtualmachinescalesets",
    "microsoft.sql/servers/databases",
    "microsoft.dbformysql/flexibleservers",
    "microsoft.dbformysql/servers",
    "microsoft.dbforpostgresql/flexibleservers",
    "microsoft.dbforpostgresql/servers",
    "microsoft.dbformariadb/servers",
    "microsoft.cache/redis",
    "microsoft.documentdb/databaseaccounts",
    "microsoft.containerservice/managedclusters",
    "microsoft.app/managedenvironments",
    "microsoft.synapse/workspaces",
}


def _derive_finops_kpi_extras(dash: "DashboardData") -> dict:
    """Complete the FinOps summary fast-path: compute RI coverage/utilization and
    budget utilization (which the cache-based fast path otherwise leaves at 0) plus
    empty-state flags so the UI can distinguish "0%" from "no data configured".

    RI coverage is derived from the cached dashboard (resource-level ri_covered +
    billing-derived reservations), so it works even when the Reservations API 403s.
    Budget KPIs come from the light Consumption Budgets API; absent budgets yield
    has_budgets=False rather than a misleading 0%.
    """
    out = {
        "ri_coverage_pct": 0.0,
        "ri_utilization_pct": 0.0,
        "has_reservations": False,
        "budget_utilization_pct": 0.0,
        "budgets_exceeded": 0,
        "budgets_at_risk": 0,
        "has_budgets": False,
    }
    # ── RI coverage / utilization from dashboard data ─────────────────────
    try:
        resources = dash.resources or []
        active = dash.active_reservations or []
        out["has_reservations"] = bool(active)
        covered_cost = sum(
            (getattr(r, "cost_current_month", 0.0) or 0.0)
            for r in resources if getattr(r, "ri_covered", False)
        )
        eligible_cost = sum(
            (getattr(r, "cost_current_month", 0.0) or 0.0)
            for r in resources
            if (getattr(r, "resource_type", "") or "").lower() in _RI_ELIGIBLE_TYPES
        )
        if eligible_cost > 0:
            out["ri_coverage_pct"] = round(min(100.0, covered_cost / eligible_cost * 100.0), 1)
        utils = [
            u for u in (res.get("utilization_pct") for res in active)
            if isinstance(u, (int, float)) and u > 0
        ]
        if utils:
            out["ri_utilization_pct"] = round(sum(utils) / len(utils), 1)
        elif active:
            # Billing-derived reservations map 1:1 to active resources → fully utilized
            out["ri_utilization_pct"] = 100.0
    except Exception as e:
        logger.debug("FinOps RI KPI derivation failed: %s", e)
    # ── Budget utilization (light Consumption API) ────────────────────────
    try:
        import services.budget_service as _budget_svc
        budgets = _budget_svc.list_budgets()
        out["has_budgets"] = bool(budgets)
        if budgets:
            util_total = 0.0
            for b in budgets:
                try:
                    v = _budget_svc.compute_budget_variance(b.id)
                    if v:
                        util_total += v.utilization_pct
                        if v.status == "exceeded":
                            out["budgets_exceeded"] += 1
                        elif v.status == "at_risk":
                            out["budgets_at_risk"] += 1
                except Exception:
                    pass
            out["budget_utilization_pct"] = round(util_total / len(budgets), 1)
    except Exception as e:
        logger.debug("FinOps budget KPI derivation failed: %s", e)
    return out


@app.get("/api/finops/summary", response_model=FinOpsKPI, tags=["FinOps"])
async def finops_summary():
    """
    Executive FinOps KPIs — MTD spend, MoM delta, budget health, RI coverage,
    tagging compliance.  Uses dashboard cache as fast-path; falls back to live
    Azure Cost Management API when cache is cold.
    """
    _require_finops()
    from datetime import datetime, timezone as _tz
    # ── Fast path: use already-loaded dashboard cache ─────────────────────
    dash: Optional[DashboardData] = _cache.get("data:*") or _cache.get("data")
    if dash is None:
        # Fresh server (snapshot-first instant load): rehydrate the dashboard from
        # the durable snapshot so the fast path works without a live build.
        dash = _ensure_dashboard_in_cache()
    if dash and dash.kpi:
        k = dash.kpi
        resources = dash.resources or []
        tag_compliance = getattr(dash, "tag_compliance_pct", None)
        savings = getattr(dash, "total_potential_savings", None) or k.total_potential_savings
        # Estimate end-of-month from MTD daily rate
        today = datetime.now(tz=_tz.utc).date()
        days_elapsed = today.day
        days_in_month = 30
        eom_forecast = (k.total_cost_current_month / days_elapsed * days_in_month) if days_elapsed > 0 else 0.0
        # Build 30-day trend from cached daily cost arrays. Combine previous +
        # current month so the window is contiguous (the current-month array
        # alone only covers the few elapsed days of the month).
        daily_cm = list(dash.total_daily_cm or [])
        daily_pm = list(dash.total_daily_pm or [])
        # The cached dashboard may have been rehydrated from a resource-scan
        # snapshot captured before cost data was available — backfill the cost
        # arrays from the durable cost snapshot so the trend is never empty.
        if _cost_series_empty(daily_cm, daily_pm):
            _csnap = persistence_svc.load_latest_cost_snapshot()
            if _csnap:
                daily_cm = list(_csnap.get("total_daily_cm") or [])
                daily_pm = list(_csnap.get("total_daily_pm") or [])
        combined = daily_pm + daily_cm
        trend_30d = combined[-30:] if len(combined) >= 30 else ([0.0] * (30 - len(combined)) + combined)
        trend_dates = [str(today - timedelta(days=29 - i)) for i in range(30)]
        # Per-subscription cost breakdown (up to 5, sorted by cost)
        subs_sorted = sorted(dash.subscriptions or [], key=lambda s: -(s.cost_current or 0))
        by_sub = [
            {
                "id":   s.subscription_id,
                "name": s.subscription_name or (s.subscription_id[:8] + "\u2026"),
                "cost": round(s.cost_current or 0.0, 2),
            }
            for s in subs_sorted[:5]
        ]
        extras = _derive_finops_kpi_extras(dash)
        return FinOpsKPI(
            total_spend_mtd=round(k.total_cost_current_month, 2),
            total_spend_last_month=round(k.total_cost_previous_month, 2),
            mom_delta_usd=round(k.mom_cost_delta, 2),
            mom_delta_pct=round(k.mom_cost_delta_pct, 1),
            forecast_eom_usd=round(eom_forecast, 2),
            savings_identified_usd=round(savings, 2),
            budget_utilization_pct=extras["budget_utilization_pct"],
            budgets_exceeded=extras["budgets_exceeded"],
            budgets_at_risk=extras["budgets_at_risk"],
            ri_coverage_pct=extras["ri_coverage_pct"],
            ri_utilization_pct=extras["ri_utilization_pct"],
            has_reservations=extras["has_reservations"],
            has_budgets=extras["has_budgets"],
            tagging_compliance_pct=round(tag_compliance or 0.0, 1),
            total_untagged=int(getattr(dash, "total_untagged", 0) or 0),
            tag_required_keys=list(REQUIRED_TAGS),
            subscription_count=k.subscription_count,
            total_resource_count=len(resources),
            cost_trend_30d=[round(v, 2) for v in trend_30d],
            cost_trend_dates=trend_dates,
            by_subscription=by_sub,
            data_source="dashboard_cache",
            generated_at=datetime.now(tz=_tz.utc).isoformat(),
        )
    # ── Warm-cache path: use pre-fetched FinOps data if available ─────────
    _wc_summary = _finops_warm_cache.get("summary")
    if _wc_summary and _finops_warm_cache.get("summary_ts"):
        logger.debug("finops_summary: serving from pre-warmed cache")
        return _wc_summary
    # ── Slow path: query Azure Cost Management live ────────────────────────
    loop = asyncio.get_event_loop()
    return await loop.run_in_executor(
        _pool,
        lambda: finops_svc.get_finops_kpi(
            subscription_ids=finops_data_svc.get_subscription_ids()
        )
    )


# ── FinOps Dashboard Filtered ─────────────────────────────────────────────────

@app.get("/api/finops/dashboard-data", tags=["FinOps"])
async def finops_dashboard_data(
    subscription_id: str = None,
    resource_group: str = None,
    time_range: str = "last_30d",
    group_by: str = "SubscriptionId",
):
    """
    Filtered dashboard data for interactive FinOps dashboard.
    Returns cost breakdown, trend, and KPIs filtered by subscription/RG/time range.
    Uses the cost explorer engine underneath.
    """
    _require_finops()
    from models.schemas import FinOpsCostExplorerQuery, FinOpsCostFilters

    loop = asyncio.get_event_loop()

    # ── Warehouse-first: paint instantly from stored data so the dashboard never
    #    hangs on throttled live Cost Management.  Falls through to live only when
    #    the warehouse has no data. ────────────────────────────────────────────
    def _wh_dash():
        try:
            from services import finops_warehouse_service as _wh
            if not _wh.has_warehouse_data():
                return None
            subs = [subscription_id] if subscription_id else None
            days = 7 if "7" in time_range else (90 if "90" in time_range else 30)
            d = _wh.get_warehouse_dashboard(subscription_ids=subs, resource_group=resource_group, days=days)
            if not isinstance(d, dict) or d.get("error"):
                return None
            gl = (group_by or "").lower()
            if "subscription" in gl:
                rows = [(r.get("subscription_id", ""), r.get("cost", 0)) for r in d.get("by_subscription", [])]
            else:
                rows = [(r.get("service_family", ""), r.get("cost", 0)) for r in d.get("by_service", [])]
            if not rows:
                return None
            total = sum(c for _, c in rows) or 1.0
            breakdown = [{"label": (n or "(unassigned)"), "cost": round(c, 2), "pct": round(c / total * 100, 1)}
                         for n, c in sorted(rows, key=lambda x: -x[1])[:10]]
            trend = d.get("daily_trend", []) or []
            return {
                "total_cost": round(sum(c for _, c in rows), 2),
                "breakdown": breakdown,
                "trend": trend,
                "group_by": group_by,
                "time_range": time_range,
                "date_from": (trend[0]["date"] if trend else ""),
                "date_to": (trend[-1]["date"] if trend else ""),
                "data_source": "finops_warehouse",
            }
        except Exception as exc:
            logger.warning("warehouse dashboard-data fallback failed: %s", exc)
            return None

    _whd = await loop.run_in_executor(_pool, _wh_dash)
    if _whd is not None and _whd.get("breakdown"):
        return _whd

    # Build filters
    filters = FinOpsCostFilters()
    if subscription_id:
        filters.subscriptions = [subscription_id]
    if resource_group:
        filters.resource_groups = [resource_group]

    # 1. Get breakdown by dimension
    breakdown_query = FinOpsCostExplorerQuery(
        time_range=time_range,
        granularity="None",
        group_by=[group_by],
        filters=filters,
        cost_type="ActualCost",
    )

    # 2. Get daily trend
    trend_query = FinOpsCostExplorerQuery(
        time_range=time_range,
        granularity="Daily",
        group_by=[],
        filters=filters,
        cost_type="ActualCost",
    )

    async def _run(q):
        return await asyncio.wait_for(
            loop.run_in_executor(_pool, lambda: finops_svc.run_cost_explorer(q)),
            timeout=12,
        )

    # Run breakdown + trend concurrently so worst-case latency is one timeout, not two.
    breakdown_result, trend_result = await asyncio.gather(
        _run(breakdown_query), _run(trend_query), return_exceptions=True
    )
    if isinstance(breakdown_result, Exception):
        logger.warning(f"FinOps breakdown query failed/slow: {breakdown_result}")
        breakdown_result = None
    if isinstance(trend_result, Exception):
        logger.warning(f"FinOps trend query failed/slow: {trend_result}")
        trend_result = None

    # Build chart-ready data
    breakdown_items = []
    if breakdown_result and breakdown_result.top_contributors:
        breakdown_items = breakdown_result.top_contributors[:10]
    elif breakdown_result and breakdown_result.data_points:
        for dp in breakdown_result.data_points:
            if dp.label:
                breakdown_items.append({"label": dp.label, "cost": dp.cost_usd, "pct": 0})
        total = sum(i["cost"] for i in breakdown_items) or 1
        for item in breakdown_items:
            item["pct"] = round(item["cost"] / total * 100, 1)
        breakdown_items = sorted(breakdown_items, key=lambda x: -x["cost"])[:10]

    trend_points = []
    if trend_result and trend_result.data_points:
        for dp in sorted(trend_result.data_points, key=lambda x: x.date or ""):
            if dp.date:
                trend_points.append({"date": dp.date, "cost": round(dp.cost_usd, 2)})

    total_cost = breakdown_result.total_usd if breakdown_result else 0.0

    return {
        "total_cost": round(total_cost, 2),
        "breakdown": breakdown_items,
        "trend": trend_points,
        "group_by": group_by,
        "time_range": time_range,
        "date_from": breakdown_result.date_from if breakdown_result else "",
        "date_to": breakdown_result.date_to if breakdown_result else "",
    }


# ── Cost Explorer ──────────────────────────────────────────────────────────────


# ── Subscription List (for filter dropdowns) ──────────────────────────────────

# Cache of accessible subscription display names (id -> name). The scope dropdown
# must list EVERY subscription the identity can switch to, not only the ones in the
# current (possibly single-subscription) scan. Resolving names hits ARM, so cache it.
_sub_names_cache: dict = {"ts": 0.0, "map": {}}


def _resolve_all_subscription_names(ttl_seconds: int = 3600) -> dict:
    """Return {subscription_id: display_name} for every subscription the identity
    can enumerate (best-effort, cached for ttl_seconds). Empty dict on failure."""
    import time as _t
    now = _t.time()
    if _sub_names_cache["map"] and (now - _sub_names_cache["ts"] < ttl_seconds):
        return _sub_names_cache["map"]
    try:
        from azure.mgmt.subscription import SubscriptionClient
        from services.azure_auth import get_credential
        client = SubscriptionClient(get_credential())
        m = {s.subscription_id: (s.display_name or s.subscription_id)
             for s in client.subscriptions.list()
             if getattr(s, "state", "Enabled") in (None, "Enabled")}
        if m:
            _sub_names_cache["map"] = m
            _sub_names_cache["ts"]  = now
        return m
    except Exception as _e:
        logger.debug("subscription-name resolve failed: %s", _e)
        return _sub_names_cache["map"] or {}


@app.get("/api/subscriptions", tags=["FinOps"])
async def list_subscriptions():
    """Subscription list for the scope dropdown.

    Returns EVERY subscription the managed identity can access (so the user can
    switch scope to any of them), enriched with live cost / resource counts from
    the current scan when present. Previously this returned only the *scanned*
    subscriptions, so a single-subscription scan collapsed the dropdown to one
    entry even though the identity could read several.
    """
    dash: Optional[DashboardData] = _cache.get("data:*") or _cache.get("data")
    scanned = {s.subscription_id: s for s in (dash.subscriptions or [])} if dash else {}
    # Authoritative ID set honours config (explicit pinned list vs auto-discover).
    try:
        all_ids = list(finops_data_svc.get_subscription_ids()) if _FINOPS_AVAILABLE else []
    except Exception:
        all_ids = []
    # Always include anything actually scanned, even if config later narrows it.
    for sid in scanned:
        if sid not in all_ids:
            all_ids.append(sid)
    if not all_ids:
        return [
            {"subscription_id": s.subscription_id,
             "subscription_name": s.subscription_name or s.subscription_id,
             "cost_current": round(s.cost_current or 0.0, 2)}
            for s in scanned.values()
        ]
    # Resolve display names only for subscriptions not already enriched by the scan.
    name_map = _resolve_all_subscription_names() if any(sid not in scanned for sid in all_ids) else {}
    out = []
    for sid in all_ids:
        s = scanned.get(sid)
        if s:
            out.append({
                "subscription_id":   sid,
                "subscription_name": s.subscription_name or name_map.get(sid) or sid,
                "cost_current":      round(s.cost_current or 0.0, 2),
            })
        else:
            out.append({
                "subscription_id":   sid,
                "subscription_name": name_map.get(sid, sid),
                "cost_current":      0.0,
            })
    return out


@app.get("/api/management-groups", tags=["FinOps"])
async def list_management_groups():
    """Management-group hierarchy for the scope selector.

    Returns every management group the identity can read — each with the subscription IDs
    nested under it (recursively) — plus the flat subscription list, so the scope dropdown can
    render an "All / Management Groups / Subscriptions" tree and a group selection expands to all
    of its child subscriptions. Uses the Management Groups REST API (needs Management Group
    Reader, granted at the tenant root); the az CLI is avoided because it tries to register a
    resource provider (write). Degrades to subscriptions-only when MG data is unavailable.
    """
    import urllib.request

    dash: Optional[DashboardData] = _cache.get("data:*") or _cache.get("data")
    scanned = {s.subscription_id: s for s in (dash.subscriptions or [])} if dash else {}
    try:
        all_ids = list(finops_data_svc.get_subscription_ids()) if _FINOPS_AVAILABLE else []
    except Exception:
        all_ids = []
    for sid in scanned:
        if sid not in all_ids:
            all_ids.append(sid)
    name_map = _resolve_all_subscription_names() if all_ids else {}

    def _sub_name(sid):
        s = scanned.get(sid)
        return (s.subscription_name if s else None) or name_map.get(sid) or sid

    subs_flat = [{"subscription_id": sid, "subscription_name": _sub_name(sid)} for sid in all_ids]

    def _mg_tree():
        try:
            from services.azure_auth import get_credential
            tok = get_credential().get_token("https://management.azure.com/.default").token

            def _get(path):
                req = urllib.request.Request("https://management.azure.com" + path,
                                             headers={"Authorization": f"Bearer {tok}"})
                with urllib.request.urlopen(req, timeout=30) as r:
                    return json.loads(r.read().decode())

            root_id = (settings_svc.get_value("ENTRA_TENANT_ID", "") or os.environ.get("ENTRA_TENANT_ID", "") or "").strip()
            if not root_id:
                listing = _get("/providers/Microsoft.Management/managementGroups?api-version=2020-05-01")
                vals = listing.get("value", [])
                root_id = vals[0]["name"] if vals else ""
            if not root_id:
                return []
            tree = _get(f"/providers/Microsoft.Management/managementGroups/{root_id}?api-version=2020-05-01&$expand=children&$recurse=true")
            accessible = set(all_ids)
            flat = []

            def walk(node, depth):
                props = node.get("properties") or node
                # Append this MG first (pre-order) so a parent lists above its children; `level`
                # (0 for the tenant root's direct children) drives the UI indentation.
                entry = {
                    "id": node.get("name"),
                    "name": props.get("displayName") or node.get("name"),
                    "level": max(depth - 1, 0),
                    "subscription_ids": [],
                }
                flat.append(entry)
                descendant = set()
                for ch in (props.get("children") or []):
                    ctype = (ch.get("type") or "").lower()
                    if "subscription" in ctype:
                        sid = ch.get("name")
                        if sid:
                            descendant.add(sid)
                    elif "managementgroup" in ctype:
                        descendant |= walk(ch, depth + 1)
                entry["subscription_ids"] = (
                    sorted([s for s in descendant if s in accessible]) if accessible else sorted(descendant)
                )
                return descendant

            walk(tree, 0)
            # Drop only the synthetic tenant root (represented by "All subscriptions"). Keep every
            # other MG so the full hierarchy is visible; MGs with no subscriptions this identity can
            # read are returned with an empty list and shown as non-selectable in the UI.
            return [m for m in flat if m["id"] != root_id]
        except Exception as exc:
            logger.info("Management-group hierarchy unavailable (needs Management Group Reader): %s", exc)
            return []

    loop = asyncio.get_event_loop()
    mgs = await loop.run_in_executor(_pool, _mg_tree)
    return {"available": bool(mgs), "management_groups": mgs, "subscriptions": subs_flat}


# ── Filter Options (for Cost Explorer multi-selects) ─────────────────────────

@app.get("/api/finops/filter-options", tags=["FinOps"])
async def finops_filter_options():
    """
    Returns unique values for every filter dimension from the dashboard cache.
    Instant — no Azure API call.  Used to populate Cost Explorer filter dropdowns.
    """
    dash: Optional[DashboardData] = _cache.get("data:*") or _cache.get("data")
    if not dash:
        return {
            "subscriptions": [], "resource_groups": [], "resource_types": [],
            "regions": [], "service_families": [], "tag_keys": [],
        }
    resources = dash.resources or []
    subs  = sorted({r.subscription_id for r in resources if r.subscription_id})
    rgs   = sorted({r.resource_group  for r in resources if r.resource_group})[:200]
    types = sorted({r.resource_type   for r in resources if r.resource_type})[:100]
    regs  = sorted({getattr(r, "location", "") or "" for r in resources} - {""})[:100]
    # Tag keys from resource tags
    tag_keys: set = set()
    for r in resources:
        tags = getattr(r, "tags", None) or {}
        if isinstance(tags, dict):
            tag_keys.update(tags.keys())
    # Resource counts per dimension for UI badges
    rg_counts = {}
    region_counts = {}
    type_counts = {}
    for r in resources:
        rg = r.resource_group
        if rg:
            rg_counts[rg] = rg_counts.get(rg, 0) + 1
        loc = getattr(r, "location", "") or ""
        if loc:
            region_counts[loc] = region_counts.get(loc, 0) + 1
        rt = r.resource_type
        if rt:
            type_counts[rt] = type_counts.get(rt, 0) + 1
    # Subscription names
    sub_name_map = {s.subscription_id: s.subscription_name or s.subscription_id for s in (dash.subscriptions or [])}
    sub_resource_counts = {}
    for r in resources:
        sid = r.subscription_id
        if sid:
            sub_resource_counts[sid] = sub_resource_counts.get(sid, 0) + 1
    sub_items = [
        {"id": sid, "name": sub_name_map.get(sid, sid), "count": sub_resource_counts.get(sid, 0)}
        for sid in subs
    ]
    return {
        "subscriptions":      sub_items,
        "resource_groups":    [{"value": rg, "label": rg, "count": rg_counts.get(rg, 0)} for rg in rgs],
        "resource_types":     [{"value": t, "label": t.split("/")[-1], "count": type_counts.get(t, 0)} for t in types],
        "regions":            [{"value": r, "label": r, "count": region_counts.get(r, 0)} for r in regs],
        "tag_keys":           sorted(tag_keys)[:50],
        "available_tag_keys": sorted(tag_keys)[:50],
        "service_families":   [],
    }



@app.post("/api/finops/cost-explorer", response_model=FinOpsCostExplorerResult, tags=["FinOps"])
async def finops_cost_explorer(query: FinOpsCostExplorerQuery):
    """
    Self-service cost analysis — identical to Azure Portal Cost Analysis.
    Supports 15-dimension filtering, 11 group-by dimensions, daily/monthly granularity,
    actual & amortized cost, and tag-based grouping.
    """
    _require_finops()
    loop = asyncio.get_event_loop()
    return await loop.run_in_executor(
        _pool,
        lambda: finops_svc.run_cost_explorer(query)
    )


# ── Cost Allocation ────────────────────────────────────────────────────────────

def _warehouse_allocation(dimension: str) -> Optional[FinOpsAllocationReport]:
    """Build a cost-allocation report from the warehouse (Azure SQL) so the tab
    loads instantly from stored data when live Cost Management is slow/throttled."""
    try:
        from models.schemas import FinOpsAllocationItem
        from services import finops_warehouse_service as _wh
        if not _wh.has_warehouse_data():
            return None
        dash = _wh.get_warehouse_dashboard(days=30)
        if not isinstance(dash, dict) or dash.get("error"):
            return None
        dl = dimension.lower()
        if "subscription" in dl:
            rows = [(r.get("subscription_id", ""), r.get("cost", 0)) for r in dash.get("by_subscription", [])]
            label = "Subscription"
        elif "environment" in dl:
            rows = [(r.get("environment", "untagged"), r.get("cost", 0)) for r in dash.get("by_environment", [])]
            label = "Environment"
        else:
            rows = [(r.get("service_family", ""), r.get("cost", 0)) for r in dash.get("by_service", [])]
            label = "Service Family"
        total = sum(c for _, c in rows) or 1.0
        items = [
            FinOpsAllocationItem(dimension_value=(n or "(unassigned)"), cost_usd=round(c, 2), cost_pct=round(c / total * 100, 1))
            for n, c in sorted(rows, key=lambda x: -x[1])
        ]
        return FinOpsAllocationReport(
            dimension=dimension, dimension_label=label, items=items,
            total_usd=round(sum(c for _, c in rows), 2),
            period_label="Last 30 days (warehouse)", data_source="finops_warehouse",
        )
    except Exception as exc:
        logger.warning("warehouse allocation fallback failed: %s", exc)
        return None


def _warehouse_tag_analytics(required_tags: Optional[List[str]] = None) -> Optional[FinOpsTagAnalyticsResult]:
    """Build a tag-analytics report from the warehouse (Azure SQL) so the Tags tab
    paints instantly from stored data instead of waiting on throttled Cost Management.
    Coverage here is cost-weighted (share of cost under a non-empty tag value)."""
    try:
        from models.schemas import FinOpsTagKeyStats
        from services import finops_warehouse_service as _wh
        from services.tag_analytics_service import DEFAULT_REQUIRED_TAGS
        if not _wh.has_warehouse_data():
            return None
        req = required_tags or DEFAULT_REQUIRED_TAGS
        dash = _wh.get_warehouse_dashboard(days=30)
        stats: List[FinOpsTagKeyStats] = []
        worst_coverage = 100.0
        any_data = False
        for tag_key in req:
            rows = _wh.get_tag_breakdown(tag_key, months=1) or []
            if rows:
                any_data = True
            tagged = sum(r.get("cost", 0) for r in rows if (r.get("tag_value") or "untagged").lower() != "untagged")
            untag = sum(r.get("cost", 0) for r in rows if (r.get("tag_value") or "untagged").lower() == "untagged")
            tag_total = tagged + untag
            top = sorted(
                [{"tag_value": r.get("tag_value"), "cost_usd": round(r.get("cost", 0), 2), "resource_count": 0,
                  "pct": round(r.get("cost", 0) / (tag_total or 1) * 100, 1)}
                 for r in rows if (r.get("tag_value") or "untagged").lower() != "untagged"],
                key=lambda x: -x["cost_usd"],
            )[:5]
            cov = round(tagged / (tag_total or 1) * 100, 1) if tag_total else 0.0
            if tag_total:
                worst_coverage = min(worst_coverage, cov)
            stats.append(FinOpsTagKeyStats(
                tag_key=tag_key, covered_resources=0, total_resources=0,
                coverage_pct=cov, total_cost_usd=round(tagged, 2),
                distinct_values=len(top), top_values=top, is_required=True,
            ))
        if not any_data:
            return None
        untagged_cost = 0.0
        if isinstance(dash, dict):
            for r in dash.get("by_environment", []):
                if (r.get("environment") or "").lower() in ("untagged", "", "unassigned", "none", "(unassigned)"):
                    untagged_cost += r.get("cost", 0)
        return FinOpsTagAnalyticsResult(
            tag_keys=stats,
            untagged_cost_usd=round(untagged_cost, 2),
            untagged_resource_count=0,
            compliance_score_pct=round(worst_coverage if any_data else 0.0, 1),
            required_tags=req,
            generated_at=datetime.now(timezone.utc).isoformat(),
            data_source="finops_warehouse",
        )
    except Exception as exc:
        logger.warning("warehouse tag analytics fallback failed: %s", exc)
        return None


def _warehouse_chargeback() -> Optional[FinOpsChargebackReport]:
    """Build a chargeback report (by CostCenter tag) from the warehouse so the
    Chargeback tab paints instantly from stored data."""
    try:
        from models.schemas import FinOpsChargebackEntry
        from services import finops_warehouse_service as _wh
        if not _wh.has_warehouse_data():
            return None
        rows = _wh.get_tag_breakdown("CostCenter", months=1) or []
        allocated = 0.0
        unallocated = 0.0
        entries: List[FinOpsChargebackEntry] = []
        for r in rows:
            cc = (r.get("tag_value") or "untagged")
            cost = round(r.get("cost", 0), 2)
            if cc.lower() in ("untagged", "", "unassigned", "none", "(unassigned)"):
                unallocated += cost
                continue
            allocated += cost
            entries.append(FinOpsChargebackEntry(
                cost_center=cc, allocated_cost_usd=cost,
                resource_count=0, subscription_count=0, coverage_pct=100.0,
            ))
        entries.sort(key=lambda e: -e.allocated_cost_usd)
        # No CostCenter tags at all → show the true "all unallocated" state from the
        # warehouse total so the tab still paints instantly (instead of erroring).
        if not entries:
            dash = _wh.get_warehouse_dashboard(days=30)
            total_cost = 0.0
            if isinstance(dash, dict):
                total_cost = sum(r.get("cost", 0) for r in dash.get("by_subscription", [])) or 0.0
            if total_cost <= 0:
                return None
            unallocated = round(total_cost, 2)
        total = allocated + unallocated
        return FinOpsChargebackReport(
            entries=entries,
            total_allocated_usd=round(allocated, 2),
            total_unallocated_usd=round(unallocated, 2),
            coverage_pct=round(allocated / (total or 1) * 100, 1),
            period_label="Last 30 days (warehouse)",
            data_source="finops_warehouse",
            generated_at=datetime.now(timezone.utc).isoformat(),
        )
    except Exception as exc:
        logger.warning("warehouse chargeback fallback failed: %s", exc)
        return None


def _warehouse_forecast(horizon_days: int = 90) -> Optional[FinOpsForecastResult]:
    """Linear-regression spend forecast from the warehouse daily trend so the
    Forecast tab paints instantly when the live Cost Management forecast (warm
    cache) is cold."""
    try:
        import calendar as _cal
        from datetime import timedelta as _td
        from models.schemas import FinOpsForecastPoint
        from services import finops_warehouse_service as _wh
        if not _wh.has_warehouse_data():
            return None
        dash = _wh.get_warehouse_dashboard(days=60)
        trend = (dash or {}).get("daily_trend", []) if isinstance(dash, dict) else []
        pts = [(str(p["date"])[:10], float(p["cost"])) for p in trend if p.get("date") and p.get("cost") is not None]
        if len(pts) < 5:
            return None
        n = len(pts)
        xs = list(range(n))
        ys = [c for _, c in pts]
        mean_x = sum(xs) / n
        mean_y = sum(ys) / n
        denom = sum((x - mean_x) ** 2 for x in xs) or 1.0
        slope = sum((xs[i] - mean_x) * (ys[i] - mean_y) for i in range(n)) / denom
        intercept = mean_y - slope * mean_x
        std = (sum((y - mean_y) ** 2 for y in ys) / n) ** 0.5
        history = [FinOpsForecastPoint(date=pts[i][0], cost_usd=round(ys[i], 2), is_forecast=False,
                                       source="linear_regression_fallback") for i in range(n)]
        last_date = datetime.strptime(pts[-1][0], "%Y-%m-%d").date()
        horizon = max(7, min(horizon_days, 365))
        forecast: List[FinOpsForecastPoint] = []
        for k in range(1, horizon + 1):
            d = last_date + _td(days=k)
            val = max(0.0, intercept + slope * (n - 1 + k))
            forecast.append(FinOpsForecastPoint(
                date=str(d), cost_usd=round(val, 2),
                confidence_lower=round(max(0.0, val - std), 2),
                confidence_upper=round(val + std, 2),
                is_forecast=True, source="linear_regression_fallback",
            ))
        today = datetime.now(timezone.utc).date()
        eom_day = _cal.monthrange(today.year, today.month)[1]
        days_left = max(0, eom_day - today.day)
        mtd = sum(ys[-min(today.day, n):])
        eom = round(mtd + mean_y * days_left, 2)
        direction = "rising" if slope > mean_y * 0.01 else ("falling" if slope < -mean_y * 0.01 else "stable")
        return FinOpsForecastResult(
            scope_label="All Subscriptions (warehouse)",
            history=history, forecast=forecast,
            horizon_days=horizon, forecast_method="linear_regression_fallback",
            total_forecast_usd=round(sum(p.cost_usd for p in forecast), 2),
            eom_forecast_usd=eom, eoq_forecast_usd=round(mean_y * 90, 2),
            trend_direction=direction,
            mom_trend_pct=round(slope / (mean_y or 1) * 100 * 30, 1),
            confidence_level="low",
            generated_at=datetime.now(timezone.utc).isoformat(),
        )
    except Exception as exc:
        logger.warning("warehouse forecast fallback failed: %s", exc)
        return None


@app.get("/api/finops/allocation", response_model=FinOpsAllocationReport, tags=["FinOps"])
async def finops_allocation(
    dimension: str = "SubscriptionId",
    time_range: str = "mtd",
):
    """
    Cost allocation by any Azure dimension. Serves instantly from the warehouse
    cache (stored database data) so the tab paints immediately and never hangs;
    only falls back to a bounded live Cost Management query when the warehouse
    has no data for the requested dimension.
    """
    _require_finops()
    loop = asyncio.get_event_loop()
    # 1. Warehouse-first: instant paint from stored data (sub-second).
    wh = await loop.run_in_executor(_pool, lambda: _warehouse_allocation(dimension))
    if wh is not None and wh.items:
        return wh
    # 2. No warehouse data → bounded live query, else 503.
    try:
        return await asyncio.wait_for(
            loop.run_in_executor(
                _pool,
                lambda: finops_svc.get_cost_allocation(dimension=dimension, time_range=time_range),
            ),
            timeout=12,
        )
    except Exception as exc:
        logger.warning("finops_allocation live failed/slow: %s", exc)
        raise HTTPException(status_code=503, detail="Cost allocation temporarily unavailable — retry shortly")


# ── Chargeback / Showback ──────────────────────────────────────────────────────

@app.get("/api/finops/chargeback", response_model=FinOpsChargebackReport, tags=["FinOps"])
async def finops_chargeback(time_range: str = "last_month"):
    """
    Chargeback report grouped by CostCenter tag.  Serves instantly from the
    warehouse (stored data) so the tab never hangs; falls back to a bounded
    live Cost Management query only when the warehouse has no CostCenter data.
    """
    _require_finops()
    loop = asyncio.get_event_loop()
    # 1. Warehouse-first: instant paint from stored data.
    wh = await loop.run_in_executor(_pool, _warehouse_chargeback)
    if wh is not None:
        return wh
    # 2. No warehouse data → bounded live query, else 503.
    try:
        return await asyncio.wait_for(
            loop.run_in_executor(_pool, lambda: finops_svc.get_chargeback_report(time_range=time_range)),
            timeout=12,
        )
    except Exception as exc:
        logger.warning("finops_chargeback live failed/slow: %s", exc)
        raise HTTPException(status_code=503, detail="Chargeback temporarily unavailable — retry shortly")


# ── Forecasting ───────────────────────────────────────────────────────────────

@app.get("/api/finops/forecast", response_model=FinOpsForecastResult, tags=["FinOps"])
async def finops_forecast(horizon: int = 90):
    """
    Spend forecast using Azure Cost Management Forecast API (same ML model
    as Azure Portal).  Falls back to linear regression if Forecast API is
    not accessible for the billing account type.
    """
    _require_finops()
    # ── Fast path: serve from warm cache (populated every 30 min) ─────────
    _wc_fc = _finops_warm_cache.get("forecast")
    if _wc_fc and _finops_warm_cache.get("forecast_ts"):
        age = datetime.now(tz=timezone.utc).timestamp() - _finops_warm_cache["forecast_ts"]
        if age < 14400:  # 4 hours
            logger.debug("finops_forecast: serving from warm cache (age %.0fs)", age)
            return _wc_fc
    loop = asyncio.get_event_loop()
    # ── Cold start: paint instantly from a warehouse linear-regression forecast.
    #    The background warmer upgrades this to the Azure ML forecast within 30 min.
    wh = await loop.run_in_executor(_pool, lambda: _warehouse_forecast(horizon))
    if wh is not None and wh.forecast:
        return wh
    try:
        return await asyncio.wait_for(
            loop.run_in_executor(
                _pool,
                lambda: forecast_finops_svc.get_forecast(
                    horizon_days=max(7, min(horizon, 365)),
                    subscription_ids=finops_data_svc.get_subscription_ids(),
                ),
            ),
            timeout=45.0,
        )
    except asyncio.TimeoutError:
        logger.warning("finops_forecast: timed out after 45s — returning empty result")
        raise HTTPException(status_code=503, detail="Forecast timed out — data will be available after background cache refresh")


# ── Commitments — Reservations & Savings Plans ────────────────────────────────

@app.get("/api/finops/commitments", response_model=FinOpsCommitmentSummary, tags=["FinOps"])
async def finops_commitments():
    """
    Live RI & savings plan data from ReservationManagementClient +
    Azure-native buy recommendations from ConsumptionManagementClient.
    Identical to Azure Portal Reservations blade.
    """
    _require_finops()
    # ── Fast path: serve from warm cache ──────────────────────────────────
    _wc_cm = _finops_warm_cache.get("commitments")
    if _wc_cm and _finops_warm_cache.get("commitments_ts"):
        age = datetime.now(tz=timezone.utc).timestamp() - _finops_warm_cache["commitments_ts"]
        if age < 14400:  # 4 hours
            logger.debug("finops_commitments: serving from warm cache (age %.0fs)", age)
            return _wc_cm
    loop = asyncio.get_event_loop()
    return await loop.run_in_executor(_pool, commitment_svc.get_commitment_summary)


# ── Savings Opportunities ──────────────────────────────────────────────────────

@app.get("/api/finops/savings", response_model=FinOpsSavingsSummary, tags=["FinOps"])
async def finops_savings():
    """
    Consolidated savings opportunities: RI purchase (Azure native recs),
    right-sizing, waste cleanup, orphan removal, and licensing.
    """
    _require_finops()
    # ── Fast path: serve from warm cache ──────────────────────────────────
    _wc_sv = _finops_warm_cache.get("savings")
    if _wc_sv and _finops_warm_cache.get("savings_ts"):
        age = datetime.now(tz=timezone.utc).timestamp() - _finops_warm_cache["savings_ts"]
        if age < 14400:  # 4 hours
            logger.debug("finops_savings: serving from warm cache (age %.0fs)", age)
            return _wc_sv
    dash_model = _cache.get("data:*") or _cache.get("data")
    try:
        dash_dict = json.loads(dash_model.model_dump_json()) if dash_model else None
    except Exception:
        dash_dict = None
    loop = asyncio.get_event_loop()
    return await loop.run_in_executor(
        _pool,
        lambda: finops_svc.get_savings_summary(dashboard_cache=dash_dict)
    )


# ── Tag Analytics ──────────────────────────────────────────────────────────────

@app.get("/api/finops/tag-analytics", response_model=FinOpsTagAnalyticsResult, tags=["FinOps"])
async def finops_tag_analytics(time_range: str = "mtd"):
    """
    Tag coverage and cost per tag value.  Serves instantly from the warehouse
    (stored database data) so the tab paints immediately and never hangs; only
    falls back to a bounded live query (Cost Management + Resource Graph) when
    the warehouse has no tag data.
    """
    _require_finops()
    loop = asyncio.get_event_loop()
    # 1. Warehouse-first: instant paint from stored data.
    wh = await loop.run_in_executor(_pool, _warehouse_tag_analytics)
    if wh is not None and wh.tag_keys:
        return wh
    # 2. No warehouse tag data → bounded live query, else 503.
    try:
        return await asyncio.wait_for(
            loop.run_in_executor(_pool, lambda: tag_analytics_svc.get_tag_analytics(time_range=time_range)),
            timeout=12,
        )
    except Exception as exc:
        logger.warning("finops_tag_analytics live failed/slow: %s", exc)
        raise HTTPException(status_code=503, detail="Tag analytics temporarily unavailable — retry shortly")


@app.get("/api/finops/tag-analytics/{tag_key}", response_model=FinOpsTagCostMatrix, tags=["FinOps"])
async def finops_tag_cost_matrix(tag_key: str, time_range: str = "mtd"):
    """
    Cost breakdown by values of a specific tag key.
    Uses Azure Cost Management group_by=[TagKey:{tag_key}] — live data.
    """
    _require_finops()
    loop = asyncio.get_event_loop()
    return await loop.run_in_executor(
        _pool,
        lambda: tag_analytics_svc.get_tag_cost_matrix(tag_key=tag_key, time_range=time_range)
    )


# ── Top Movers ────────────────────────────────────────────────────────────────

@app.get("/api/finops/top-movers", response_model=List[FinOpsTopMover], tags=["FinOps"])
async def finops_top_movers(
    dimension: str = "ResourceGroupName",
    limit: int = 20,
):
    """
    Resources/groups with largest cost change vs. prior 30-day period.
    Live Azure Cost Management data — same as Azure Portal Cost Change analysis.
    """
    _require_finops()
    loop = asyncio.get_event_loop()
    return await loop.run_in_executor(
        _pool,
        lambda: finops_svc.get_top_movers(
            dimension=dimension,
            limit=max(1, min(limit, 100))
        )
    )


# ── Budget Management ─────────────────────────────────────────────────────────

@app.get("/api/finops/budgets", response_model=List[FinOpsBudgetDefinition], tags=["FinOps"])
async def finops_list_budgets(sync_azure: bool = False):
    """
    List all budgets (Azure native + custom).
    Pass ?sync_azure=true to pull latest budgets from Azure Portal first.
    """
    _require_finops()
    if sync_azure:
        loop = asyncio.get_event_loop()
        await loop.run_in_executor(
            _pool,
            lambda: budget_svc.sync_azure_budgets()
        )
    return budget_svc.list_budgets()


class BudgetCreateRequest(BaseModel):
    name:              str
    scope_type:        str   = "all"
    scope_id:          str   = ""
    amount_usd:        float = 0.0
    period:            str   = "Monthly"
    start_date:        str   = ""
    alert_thresholds:  List[float] = [50.0, 75.0, 90.0, 100.0]
    owner_email:       str   = ""
    cost_center:       str   = ""
    tag_filters:       Dict[str, str] = {}


@app.post("/api/finops/budgets", response_model=FinOpsBudgetDefinition, tags=["FinOps"])
async def finops_create_budget(req: BudgetCreateRequest):
    """Create a custom budget (stored locally in SQLite)."""
    _require_finops()
    return budget_svc.create_budget(
        name=req.name,
        scope_type=req.scope_type,
        scope_id=req.scope_id,
        amount_usd=req.amount_usd,
        period=req.period,
        start_date=req.start_date,
        alert_thresholds=req.alert_thresholds,
        owner_email=req.owner_email,
        cost_center=req.cost_center,
        tag_filters=req.tag_filters,
    )


@app.get("/api/finops/budgets/alerts", response_model=List[FinOpsBudgetAlert], tags=["FinOps"])
async def finops_budget_alerts():
    """Return all triggered budget alerts from log."""
    _require_finops()
    return budget_svc.get_budget_alerts()


@app.get("/api/finops/budgets/{budget_id}", response_model=FinOpsBudgetDefinition, tags=["FinOps"])
async def finops_get_budget(budget_id: str):
    """Get a single budget definition."""
    _require_finops()
    b = budget_svc.get_budget(budget_id)
    if not b:
        raise HTTPException(status_code=404, detail="Budget not found")
    return b


class BudgetUpdateRequest(BaseModel):
    name:              Optional[str]         = None
    amount_usd:        Optional[float]       = None
    period:            Optional[str]         = None
    alert_thresholds:  Optional[List[float]] = None
    owner_email:       Optional[str]         = None
    cost_center:       Optional[str]         = None


@app.put("/api/finops/budgets/{budget_id}", response_model=FinOpsBudgetDefinition, tags=["FinOps"])
async def finops_update_budget(budget_id: str, req: BudgetUpdateRequest):
    """Update a budget definition."""
    _require_finops()
    updates = {k: v for k, v in req.model_dump().items() if v is not None}
    b = budget_svc.update_budget(budget_id, **updates)
    if not b:
        raise HTTPException(status_code=404, detail="Budget not found")
    return b


@app.delete("/api/finops/budgets/{budget_id}", status_code=204, tags=["FinOps"])
async def finops_delete_budget(budget_id: str):
    """Delete a custom budget."""
    _require_finops()
    ok = budget_svc.delete_budget(budget_id)
    if not ok:
        raise HTTPException(status_code=404, detail="Budget not found")


@app.get("/api/finops/budgets/{budget_id}/variance", response_model=FinOpsBudgetVariance, tags=["FinOps"])
async def finops_budget_variance(budget_id: str):
    """
    Compute live budget vs. actual spend variance.
    Actual spend fetched directly from Azure Cost Management API.
    """
    _require_finops()
    loop = asyncio.get_event_loop()
    variance = await loop.run_in_executor(
        _pool,
        lambda: budget_svc.compute_budget_variance(budget_id)
    )
    if not variance:
        raise HTTPException(status_code=404, detail="Budget not found")
    return variance


# ── Exports ────────────────────────────────────────────────────────────────────

def _csv(value) -> str:
    """Escape a single value for CSV output (quote if it contains a comma,
    quote, or newline). Keeps export rows well-formed even when service /
    resource names contain commas."""
    s = "" if value is None else str(value)
    if any(ch in s for ch in (",", '"', "\n", "\r")):
        return '"' + s.replace('"', '""') + '"'
    return s


def _field(obj, key, default=None):
    """Read a field from an item that may be a dict OR a pydantic/dataclass
    object — export renderers must not care which shape the service returned."""
    if isinstance(obj, dict):
        return obj.get(key, default)
    return getattr(obj, key, default)


def _num(value, default=0.0) -> float:
    """Coerce a possibly-None / string value to float, never raising."""
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


@app.get("/api/finops/export/csv", tags=["FinOps"])
async def finops_export_csv(
    dimension: str = "SubscriptionId",
    time_range: str = "last_month",
):
    """
    Export cost allocation data as CSV.
    Data sourced live from Azure Cost Management API.
    """
    _require_finops()
    loop = asyncio.get_event_loop()
    try:
        report: Optional[FinOpsAllocationReport] = await asyncio.wait_for(
            loop.run_in_executor(
                _pool,
                lambda: finops_svc.get_cost_allocation(dimension=dimension, time_range=time_range)
            ),
            timeout=35,
        )
    except Exception as exc:
        logger.warning("CSV allocation export fetch failed/slow: %s", exc)
        report = None

    # Warehouse fallback (fast, offline Azure SQL) when live Cost Management is
    # throttled or returns nothing — so the export still carries real numbers.
    wh_rows = None
    if not report or not getattr(report, "items", None):
        try:
            from services import finops_warehouse_service as _wh
            if _wh.has_warehouse_data():
                dash = await asyncio.wait_for(
                    loop.run_in_executor(_pool, lambda: _wh.get_warehouse_dashboard(days=30)),
                    timeout=15,
                )
                if isinstance(dash, dict) and not dash.get("error"):
                    dl = dimension.lower()
                    if "subscription" in dl:
                        wh_rows = [(r.get("subscription_id", ""), r.get("cost", 0)) for r in dash.get("by_subscription", [])]
                    elif "environment" in dl:
                        wh_rows = [(r.get("environment", "untagged"), r.get("cost", 0)) for r in dash.get("by_environment", [])]
                    else:
                        wh_rows = [(r.get("service_family", ""), r.get("cost", 0)) for r in dash.get("by_service", [])]
        except Exception as exc:
            logger.warning("CSV allocation warehouse fallback failed: %s", exc)
            wh_rows = None

    # Assemble the full CSV up-front (not a streaming generator) so any
    # formatting error is caught here and we always return a complete, valid file
    # instead of a half-streamed/blank download.
    rows = ["Dimension,Cost (USD),% of Total,MoM Delta %"]
    try:
        if report and getattr(report, "items", None):
            for item in report.items:
                try:
                    rows.append(f"{_csv(_field(item, 'dimension_value', ''))},{_num(_field(item, 'cost_usd')):.2f},{_num(_field(item, 'cost_pct')):.1f},{_num(_field(item, 'mom_delta_pct')):.1f}")
                except Exception:
                    continue
            _un = _num(getattr(report, "unallocated_usd", 0))
            if _un > 0:
                rows.append(f"(Unallocated),{_un:.2f},{_num(getattr(report, 'unallocated_pct', 0)):.1f},")
            rows.append("")
            rows.append(f"Total,{_num(getattr(report, 'total_usd', 0)):.2f},100.0,")
            rows.append(f"Data Source,{_csv(getattr(report, 'data_source', 'Azure Cost Management'))},,")
        elif wh_rows:
            total = sum(_num(c) for _, c in wh_rows) or 1
            for name, cost in sorted(wh_rows, key=lambda x: -_num(x[1])):
                rows.append(f"{_csv(name)},{_num(cost):.2f},{(_num(cost) / total * 100):.1f},")
            rows.append("")
            rows.append(f"Total,{total:.2f},100.0,")
            rows.append("Data Source,FinOps Warehouse (Azure SQL),,")
        else:
            rows.append("(No data available - Azure Cost Management may be throttled. Please retry shortly.),,,")
    except Exception as exc:
        logger.warning("CSV allocation render failed: %s", exc)
        rows = ["Dimension,Cost (USD),% of Total,MoM Delta %", "(Export error - please retry),,,"]

    csv_text = "\r\n".join(rows) + "\r\n"
    fname = f"azure-cost-{dimension.lower()}-{time_range}.csv"
    from fastapi.responses import Response
    return Response(content=csv_text, media_type="text/csv",
                    headers={"Content-Disposition": f'attachment; filename="{fname}"'})


@app.get("/api/finops/export/chargeback-csv", tags=["FinOps"])
async def finops_export_chargeback_csv(time_range: str = "last_month"):
    """
    Export chargeback report as CSV for finance team.
    Data sourced from Azure Cost Management API grouped by CostCenter tag.
    """
    _require_finops()
    loop = asyncio.get_event_loop()
    try:
        report: Optional[FinOpsChargebackReport] = await asyncio.wait_for(
            loop.run_in_executor(
                _pool,
                lambda: finops_svc.get_chargeback_report(time_range=time_range)
            ),
            timeout=35,
        )
    except Exception as exc:
        logger.warning("CSV chargeback export fetch failed/slow: %s", exc)
        report = None

    rows = ["Cost Center,Allocated Cost (USD),Coverage %,Period"]
    try:
        if report and getattr(report, "entries", None):
            for entry in report.entries:
                rows.append(f"{_csv(entry.cost_center)},{(entry.allocated_cost_usd or 0):.2f},{(entry.coverage_pct or 0):.1f},{_csv(getattr(report, 'period_label', ''))}")
            rows.append("")
            rows.append(f"Total Allocated,{(report.total_allocated_usd or 0):.2f},,")
            rows.append(f"Unallocated,{(report.total_unallocated_usd or 0):.2f},,")
            rows.append(f"Coverage %,{(report.coverage_pct or 0):.1f},,")
            rows.append(f"Data Source,{_csv(getattr(report, 'data_source', 'Azure Cost Management'))},,")
        else:
            rows.append("(No data available - Azure Cost Management may be throttled. Please retry shortly.),,,")
    except Exception as exc:
        logger.warning("CSV chargeback render failed: %s", exc)
        rows = ["Cost Center,Allocated Cost (USD),Coverage %,Period", "(Export error - please retry),,,"]

    csv_text = "\r\n".join(rows) + "\r\n"
    from fastapi.responses import Response
    return Response(content=csv_text, media_type="text/csv",
                    headers={"Content-Disposition": f'attachment; filename="azure-chargeback-{time_range}.csv"'})


# ── FOCUS (FinOps Open Cost and Usage Specification) ──────────────────────────

@app.get("/api/finops/focus", tags=["FinOps"])
async def finops_focus(days: int = 30, limit: int = 2000, subscription_id: Optional[str] = None):
    """
    Return cost data normalized to the FOCUS 1.2 specification (the open FinOps
    standard used by the Microsoft FinOps toolkit / FinOps Hubs).
    Sourced from the warehouse — fast, offline, no Cost Management throttling.
    """
    _require_finops()
    from services import focus_service as _focus
    loop = asyncio.get_event_loop()
    subs = [subscription_id] if subscription_id else None
    try:
        records = await asyncio.wait_for(
            loop.run_in_executor(_pool, lambda: _focus.get_focus_records(subscription_ids=subs, days=days, limit=limit)),
            timeout=25,
        )
    except Exception as exc:
        logger.warning("FOCUS query failed/slow: %s", exc)
        records = []
    summary = _focus.focus_summary(records)
    return {"summary": summary, "records": records}


@app.get("/api/finops/export/focus-csv", tags=["FinOps"])
async def finops_export_focus_csv(days: int = 30, limit: int = 50000, subscription_id: Optional[str] = None):
    """Download a FOCUS 1.2-conformant CSV (portable, toolkit-compatible)."""
    _require_finops()
    from services import focus_service as _focus
    loop = asyncio.get_event_loop()
    subs = [subscription_id] if subscription_id else None
    try:
        records = await asyncio.wait_for(
            loop.run_in_executor(_pool, lambda: _focus.get_focus_records(subscription_ids=subs, days=days, limit=limit)),
            timeout=30,
        )
    except Exception as exc:
        logger.warning("FOCUS CSV export failed/slow: %s", exc)
        records = []

    cols = _focus.FOCUS_COLUMNS
    lines = [",".join(cols)]
    try:
        for rec in records:
            lines.append(",".join(_csv(rec.get(c, "")) for c in cols))
        if not records:
            lines.append("(No warehouse data yet - the FinOps warehouse populates within ~12h of first run),,,")
    except Exception as exc:
        logger.warning("FOCUS CSV render failed: %s", exc)
    csv_text = "\r\n".join(lines) + "\r\n"
    from fastapi.responses import Response
    fname = f"azure-focus-1.2-{datetime.now(timezone.utc).strftime('%Y-%m-%d')}.csv"
    return Response(content=csv_text, media_type="text/csv",
                    headers={"Content-Disposition": f'attachment; filename="{fname}"'})


# ── FinOps AI insights (per-view analysis + recommendations) ──────────────────

class FinOpsAIRequest(BaseModel):
    view: str
    data: Dict[str, Any] = {}
    filters: Optional[Dict[str, Any]] = None
    force_refresh: bool = False
    scope: Optional[str] = None


@app.post("/api/finops/ai/insights", tags=["FinOps"])
async def finops_ai_insights(req: FinOpsAIRequest):
    """
    AI cost analysis + recommendations for a FinOps view. Cached in Redis per
    (view + data fingerprint); pass force_refresh=true for a fresh generation.
    """
    _require_finops()
    from services import finops_ai_service as _ai
    loop = asyncio.get_event_loop()
    try:
        result = await asyncio.wait_for(
            loop.run_in_executor(
                _ai_pool,
                lambda: _ai.get_finops_insights(req.view, req.data, req.filters, req.force_refresh, req.scope),
            ),
            timeout=60,
        )
    except Exception as exc:
        logger.warning("FinOps AI insights failed/slow: %s", exc)
        raise HTTPException(status_code=503, detail=f"AI analysis unavailable: {exc}")
    return result


# ── On-prem software governance (allow/block/required + AI intelligence) ──────

class SoftwarePolicyModel(BaseModel):
    required: List[str] = []
    blocked: List[str] = []
    allowed: List[str] = []


class SoftwareEvalRequest(BaseModel):
    inventory: List[Any] = []
    policy: Optional[Dict[str, Any]] = None


class SoftwareAIRequest(BaseModel):
    software: List[Any] = []
    force_refresh: bool = False
    from_fleet: bool = False


@app.get("/api/onprem/software-policy", tags=["OnPrem"])
async def onprem_get_software_policy():
    from services import software_policy_service as sp
    return sp.get_policy()


@app.put("/api/onprem/software-policy", tags=["OnPrem"])
async def onprem_set_software_policy(policy: SoftwarePolicyModel):
    from services import software_policy_service as sp
    return sp.set_policy(policy.model_dump())


@app.post("/api/onprem/software-policy/evaluate", tags=["OnPrem"])
async def onprem_eval_software(req: SoftwareEvalRequest):
    from services import software_policy_service as sp
    return sp.evaluate(req.inventory, req.policy)


@app.post("/api/onprem/software-ai", tags=["OnPrem"])
async def onprem_software_ai(req: SoftwareAIRequest):
    from services import software_policy_service as sp
    loop = asyncio.get_event_loop()
    try:
        return await asyncio.wait_for(
            loop.run_in_executor(_ai_pool, lambda: sp.get_software_intelligence(req.software, req.force_refresh, req.from_fleet)),
            timeout=90,
        )
    except Exception as exc:
        logger.warning("Software AI failed/slow: %s", exc)
        raise HTTPException(status_code=503, detail=f"AI analysis unavailable: {exc}")


@app.get("/api/onprem/software/fleet", tags=["OnPrem"])
async def onprem_software_fleet():
    """Fleet-wide software governance over the real on-prem/Arc inventory:
    per-server compliance, a categorized software catalog (risk/license/EOL),
    and estate roll-up signals."""
    from services import software_policy_service as sp
    return await asyncio.to_thread(sp.get_fleet_governance)


@app.get("/api/onprem/software/templates", tags=["OnPrem"])
async def onprem_software_templates():
    """One-click policy templates (security baseline, no remote access, Azure-ready, license audit)."""
    from services import software_policy_service as sp
    return {"templates": sp.get_policy_templates()}


# ── Scopes (subscriptions + management groups) ────────────────────────────────

@app.get("/api/finops/scopes", tags=["FinOps"])
async def finops_scopes():
    """Available scopes: subscriptions from cache + management groups if accessible."""
    dash: Optional[DashboardData] = _cache.get("data:*") or _cache.get("data")
    subs = []
    if dash:
        subs = [
            {"id": s.subscription_id, "name": s.subscription_name or s.subscription_id, "type": "subscription"}
            for s in (dash.subscriptions or [])
        ]
    if not subs and _FINOPS_AVAILABLE:
        subs = [{"id": sid, "name": sid, "type": "subscription"}
                for sid in finops_data_svc.get_subscription_ids()]
    # Try management groups (requires ManagementGroupsAPI access)
    mgs: list = []
    try:
        from azure.mgmt.managementgroups import ManagementGroupsAPI
        from services.azure_auth import get_credential
        mg_client = ManagementGroupsAPI(get_credential())
        for mg in mg_client.management_groups.list():
            mgs.append({"id": mg.name, "name": mg.display_name or mg.name, "type": "management_group"})
    except Exception:
        pass
    return {"subscriptions": subs, "management_groups": mgs}


# ── Azure Advisor Cost Recommendations ────────────────────────────────────────

@app.get("/api/finops/advisor-cost", tags=["FinOps"])
async def finops_advisor_cost():
    """Azure Advisor cost recommendations extracted from the dashboard resource cache."""
    dash: Optional[DashboardData] = _cache.get("data:*") or _cache.get("data")
    if not dash:
        return {"items": [], "total_savings_monthly": 0.0, "item_count": 0,
                "data_source": "cache_unavailable"}
    sub_name_map = {s.subscription_id: s.subscription_name or s.subscription_id
                    for s in (dash.subscriptions or [])}
    items = []
    for r in (dash.resources or []):
        for rec in (r.advisor_recommendations or []):
            if rec.category.lower() == "cost":
                items.append({
                    "resource_id":               r.resource_id,
                    "resource_name":             r.resource_name,
                    "resource_type":             r.resource_type,
                    "resource_group":            r.resource_group,
                    "subscription_id":           r.subscription_id,
                    "subscription_name":         sub_name_map.get(r.subscription_id, (r.subscription_id or "")[:8] + "…"),
                    "impact":                    rec.impact,
                    "recommendation":            rec.short_description,
                    "potential_savings_monthly": round(rec.potential_savings, 2),
                })
    items.sort(key=lambda x: -x["potential_savings_monthly"])
    return {
        "items": items,
        "total_savings_monthly": round(sum(i["potential_savings_monthly"] for i in items), 2),
        "item_count": len(items),
        "data_source": "dashboard_cache",
    }


# ── Resource Optimization (oversized + underutilized) ────────────────────────

@app.get("/api/finops/resource-optimization", tags=["FinOps"])
async def finops_resource_optimization():
    """Oversized and underutilized resources from dashboard cache for FinOps view."""
    dash: Optional[DashboardData] = _cache.get("data:*") or _cache.get("data")
    if not dash:
        return {"oversized": [], "underutilized": [], "orphaned": [],
                "oversized_count": 0, "underutilized_count": 0, "orphaned_count": 0,
                "total_oversized_savings": 0.0}
    sub_name_map = {s.subscription_id: s.subscription_name or s.subscription_id
                    for s in (dash.subscriptions or [])}
    oversized, underutilized = [], []
    for r in (dash.resources or []):
        sub_name = sub_name_map.get(r.subscription_id or "", (r.subscription_id or "")[:8])
        base = {
            "resource_id":        r.resource_id,
            "resource_name":      r.resource_name,
            "resource_type":      r.resource_type,
            "resource_group":     r.resource_group,
            "subscription_name":  sub_name,
            "cost_current_month": round(r.cost_current_month, 2),
            "location":           getattr(r, "location", ""),
        }
        if r.rightsize_sku and r.cost_current_month > 0:
            oversized.append({
                **base,
                "current_sku":    r.sku,
                "recommended_sku": r.rightsize_sku,
                "savings_pct":    round(r.rightsize_savings_pct, 1),
                "monthly_savings": round(r.cost_current_month * r.rightsize_savings_pct / 100, 2),
                "avg_cpu_pct":    r.avg_cpu_pct,
                "avg_memory_pct": r.avg_memory_pct,
                "score":          r.final_score,
            })
        elif (r.final_score is not None and r.final_score < 25
              and r.cost_current_month >= 5 and not r.rightsize_sku):
            underutilized.append({
                **base,
                "utilization_score": round(r.final_score, 1),
                "avg_cpu_pct":    r.avg_cpu_pct,
                "avg_memory_pct": r.avg_memory_pct,
                "score_label":    r.score_label.value if r.score_label else "Unknown",
                "days_since_active": r.days_since_active,
                "recommendation": r.recommendation or "Review for rightsizing or decommission",
            })
    oversized.sort(key=lambda x: -x["monthly_savings"])
    underutilized.sort(key=lambda x: -x["cost_current_month"])
    orphaned = [
        {
            "resource_name":    getattr(o, "resource_name", ""),
            "resource_type":    getattr(o, "resource_type", ""),
            "resource_group":   getattr(o, "resource_group", ""),
            "cost_current_month": round(getattr(o, "cost_usd_month", 0), 2),
            "orphan_reason":    getattr(o, "orphan_reason", ""),
            "estimated_monthly_savings": round(getattr(o, "estimated_monthly_savings", 0), 2),
        }
        for o in sorted(dash.orphans or [], key=lambda x: -getattr(x, "cost_usd_month", 0))[:50]
    ]
    return {
        "oversized":            oversized[:50],
        "underutilized":        underutilized[:100],
        "orphaned":             orphaned,
        "oversized_count":      len(oversized),
        "underutilized_count":  len(underutilized),
        "orphaned_count":       len(orphaned),
        "total_oversized_savings": round(sum(x["monthly_savings"] for x in oversized), 2),
    }


# ── XLSX Export (cost-explorer query → formatted XLSX) ───────────────────────

@app.post("/api/finops/export/xlsx", tags=["FinOps"])
async def finops_export_xlsx(query: FinOpsCostExplorerQuery):
    """Export a cost-explorer query result as a formatted XLSX file."""
    _require_finops()
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl is not installed on the server")

    from io import BytesIO
    loop = asyncio.get_event_loop()
    try:
        result = await asyncio.wait_for(
            loop.run_in_executor(_pool, lambda: finops_svc.run_cost_explorer(query)),
            timeout=35,
        )
    except Exception as exc:
        logger.warning("XLSX cost-explorer export fetch failed/slow: %s", exc)
        result = None

    # Never hard-fail the download: if the live query is throttled/empty, return a
    # small valid workbook explaining why, so the user always gets a file.
    if result is None or not getattr(result, "data_points", None) and not getattr(result, "top_contributors", None):
        wb_err = openpyxl.Workbook()
        ws = wb_err.active; ws.title = "Export"
        ws["A1"] = "Azure cost export unavailable"
        ws["A2"] = ("Azure Cost Management returned no data or was throttled (HTTP 429). "
                    "Please retry in a few moments.")
        ws["A1"].font = Font(bold=True, size=12)
        buf = BytesIO(); wb_err.save(buf); buf.seek(0)
        return StreamingResponse(buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": 'attachment; filename="azure-cost-export-unavailable.xlsx"'})

    HFILL = PatternFill("solid", fgColor="1E3A5F")
    HFONT = Font(bold=True, color="FFFFFF", size=11)
    SFONT = Font(bold=True, color="2D8CFF", size=9)

    def _header(ws, headers):
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.fill = HFILL; c.font = HFONT
            c.alignment = Alignment(horizontal="center")

    def _autowidth(ws):
        for col in ws.columns:
            w = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 4, 50)

    wb = openpyxl.Workbook()

    # Sheet 1: Cost by Date
    ws1 = wb.active; ws1.title = "Cost by Date"
    labels = sorted({k for dp in (result.data_points or []) for k in (dp.breakdown or {}).keys()}) or ["Total"]
    _header(ws1, ["Date"] + labels)
    for ri, dp in enumerate(result.data_points or [], 2):
        ws1.cell(row=ri, column=1, value=dp.date)
        if dp.breakdown:
            for ci, lbl in enumerate(labels, 2):
                ws1.cell(row=ri, column=ci, value=round(dp.breakdown.get(lbl, 0), 2))
        else:
            ws1.cell(row=ri, column=2, value=round(dp.cost_usd, 2))

    # Sheet 2: Top Contributors
    ws2 = wb.create_sheet("Top Contributors")
    _header(ws2, ["Dimension", "Cost (USD)", "% of Total"])
    for ri, tc in enumerate(result.top_contributors or [], 2):
        ws2.cell(row=ri, column=1, value=tc.get("label", ""))
        ws2.cell(row=ri, column=2, value=round(tc.get("cost", 0), 2))
        ws2.cell(row=ri, column=3, value=tc.get("pct", 0))

    # Sheet 3: Summary
    ws3 = wb.create_sheet("Summary")
    _header(ws3, ["Metric", "Value"])
    for ri, (k, v) in enumerate([
        ("Total Cost (USD)", f"${result.total_usd:,.2f}"),
        ("Date From",        result.date_from),
        ("Date To",          result.date_to),
        ("Granularity",      result.granularity),
        ("Cost Type",        result.cost_type),
        ("Currency",         result.currency or "USD"),
        ("Data Source",      result.data_source or "Azure Cost Management"),
        ("Generated",        datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")),
    ], 2):
        ws3.cell(row=ri, column=1, value=k)
        ws3.cell(row=ri, column=2, value=v)

    for ws in [ws1, ws2, ws3]:
        _autowidth(ws)

    buf = BytesIO(); wb.save(buf); buf.seek(0)
    fname = f"azure-cost-{result.date_from}-to-{result.date_to}.xlsx"
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'})


# ── Full FinOps Report XLSX ────────────────────────────────────────────────────

@app.get("/api/finops/report/xlsx", tags=["FinOps"])
async def finops_report_xlsx():
    """Download the complete FinOps report as a multi-sheet XLSX."""
    _require_finops()
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl is not installed on the server")

    from io import BytesIO
    loop = asyncio.get_event_loop()
    sub_ids = finops_data_svc.get_subscription_ids()

    # ── Use pre-warmed FinOps cache when available (instant) ──────────────
    _fc = _finops_warm_cache
    _fc_kpi   = _fc.get("summary")   if _fc.get("summary_ts")   else None
    _fc_alloc = _fc.get("alloc_ServiceFamily") if _fc.get("alloc_ServiceFamily_ts") else None
    _fc_cb    = _fc.get("chargeback") if _fc.get("chargeback_ts") else None

    # Fetch data: use pre-warmed cache if available, else fetch live (concurrently)
    futures = {}
    if not _fc_kpi:
        futures["kpi"] = loop.run_in_executor(_pool, lambda: finops_svc.get_finops_kpi(sub_ids))
    if not _fc_alloc:
        futures["alloc"] = loop.run_in_executor(_pool, lambda: finops_svc.get_cost_allocation("ServiceFamily", "last_30d", sub_ids))
    if not _fc_cb:
        futures["cb"] = loop.run_in_executor(_pool, lambda: finops_svc.get_chargeback_report("last_30d", sub_ids))
    futures["savings"] = loop.run_in_executor(_pool, lambda: finops_svc.get_savings_summary(_cache.get("data:*") or _cache.get("data")))

    results = {}
    for key, fut in futures.items():
        try:
            results[key] = await asyncio.wait_for(fut, timeout=40)
        except Exception as e:
            logger.warning("Report XLSX: %s fetch failed/slow: %s", key, e)
            results[key] = e

    kpi        = _fc_kpi   or results.get("kpi")
    alloc      = _fc_alloc or results.get("alloc")
    chargeback = _fc_cb    or results.get("cb")
    savings    = results.get("savings")

    HFILL = PatternFill("solid", fgColor="1E3A5F")
    HFONT = Font(bold=True, color="FFFFFF", size=11)
    BFILL = PatternFill("solid", fgColor="0F1C2E")
    TITLE = Font(bold=True, color="2D8CFF", size=14)

    def _header(ws, headers):
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.fill = HFILL; c.font = HFONT
            c.alignment = Alignment(horizontal="center")

    def _autowidth(ws):
        for col in ws.columns:
            w = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 4, 50)

    wb = openpyxl.Workbook()

    # ── Sheet 1: Executive Summary ────────────────────────────────────────────
    ws_kpi = wb.active; ws_kpi.title = "Executive Summary"
    ws_kpi["A1"] = "Azure FinOps Report"; ws_kpi["A1"].font = TITLE
    ws_kpi["A2"] = f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"
    ws_kpi["A2"].font = Font(color="64748B", italic=True, size=9)
    _header(ws_kpi, ["KPI", "Value", "Status"])
    if not isinstance(kpi, Exception):
        kpi_rows = [
            ("Total Spend (MTD)",      f"${kpi.total_spend_mtd:,.2f}",        ""),
            ("Last Month Spend",       f"${kpi.total_spend_last_month:,.2f}", ""),
            ("MoM Delta",              f"{kpi.mom_delta_pct:+.1f}%",           "▲ Over" if kpi.mom_delta_pct > 5 else "▼ Under" if kpi.mom_delta_pct < -5 else "Stable"),
            ("EOM Forecast",           f"${kpi.forecast_eom_usd:,.2f}",       ""),
            ("Budget Utilization",     f"{kpi.budget_utilization_pct:.1f}%",  "⚠ At Risk" if kpi.budget_utilization_pct >= 80 else "OK"),
            ("Budgets Exceeded",       str(kpi.budgets_exceeded),             "⚠" if kpi.budgets_exceeded > 0 else "✓"),
            ("Savings Identified",     f"${kpi.savings_identified_usd:,.2f}", ""),
            ("RI Coverage",            f"{kpi.ri_coverage_pct:.1f}%",         "Low" if kpi.ri_coverage_pct < 60 else "Good"),
            ("RI Utilization",         f"{kpi.ri_utilization_pct:.1f}%",      "Low" if kpi.ri_utilization_pct < 70 else "Good"),
            ("Tag Compliance",         f"{kpi.tagging_compliance_pct:.1f}%",  "Poor" if kpi.tagging_compliance_pct < 70 else "Good"),
            ("Anomalies Detected",     str(kpi.anomaly_count),                "⚠" if kpi.anomaly_count > 0 else "✓"),
        ]
        for ri, (k, v, s) in enumerate(kpi_rows, 4):
            ws_kpi.cell(row=ri, column=1, value=k).font = Font(bold=True)
            ws_kpi.cell(row=ri, column=2, value=v)
            ws_kpi.cell(row=ri, column=3, value=s)

    # ── Sheet 2: Cost by Service ──────────────────────────────────────────────
    ws_alloc = wb.create_sheet("Cost by Service")
    _header(ws_alloc, ["Service Family", "Cost (USD)", "% of Total", "MoM Delta %"])
    if not isinstance(alloc, Exception):
        for ri, item in enumerate(alloc.items or [], 2):
            ws_alloc.cell(row=ri, column=1, value=item.dimension_value)
            ws_alloc.cell(row=ri, column=2, value=round(item.cost_usd, 2))
            ws_alloc.cell(row=ri, column=3, value=item.cost_pct)
            ws_alloc.cell(row=ri, column=4, value=getattr(item, "mom_delta_pct", None))
        total_row = len(alloc.items or []) + 2
        ws_alloc.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
        ws_alloc.cell(row=total_row, column=2, value=round(alloc.total_usd, 2)).font = Font(bold=True)

    # ── Sheet 3: Chargeback ───────────────────────────────────────────────────
    ws_cb = wb.create_sheet("Chargeback")
    _header(ws_cb, ["Cost Center", "Allocated Cost (USD)", "Coverage %", "Resource Count", "Subscription Count"])
    if not isinstance(chargeback, Exception):
        for ri, entry in enumerate(chargeback.entries or [], 2):
            ws_cb.cell(row=ri, column=1, value=entry.cost_center)
            ws_cb.cell(row=ri, column=2, value=round(entry.allocated_cost_usd, 2))
            ws_cb.cell(row=ri, column=3, value=round(entry.coverage_pct, 1))
            ws_cb.cell(row=ri, column=4, value=entry.resource_count)
            ws_cb.cell(row=ri, column=5, value=getattr(entry, "subscription_count", 0))

    # ── Sheet 4: Savings Opportunities ───────────────────────────────────────
    ws_sav = wb.create_sheet("Savings Opportunities")
    _header(ws_sav, ["Resource", "Category", "Action", "Monthly Savings (USD)", "Effort", "Confidence"])
    if not isinstance(savings, Exception) and savings:
        for ri, opp in enumerate(savings.opportunities or [], 2):
            ws_sav.cell(row=ri, column=1, value=opp.resource_name)
            ws_sav.cell(row=ri, column=2, value=opp.category)
            ws_sav.cell(row=ri, column=3, value=opp.action)
            ws_sav.cell(row=ri, column=4, value=round(opp.potential_savings_usd, 2))
            ws_sav.cell(row=ri, column=5, value=opp.effort)
            ws_sav.cell(row=ri, column=6, value=opp.confidence)

    # ── Sheet 5: Azure Advisor Recommendations ────────────────────────────────
    ws_adv = wb.create_sheet("Advisor Recommendations")
    _header(ws_adv, ["Resource", "Type", "Resource Group", "Impact", "Recommendation", "Monthly Savings (USD)"])
    dash: Optional[DashboardData] = _cache.get("data:*") or _cache.get("data")
    ri = 2
    if dash:
        for r in (dash.resources or []):
            for rec in (r.advisor_recommendations or []):
                if rec.category.lower() == "cost":
                    ws_adv.cell(row=ri, column=1, value=r.resource_name)
                    ws_adv.cell(row=ri, column=2, value=r.resource_type)
                    ws_adv.cell(row=ri, column=3, value=r.resource_group)
                    ws_adv.cell(row=ri, column=4, value=rec.impact)
                    ws_adv.cell(row=ri, column=5, value=rec.short_description)
                    ws_adv.cell(row=ri, column=6, value=round(rec.potential_savings, 2))
                    ri += 1

    # ── Sheet 6: Resource Optimization ───────────────────────────────────────
    ws_opt = wb.create_sheet("Resource Optimization")
    _header(ws_opt, ["Resource", "Type", "Current SKU", "Recommended SKU", "CPU %", "Monthly Savings (USD)", "Score"])
    if dash:
        ri = 2
        for r in (dash.resources or []):
            if r.rightsize_sku and r.cost_current_month > 0:
                ws_opt.cell(row=ri, column=1, value=r.resource_name)
                ws_opt.cell(row=ri, column=2, value=r.resource_type)
                ws_opt.cell(row=ri, column=3, value=r.sku)
                ws_opt.cell(row=ri, column=4, value=r.rightsize_sku)
                ws_opt.cell(row=ri, column=5, value=r.avg_cpu_pct)
                ws_opt.cell(row=ri, column=6, value=round(r.cost_current_month * r.rightsize_savings_pct / 100, 2))
                ws_opt.cell(row=ri, column=7, value=round(r.final_score, 1))
                ri += 1

    for ws in wb.worksheets:
        _autowidth(ws)

    buf = BytesIO(); wb.save(buf); buf.seek(0)
    fname = f"finops-report-{datetime.now(timezone.utc).strftime('%Y-%m-%d')}.xlsx"
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'})


# ── FinOps Cache Management ────────────────────────────────────────────────────

@app.post("/api/finops/cache/warmup", tags=["FinOps"])
async def finops_trigger_warmup():
    """Manually trigger FinOps warm-cache population (runs in background)."""
    _require_finops()
    asyncio.ensure_future(_finops_cache_warmup())
    return {"status": "warmup_started", "message": "Cache warmup triggered — data will be available within ~2 minutes"}


@app.post("/api/finops/cache/clear", tags=["FinOps"])
async def finops_clear_cache():
    """Force-clear the FinOps query cache so next requests fetch fresh Azure data."""
    _require_finops()
    removed = finops_data_svc.clear_finops_cache()
    _finops_warm_cache.clear()
    return {"cleared": removed, "message": f"Cleared {removed} cached FinOps query entries + warm cache"}


@app.get("/api/finops/cache/status", tags=["FinOps"])
async def finops_cache_status():
    """Returns the status of the pre-warmed FinOps cache."""
    ts = _finops_warm_cache.get("last_warmup_ts")
    age = (datetime.now(tz=timezone.utc).timestamp() - ts) if ts else None
    return {
        "last_warmup": _finops_warm_cache.get("last_warmup"),
        "age_seconds": round(age, 0) if age else None,
        "stale": age > _FINOPS_WARMUP_INTERVAL_SECONDS if age else True,
        "warmup_interval_seconds": _FINOPS_WARMUP_INTERVAL_SECONDS,
        "cached_keys": [k for k in _finops_warm_cache if not k.endswith("_ts")],
    }


# ═══════════════════════════════════════════════════════════════════════════════
# FINOPS WAREHOUSE — OFFLINE-FIRST NIGHTLY DATA WAREHOUSE ROUTES
# All dashboard reads come from Azure SQL — zero live Azure API calls.
# Data collected every night at midnight UTC via background ETL job.
# ═══════════════════════════════════════════════════════════════════════════════

def _require_warehouse():
    if not _FINOPS_WAREHOUSE_AVAILABLE:
        raise HTTPException(status_code=503, detail="FinOps Warehouse module unavailable — check logs")


async def _run_warehouse_etl_async(triggered_by: str = "manual"):
    """Run the ETL job in a thread pool and clear the task reference when done."""
    global _warehouse_etl_task
    loop = asyncio.get_event_loop()
    try:
        sub_ids = finops_data_svc.get_subscription_ids() if _FINOPS_AVAILABLE else []
        await loop.run_in_executor(_pool, lambda: finops_warehouse_svc.run_full_etl(sub_ids, triggered_by))
        # Fresh data landed — invalidate the warehouse read caches so the next UI
        # request reflects it immediately instead of a stale cached aggregate.
        _fw_bump_version()
    except Exception as e:
        logger.error("Warehouse ETL async wrapper error: %s", e)
    finally:
        _warehouse_etl_task = None


# ── FinOps Warehouse read cache (Redis L2) ────────────────────────────────────
# The warehouse SQL tables are refreshed by the ETL only every ~12h, yet these
# read endpoints run heavy aggregation queries against Azure SQL on EVERY UI
# request. Cache each result in Redis (cache-aside; graceful no-op without Redis)
# so repeat opens skip the SQL round-trip. Keys are namespaced by an ETL version
# token that is bumped when a collection completes, so fresh data is served right
# after an ETL; the TTL is only a backstop.
_FW_CACHE_TTL = 900  # seconds (15 min)


def _fw_cache_version() -> str:
    try:
        v = cache_svc.get_json("fw:ver")
        return str(v) if v else "0"
    except Exception:
        return "0"


def _fw_bump_version() -> None:
    try:
        cache_svc.set_json("fw:ver", int(_time_mod.time()))
    except Exception:
        pass


def _fw_cache_key(name: str, *parts) -> str:
    safe = ":".join("" if p is None else str(p) for p in parts)
    return f"fw:{name}:{_fw_cache_version()}:{safe}"


@app.get("/api/finops/warehouse/status", tags=["FinOps Warehouse"])
async def warehouse_status():
    """Return metadata about the most recent ETL run and data freshness."""
    _require_warehouse()
    status = finops_warehouse_svc.get_last_run_status()
    status["etl_running"] = finops_warehouse_svc.is_etl_running()
    status["has_data"] = finops_warehouse_svc.has_warehouse_data()
    return status


@app.post("/api/finops/warehouse/trigger", tags=["FinOps Warehouse"])
async def warehouse_trigger():
    """Manually trigger the nightly ETL data collection job."""
    global _warehouse_etl_task
    _require_warehouse()
    if finops_warehouse_svc.is_etl_running():
        return {"ok": False, "message": "ETL already running — check /api/finops/warehouse/status"}
    _warehouse_etl_task = asyncio.create_task(_run_warehouse_etl_async("manual"))
    return {"ok": True, "message": "ETL collection started — data will be available in a few minutes"}


@app.get("/api/finops/warehouse/dashboard", tags=["FinOps Warehouse"])
async def warehouse_dashboard(
    subscription_id: Optional[str] = None,
    resource_group: Optional[str] = None,
    days: int = 30,
):
    """
    Return all KPIs, charts, and anomalies for the FinOps Warehouse dashboard.
    All data is served from Azure SQL — no live Azure API calls.
    """
    _require_warehouse()
    sub_ids = [subscription_id] if subscription_id else None
    loop = asyncio.get_event_loop()
    _ck = _fw_cache_key("dashboard", subscription_id, resource_group, days)
    data = cache_svc.get_json(_ck)
    if data is None:
        data = await loop.run_in_executor(
            _pool,
            lambda: finops_warehouse_svc.get_warehouse_dashboard(sub_ids, resource_group, days),
        )
        cache_svc.set_json(_ck, data, ttl_seconds=_FW_CACHE_TTL)
    run_status = finops_warehouse_svc.get_last_run_status()
    data["data_freshness"] = {
        "completed_at": run_status.get("completed_at"),
        "data_age_hours": run_status.get("data_age_hours"),
        "status": run_status.get("status"),
        "etl_running": finops_warehouse_svc.is_etl_running(),
    }
    return data


@app.get("/api/finops/warehouse/resources", tags=["FinOps Warehouse"])
async def warehouse_resources(
    subscription_id: Optional[str] = None,
    resource_group: Optional[str] = None,
    resource_type: Optional[str] = None,
    service_family: Optional[str] = None,
    days: int = 30,
    page: int = 1,
    page_size: int = 50,
    sort_by: str = "cost",
    sort_dir: str = "desc",
):
    """Paginated, filterable resource-level cost table from the warehouse."""
    _require_warehouse()
    sub_ids = [subscription_id] if subscription_id else None
    loop = asyncio.get_event_loop()
    _ck = _fw_cache_key("resources", subscription_id, resource_group, resource_type,
                        service_family, days, page, page_size, sort_by, sort_dir)
    _hit = cache_svc.get_json(_ck)
    if _hit is not None:
        return _hit
    result = await loop.run_in_executor(
        _pool,
        lambda: finops_warehouse_svc.get_resource_costs(
            sub_ids, resource_group, resource_type, service_family,
            days, page, page_size, sort_by, sort_dir,
        ),
    )
    cache_svc.set_json(_ck, result, ttl_seconds=_FW_CACHE_TTL)
    return result


@app.get("/api/finops/warehouse/anomalies", tags=["FinOps Warehouse"])
async def warehouse_anomalies(
    severity: Optional[str] = None,
    status: str = "open",
    limit: int = 50,
):
    """Return detected cost anomalies (spikes vs 7-day rolling average)."""
    _require_warehouse()
    loop = asyncio.get_event_loop()
    _ck = _fw_cache_key("anomalies", severity, status, limit)
    _hit = cache_svc.get_json(_ck)
    if _hit is not None:
        return _hit
    result = await loop.run_in_executor(
        _pool,
        lambda: finops_warehouse_svc.get_anomalies(severity, status, limit),
    )
    cache_svc.set_json(_ck, result, ttl_seconds=_FW_CACHE_TTL)
    return result


@app.get("/api/finops/warehouse/by-service", tags=["FinOps Warehouse"])
async def warehouse_by_service(
    subscription_id: Optional[str] = None,
    months: int = 6,
):
    """Monthly cost breakdown by service family for the last N months."""
    _require_warehouse()
    sub_ids = [subscription_id] if subscription_id else None
    loop = asyncio.get_event_loop()
    _ck = _fw_cache_key("by-service", subscription_id, months)
    _hit = cache_svc.get_json(_ck)
    if _hit is not None:
        return _hit
    result = await loop.run_in_executor(
        _pool,
        lambda: finops_warehouse_svc.get_service_breakdown(sub_ids, months),
    )
    cache_svc.set_json(_ck, result, ttl_seconds=_FW_CACHE_TTL)
    return result


@app.get("/api/finops/warehouse/by-tag", tags=["FinOps Warehouse"])
async def warehouse_by_tag(
    tag_key: str = "Environment",
    subscription_id: Optional[str] = None,
    months: int = 3,
):
    """Monthly cost breakdown by tag value for a specific tag key."""
    _require_warehouse()
    sub_ids = [subscription_id] if subscription_id else None
    loop = asyncio.get_event_loop()
    _ck = _fw_cache_key("by-tag", tag_key, subscription_id, months)
    _hit = cache_svc.get_json(_ck)
    if _hit is not None:
        return _hit
    result = await loop.run_in_executor(
        _pool,
        lambda: finops_warehouse_svc.get_tag_breakdown(tag_key, sub_ids, months),
    )
    cache_svc.set_json(_ck, result, ttl_seconds=_FW_CACHE_TTL)
    return result


@app.get("/api/database/info", tags=["System"])
async def database_info():
    """Return current database provider configuration."""
    from services.database import get_db_info
    return get_db_info()


@app.post("/api/metrics-snapshot/refresh", tags=["System"])
async def metrics_snapshot_refresh():
    """Trigger an immediate background utilisation-metrics snapshot run."""
    if _metrics_snapshot_state.get("running"):
        return {"ok": True, "message": "Metrics snapshot already running", "running": True}
    asyncio.create_task(_metrics_snapshot_run())
    return {"ok": True, "message": "Metrics snapshot started"}


@app.get("/api/metrics-snapshot/status", tags=["System"])
async def metrics_snapshot_status():
    """Return the status/age of the last utilisation-metrics snapshot."""
    ts = _metrics_snapshot_state.get("last_run_ts")
    age_hours = ((datetime.now(tz=timezone.utc).timestamp() - ts) / 3600.0) if ts else None
    interval_hours = float(settings_svc.get_value("METRICS_SNAPSHOT_INTERVAL_HOURS", 6.0))
    return {
        "last_run": _metrics_snapshot_state.get("last_run"),
        "age_hours": round(age_hours, 2) if age_hours is not None else None,
        "resource_count": _metrics_snapshot_state.get("resource_count", 0),
        "running": _metrics_snapshot_state.get("running", False),
        "last_error": _metrics_snapshot_state.get("last_error"),
        "interval_hours": interval_hours,
        "stale": (age_hours is None) or (age_hours > interval_hours if interval_hours > 0 else False),
    }


@app.post("/api/finops/cost-snapshot/refresh", tags=["FinOps"])
async def cost_snapshot_refresh():
    """Trigger an immediate background cost-bundle snapshot capture."""
    if _cost_snapshot_state.get("running"):
        return {"ok": True, "message": "Cost snapshot already running", "running": True}
    asyncio.create_task(_cost_snapshot_run())
    return {"ok": True, "message": "Cost snapshot started"}


@app.get("/api/finops/cost-snapshot/status", tags=["FinOps"])
async def cost_snapshot_status():
    """Return the status/age of the last cost-bundle snapshot."""
    ts = _cost_snapshot_state.get("last_run_ts")
    age_hours = ((datetime.now(tz=timezone.utc).timestamp() - ts) / 3600.0) if ts else None
    interval_hours = float(settings_svc.get_value("COST_SNAPSHOT_INTERVAL_HOURS", 12.0))
    # Also surface the captured_at of the persisted snapshot (may predate this process).
    captured_at = None
    try:
        snap = persistence_svc.load_latest_cost_snapshot()
        if snap:
            captured_at = snap.get("captured_at")
    except Exception:
        pass
    return {
        "last_run": _cost_snapshot_state.get("last_run"),
        "captured_at": captured_at,
        "age_hours": round(age_hours, 2) if age_hours is not None else None,
        "running": _cost_snapshot_state.get("running", False),
        "last_error": _cost_snapshot_state.get("last_error"),
        "last_summary": _cost_snapshot_state.get("last_summary"),
        "interval_hours": interval_hours,
    }


@app.on_event("startup")
async def start_auto_refresh_scheduler() -> None:
    global _auto_refresh_task
    # ── Initialize database abstraction layer ────────────────────────────────
    try:
        from services.database import init_db, get_db_info
        init_db()
        db_info = get_db_info()
        logger.info("Database initialized: provider=%s", db_info["provider"])
    except Exception as e:
        logger.error("Database initialization failed: %s", e)

    # ── Warm the in-process dashboard cache from Redis (if available) ─────────
    # When a shared Redis L2 is configured, a freshly-started process (or a new
    # replica) hydrates the last rebuilt dashboard instantly — Waste Quadrant
    # and cost trend populated — instead of rebuilding under 429 throttling.
    try:
        if cache_svc.is_enabled():
            _warm = cache_svc.get_json("dash:latest")
            if _warm:
                from models.schemas import DashboardData as _DD
                _wm = _DD(**{k: v for k, v in _warm.items() if k in _DD.model_fields})
                _wts = datetime.now(tz=timezone.utc).timestamp()
                _cache["data:*"] = _wm
                _cache["data:*:ts"] = _wts
                _cache["data"] = _wm
                _cache["cached_at"] = _wts
                logger.info("Startup: hydrated dashboard from Redis L2 cache (warm open)")
    except Exception as _rhe:
        logger.warning("Startup: Redis dashboard hydrate skipped: %s", _rhe)
    # ── Live-data model (Phase 4): bulk-scan restore removed ─────────────────
    # The dashboard now loads on demand, per panel, directly from live Azure
    # APIs. No persisted dashboard scan is downloaded or restored at startup,
    # so the portal opens instantly instead of blocking for minutes.
    try:
        persisted = None
        if persisted:
            from models.schemas import DashboardData
            restored = DashboardData(**{k: v for k, v in persisted.items() if k in DashboardData.model_fields})
            # Don't carry over old cost warnings from a previous scan — they are stale
            # and confusing on startup. A new scan will re-evaluate and set them if needed.
            restored.cost_data_warning = None
            # Use the scan's original last_refreshed as the cache timestamp.
            # If we used now() here, /api/cache/status would report a timestamp
            # that is hours newer than data.last_refreshed, causing the frontend
            # cacheStatus effect to think the cache was just updated and
            # immediately trigger an unnecessary re-scan on every page refresh.
            _original_ts: float | None = None
            _lr_str = persisted.get("last_refreshed")
            if _lr_str:
                try:
                    _original_ts = datetime.fromisoformat(
                        _lr_str.replace("Z", "+00:00")
                    ).timestamp()
                except Exception:
                    pass
            cache_ts = _original_ts if _original_ts else datetime.now(tz=timezone.utc).timestamp()

            # ── Recompute strategic features if absent in the persisted scan ─
            if not restored.waf_scorecard or not restored.security_gaps or not restored.modernization_opportunities:
                logger.info("Startup: recomputing strategic features for cached scan…")
                try:
                    from services.security_service      import identify_security_gaps
                    from services.modernization_service import detect_modernization_opportunities
                    from services.waf_service           import compute_waf_scorecard
                    from models.schemas                 import KPIData, OrphanResource, RightSizeOpportunity
                    _sec_gaps = restored.security_gaps or identify_security_gaps(restored.resources)
                    _mod_opps = restored.modernization_opportunities or detect_modernization_opportunities(restored.resources)
                    if not restored.waf_scorecard:
                        restored.waf_scorecard = compute_waf_scorecard(
                            restored.resources,
                            restored.kpi,
                            restored.orphans,
                            restored.rightsize_opportunities,
                            _sec_gaps,
                        )
                    if not restored.security_gaps:
                        restored.security_gaps = _sec_gaps
                    if not restored.modernization_opportunities:
                        restored.modernization_opportunities = _mod_opps
                    logger.info(
                        "Startup: strategic features recomputed — %d security gaps, %d modernization opps",
                        len(restored.security_gaps), len(restored.modernization_opportunities),
                    )
                except Exception as _fe:
                    logger.warning("Startup: could not recompute strategic features: %s", _fe)

            # ── Recompute innovation / maturity / licensing if absent ──────
            if not restored.innovation_gaps or not restored.cloud_maturity or not restored.licensing_opportunities:
                logger.info("Startup: recomputing innovation/maturity/licensing features…")
                try:
                    from services.innovation_service import detect_innovation_gaps
                    from services.maturity_service   import compute_cloud_maturity
                    from services.licensing_service  import detect_licensing_opportunities
                    if not restored.innovation_gaps:
                        _inv_gaps, _svc_scores = detect_innovation_gaps(restored.resources)
                        restored.innovation_gaps         = _inv_gaps
                        restored.service_adoption_scores = _svc_scores
                    if not restored.cloud_maturity:
                        restored.cloud_maturity = compute_cloud_maturity(
                            restored.resources, restored.security_gaps, restored.waf_scorecard
                        )
                    if not restored.licensing_opportunities:
                        restored.licensing_opportunities = detect_licensing_opportunities(restored.resources)
                    logger.info(
                        "Startup: innovation/maturity/licensing recomputed — %d gaps, %d licensing opps",
                        len(restored.innovation_gaps), len(restored.licensing_opportunities),
                    )
                except Exception as _fe2:
                    logger.warning("Startup: could not recompute innovation/maturity/licensing: %s", _fe2)

            # ── Recompute ACR opportunities if absent ─────────────────────
            if not restored.acr_opportunities:
                try:
                    from services.acr_service import analyze_acr_opportunities
                    restored.acr_opportunities = analyze_acr_opportunities(restored.resources)
                    logger.info(
                        "Startup: ACR opportunities recomputed — %d gaps, $%.0f/mo potential",
                        restored.acr_opportunities.total_gaps,
                        restored.acr_opportunities.estimated_total_monthly_acr,
                    )
                except Exception as _ae:
                    logger.warning("Startup: could not recompute ACR opportunities: %s", _ae)

            # ── Recompute backup coverage if absent ───────────────────────
            if not restored.backup_coverage:
                try:
                    from services.backup_service import analyze_backup_coverage
                    restored.backup_coverage = analyze_backup_coverage(restored.resources)
                    logger.info(
                        "Startup: backup coverage recomputed — %d gaps (%d critical, %d high)",
                        restored.backup_coverage.total_gaps,
                        restored.backup_coverage.critical_gaps,
                        restored.backup_coverage.high_gaps,
                    )
                except Exception as _be:
                    logger.warning("Startup: could not recompute backup coverage: %s", _be)

            _cache["data:*"] = restored
            _cache["data:*:ts"] = cache_ts
            _cache["data"] = restored
            _cache["cached_at"] = cache_ts
            logger.info(
                "Startup: restored persisted scan into cache (%d resources)",
                len(persisted.get("resources", [])),
            )

            # ── Build dependency graph from cached resources (offline) ─────
            try:
                _raw_resources = [r.model_dump() for r in restored.resources]
                _dep_graph = build_dependency_graph(
                    _raw_resources,
                    vm_attachments=None,       # not persisted; will rediscover from properties
                    use_resource_graph=False,   # offline startup — no API calls
                )
                _cache["dependency_graph"] = _dep_graph
                logger.info(
                    "Startup: dependency graph built — %d nodes, %d edges, %d clusters, %d SPOF",
                    _dep_graph.node_count, _dep_graph.edge_count,
                    _dep_graph.cluster_count, len(_dep_graph.spof),
                )
            except Exception as _dge:
                logger.warning("Startup: could not build dependency graph: %s", _dge)

            # If narrative is missing, regenerate it in the background
            if not restored.ai_narrative and get_active_provider() != "none":
                asyncio.create_task(_regenerate_ai_narrative())
    except Exception as _se:
        logger.warning("Startup: could not restore persisted scan: %s", _se)

    # ── Live-data model (Phase 4): auto-refresh + FinOps warmup loops removed ─
    # No background bulk scans run. Panels fetch live data on demand.

    # ── Auto-resume interrupted APEX workflows ────────────────────────────────
    # If the server restarted while an APEX workflow was running, the background
    # thread died but the DB still shows status='running' or 'stale'. Detect and resume.
    try:
        assessment_svc = get_assessment_service()
        # First mark any long-running as stale
        stale = assessment_svc.detect_and_recover_stale_workflows()
        # Then resume ALL workflows that are in running OR stale state
        _conn = assessment_svc._get_connection()
        _rows = _conn.execute(
            "SELECT workflow_id, assessment_id FROM assessment_apex_workflow WHERE status IN ('running', 'stale')"
        ).fetchall()
        _conn.close()
        for wf_id, assessment_id in _rows:
            try:
                assessment_svc.resume_apex_workflow(assessment_id)
                logger.info("Startup: auto-resumed APEX workflow %s for assessment %s", wf_id, assessment_id)
            except Exception as _re:
                logger.warning("Startup: could not resume APEX workflow %s: %s", wf_id, _re)
    except Exception as _awe:
        logger.warning("Startup: APEX workflow auto-resume check failed: %s", _awe)

    # ── Start on-prem discovery engine if configured ──────────────────────────
    try:
        from services.onprem_engine import auto_start_if_configured
        await auto_start_if_configured()
    except Exception as _ope:
        logger.warning("Startup: on-prem engine auto-start skipped: %s", _ope)

    # ── Start on-prem scheduled-monitoring loop ───────────────────────────────
    # Periodically re-scans added servers per the saved schedule (daily / every
    # N hours). Manual "Scan now" also flows through the same engine.
    try:
        from services.onprem_scheduler import start_scheduler
        start_scheduler()
        logger.info("Startup: on-prem scheduled-monitoring loop scheduled")
    except Exception as _opse:
        logger.warning("Startup: on-prem scheduler start skipped: %s", _opse)

    # ── Start utilisation-metrics snapshot background job ─────────────────────
    # Keeps the resource_metrics store fresh so the dashboard fast-open path can
    # hydrate utilisation (Waste Quadrant / scoring) without a live Monitor pull.
    try:
        asyncio.create_task(_metrics_snapshot_loop())
        logger.info("Startup: utilisation-metrics snapshot loop scheduled")
    except Exception as _mse:
        logger.warning("Startup: could not start metrics snapshot loop: %s", _mse)

    # ── Start cost-bundle snapshot background job ─────────────────────────────
    # Periodically persists the tenant total-daily cost series + FinOps KPIs so
    # the home SpendTrend / FinOps views render from the DB instead of live
    # 429-throttled Cost Management calls.
    try:
        asyncio.create_task(_cost_snapshot_loop())
        logger.info("Startup: cost-bundle snapshot loop scheduled")
    except Exception as _cse:
        logger.warning("Startup: could not start cost snapshot loop: %s", _cse)

    # ── Dashboard resource-scan self-heal (fresh-deploy RBAC race) ────────────
    # If the durable dashboard snapshot is missing or empty (a first scan that ran
    # before the managed identity's Reader role propagated persisted 0 resources),
    # re-scan with backoff until resources land so the portal self-populates instead
    # of showing "0 resources" until a manual refresh. No-op on a healthy restart.
    try:
        asyncio.create_task(_dashboard_snapshot_catchup())
        logger.info("Startup: dashboard snapshot catch-up scheduled")
    except Exception as _dse:
        logger.warning("Startup: could not start dashboard snapshot catch-up: %s", _dse)

    # ── Start FinOps warm-cache loop ──────────────────────────────────────────
    # Pre-fetches forecast, savings, commitments into memory every 30 minutes
    # so FinOps Overview page loads in <3s instead of 5+ minutes.
    try:
        if _FINOPS_AVAILABLE:
            asyncio.create_task(_finops_cache_warmup_loop())
            logger.info("Startup: FinOps warm-cache loop scheduled")
    except Exception as _fwse:
        logger.warning("Startup: could not start FinOps warm-cache loop: %s", _fwse)

    # ── Start FinOps Warehouse nightly ETL scheduler ──────────────────────────
    # Downloads all Azure cost data into Azure SQL every night at midnight UTC.
    # On first startup, triggers an immediate run if the warehouse is empty.
    try:
        if _FINOPS_WAREHOUSE_AVAILABLE:
            asyncio.create_task(_finops_warehouse_scheduler())
            logger.info("Startup: FinOps Warehouse ETL scheduler registered")
        else:
            logger.warning("Startup: FinOps Warehouse scheduler skipped — module unavailable")
    except Exception as _wse:
        logger.warning("Startup: could not start FinOps Warehouse scheduler: %s", _wse)

    # ── Run FinOps Warehouse migration on startup ─────────────────────────────
    try:
        if _FINOPS_WAREHOUSE_AVAILABLE:
            from migrations.five_finops_warehouse import run_migration as _wh_migrate
            _wh_migrate()
    except Exception:
        try:
            import importlib.util, sys as _sys, os as _os
            _mig_path = _os.path.join(_os.path.dirname(__file__), "migrations", "005_finops_warehouse.py")
            _spec = importlib.util.spec_from_file_location("wh_migration", _mig_path)
            _mod = importlib.util.module_from_spec(_spec)
            _spec.loader.exec_module(_mod)
            _mod.run_migration()
            logger.info("Startup: FinOps Warehouse schema migration applied")
        except Exception as _mig_err:
            logger.warning("Startup: FinOps Warehouse migration skipped: %s", _mig_err)



# ── AI narrative refresh endpoint ─────────────────────────────────────────────

@app.post("/api/ai/refresh-narrative")
async def refresh_ai_narrative():
    """Regenerate AI narrative for the current cached data without a full re-scan."""
    if get_active_provider() == "none":
        raise HTTPException(status_code=400, detail="No AI provider configured.")
    cached = _cache.get("data:*") or _cache.get("data")
    if not cached:
        raise HTTPException(status_code=404, detail="No cached data to analyse.")
    asyncio.create_task(_regenerate_ai_narrative())
    return {"ok": True, "message": "AI narrative regeneration started."}


@app.get("/api/resource/{resource_id:path}/metrics")
async def get_single_resource_metrics(resource_id: str):
    """On-demand live metrics for a single resource (loaded when expanded).

    This is the live replacement for the old bulk metrics pull: instead of
    pulling 30-day Monitor metrics for every resource at open time, we fetch
    them for one resource only when the user expands it.
    """
    data: Optional[DashboardData] = _cache.get("data:*") or _cache.get("data")
    rid_lower = resource_id.lower()
    resource = None
    if data:
        resource = next((r for r in data.resources if r.resource_id.lower() == rid_lower), None)

    rtype = resource.resource_type if resource else ""
    sub_id = getattr(resource, "subscription_id", "") if resource else ""

    loop = asyncio.get_event_loop()
    try:
        res = await loop.run_in_executor(
            None, partial(get_resource_metrics, resource_id, rtype, sub_id or "")
        )
    except Exception as exc:
        logger.warning("On-demand metrics fetch failed for %s: %s", resource_id, exc)
        raise HTTPException(status_code=502, detail=f"Could not fetch metrics: {exc}")

    payload = {
        "resource_id": resource_id,
        "primary_utilization": res.primary_utilization,
        "peak_utilization": res.peak_utilization,
        "cpu": res.cpu,
        "memory": res.memory,
        "disk": res.disk,
        "network": res.network,
        "has_any_activity": res.has_any_activity,
        "raw": res.raw,
        "raw_absolute": res.raw_absolute,
    }

    # Cache the freshly pulled metric so a later full scan / reload can reuse it.
    try:
        from types import SimpleNamespace
        persistence_svc.save_resource_metrics({rid_lower: SimpleNamespace(**{
            k: v for k, v in payload.items() if k != "resource_id"
        })})
    except Exception as _se:
        logger.debug("Could not persist on-demand metric for %s: %s", resource_id, _se)

    # Patch the cached resource so the panel reflects the live utilization.
    if resource is not None and res.primary_utilization is not None:
        try:
            resource.utilization_pct = res.primary_utilization
        except Exception:
            pass

    return payload


@app.post("/api/resource/{resource_id:path}/ai-analyze")
async def analyze_single_resource(resource_id: str):
    """
    Run on-demand AI analysis for a single resource from the cache.
    Updates the cached resource in-place and returns the verdict.
    Useful when a scan was saved before AI was configured.
    """
    if get_active_provider() == "none":
        raise HTTPException(status_code=400, detail="No AI provider configured — set an API key in Settings.")

    data: Optional[DashboardData] = _cache.get("data:*") or _cache.get("data")
    if not data:
        raise HTTPException(status_code=404, detail="No cached scan data available.")

    # Normalise the incoming resource_id for comparison
    rid_lower = resource_id.lower()
    resource  = next((r for r in data.resources if r.resource_id.lower() == rid_lower), None)
    if not resource:
        raise HTTPException(status_code=404, detail="Resource not found in current scan cache.")

    if resource.is_infrastructure:
        return {"ok": False, "reason": "Infrastructure resources are not sent for AI review."}

    # Build the minimal dict that get_ai_verdicts expects
    resource_dict = {
        "resource_id":              resource.resource_id,
        "resource_name":            resource.resource_name,
        "resource_type":            resource.resource_type,
        "resource_group":           resource.resource_group,
        "final_score":              resource.final_score,
        "score_label":              resource.score_label.value if resource.score_label else "Unknown",
        "cost_current_month":       resource.cost_current_month,
        "primary_utilization_pct":  resource.primary_utilization_pct,
        "data_confidence":          resource.data_confidence or "medium",
        "is_orphan":                resource.is_orphan,
        "orphan_reason":            resource.orphan_reason or "",
        "has_any_activity":         resource.has_any_activity,
        "days_since_active":        resource.days_since_active or 0,
        "activity_log_count":       resource.activity_log_count or 0,
        "advisor_recommendations":  [
            {"category": a.category, "impact": a.impact, "short_description": a.short_description or ""}
            for a in (resource.advisor_recommendations or [])
        ],
        "is_infrastructure":        resource.is_infrastructure,
    }

    loop     = asyncio.get_event_loop()
    executor = ThreadPoolExecutor(max_workers=1)
    try:
        verdicts = await loop.run_in_executor(executor, partial(get_ai_verdicts, [resource_dict]))
    except Exception as exc:
        logger.exception("On-demand AI analysis failed for %s", resource_id)
        raise HTTPException(status_code=500, detail=f"AI analysis failed: {exc}")

    if not verdicts:
        raise HTTPException(status_code=500, detail="AI returned no verdict.")

    v = verdicts[0]
    if v.error:
        raise HTTPException(status_code=500, detail=f"AI error: {v.error}")

    # Patch the cached resource list in-place
    for i, r in enumerate(data.resources):
        if r.resource_id.lower() == rid_lower:
            data.resources[i] = r.model_copy(update={
                "ai_confidence":      v.confidence,
                "ai_action":          v.action,
                "ai_explanation":     v.explanation,
                "ai_score_adjustment": v.score_adjustment,
            })
            break

    logger.info("On-demand AI analysis complete for %s: %s / %s", resource.resource_name, v.action, v.confidence)
    return {
        "ok":               True,
        "resource_id":      resource_id,
        "ai_confidence":    v.confidence,
        "ai_action":        v.action,
        "ai_explanation":   v.explanation,
        "score_adjustment": v.score_adjustment,
    }


# ── Cache status endpoint ─────────────────────────────────────────────────────

@app.get("/api/cache/status", response_model=CacheStatus)
async def cache_status():
    """Returns when data was last fetched, whether a refresh is running, and when the next one is scheduled."""
    # Find the most recently cached scan
    last_ts: Optional[float] = None
    for key, val in _cache.items():
        if key.endswith(":ts"):
            if last_ts is None or val > last_ts:
                last_ts = val

    # After a fresh process start / replica restart the in-memory cache is empty even though a
    # durable snapshot exists on disk. Rehydrate it (best-effort) so the portal's background
    # recovery poll sees the last successful scan instead of reporting "no data" — without this an
    # already-open page can't auto-recover after the server restarts until the user reloads.
    if last_ts is None:
        try:
            if _ensure_dashboard_in_cache() is not None:
                for key, val in _cache.items():
                    if key.endswith(":ts") and (last_ts is None or val > last_ts):
                        last_ts = val
        except Exception:
            pass

    interval_hours = int(settings_svc.get_value("auto_refresh_interval_hours", 0))

    # Stale-guard: never report "refreshing" forever. If a scan has been running
    # longer than 4 minutes (slow/hung live build), surface it as not-refreshing
    # so the UI spinner clears; the background loop's finally still resets the flag.
    refreshing = _is_refreshing
    if refreshing and _refresh_started_ts:
        if datetime.now(tz=timezone.utc).timestamp() - _refresh_started_ts > 240:
            refreshing = False

    return CacheStatus(
        data_available    = bool(_cache),
        last_refreshed    = datetime.fromtimestamp(last_ts, tz=timezone.utc).isoformat() if last_ts else None,
        is_refreshing     = refreshing,
        next_refresh      = datetime.fromtimestamp(_next_refresh_ts, tz=timezone.utc).isoformat() if _next_refresh_ts else None,
        auto_refresh_interval_hours = interval_hours,
    )


@app.get("/api/cache/redis-status")
async def redis_cache_status():
    """Reports whether the shared Redis L2 cache / distributed lock is active.

    Never returns secrets — only enabled/url-configured flags, last error and
    the Redis server version when connected.
    """
    return cache_svc.status()


@app.get("/api/dashboard/cached")
async def get_cached_dashboard_instant():
    """
    Return the most recently cached dashboard payload without triggering a new scan.
    Returns 204 (no content) when no cached data is available.
    """
    from fastapi.responses import Response as _Response
    # Find the freshest cache entry
    best_ts: float = 0.0
    best_data: Optional[DashboardData] = None
    for key, val in _cache.items():
        if not key.endswith(":ts"):
            continue
        data_key = key[:-3]  # strip ":ts"
        if data_key in _cache and val > best_ts:
            best_ts = val
            best_data = _cache[data_key]
    # Also check the legacy REST-endpoint cache slot
    if "data" in _cache and _cache.get("cached_at", 0) > best_ts:
        best_data = _cache["data"]

    def _stamp(payload):
        # The frontend rejects any cached payload whose _snapshot_schema doesn't
        # match its own version (guards against rendering an old-shape snapshot).
        # In-memory cache entries are built by the CURRENT code, so their shape
        # always matches — they just never get the stamp that save_dashboard adds
        # to the durable copy. Stamp it here so the portal opens instantly from
        # the live in-memory cache instead of forcing a foreground rebuild.
        if isinstance(payload, dict):
            payload["_snapshot_schema"] = persistence_svc.SNAPSHOT_SCHEMA_VERSION
        return payload

    if best_data is None:
        # In-memory cache is empty (e.g. fresh process start) — fall back to the
        # durable on-disk snapshot so the portal still opens instantly with the
        # last successful scan instead of forcing a full live build.
        #
        # Also rehydrate the in-memory cache from that snapshot so on-demand
        # endpoints (AI module analysis, migration assessment, health score, …)
        # have resource data to work with — otherwise they return 404
        # "No resource data" even though the user is looking at a full dashboard.
        _ensure_dashboard_in_cache()
        # Return the snapshot AS-IS for the response: a genuinely old-shape
        # snapshot keeps its (missing/old) stamp so the frontend's version check
        # can still reject it and rebuild rather than render a stale payload shape.
        # Prefer the Redis L2 copy (populated by _ensure_dashboard_in_cache) so we
        # don't re-read the large payload from Azure SQL a second time.
        snap = cache_svc.get_json("dash:latest") or persistence_svc.load_latest_dashboard()
        if snap:
            return snap
        return _Response(status_code=204)

    # In-memory hit — serialize the model to a plain dict and stamp the current
    # schema version so the frontend accepts it (same payload shape as the
    # durable snapshot, just sourced from the live cache).
    if hasattr(best_data, "model_dump"):
        payload = json.loads(best_data.model_dump_json())
    elif isinstance(best_data, dict):
        payload = best_data
    else:
        payload = best_data
    return _stamp(payload)


# ── Project API (portal-first: save resource selections as named projects) ────

class ProjectCreate(BaseModel):
    # Core fields (backward compatible with old UI)
    name: Optional[str] = None  # Old UI field
    resource_ids: Optional[List[str]] = None  # Old UI field
    description: str = ""
    color: str = "#3b82f6"
    icon: str = "📁"
    
    # APEX BCDR fields (new)
    project_id: Optional[str] = None
    project_name: Optional[str] = None
    business_unit: Optional[str] = None
    criticality: Optional[str] = None
    rto_target: Optional[str] = None
    rpo_target: Optional[str] = None
    environment: Optional[str] = None
    dr_tier: Optional[str] = None
    owner: Optional[str] = None


class ProjectUpdate(BaseModel):
    name: Optional[str] = None
    resource_ids: Optional[List[str]] = None
    description: Optional[str] = None
    color: Optional[str] = None
    icon: Optional[str] = None


class ResourcesAddRemove(BaseModel):
    resource_ids: List[str]


@app.get("/api/projects")
async def get_projects():
    """List all saved BCDR projects."""
    from services.database import get_raw_connection, is_azure_sql
    projects = []
    try:
        conn = get_raw_connection()
        try:
            cursor = conn.cursor()
            if is_azure_sql():
                columns = [r[0] for r in cursor.execute(
                    "SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS WHERE TABLE_NAME='projects'"
                ).fetchall()]
            else:
                cursor.execute("PRAGMA table_info(projects)")
                columns = [col[1] for col in cursor.fetchall()]
            if 'project_id' in columns:
                cursor.execute("""
                    SELECT project_id, project_name, description, business_unit,
                           criticality, rto_target, rpo_target, environment, dr_tier, owner,
                           created_at, updated_at
                    FROM projects
                    WHERE project_id IS NOT NULL
                    ORDER BY created_at DESC
                """)
                rows = cursor.fetchall()
                for row in rows:
                    projects.append({
                        "project_id": row[0],
                        "project_name": row[1],
                        "description": row[2],
                        "business_unit": row[3],
                        "criticality": row[4],
                        "rto_target": row[5],
                        "rpo_target": row[6],
                        "environment": row[7],
                        "dr_tier": row[8],
                        "owner": row[9],
                        "created_at": row[10],
                        "updated_at": row[11],
                    })
        finally:
            conn.close()
    except Exception as e:
        logger.warning(f"Error loading projects: {e}")
    return {"projects": projects}


@app.post("/api/projects", status_code=201)
async def create_project(body: ProjectCreate):
    """Create a new project from a selection of resource IDs (old UI) or APEX BCDR project (new UI)."""
    # Support both old UI (name + resource_ids) and new APEX UI (project_id + project_name)
    project_name = body.project_name or body.name or ""
    project_id = body.project_id or None
    
    if not project_name.strip():
        raise HTTPException(status_code=400, detail="Project name is required")
    
    # If this is an APEX project, store in APEX database
    if project_id:
        from services.database import get_raw_connection
        conn = get_raw_connection()
        try:
            cursor = conn.cursor()
            cursor.execute("""
                INSERT INTO projects (
                    project_id, project_name, description, business_unit, 
                    criticality, rto_target, rpo_target, environment, dr_tier, owner
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                project_id, project_name, body.description, body.business_unit,
                body.criticality, body.rto_target, body.rpo_target, 
                body.environment, body.dr_tier, body.owner
            ))
            conn.commit()
            
            # Also store resource associations if provided
            if body.resource_ids:
                for resource_id in body.resource_ids:
                    cursor.execute("""
                        INSERT INTO project_resources (project_id, resource_id, role)
                        VALUES (?, ?, ?)
                    """, (project_id, resource_id, "primary"))
                conn.commit()
        finally:
            conn.close()
        
        return {
            "project_id": project_id,
            "project_name": project_name,
            "description": body.description,
            "business_unit": body.business_unit,
            "criticality": body.criticality,
            "rto_target": body.rto_target,
            "rpo_target": body.rpo_target,
            "environment": body.environment,
            "dr_tier": body.dr_tier,
            "owner": body.owner,
            "created_at": datetime.utcnow().isoformat()
        }
    else:
        # Old UI - use legacy project service
        return project_svc.create_project(
            name=project_name,
            resource_ids=body.resource_ids or [],
            description=body.description,
            color=body.color,
            icon=body.icon,
        )


@app.get("/api/projects/{project_id}")
async def get_project(project_id: str):
    p = project_svc.get_project(project_id)
    if not p:
        raise HTTPException(status_code=404, detail="Project not found")
    return p


@app.put("/api/projects/{project_id}")
async def update_project(project_id: str, body: ProjectUpdate):
    p = project_svc.update_project(
        project_id,
        name=body.name,
        resource_ids=body.resource_ids,
        description=body.description,
        color=body.color,
        icon=body.icon,
    )
    if not p:
        raise HTTPException(status_code=404, detail="Project not found")
    return p


@app.delete("/api/projects/{project_id}", status_code=204)
async def delete_project(project_id: str):
    # Try APEX project first
    from services.database import get_raw_connection, is_azure_sql
    try:
        conn = get_raw_connection()
        try:
            cursor = conn.cursor()
            if is_azure_sql():
                columns = [r[0] for r in cursor.execute(
                    "SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS WHERE TABLE_NAME='projects'"
                ).fetchall()]
            else:
                cursor.execute("PRAGMA table_info(projects)")
                columns = [col[1] for col in cursor.fetchall()]
            if 'project_id' in columns:
                cursor.execute("DELETE FROM projects WHERE project_id = ?", (project_id,))
                if cursor.rowcount > 0:
                    conn.commit()
                    return
        finally:
            conn.close()
    except Exception:
        pass
    # Fallback to legacy project service
    if not project_svc.delete_project(project_id):
        raise HTTPException(status_code=404, detail="Project not found")


@app.post("/api/projects/{project_id}/resources")
async def add_resources(project_id: str, body: ResourcesAddRemove):
    """Append resources to an existing project."""
    p = project_svc.add_resources_to_project(project_id, body.resource_ids)
    if not p:
        raise HTTPException(status_code=404, detail="Project not found")
    return p


@app.delete("/api/projects/{project_id}/resources")
async def remove_resources(project_id: str, body: ResourcesAddRemove):
    """Remove specific resources from a project."""
    p = project_svc.remove_resources_from_project(project_id, body.resource_ids)
    if not p:
        raise HTTPException(status_code=404, detail="Project not found")
    return p


# ── Dependency Graph API ────────────────────────────────────────────────────────

@app.get("/api/dependencies")
async def get_dependencies():
    """Return the full dependency graph (nodes, edges, clusters, SPOF)."""
    graph = _cache.get("dependency_graph")
    if not graph:
        return {"error": "Dependency graph not yet computed. Trigger a scan first.", "node_count": 0, "edge_count": 0}
    return graph.model_dump()


@app.get("/api/dependencies/summary")
async def get_dependencies_summary():
    """Lightweight summary of the dependency graph."""
    graph = _cache.get("dependency_graph")
    if not graph:
        return {"node_count": 0, "edge_count": 0, "cluster_count": 0, "spof_count": 0}
    return get_graph_summary(graph).model_dump()


@app.get("/api/dependencies/clusters")
async def get_dependency_clusters():
    """Auto-detected workload clusters from dependency graph."""
    graph = _cache.get("dependency_graph")
    if not graph:
        return []
    return [c.model_dump() for c in graph.clusters]


@app.get("/api/dependencies/spof")
async def get_spof():
    """Single points of failure identified by the dependency graph."""
    graph = _cache.get("dependency_graph")
    if not graph:
        return []
    return [s.model_dump() for s in graph.spof]


@app.get("/api/dependencies/{resource_id:path}/blast-radius")
async def get_resource_blast_radius(resource_id: str):
    """Calculate blast radius for a specific resource."""
    graph = _cache.get("dependency_graph")
    if not graph:
        raise HTTPException(status_code=404, detail="Dependency graph not computed")
    try:
        return get_blast_radius(resource_id, graph).model_dump()
    except Exception as exc:
        raise HTTPException(status_code=404, detail=str(exc))


@app.get("/api/dependencies/{resource_id:path}")
async def get_single_resource_dependencies(resource_id: str):
    """Get upstream/downstream dependencies for a single resource."""
    graph = _cache.get("dependency_graph")
    if not graph:
        raise HTTPException(status_code=404, detail="Dependency graph not computed")
    return get_resource_dependencies(resource_id, graph)


# ── BCDR Assessment API ────────────────────────────────────────────────────────

def _get_bcdr_assessments() -> list:
    """Build BCDR assessments from cached resources, with caching."""
    if "bcdr_assessments" in _cache:
        return _cache["bcdr_assessments"]
    data = _cache.get("data:*") or _cache.get("data")
    if not data:
        return []
    try:
        resources = [r.model_dump() for r in (data.resources or [])]
        settings  = settings_svc.get()
        sub_names: dict[str, str] = {}
        for s in settings.get("subscriptions", []):
            sub_names[s.get("id", "")] = s.get("name", "")
        assessments = []
        for r in resources:
            sub_name = sub_names.get(r.get("subscription_id", ""), "")
            assessments.append(
                assess_all_resources([r], subscription_name=sub_name)[0]
            )
        _cache["bcdr_assessments"] = assessments
        return assessments
    except Exception as exc:
        logger.warning("BCDR assessments build failed: %s", exc)
        return []


@app.get("/api/bcdr/dashboard")
async def bcdr_dashboard():
    """BCDR dashboard summary — zone breakdown, risk heatmap, tier breakdown."""
    assessments = _get_bcdr_assessments()
    return build_bcdr_dashboard_summary(assessments)


@app.get("/api/bcdr/assessments")
async def bcdr_assessments(
    location: Optional[str] = None,
    zone_status: Optional[str] = None,
    tier: Optional[str] = None,
    priority: Optional[str] = None,
    quick_wins_only: bool = False,
    limit: int = 200,
    offset: int = 0,
):
    """Full 19-column BCDR assessment with filtering and pagination."""
    assessments = _get_bcdr_assessments()
    recs = generate_all_recommendations(assessments)

    # Apply filters
    if location:
        recs = [r for r in recs if r.location == location.lower()]
    if zone_status:
        recs = [r for r in recs if r.zone_status.lower() == zone_status.lower()]
    if tier:
        recs = [r for r in recs if r.workload_tier.lower() == tier.lower()]
    if priority:
        recs = [r for r in recs if r.sa_priority == priority.upper()]
    if quick_wins_only:
        recs = [r for r in recs if r.sa_quick_win == "Yes"]

    # Sort by risk score descending
    recs.sort(key=lambda r: r.zone_risk_score, reverse=True)

    total = len(recs)
    page  = recs[offset:offset + limit]
    return {
        "total":   total,
        "offset":  offset,
        "limit":   limit,
        "items":   [r.to_dict() for r in page],
        "summary": build_priority_summary(recs),
    }


@app.get("/api/bcdr/quick-wins")
async def bcdr_quick_wins():
    """Top quick-win BCDR actions (QuickWin=Yes, sorted by priority)."""
    assessments = _get_bcdr_assessments()
    recs = generate_all_recommendations(assessments)
    return {"items": build_quick_wins(recs), "total": len([r for r in recs if r.sa_quick_win == "Yes"])}


@app.get("/api/bcdr/resource/{resource_id:path}")
async def bcdr_single_resource(resource_id: str):
    """Full BCDR recommendation for a single resource."""
    assessments = _get_bcdr_assessments()
    match = next((a for a in assessments if a.resource_id == resource_id), None)
    if not match:
        raise HTTPException(status_code=404, detail="Resource not found in BCDR assessment cache")
    return generate_all_recommendations([match])[0].to_dict()


@app.post("/api/bcdr/refresh")
async def bcdr_refresh():
    """Clear BCDR assessment cache and force rebuild on next request."""
    _cache.pop("bcdr_assessments", None)
    assessments = _get_bcdr_assessments()
    return {"ok": True, "total": len(assessments)}


@app.get("/api/bcdr/business-impact")
async def bcdr_business_impact():
    """Business Impact Analysis — criticality tiers, impact scores, downtime costs."""
    data = _cache.get("data:*") or _cache.get("data")
    if not data:
        raise HTTPException(status_code=404, detail="No scan data available")
    resources = [r.model_dump() for r in (data.resources or [])]
    assessments = _get_bcdr_assessments()
    return build_business_impact_analysis(resources, assessments)


@app.get("/api/bcdr/recovery-sequence")
async def bcdr_recovery_sequence():
    """Recovery Sequence Planner — dependency-aware recovery ordering."""
    data = _cache.get("data:*") or _cache.get("data")
    if not data:
        raise HTTPException(status_code=404, detail="No scan data available")
    resources = [r.model_dump() for r in (data.resources or [])]
    assessments = _get_bcdr_assessments()
    return build_recovery_sequence_plan(resources, assessments)


# ── BCDR Metadata (Phase 1 Planning) API ─────────────────────────────────────

class BCDRMetadataBody(BaseModel):
    criticality:       Optional[str] = None
    dr_tier:           Optional[str] = None
    rto_target:        Optional[str] = None
    rpo_target:        Optional[str] = None
    business_function: Optional[str] = None
    notes:             Optional[str] = None

class BulkBCDRMetadataBody(BaseModel):
    updates: List[Dict[str, Any]]


@app.get("/api/bcdr/metadata")
async def get_all_bcdr_metadata():
    """Get all BCDR metadata for all resources."""
    if not bcdr_meta_svc:
        return {"error": "bcdr_metadata_service not available"}
    
    return bcdr_meta_svc.get_all_bcdr_metadata()


@app.get("/api/bcdr/metadata/{resource_id:path}")
async def get_resource_bcdr_metadata(resource_id: str):
    """Get BCDR metadata for a single resource."""
    if not bcdr_meta_svc:
        raise HTTPException(status_code=503, detail="bcdr_metadata_service not available")
    
    metadata = bcdr_meta_svc.get_bcdr_metadata(resource_id)
    if not metadata:
        return {"resource_id": resource_id, "metadata": None}
    
    return metadata


@app.post("/api/bcdr/metadata/{resource_id:path}")
async def save_resource_bcdr_metadata(resource_id: str, body: BCDRMetadataBody):
    """Save or update BCDR metadata for a resource."""
    if not bcdr_meta_svc:
        raise HTTPException(status_code=503, detail="bcdr_metadata_service not available")
    
    metadata = body.dict()
    result = bcdr_meta_svc.save_bcdr_metadata(resource_id, metadata)
    return result


@app.post("/api/bcdr/metadata/bulk")
async def bulk_save_bcdr_metadata(body: BulkBCDRMetadataBody):
    """Bulk save BCDR metadata for multiple resources."""
    if not bcdr_meta_svc:
        raise HTTPException(status_code=503, detail="bcdr_metadata_service not available")
    
    count = bcdr_meta_svc.bulk_save_bcdr_metadata(body.updates)
    return {"ok": True, "updated_count": count}


@app.delete("/api/bcdr/metadata/{resource_id:path}")
async def delete_resource_bcdr_metadata(resource_id: str):
    """Delete BCDR metadata for a resource."""
    if not bcdr_meta_svc:
        raise HTTPException(status_code=503, detail="bcdr_metadata_service not available")
    
    success = bcdr_meta_svc.delete_bcdr_metadata(resource_id)
    return {"ok": success}


@app.get("/api/bcdr/metadata/stats")
async def get_bcdr_metadata_stats():
    """Get statistics about BCDR metadata coverage."""
    if not bcdr_meta_svc:
        return {"error": "bcdr_metadata_service not available"}
    
    return bcdr_meta_svc.get_bcdr_metadata_stats()


# ── BCDR Deliverables API (Timeline, Testing Plan, Compliance, Strategy, Excel) ──

@app.get("/api/bcdr/executive-summary")
async def bcdr_executive_summary():
    """Executive summary of BCDR assessment — aggregated KPIs and key findings."""
    assessments = _get_bcdr_assessments()
    recs = generate_all_recommendations(assessments)
    return build_executive_summary(assessments, recs)


@app.get("/api/bcdr/timeline")
async def bcdr_timeline():
    """6-phase BCDR implementation timeline / action plan."""
    return get_timeline_action_plan()


@app.get("/api/bcdr/testing-plan")
async def bcdr_testing_plan():
    """DR testing plan template with checklists, success criteria, quarterly schedule, and RACI."""
    return get_dr_testing_plan()


@app.get("/api/bcdr/compliance")
async def bcdr_compliance():
    """Qatar BCDR compliance checklist — 7 categories, 40+ items, sign-off tracking."""
    return get_compliance_checklist()


@app.get("/api/bcdr/strategy-reference")
async def bcdr_strategy_ref():
    """BCDR strategy reference guide — 7 DR patterns, Qatar constraints, decision matrix."""
    return get_strategy_reference()


@app.get("/api/bcdr/excel-report")
async def bcdr_excel_report():
    """Generate and download multi-sheet BCDR assessment Excel report."""
    assessments = _get_bcdr_assessments()
    recs = generate_all_recommendations(assessments)
    try:
        excel_bytes = generate_excel_report(assessments, recs)
    except RuntimeError as e:
        raise HTTPException(status_code=500, detail=str(e))
    filename = f"BCDR_Assessment_Report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        excel_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ── Custom Tagging API ────────────────────────────────────────────────────────

class TagSchemaEntry(BaseModel):
    tag_key:      str
    display_name: str
    tag_type:     str = "text"        # text | enum | bool | number
    enum_values:  list = []
    category:     str = "Custom"
    is_required:  bool = False
    color:        str = "#6b7280"

class ResourceTagsBody(BaseModel):
    tags: Dict[str, str]
    merge: bool = False  # If True, merge with existing tags; if False, replace all

class BulkTagBody(BaseModel):
    resource_ids: List[str]
    tags:         Dict[str, str]
    merge: bool = False  # If True, merge with existing tags; if False, replace all

class TagImportBody(BaseModel):
    csv_text: str


def _require_tagging():
    if tagging_svc is None:
        raise HTTPException(status_code=503, detail="Tagging service unavailable")

def _require_ai():
    if ai_infra_svc is None:
        raise HTTPException(status_code=503, detail="AI service unavailable")


# ── Live Azure Tag Keys & Values (for searchable tag pickers) ────────────────

_tag_keys_cache: dict = {"data": None, "ts": 0}
_tag_values_cache: dict = {}  # {key: {"data": [...], "ts": float}}
_TAG_CACHE_TTL = 300  # 5 minutes

@app.get("/api/tags/keys", tags=["Tags"])
async def get_live_tag_keys():
    """
    Return all unique tag keys across Azure resources.
    Pulls from dashboard cache for speed, merges with custom tags.
    5-minute in-memory cache to avoid repeated processing.
    """
    import time
    now = time.time()
    if _tag_keys_cache["data"] is not None and (now - _tag_keys_cache["ts"]) < _TAG_CACHE_TTL:
        return {"tag_keys": _tag_keys_cache["data"], "cached": True}

    tag_keys: set = set()

    # 1. From dashboard cache (Azure resource tags)
    dash: Optional[DashboardData] = _cache.get("data:*") or _cache.get("data")
    if dash and dash.resources:
        for r in dash.resources:
            tags = getattr(r, "tags", None) or {}
            if isinstance(tags, dict):
                tag_keys.update(tags.keys())

    # 2. From custom tagging service
    if tagging_svc:
        try:
            stats = tagging_svc.get_tag_statistics()
            tag_keys.update(stats.get("key_counts", {}).keys())
        except Exception:
            pass

    sorted_keys = sorted(tag_keys)
    _tag_keys_cache["data"] = sorted_keys
    _tag_keys_cache["ts"] = now
    return {"tag_keys": sorted_keys, "cached": False}


@app.get("/api/tags/values/{tag_key}", tags=["Tags"])
async def get_live_tag_values(tag_key: str):
    """
    Return all unique values for a specific tag key across Azure resources.
    Pulls from dashboard cache. 5-minute in-memory cache per key.
    """
    import time
    now = time.time()
    cached = _tag_values_cache.get(tag_key)
    if cached and (now - cached["ts"]) < _TAG_CACHE_TTL:
        return {"tag_key": tag_key, "values": cached["data"], "count": len(cached["data"]), "cached": True}

    values: set = set()

    # From dashboard cache
    dash: Optional[DashboardData] = _cache.get("data:*") or _cache.get("data")
    if dash and dash.resources:
        for r in dash.resources:
            tags = getattr(r, "tags", None) or {}
            if isinstance(tags, dict):
                val = tags.get(tag_key)
                if val is not None and str(val).strip():
                    values.add(str(val).strip())

    # From custom tags
    if tagging_svc:
        try:
            all_custom = tagging_svc.get_all_custom_tags()
            for rid, tags in all_custom.items():
                val = tags.get(tag_key)
                if val is not None and str(val).strip():
                    values.add(str(val).strip())
        except Exception:
            pass

    sorted_values = sorted(values)
    _tag_values_cache[tag_key] = {"data": sorted_values, "ts": now}
    return {"tag_key": tag_key, "values": sorted_values, "count": len(sorted_values), "cached": False}


@app.get("/api/tags/schema")
async def get_tag_schema():
    """Return all custom tag key definitions."""
    _require_tagging()
    return tagging_svc.get_tag_schema()


@app.post("/api/tags/schema")
async def upsert_tag_schema(entry: TagSchemaEntry):
    """Create or update a tag key definition."""
    _require_tagging()
    return tagging_svc.upsert_tag_schema(entry.model_dump())


@app.delete("/api/tags/schema/{tag_key}")
async def delete_tag_schema(tag_key: str):
    """Delete a tag key and all resource values for it."""
    _require_tagging()
    tagging_svc.delete_tag_schema(tag_key)
    return {"ok": True}


@app.get("/api/tags/resource/{resource_id:path}")
async def get_resource_tags(resource_id: str):
    """Get custom tags for a single resource."""
    return tagging_svc.get_custom_tags(resource_id)


@app.put("/api/tags/resource/{resource_id:path}")
async def set_resource_tags(resource_id: str, body: ResourceTagsBody):
    """Set custom tags for a resource (replace or merge mode)."""
    if body.merge:
        # Merge mode: get existing tags and update with new ones
        existing = tagging_svc.get_resource_tags(resource_id)
        merged = {**existing, **body.tags}
        tagging_svc.set_resource_tags(resource_id, merged)
    else:
        # Replace mode: replace all tags
        tagging_svc.set_resource_tags(resource_id, body.tags)
    return {"ok": True}


@app.post("/api/tags/bulk")
async def bulk_tag_resources(body: BulkTagBody):
    """Apply tags to multiple resources at once (replace or merge mode)."""
    for rid in body.resource_ids:
        if body.merge:
            existing = tagging_svc.get_resource_tags(rid)
            merged = {**existing, **body.tags}
            tagging_svc.set_resource_tags(rid, merged)
        else:
            tagging_svc.set_resource_tags(rid, body.tags)
    return {"ok": True, "tags_set": len(body.resource_ids)}


@app.get("/api/tags/all")
async def get_all_tags(resource_ids: Optional[str] = None):
    """
    Return all custom tags.
    Pass ?resource_ids=id1,id2 to filter to specific resources.
    """
    ids = [i.strip() for i in resource_ids.split(",")] if resource_ids else None
    return tagging_svc.get_all_custom_tags(ids)


@app.get("/api/tags/stats")
async def get_tag_stats():
    """Tag usage statistics."""
    return tagging_svc.get_tag_statistics()


@app.get("/api/tags/export")
async def export_tags_csv():
    """Export all custom tags as CSV."""
    from fastapi.responses import Response
    csv_text = tagging_svc.export_tags_csv()
    return Response(content=csv_text, media_type="text/csv",
                    headers={"Content-Disposition": "attachment; filename=resource-tags.csv"})


@app.post("/api/tags/import")
async def import_tags_csv(body: TagImportBody):
    """Import custom tags from CSV text."""
    return tagging_svc.import_tags_csv(body.csv_text)


# ── Infrastructure Intelligence Summary API ──────────────────────────────────

@app.get("/api/infra/summary")
async def infra_summary():
    """
    Comprehensive infrastructure summary — data-driven, no AI required.
    Powers the Infrastructure Dashboard with KPIs, breakdowns, and health indicators.
    """
    resources = _get_resources_list()
    if not resources:
        return {"error": "No resource data — run a scan first", "has_data": False}

    total_cost = sum(r.get("cost_current_month", 0) for r in resources)
    prev_cost = sum(r.get("cost_previous_month", 0) for r in resources)

    # Location breakdown
    by_location = {}
    for r in resources:
        loc = r.get("location", "unknown")
        if loc not in by_location:
            by_location[loc] = {"count": 0, "cost": 0}
        by_location[loc]["count"] += 1
        by_location[loc]["cost"] += r.get("cost_current_month", 0)

    # Type breakdown
    by_type = {}
    for r in resources:
        t = r.get("resource_type", "unknown")
        short = t.split("/")[-1] if "/" in t else t
        if short not in by_type:
            by_type[short] = {"count": 0, "cost": 0, "full_type": t}
        by_type[short]["count"] += 1
        by_type[short]["cost"] += r.get("cost_current_month", 0)

    # Resource Group breakdown
    by_rg = {}
    for r in resources:
        rg = r.get("resource_group", "unknown")
        if rg not in by_rg:
            by_rg[rg] = {"count": 0, "cost": 0, "types": set()}
        by_rg[rg]["count"] += 1
        by_rg[rg]["cost"] += r.get("cost_current_month", 0)
        by_rg[rg]["types"].add(r.get("resource_type", "").split("/")[-1])
    # Convert sets to lists for JSON
    for rg in by_rg.values():
        rg["types"] = list(rg["types"])

    # Subscription breakdown
    by_sub = {}
    for r in resources:
        sub = r.get("subscription_id", "unknown")
        sub_name = r.get("subscription_name", sub[:8] + "..." if len(sub) > 8 else sub)
        if sub not in by_sub:
            by_sub[sub] = {"count": 0, "cost": 0, "name": sub_name}
        by_sub[sub]["count"] += 1
        by_sub[sub]["cost"] += r.get("cost_current_month", 0)

    # Health indicators
    waste_count = sum(1 for r in resources if r.get("final_score", 100) <= 25)
    likely_waste = sum(1 for r in resources if 25 < r.get("final_score", 100) <= 50)
    underutilized = sum(1 for r in resources if 50 < r.get("final_score", 100) <= 75)
    well_used = sum(1 for r in resources if r.get("final_score", 100) > 75)
    no_metrics = sum(1 for r in resources if r.get("primary_utilization_pct") is None)

    # Security & Compliance
    has_backup = sum(1 for r in resources if r.get("has_backup"))
    has_lock = sum(1 for r in resources if r.get("has_lock"))
    is_orphan = sum(1 for r in resources if r.get("is_orphan"))
    tagged = sum(1 for r in resources if r.get("tags") and len(r.get("tags", {})) > 0)
    deallocated = sum(1 for r in resources if r.get("power_state") in ("deallocated", "stopped"))

    # Cost insights
    top_cost_resources = sorted(resources, key=lambda r: -(r.get("cost_current_month", 0)))[:10]
    top_cost = [{
        "name": r.get("resource_name", ""),
        "type": r.get("resource_type", "").split("/")[-1],
        "cost": round(r.get("cost_current_month", 0), 2),
        "rg": r.get("resource_group", ""),
        "score": r.get("final_score", 0),
        "score_label": r.get("score_label", "Unknown"),
    } for r in top_cost_resources]

    # Waste cost estimation
    waste_cost = sum(r.get("cost_current_month", 0) for r in resources if r.get("final_score", 100) <= 25)
    likely_waste_cost = sum(r.get("cost_current_month", 0) for r in resources if 25 < r.get("final_score", 100) <= 50)

    # Score distribution
    score_dist = {
        "confirmed_waste": waste_count,
        "likely_waste": likely_waste,
        "underutilized": underutilized,
        "well_used": well_used,
        "no_metrics": no_metrics,
    }

    # Advisor recommendations count
    advisor_count = sum(len(r.get("advisor_recommendations", [])) for r in resources)

    return {
        "has_data": True,
        "total_resources": len(resources),
        "total_cost": round(total_cost, 2),
        "prev_cost": round(prev_cost, 2),
        "cost_trend_pct": round(((total_cost - prev_cost) / max(prev_cost, 1)) * 100, 1) if prev_cost > 0 else 0,
        "by_location": by_location,
        "by_type": dict(sorted(by_type.items(), key=lambda x: -x[1]["cost"])[:20]),
        "by_resource_group": dict(sorted(by_rg.items(), key=lambda x: -x[1]["cost"])[:20]),
        "by_subscription": by_sub,
        "health": {
            "backup_coverage_pct": round(has_backup / max(len(resources), 1) * 100, 1),
            "lock_coverage_pct": round(has_lock / max(len(resources), 1) * 100, 1),
            "tag_compliance_pct": round(tagged / max(len(resources), 1) * 100, 1),
            "orphan_count": is_orphan,
            "deallocated_count": deallocated,
            "advisor_recommendations": advisor_count,
        },
        "score_distribution": score_dist,
        "top_cost_resources": top_cost,
        "waste_summary": {
            "confirmed_waste_cost": round(waste_cost, 2),
            "likely_waste_cost": round(likely_waste_cost, 2),
            "total_potential_savings": round(waste_cost + likely_waste_cost * 0.5, 2),
        },
    }


# ── AI Infrastructure Intelligence API ───────────────────────────────────────

def _ensure_dashboard_in_cache() -> Optional[DashboardData]:
    """Return the in-memory dashboard, rehydrating it from the latest durable
    snapshot when the in-process cache is empty.

    With the snapshot-first instant-load model the portal renders straight from
    the persisted snapshot without running a live build, so the in-memory
    ``_cache`` can be empty even though the user sees a full dashboard. Every
    on-demand endpoint (AI module analysis, migration assessment, health score,
    etc.) reads resources from ``_cache``; without this rehydration those
    endpoints return 404 "No resource data" after a fresh process start.
    """
    data: Optional[DashboardData] = _cache.get("data:*") or _cache.get("data")
    # Only trust the in-memory dashboard when it actually carries resources. A snapshot-first
    # instant-load shell (or a scoped/partial build) can be cached with KPIs but an EMPTY
    # resources array; returning it here makes every resource-dependent endpoint (AI analysis,
    # assessments, health score) 404 "No resource data" even though the durable snapshot holds the
    # full resource list. When resources are missing, fall through and rehydrate from the snapshot.
    if data and (getattr(data, "resources", None) or []):
        return data
    try:
        # L2 (Redis) cache-aside in front of the durable SQL snapshot: a cold
        # process / second replica gets the full dashboard from Redis instead of
        # re-reading the large payload from Azure SQL. Falls back to SQL (and then
        # populates Redis) when the cache is cold or unavailable.
        snap = cache_svc.get_json("dash:latest")
        if not snap:
            snap = persistence_svc.load_latest_dashboard()
            if snap:
                try:
                    _disp_ttl = float(settings_svc.get_value("metrics_display_ttl_hours", 24.0))
                    cache_svc.set_json("dash:latest", snap, ttl_seconds=int(_disp_ttl * 3600))
                except Exception:
                    pass
        if not snap:
            return data
        restored = DashboardData(**{k: v for k, v in snap.items() if k in DashboardData.model_fields})
        # The dashboard snapshot may have been captured before cost data was
        # available (cost is fetched on a separate, throttle-tolerant cadence).
        # Backfill the daily cost arrays from the durable cost snapshot so the
        # home Spend Trend and FinOps trend are never empty on the fast path.
        try:
            if _cost_series_empty(getattr(restored, "total_daily_cm", None), getattr(restored, "total_daily_pm", None)):
                _csnap = persistence_svc.load_latest_cost_snapshot()
                if _csnap:
                    restored.total_daily_cm = _csnap.get("total_daily_cm") or restored.total_daily_cm
                    restored.total_daily_pm = _csnap.get("total_daily_pm") or restored.total_daily_pm
                    # Patch KPI spend figures too when the scan snapshot predates
                    # cost capture, so the home spend cards aren't left at $0.
                    _k = getattr(restored, "kpi", None)
                    if _k is not None and not (
                        getattr(_k, "total_cost_current_month", 0) or getattr(_k, "total_cost_previous_month", 0)
                    ):
                        _curr = round(sum(_csnap.get("total_daily_cm") or []), 2)
                        _prev = round(sum(_csnap.get("total_daily_pm") or []), 2)
                        _k.total_cost_current_month = _curr
                        _k.total_cost_previous_month = _prev
                        _k.mom_cost_delta = round(_curr - _prev, 2)
                        _k.mom_cost_delta_pct = round((_curr - _prev) / _prev * 100, 2) if _prev else 0.0
        except Exception as _ce:
            logger.debug("Cost-array backfill on rehydrate skipped: %s", _ce)
        ts: Optional[float] = None
        _lr = snap.get("last_refreshed")
        if _lr:
            try:
                ts = datetime.fromisoformat(str(_lr).replace("Z", "+00:00")).timestamp()
            except Exception:
                ts = None
        if ts is None:
            ts = datetime.now(tz=timezone.utc).timestamp()
        _cache["data:*"] = restored
        _cache["data:*:ts"] = ts
        _cache["data"] = restored
        _cache["cached_at"] = ts
        logger.info(
            "Rehydrated dashboard from snapshot into cache (%d resources)",
            len(restored.resources or []),
        )
        return restored
    except Exception as e:
        logger.warning("Could not rehydrate dashboard from snapshot: %s", e)
        return data


def _get_resources_list() -> List[dict]:
    """Extract resource dicts from cache (rehydrating from snapshot if needed)."""
    data = _ensure_dashboard_in_cache()
    if not data:
        return []
    return [r.model_dump() for r in (data.resources or [])]


def _get_resources_objects():
    """Extract raw ResourceMetrics objects from cache (for services that expect attribute access)."""
    data = _ensure_dashboard_in_cache()
    if not data:
        return []
    return data.resources or []


@app.get("/api/ai/status")
async def ai_infra_status():
    """AI infrastructure service availability and model info."""
    info = ai_infra_svc.get_provider_info()
    return info


# ── Architecture Map (embedded engine) integration ─────────────────────────
# The embedded diagram engine ships its own light theme + branding. To present
# it as a native part of the product we bake a dark reskin + our own mark + de-
# branded copy directly into the container's served files (idempotent, survives
# SPA navigation, no white flash). Source of truth: backend/assets/zuremap_brand.css.
_ARCHMAP_CONTAINER = "zuremap"
_ARCHMAP_BROWSER_DIR = "/app/dist/zuremap/browser"
_archmap_rebranded = False


def _archmap_rebrand() -> bool:
    """Inject the dark brand skin + de-brand the embedded engine's static files.

    Best-effort and idempotent: re-running strips the previous skin and reapplies
    a fresh one. Returns False (never raises) if Docker/the container is absent.
    """
    import subprocess, os, re

    css_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "assets", "zuremap_brand.css")
    try:
        with open(css_path, "r", encoding="utf-8") as fh:
            css = fh.read()
    except OSError:
        return False

    try:
        # 1) Read the served index.html from the container
        r = subprocess.run(
            ["docker", "exec", _ARCHMAP_CONTAINER, "cat", f"{_ARCHMAP_BROWSER_DIR}/index.html"],
            capture_output=True, text=True, timeout=10,
        )
        if r.returncode != 0 or "<head>" not in r.stdout:
            return False
        html = r.stdout

        # 2) Strip any previous skin (link or inline) and inject a fresh inline
        #    <style> — inline keeps it origin/path independent (works direct & proxied).
        html = re.sub(r'<style id="brand-skin">.*?</style>', "", html, flags=re.S)
        html = re.sub(r'<link id="brand-skin"[^>]*>', "", html)
        html = html.replace("<title>Zuremap</title>", "<title>Architecture Map</title>")
        style_block = '<style id="brand-skin">' + css + "</style></head>"
        html = html.replace("</head>", style_block, 1)

        # 3) Write the patched index.html back into the container
        subprocess.run(
            ["docker", "exec", "-i", _ARCHMAP_CONTAINER, "sh", "-c",
             f"cat > {_ARCHMAP_BROWSER_DIR}/index.html"],
            input=html, text=True, timeout=10,
        )

        # 4) De-brand visible copy (CapCase 'ZureMap' is display-only; lowercase
        #    'zuremap' autosave/import keys are left untouched) + normalise currency.
        subprocess.run(
            ["docker", "exec", _ARCHMAP_CONTAINER, "sh", "-c",
             f"cd {_ARCHMAP_BROWSER_DIR} && "
             "sed -i 's/ZureMap/Architecture Map/g' chunk-*.js main-*.js 2>/dev/null; "
             "sed -i 's/baseCurrency:\"EUR\"/baseCurrency:\"USD\"/g' chunk-*.js 2>/dev/null; "
             "sed -i 's#<title>Zuremap</title>#<title>Architecture Map</title>#g' index.html 2>/dev/null; "
             "true"],
            capture_output=True, text=True, timeout=15,
        )
        return True
    except (FileNotFoundError, subprocess.TimeoutExpired, OSError):
        return False


# ── ZureMap Architecture Diagram Integration ────────────────────────────────
# ── Reverse proxy for the embedded Architecture Map engine (combined container) ──
# In the SINGLE combined image there is NO NGINX, so the backend itself must proxy
# /zuremap/* to the in-container engine on :3001. WITHOUT this route, /zuremap/* falls
# through to the SPA catch-all ("/{full_path:path}") and returns the PORTAL's index.html
# — which renders our own app recursively inside the Architecture Map iframe. Registered
# BEFORE the catch-all so it wins. (In docker-compose, NGINX proxies /zuremap/ directly.)
_ZM_UPSTREAM = "http://localhost:3001"
_ZM_HOP_BY_HOP = {"connection", "keep-alive", "proxy-authenticate", "proxy-authorization",
                  "te", "trailers", "transfer-encoding", "upgrade", "content-encoding", "content-length"}

@app.api_route("/zuremap", methods=["GET", "HEAD"], include_in_schema=False)
@app.api_route("/zuremap/{path:path}",
               methods=["GET", "POST", "PUT", "DELETE", "PATCH", "HEAD", "OPTIONS"],
               include_in_schema=False)
async def zuremap_proxy(request: Request, path: str = ""):
    """Reverse-proxy /zuremap/* to the in-container Architecture Map engine (:3001).
    Auth is enforced upstream by the _entra_auth_gate middleware (zm_sess cookie)."""
    import httpx
    target = f"{_ZM_UPSTREAM}/{path}"
    if request.url.query:
        target += f"?{request.url.query}"
    body = await request.body()
    fwd_headers = {k: v for k, v in request.headers.items()
                   if k.lower() != "host" and k.lower() not in _ZM_HOP_BY_HOP}
    try:
        async with httpx.AsyncClient(timeout=60.0) as client:
            up = await client.request(request.method, target, content=body,
                                      headers=fwd_headers, follow_redirects=False)
    except Exception as exc:
        # Never fall through to the SPA (that would render the portal in the iframe).
        return JSONResponse({"detail": f"Architecture Map engine not reachable: {exc}"}, status_code=502)
    resp_headers = {k: v for k, v in up.headers.items()
                    if k.lower() not in _ZM_HOP_BY_HOP and k.lower() != "content-type"}
    return Response(content=up.content, status_code=up.status_code,
                    headers=resp_headers, media_type=up.headers.get("content-type"))


@app.get("/api/zuremap/status")
async def zuremap_status():
    """Check if ZureMap container is reachable and authenticated."""
    import httpx
    # When embedded in the combined container (ZUREMAP_EMBED=proxy), the engine runs on
    # localhost:3001 INSIDE the container, but the user's browser is REMOTE — so it must
    # load the engine through the SAME-ORIGIN /zuremap/ reverse proxy, NOT http://localhost:3001
    # (which would resolve to the client's own machine and fail). Force mode='proxy' in embed
    # deployments so the frontend never points the iframe at the client's localhost.
    embed_proxy = os.environ.get("ZUREMAP_EMBED", "").strip().lower() == "proxy"
    # Probe localhost first: for a locally-run backend that's the mapped port and
    # responds instantly, whereas the docker service name "zuremap" does NOT
    # resolve from the host and stalls on DNS. Inside docker-compose, localhost
    # is refused immediately and we fall through to the service name.
    urls_to_try = [
        ("http://localhost:3001", "local"),  # Local dev / host-run backend
        ("http://zuremap:3001", "docker"),   # Docker Compose service name
    ]
    for url, mode in urls_to_try:
        try:
            async with httpx.AsyncClient(timeout=2.0) as client:
                resp = await client.get(url)
                if resp.status_code < 500:
                    # Check login status
                    logged_in = False
                    try:
                        login_resp = await client.get(f"{url}/api/az/login-status")
                        login_data = login_resp.json()
                        logged_in = login_data.get("loggedIn", False)
                    except Exception:
                        pass
                    # Apply the brand skin once per process (fire-and-forget so the
                    # status response stays fast; idempotent and best-effort).
                    global _archmap_rebranded
                    if not _archmap_rebranded:
                        _archmap_rebranded = True
                        try:
                            asyncio.create_task(asyncio.to_thread(_archmap_rebrand))
                        except Exception:
                            pass
                    # In embed deployments always advertise 'proxy' so the SPA uses /zuremap/.
                    effective_mode = "proxy" if embed_proxy else mode
                    return {"available": True, "url": "/zuremap/", "mode": effective_mode, "loggedIn": logged_in}
        except Exception:
            continue
    return {"available": False, "url": None, "mode": None, "loggedIn": False}


@app.post("/api/zuremap/auth")
async def zuremap_auth():
    """Auto-login ZureMap using the backend's service principal credentials."""
    import httpx
    import subprocess
    tid = settings_svc.get_value("AZURE_TENANT_ID", "")
    cid = settings_svc.get_value("AZURE_CLIENT_ID", "")
    sec = settings_svc.get_value("AZURE_CLIENT_SECRET", "")
    if not all([tid, cid, sec]):
        raise HTTPException(status_code=400, detail="Azure service principal not configured in settings")

    # Try az login inside the ZureMap container via docker exec
    try:
        result = subprocess.run(
            ["docker", "exec", "zuremap", "az", "login",
             "--service-principal", "--username", cid,
             "--password", sec, "--tenant", tid, "--output", "none"],
            capture_output=True, text=True, timeout=30
        )
        if result.returncode == 0:
            # Ensure resource-graph extension is installed
            subprocess.run(
                ["docker", "exec", "zuremap", "az", "extension", "add",
                 "--name", "resource-graph", "--yes"],
                capture_output=True, text=True, timeout=30
            )
            # Apply the dark brand skin + de-branding + currency normalisation.
            global _archmap_rebranded
            _archmap_rebranded = True
            _archmap_rebrand()
            return {"ok": True, "message": "Architecture engine authenticated with service principal"}
        else:
            raise HTTPException(status_code=500, detail=f"az login failed: {result.stderr[:300]}")
    except FileNotFoundError:
        raise HTTPException(status_code=503, detail="Docker CLI not available on this host")
    except subprocess.TimeoutExpired:
        raise HTTPException(status_code=504, detail="az login timed out")


@app.post("/api/zuremap/rebrand")
async def zuremap_rebrand():
    """Re-apply the dark brand skin + de-branding to the embedded engine."""
    global _archmap_rebranded
    ok = await asyncio.to_thread(_archmap_rebrand)
    _archmap_rebranded = ok or _archmap_rebranded
    if not ok:
        raise HTTPException(status_code=503, detail="Architecture engine container not reachable")
    return {"ok": True, "message": "Brand skin applied"}


@app.post("/api/zuremap/session", include_in_schema=False)
async def zuremap_session(request: Request):
    """Issue the short-lived cookie that gates the same-origin /zuremap proxy.

    Reached only AFTER passing the /api Bearer gate (or in open/local mode), so the
    caller is already authenticated. The embedded engine's iframe then sends this
    cookie automatically on its /zuremap/* requests (it cannot send our Bearer).
    """
    cookie = auth_svc.make_zuremap_cookie()
    resp = JSONResponse({"ok": True})
    secure = request.url.scheme == "https"  # cookie stays usable over http on localhost
    resp.set_cookie("zm_sess", cookie, max_age=28800, httponly=True,
                    secure=secure, samesite="lax", path="/zuremap")
    return resp


@app.get("/api/ai/workload")
async def ai_workload_analysis(refresh: bool = False):
    """Holistic Claude-powered workload analysis (cached 6h)."""
    resources = _get_resources_list()
    if not resources:
        raise HTTPException(status_code=404, detail="No resource data — run a scan first")

    dep_summary  = get_graph_summary(_cache.get("dependency_graph")) if _cache.get("dependency_graph") else {}
    if hasattr(dep_summary, "model_dump"):
        dep_summary = dep_summary.model_dump()
    bcdr_summary = build_bcdr_dashboard_summary(_get_bcdr_assessments())
    custom_tags  = tagging_svc.get_all_custom_tags()

    return ai_infra_svc.analyze_workload(
        resources    = resources,
        dep_summary  = dep_summary if isinstance(dep_summary, dict) else {},
        bcdr_summary = bcdr_summary,
        custom_tags  = custom_tags,
        force_refresh= refresh,
    )


@app.get("/api/ai/workload/stream")
async def ai_workload_stream(refresh: bool = False):
    """Streaming Claude workload analysis via SSE."""
    resources = _get_resources_list()
    if not resources:
        raise HTTPException(status_code=404, detail="No resource data")

    dep_summary  = {}
    bcdr_summary = build_bcdr_dashboard_summary(_get_bcdr_assessments())

    async def _generate():
        loop = asyncio.get_event_loop()
        gen  = ai_infra_svc.stream_workload_analysis(resources, dep_summary, bcdr_summary)
        for chunk in gen:
            yield chunk

    return StreamingResponse(_generate(), media_type="text/event-stream",
                              headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


@app.get("/api/ai/resource/{resource_id:path}")
async def ai_resource_analysis(resource_id: str):
    """Deep-dive AI analysis for a single resource."""
    resources = _get_resources_list()
    resource  = next((r for r in resources if r.get("resource_id", "").lower() == resource_id.lower()), None)
    if not resource:
        raise HTTPException(status_code=404, detail="Resource not found")

    deps        = get_resource_dependencies(resource_id, _cache.get("dependency_graph")) if _cache.get("dependency_graph") else {}
    custom_tags = tagging_svc.get_custom_tags(resource_id)
    return ai_infra_svc.analyze_single_resource(resource, deps, custom_tags)


# ── Networking APIs ───────────────────────────────────────────────────────────

@app.get("/api/networking/dashboard")
async def networking_dashboard():
    """Comprehensive networking assessment dashboard."""
    from services.networking_service import build_networking_dashboard
    resources = _get_resources_list()
    if not resources:
        raise HTTPException(status_code=404, detail="No resource data. Run a scan first.")
    return build_networking_dashboard(resources)


@app.get("/api/networking/resource/{resource_id:path}")
async def get_networking_for_resource(resource_id: str):
    """Networking findings scoped to a single resource."""
    from services.networking_service import build_networking_dashboard
    resources = _get_resources_list()
    if not resources:
        raise HTTPException(status_code=404, detail="No resource data. Run a scan first.")

    rid_lower = resource_id.lower()
    resource_name = resource_id.split("/")[-1].lower()
    # Find the resource name from cache for better matching
    for r in resources:
        r_dict = r if isinstance(r, dict) else r.__dict__
        if r_dict.get("resource_id", "").lower() == rid_lower or \
           r_dict.get("resource_id", "").lower().endswith(rid_lower):
            resource_name = r_dict.get("resource_name", resource_name).lower()
            break

    net = build_networking_dashboard(resources)

    def _resource_matches(item: dict) -> bool:
        item_name = (item.get("resource_name") or item.get("name") or "").lower()
        item_id = (item.get("resource_id") or "").lower()
        return resource_name in item_name or item_name == resource_name or \
               (item_id and (item_id == rid_lower or item_id.endswith(rid_lower)))

    # Filter findings to those mentioning this resource
    findings = []
    for category in net.get("categories", []):
        for finding in category.get("findings", []):
            if isinstance(finding, dict) and _resource_matches(finding):
                findings.append({**finding, "category": category.get("name", "")})
    # Also check top-level findings list
    for finding in net.get("findings", []):
        if isinstance(finding, dict) and _resource_matches(finding):
            findings.append(finding)

    return {
        "resource_id": resource_id,
        "resource_name": resource_name,
        "findings": findings,
        "total": len(findings),
    }


@app.get("/api/networking/topology")
async def networking_topology():
    """Deep topology analysis via Azure Resource Graph."""
    from services.networking_service import build_advanced_topology
    sub_ids = None
    try:
        from services.azure_auth import get_credential
        cred = get_credential()
        if cred and hasattr(cred, '_subscription_ids'):
            sub_ids = cred._subscription_ids
    except Exception:
        pass
    if not sub_ids:
        data = _cache.get("latest")
        if data and data.resources:
            sub_ids = list(set(r.subscription_id for r in data.resources if r.subscription_id))
    return await asyncio.get_event_loop().run_in_executor(
        _pool, build_advanced_topology, sub_ids
    )


@app.get("/api/ai/networking")
async def ai_networking_analysis(refresh: bool = False):
    """AI-powered deep networking architecture and security analysis."""
    from services.networking_service import build_networking_dashboard, build_advanced_topology
    resources = _get_resources_list()
    if not resources:
        raise HTTPException(status_code=404, detail="No resource data. Run a scan first.")
    
    # First get the structured networking summary
    net_summary = build_networking_dashboard(resources)
    if net_summary.get("empty"):
        return {"error": "No networking resources found in the estate.", "available": False}
    
    # Get subscription IDs for deep topology query
    sub_ids = None
    data = _cache.get("latest")
    if data and data.resources:
        sub_ids = list(set(r.subscription_id for r in data.resources if r.subscription_id))

    # Build advanced topology via Azure Resource Graph
    topology = None
    try:
        topology = await asyncio.get_event_loop().run_in_executor(
            _pool, build_advanced_topology, sub_ids
        )
    except Exception as exc:
        logger.warning("Advanced topology fetch failed (non-fatal): %s", exc)

    # Feed into AI analysis with topology context
    return ai_infra_svc.analyze_networking_ai(net_summary, resources, force_refresh=refresh, topology=topology)


# ── Licensing & Reservation AI Analysis ──────────────────────────────────────

@app.get("/api/licensing/reservation-analysis")
async def licensing_reservation_analysis():
    """Comprehensive reservation coverage analysis with purchase planning."""
    resources = _get_resources_list()
    if not resources:
        raise HTTPException(status_code=404, detail="No resource data. Run a scan first.")
    return await asyncio.get_event_loop().run_in_executor(
        _pool, build_reservation_analysis, resources
    )


@app.get("/api/ai/licensing")
async def ai_licensing_analysis(refresh: bool = False):
    """AI-powered licensing & reservation optimization analysis."""
    resources = _get_resources_list()
    if not resources:
        raise HTTPException(status_code=404, detail="No resource data. Run a scan first.")

    # licensing_service functions expect ResourceMetrics objects (attribute access)
    resource_objects = _get_resources_objects()

    # Get licensing opportunities
    opps = detect_licensing_opportunities(resource_objects)
    opps_dicts = [o.dict() if hasattr(o, "dict") else o for o in opps]

    # Get reservation analysis
    res_analysis = build_reservation_analysis(resource_objects)

    # Feed into AI
    return ai_infra_svc.analyze_licensing_ai(
        opps_dicts, res_analysis, resources, force_refresh=refresh
    )


# ── Security Posture APIs (Defender for Cloud + Advisor) ─────────────────────

@app.get("/api/security/posture")
async def security_posture():
    """
    Full security posture from Microsoft Defender for Cloud and Azure Advisor.
    Includes secure score, assessments, recommendations, and compliance data.
    """
    try:
        data = await asyncio.get_event_loop().run_in_executor(
            _pool, get_full_security_posture
        )
        return data
    except Exception as exc:
        logger.error("Security posture fetch failed: %s", exc)
        raise HTTPException(status_code=500, detail=str(exc))


@app.get("/api/security/enhanced")
async def security_enhanced():
    """
    Enhanced security data combining internal gap analysis with Defender findings.
    Merges our heuristic gaps with real Azure security signals.
    """
    try:
        # Get internal security gaps from cached dashboard
        internal_gaps = []
        cached = _cache.get("dashboard_data")
        if cached and hasattr(cached, "security_gaps"):
            internal_gaps = [g.dict() if hasattr(g, "dict") else g for g in cached.security_gaps]
        elif cached and isinstance(cached, dict):
            internal_gaps = cached.get("security_gaps", [])

        # Get Defender + Advisor findings
        defender_data = await asyncio.get_event_loop().run_in_executor(
            _pool, get_full_security_posture
        )

        # Get Arc machines for hybrid security (cached 15 min)
        arc_security = []
        if arc_svc:
            _arc_cached = _arc_security_cache.get("data")
            if _arc_cached is not None and (_time_mod.monotonic() - _arc_security_cache.get("ts", 0)) < _ARC_SECURITY_TTL:
                arc_security = _arc_cached
            else:
                try:
                    arc_data = await asyncio.get_event_loop().run_in_executor(
                        _pool, arc_svc.get_arc_summary
                    )
                    arc_machines = arc_data.get("machines", [])
                    for m in arc_machines:
                        coverage = m.get("coverage", {})
                        if not coverage.get("security"):
                            arc_security.append({
                                "id": m.get("id", ""),
                                "title": "No Security Extension on Arc Machine",
                                "description": f"Arc machine '{m.get('name')}' ({m.get('osName', 'Unknown OS')}) has no security monitoring agent installed.",
                                "severity": "high",
                                "category": "Arc Security",
                                "resource_id": m.get("id", ""),
                                "resource_name": m.get("name", ""),
                                "resource_type": "microsoft.hybridcompute/machines",
                                "resource_group": m.get("resourceGroup", ""),
                                "subscription_id": m.get("subscriptionId", ""),
                                "source": "arc_analysis",
                                "remediation": "Install Microsoft Defender for Servers extension on this Arc-enabled machine.",
                                "status": "Active",
                            })
                        if m.get("status", "").lower() != "connected":
                            arc_security.append({
                                "id": m.get("id", "") + "_disconnected",
                                "title": "Arc Machine Disconnected",
                                "description": f"Arc machine '{m.get('name')}' is {m.get('status', 'unknown')} — cannot receive security updates or policies.",
                                "severity": "critical" if m.get("status", "").lower() == "expired" else "high",
                                "category": "Arc Connectivity",
                                "resource_id": m.get("id", ""),
                                "resource_name": m.get("name", ""),
                                "resource_type": "microsoft.hybridcompute/machines",
                                "resource_group": m.get("resourceGroup", ""),
                                "subscription_id": m.get("subscriptionId", ""),
                                "source": "arc_analysis",
                                "remediation": "Reconnect the Arc agent. Check network connectivity to Azure endpoints.",
                                "status": "Active",
                            })
                    _arc_security_cache["data"] = arc_security
                    _arc_security_cache["ts"] = _time_mod.monotonic()
                except Exception as arc_exc:
                    logger.warning("Arc security enrichment failed: %s", arc_exc)

        # ── Inject on-premises security findings ──
        onprem_findings = []
        try:
            from services.onprem_service import generate_onprem_security_findings
            onprem_findings = generate_onprem_security_findings()
        except Exception as ope:
            logger.warning("On-prem security enrichment failed: %s", ope)

        # Trim heavy fields — SecurityPanel only uses findings (top 500),
        # alerts (first 3), secure_score, and defender_plans.
        # sub_assessments, regulatory_controls, controls are not rendered here.
        slim_defender = {
            "secure_score":          defender_data.get("secure_score"),
            "total_recommendations": defender_data.get("total_recommendations"),
            "total_alerts":          defender_data.get("total_alerts"),
            "total_vulnerabilities": defender_data.get("total_vulnerabilities"),
            "severity_counts":       defender_data.get("severity_counts"),
            "alert_severity_counts": defender_data.get("alert_severity_counts"),
            "findings":              defender_data.get("findings", [])[:500],
            "finding_count":         defender_data.get("finding_count"),
            "alerts":                defender_data.get("alerts", [])[:20],
            "defender_plans":        defender_data.get("defender_plans"),
            "compliance":            defender_data.get("compliance"),
            "advisor_count":         defender_data.get("advisor_count"),
            "defender_count":        defender_data.get("defender_count"),
            "charts":                defender_data.get("charts"),
        }
        response = {
            "internal_gaps": internal_gaps,
            "defender": slim_defender,
            "arc_findings": arc_security,
            "onprem_findings": onprem_findings,
            "total_findings": len(internal_gaps) + slim_defender.get("finding_count", 0) + len(arc_security) + len(onprem_findings),
        }

        # Auto-persist findings to DB (fire-and-forget)
        try:
            all_for_persist = []
            for g in internal_gaps:
                gd = dict(g) if not isinstance(g, dict) else g
                gd.setdefault("source", "internal")
                gd.setdefault("category", gd.get("gap_type", ""))
                all_for_persist.append(gd)
            for f in slim_defender.get("findings", []):
                all_for_persist.append(f if isinstance(f, dict) else dict(f))
            for f in arc_security:
                all_for_persist.append(f if isinstance(f, dict) else dict(f))
            for f in onprem_findings:
                all_for_persist.append(f if isinstance(f, dict) else dict(f))
            if all_for_persist:
                await asyncio.get_event_loop().run_in_executor(
                    _pool, lambda: persist_security_findings(all_for_persist, "auto")
                )
        except Exception as pe:
            logger.warning("Auto-persist security findings failed (non-critical): %s", pe)

        return response
    except Exception as exc:
        logger.error("Enhanced security fetch failed: %s", exc)
        raise HTTPException(status_code=500, detail=str(exc))


@app.get("/api/security/resource/{resource_id:path}")
async def get_security_for_resource(resource_id: str):
    """Security findings scoped to a single resource (by resource_id or name suffix)."""
    try:
        # Internal gaps from dashboard cache
        internal_gaps = []
        cached = _cache.get("dashboard_data")
        if cached and hasattr(cached, "security_gaps"):
            all_gaps = [g.dict() if hasattr(g, "dict") else g for g in cached.security_gaps]
        elif _cache.get("data"):
            from services.security_service import identify_security_gaps
            res_dicts = [r.model_dump() for r in _cache["data"].resources]
            all_gaps = identify_security_gaps(res_dicts)
            all_gaps = [g if isinstance(g, dict) else g.__dict__ for g in all_gaps]
        else:
            all_gaps = []
        rid_lower = resource_id.lower()
        for gap in all_gaps:
            gid = (gap.get("resource_id") or "").lower()
            gname = (gap.get("resource_name") or "").lower()
            if (gid and (gid == rid_lower or gid.endswith(rid_lower) or rid_lower.endswith(gid[-30:]))) or \
               (gname and gname in rid_lower):
                internal_gaps.append(gap)

        # Defender findings
        defender_findings = []
        try:
            defender_data = await asyncio.get_event_loop().run_in_executor(_pool, get_full_security_posture)
            for f in defender_data.get("findings", []):
                fid = (f.get("resource_id") or "").lower()
                fname = (f.get("resource_name") or "").lower()
                if (fid and (fid == rid_lower or fid.endswith(rid_lower))) or \
                   (fname and fname in rid_lower):
                    defender_findings.append(f)
        except Exception:
            pass

        return {
            "resource_id": resource_id,
            "internal_gaps": internal_gaps,
            "defender_findings": defender_findings,
            "total_findings": len(internal_gaps) + len(defender_findings),
        }
    except Exception as exc:
        logger.error("Security resource fetch failed: %s", exc)
        raise HTTPException(status_code=500, detail=str(exc))


@app.get("/api/security/zero-trust")
async def security_zero_trust():
    """Zero Trust scorecard across 6 pillars."""
    try:
        cached = _cache.get("data:*") or _cache.get("data")
        if not cached:
            raise HTTPException(status_code=404, detail="No scan data available")
        resources = cached.resources if hasattr(cached, "resources") else cached.get("resources", [])
        gaps = cached.security_gaps if hasattr(cached, "security_gaps") else cached.get("security_gaps", [])

        # Try to get Defender data for enrichment
        defender_info = {}
        try:
            defender_data = await asyncio.get_event_loop().run_in_executor(
                _pool, get_full_security_posture
            )
            defender_info = {
                "secure_score": defender_data.get("secure_score"),
                "defender_plans": defender_data.get("defender_plans"),
            }
        except Exception:
            pass

        result = build_zero_trust_scorecard(resources, gaps, defender_info)
        return result
    except HTTPException:
        raise
    except Exception as exc:
        logger.error("Zero Trust scorecard failed: %s", exc)
        raise HTTPException(status_code=500, detail=str(exc))


@app.get("/api/security/attack-surface")
async def security_attack_surface():
    """Attack surface analysis — public endpoints, exposed services."""
    try:
        cached = _cache.get("data:*") or _cache.get("data")
        if not cached:
            raise HTTPException(status_code=404, detail="No scan data available")
        resources = cached.resources if hasattr(cached, "resources") else cached.get("resources", [])
        gaps = cached.security_gaps if hasattr(cached, "security_gaps") else cached.get("security_gaps", [])
        result = build_attack_surface_analysis(resources, gaps)
        return result
    except HTTPException:
        raise
    except Exception as exc:
        logger.error("Attack surface analysis failed: %s", exc)
        raise HTTPException(status_code=500, detail=str(exc))


# ── Security Findings (DB-persisted) ─────────────────────────────────────────

@app.get("/api/security/findings")
async def list_security_findings(
    severity: str = "all",
    source: str = "all",
    resource_type: str = "all",
    resource_group: str = "all",
    subscription: str = "all",
    category: str = "all",
    status: str = "active",
    search: str = "",
    sort_by: str = "severity",
    sort_dir: str = "asc",
    page: int = 0,
    page_size: int = 50,
):
    """
    Query persisted security findings with server-side filtering, sorting, and pagination.
    Supports all filter dimensions: severity, source, resource_type, resource_group, subscription, category.
    """
    try:
        result = await asyncio.get_event_loop().run_in_executor(
            _pool,
            lambda: query_security_findings(
                severity=severity if severity != "all" else None,
                source=source if source != "all" else None,
                resource_type=resource_type if resource_type != "all" else None,
                resource_group=resource_group if resource_group != "all" else None,
                subscription=subscription if subscription != "all" else None,
                category=category if category != "all" else None,
                status=status if status != "all" else None,
                search=search or None,
                sort_by=sort_by,
                sort_dir=sort_dir,
                page=page,
                page_size=min(page_size, 200),
            )
        )
        return result
    except Exception as exc:
        logger.error("Security findings query failed: %s", exc)
        raise HTTPException(status_code=500, detail=str(exc))


@app.get("/api/security/findings/export")
async def export_security_findings(
    severity: str = "all",
    source: str = "all",
    resource_type: str = "all",
    resource_group: str = "all",
    subscription: str = "all",
    category: str = "all",
    status: str = "active",
    search: str = "",
    format: str = "csv",
):
    """Export filtered security findings as CSV."""
    try:
        csv_data = await asyncio.get_event_loop().run_in_executor(
            _pool,
            lambda: export_security_csv(
                severity=severity if severity != "all" else None,
                source=source if source != "all" else None,
                resource_type=resource_type if resource_type != "all" else None,
                resource_group=resource_group if resource_group != "all" else None,
                subscription=subscription if subscription != "all" else None,
                category=category if category != "all" else None,
                status=status if status != "all" else None,
                search=search or None,
            )
        )
        from starlette.responses import Response
        return Response(
            content=csv_data,
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=security_findings.csv"},
        )
    except Exception as exc:
        logger.error("Security findings export failed: %s", exc)
        raise HTTPException(status_code=500, detail=str(exc))


@app.post("/api/security/findings/refresh")
async def refresh_security_findings():
    """
    Re-scan and persist security findings from all sources (internal gaps,
    Defender for Cloud, Azure Advisor, Arc).
    """
    try:
        all_findings = []

        # 1. Internal gaps from cached dashboard
        cached = _cache.get("dashboard_data")
        if cached:
            gaps = cached.security_gaps if hasattr(cached, "security_gaps") else cached.get("security_gaps", [])
            for g in gaps:
                gd = g.dict() if hasattr(g, "dict") else (dict(g) if not isinstance(g, dict) else g)
                gd["source"] = "internal"
                gd["category"] = gd.get("gap_type", "")
                all_findings.append(gd)

        # 2. Defender + Advisor findings
        try:
            defender_data = await asyncio.get_event_loop().run_in_executor(
                _pool, get_full_security_posture
            )
            for f in defender_data.get("findings", []):
                all_findings.append(f if isinstance(f, dict) else dict(f))
        except Exception as de:
            logger.warning("Defender findings not available for persist: %s", de)

        # 3. Arc findings
        if arc_svc:
            try:
                arc_data = await asyncio.get_event_loop().run_in_executor(
                    _pool, arc_svc.get_arc_summary
                )
                for m in arc_data.get("machines", []):
                    coverage = m.get("coverage", {})
                    if not coverage.get("security"):
                        all_findings.append({
                            "title": "No Security Extension on Arc Machine",
                            "description": f"Arc machine '{m.get('name')}' has no security monitoring agent.",
                            "severity": "high",
                            "category": "Arc Security",
                            "resource_id": m.get("id", ""),
                            "resource_name": m.get("name", ""),
                            "resource_type": "microsoft.hybridcompute/machines",
                            "resource_group": m.get("resourceGroup", ""),
                            "subscription_id": m.get("subscriptionId", ""),
                            "source": "arc_analysis",
                            "remediation": "Install Microsoft Defender for Servers extension.",
                        })
            except Exception as ae:
                logger.warning("Arc findings not available for persist: %s", ae)

        # Persist to DB
        result = await asyncio.get_event_loop().run_in_executor(
            _pool, lambda: persist_security_findings(all_findings)
        )
        return result
    except Exception as exc:
        logger.error("Security findings refresh failed: %s", exc)
        raise HTTPException(status_code=500, detail=str(exc))


@app.get("/api/security/findings/summary")
async def security_findings_summary():
    """Get aggregate statistics for active security findings."""
    try:
        result = await asyncio.get_event_loop().run_in_executor(
            _pool, get_security_summary
        )
        return result
    except Exception as exc:
        logger.error("Security findings summary failed: %s", exc)
        raise HTTPException(status_code=500, detail=str(exc))


@app.get("/api/security/findings/scans")
async def security_scan_history(limit: int = 10):
    """Get recent security scan history."""
    try:
        result = await asyncio.get_event_loop().run_in_executor(
            _pool, lambda: get_security_scan_history(min(limit, 50))
        )
        return result
    except Exception as exc:
        logger.error("Security scan history failed: %s", exc)
        raise HTTPException(status_code=500, detail=str(exc))


# ── Enhanced Backup & DR APIs ────────────────────────────────────────────────

@app.get("/api/backup/enhanced")
async def backup_enhanced(refresh: bool = False):
    """
    Comprehensive backup & DR analysis via Resource Graph.
    Includes vault health, RoC advisory, unprotected resources,
    backup job failures, ASR replication status, and chart data.
    Results are cached server-side (15-min TTL). Pass ?refresh=true to force.
    """
    try:
        data = await asyncio.get_event_loop().run_in_executor(
            _pool, lambda: get_enhanced_backup_analysis(refresh=refresh)
        )
        return data
    except Exception as exc:
        logger.error("Enhanced backup analysis failed: %s", exc)
        raise HTTPException(status_code=500, detail=str(exc))


@app.get("/api/backup/rpo-rto-matrix")
async def backup_rpo_rto_matrix():
    """RPO/RTO compliance matrix for all backup-eligible resources."""
    try:
        enhanced = await asyncio.get_event_loop().run_in_executor(
            _pool, lambda: get_enhanced_backup_analysis(refresh=False)
        )
        return build_rpo_rto_matrix(enhanced)
    except Exception as exc:
        logger.error("RPO/RTO matrix failed: %s", exc)
        raise HTTPException(status_code=500, detail=str(exc))

@app.get("/api/backup/ransomware-readiness")
async def backup_ransomware_readiness():
    """Ransomware readiness assessment based on vault security posture."""
    try:
        enhanced = await asyncio.get_event_loop().run_in_executor(
            _pool, lambda: get_enhanced_backup_analysis(refresh=False)
        )
        return build_ransomware_readiness(enhanced)
    except Exception as exc:
        logger.error("Ransomware readiness failed: %s", exc)
        raise HTTPException(status_code=500, detail=str(exc))


@app.get("/api/resilience/analysis")
async def resilience_analysis():
    """
    Server-side resilience analysis with full resource attribution.
    Runs the same 6 rules as the client-side ResiliencePanel but enriches
    each finding with complete resource_id, resource_group, subscription_id
    so findings can link to the Resource Detail Drawer.
    """
    try:
        from services.resilience_service import analyze_resilience
        resources = _get_resources_list()
        if not resources:
            return {"score": 0, "risk_level": "Unknown", "total_findings": 0, "findings": []}
        return analyze_resilience(resources)
    except Exception as exc:
        logger.error("Resilience analysis failed: %s", exc)
        raise HTTPException(status_code=500, detail=str(exc))

# ── Azure Arc / On-Premise APIs ──────────────────────────────────────────────

@app.get("/api/arc/summary")
async def arc_summary():
    """Comprehensive Azure Arc on-premise infrastructure summary."""
    if not arc_svc:
        raise HTTPException(status_code=503, detail="Arc service unavailable")
    try:
        data = await asyncio.get_event_loop().run_in_executor(
            _pool, arc_svc.get_arc_summary
        )
        return data
    except Exception as exc:
        logger.error("Arc summary failed: %s", exc)
        raise HTTPException(status_code=500, detail=str(exc))


@app.get("/api/arc/machines")
async def arc_machines(
    subscription: Optional[str] = None,
    resource_group: Optional[str] = None,
    os_type: Optional[str] = None,
    status: Optional[str] = None,
):
    """List Arc machines with optional filters."""
    if not arc_svc:
        raise HTTPException(status_code=503, detail="Arc service unavailable")
    try:
        sub_ids = [subscription] if subscription else None
        rgs = [resource_group] if resource_group else None
        machines = await asyncio.get_event_loop().run_in_executor(
            _pool,
            partial(arc_svc.get_arc_machines_filtered,
                    subscription_ids=sub_ids, resource_groups=rgs,
                    os_type=os_type, status=status)
        )
        return {"machines": machines, "total": len(machines)}
    except Exception as exc:
        logger.error("Arc machines failed: %s", exc)
        raise HTTPException(status_code=500, detail=str(exc))


@app.get("/api/arc/sql")
async def arc_sql_overview():
    """Arc SQL Server instances, databases, and availability groups."""
    if not arc_svc:
        raise HTTPException(status_code=503, detail="Arc service unavailable")
    try:
        instances = await asyncio.get_event_loop().run_in_executor(
            _pool, arc_svc.discover_arc_sql_instances
        )
        databases = await asyncio.get_event_loop().run_in_executor(
            _pool, arc_svc.discover_arc_sql_databases
        )
        ags = await asyncio.get_event_loop().run_in_executor(
            _pool, arc_svc.discover_arc_sql_availability_groups
        )
        # Enrich instances with their databases
        db_map = {}
        for db in databases:
            inst_id = (db.get("sqlInstanceId") or "").lower()
            if inst_id:
                db_map.setdefault(inst_id, []).append(db)
        for inst in instances:
            inst["databases"] = db_map.get(inst.get("id", "").lower(), [])
        return {
            "instances": instances,
            "databases": databases,
            "availability_groups": ags,
            "total_instances": len(instances),
            "total_databases": len(databases),
            "total_ags": len(ags),
        }
    except Exception as exc:
        logger.error("Arc SQL failed: %s", exc)
        raise HTTPException(status_code=500, detail=str(exc))


@app.post("/api/arc/ai-analysis")
async def arc_ai_analysis(request: Request):
    """AI-powered analysis for Arc machines — security, BCDR, governance recommendations."""
    if not ai_infra_svc:
        raise HTTPException(status_code=503, detail="AI service unavailable")
    if not arc_svc:
        raise HTTPException(status_code=503, detail="Arc service unavailable")
    body = await request.json()
    machine_ids = body.get("machine_ids", [])
    try:
        summary = await asyncio.get_event_loop().run_in_executor(
            _pool, arc_svc.get_arc_summary
        )
        # Filter to selected machines if specified
        machines = summary.get("machines", [])
        if machine_ids:
            id_set = set(mid.lower() for mid in machine_ids)
            machines = [m for m in machines if m.get("id", "").lower() in id_set]

        # Build context for AI
        machine_context = []
        for m in machines[:50]:  # cap for token limits
            mc = {
                "name": m.get("name"),
                "os": m.get("osName") or m.get("osType"),
                "os_version": m.get("osVersion"),
                "status": m.get("status"),
                "location": m.get("location"),
                "cores": m.get("cores"),
                "memory_gb": round(m.get("totalMemoryGB") or 0, 1),
                "domain": m.get("domainName"),
                "agent_version": m.get("agentVersion"),
                "extensions": [c["label"] for c in m.get("classified_extensions", [])],
                "coverage": m.get("coverage", {}),
                "sql_instances": len(m.get("sql_instances", [])),
                "tags": m.get("tags", {}),
            }
            machine_context.append(mc)

        bcdr = summary.get("bcdr", {})
        coverage = summary.get("coverage", {})

        system_prompt = """You are an expert Azure Arc and on-premise infrastructure analyst.
Analyze the provided Arc-enabled machines and provide comprehensive recommendations.
Return JSON with this structure:
{
  "overall_health": "Critical|At Risk|Fair|Good|Excellent",
  "risk_score": <0-100>,
  "executive_summary": "<2-3 sentence overview>",
  "findings": [
    {"category": "Security|Monitoring|BCDR|Governance|Patching|Performance",
     "severity": "Critical|High|Medium|Low",
     "finding": "<description>",
     "recommendation": "<actionable fix>",
     "affected_machines": <count>}
  ],
  "bcdr_assessment": {
    "readiness_level": "Critical|Low|Medium|High",
    "summary": "<BCDR readiness summary>",
    "recommendations": ["<list of BCDR recommendations>"]
  },
  "quick_wins": [
    {"title": "<quick win>", "description": "<details>", "impact": "High|Medium|Low",
     "effort": "Low|Medium|High", "affected_machines": <count>}
  ],
  "modernization_opportunities": [
    {"title": "<opportunity>", "description": "<details>", "benefit": "<expected benefit>"}
  ]
}"""

        import json as _json
        user_prompt = f"""Analyze these {len(machine_context)} Azure Arc-enabled on-premise machines:

MACHINES:
{_json.dumps(machine_context, indent=2)}

COVERAGE STATS:
- Monitoring: {coverage.get('monitoring_pct', 0)}%
- Security: {coverage.get('security_pct', 0)}%
- Patching: {coverage.get('patching_pct', 0)}%
- Change Tracking: {coverage.get('change_tracking_pct', 0)}%

BCDR SUMMARY:
- Overall Score: {bcdr.get('overall_score', 0)}
- Databases with backup: {bcdr.get('databases_with_backup', 0)}/{bcdr.get('total_databases', 0)}
- AG Coverage: {bcdr.get('ag_coverage_pct', 0)}%
- Known Risks: {len(bcdr.get('risks', []))}

Provide a thorough analysis covering security posture, monitoring gaps, BCDR readiness,
governance compliance, and modernization opportunities."""

        result_text = await asyncio.get_event_loop().run_in_executor(
            _pool,
            partial(ai_infra_svc._call_ai, system_prompt, user_prompt, 4000)
        )
        try:
            result = _json.loads(result_text)
        except _json.JSONDecodeError:
            result = {"raw_analysis": result_text}
        return result
    except Exception as exc:
        logger.error("Arc AI analysis failed: %s", exc)
        raise HTTPException(status_code=500, detail=str(exc))


@app.get("/api/ai/dependency/{resource_id:path}")
async def ai_dependency_analysis(resource_id: str):
    """AI blast-radius and SPOF analysis for a resource."""
    resources = _get_resources_list()
    resource  = next((r for r in resources if r.get("resource_id", "").lower() == resource_id.lower()), None)
    if not resource:
        raise HTTPException(status_code=404, detail="Resource not found")

    graph = _cache.get("dependency_graph")
    blast = get_blast_radius(resource_id, graph) if graph else {}
    dep_summary = get_graph_summary(graph) if graph else {}
    if hasattr(dep_summary, "model_dump"):
        dep_summary = dep_summary.model_dump()
    return ai_infra_svc.analyze_dependency_impact(resource, blast, dep_summary)


@app.get("/api/ai/roadmap")
async def ai_optimization_roadmap(refresh: bool = False):
    """AI-generated prioritized optimization roadmap."""
    resources = _get_resources_list()
    if not resources:
        raise HTTPException(status_code=404, detail="No resource data")
    return ai_infra_svc.generate_optimization_roadmap(resources, force_refresh=refresh)


@app.get("/api/ai/cloud-adoption")
async def ai_cloud_adoption_analysis(refresh: bool = False):
    """AI-generated cloud adoption & modernization analysis with migration paths."""
    resources = _get_resources_list()
    if not resources:
        raise HTTPException(status_code=404, detail="No resource data — run a scan first")

    # Get ACR opportunities for context
    acr_opps = None
    if _cache.get("dashboard_data") and hasattr(_cache["dashboard_data"], "acr_opportunities"):
        ao = _cache["dashboard_data"].acr_opportunities
        if ao:
            acr_opps = ao.dict() if hasattr(ao, "dict") else ao

    return ai_infra_svc.analyze_cloud_adoption(
        resources=resources,
        acr_opportunities=acr_opps,
        force_refresh=refresh,
    )


@app.get("/api/ai/bcdr")
async def ai_bcdr_analysis(refresh: bool = False, mode: str = "comprehensive"):
    """
    AI-generated BCDR analysis.
    
    Modes:
    - comprehensive: Full environment analysis with scoring and roadmap (default)
    - resources: Individual resource recommendations (legacy 19-column format)
    """
    resources = _get_resources_list()
    if not resources:
        raise HTTPException(status_code=404, detail="No resource data")
    
    if mode == "comprehensive":
        # Full environment analysis with AI
        report = ai_infra_svc.analyze_environment_bcdr(resources, force_refresh=refresh)
        return report
    else:
        # Legacy mode: individual resource analysis
        assessments = _get_bcdr_assessments()
        zone_dicts  = [a.__dict__ if hasattr(a, "__dict__") else a for a in assessments]
        items = ai_infra_svc.analyze_resource_bcdr(resources, zone_dicts, force_refresh=refresh)
        return {"items": items, "total": len(items), "ai_generated": True}


@app.post("/api/ai/search")
async def ai_semantic_search(body: dict):
    """Semantic / natural language resource search via AI."""
    query     = body.get("query", "").strip()
    if not query:
        raise HTTPException(status_code=400, detail="query is required")
    resources = _get_resources_list()
    top_k     = min(body.get("top_k", 20), 50)
    results   = ai_infra_svc.semantic_search_resources(query, resources, top_k)
    return {"query": query, "results": results, "total": len(results)}


# ── AI Module Analysis Endpoints ──────────────────────────────────────────────

def _get_arc_data() -> dict:
    """Helper to get Arc summary data for AI analysis."""
    try:
        return arc_svc.get_arc_summary()
    except Exception as e:
        logger.warning("Failed to get Arc data for AI analysis: %s", e)
        return {}

@app.get("/api/ai/maturity")
async def ai_maturity_analysis(refresh: bool = False):
    """AI-powered cloud maturity analysis across Azure + Arc estate."""
    from services.ai_module_analysis_service import analyze_cloud_maturity_ai
    resources = _get_resources_list()
    if not resources:
        raise HTTPException(status_code=404, detail="No resource data")
    arc_data = _get_arc_data()
    return analyze_cloud_maturity_ai(resources, arc_data=arc_data, force_refresh=refresh)

@app.get("/api/ai/security")
async def ai_security_analysis(refresh: bool = False):
    """AI-powered security posture analysis across Azure + Arc estate."""
    from services.ai_module_analysis_service import analyze_security_ai
    resources = _get_resources_list()
    if not resources:
        raise HTTPException(status_code=404, detail="No resource data")
    arc_data = _get_arc_data()
    # Get Defender data too
    defender_data = None
    try:
        from services.defender_service import get_full_security_posture
        defender_data = get_full_security_posture()
    except Exception as e:
        logger.warning("Defender data unavailable for AI security: %s", e)
    return analyze_security_ai(resources, arc_data=arc_data, defender_data=defender_data, force_refresh=refresh)

@app.get("/api/ai/innovation")
async def ai_innovation_analysis(refresh: bool = False):
    """AI-powered innovation opportunity analysis across Azure + Arc estate."""
    from services.ai_module_analysis_service import analyze_innovation_ai
    resources = _get_resources_list()
    if not resources:
        raise HTTPException(status_code=404, detail="No resource data")
    arc_data = _get_arc_data()
    return analyze_innovation_ai(resources, arc_data=arc_data, force_refresh=refresh)

@app.get("/api/ai/migration")
async def ai_migration_analysis(refresh: bool = False):
    """AI-powered migration & modernization analysis across Azure + Arc estate."""
    import asyncio
    from services.ai_module_analysis_service import analyze_migration_ai
    resources = _get_resources_list()
    if not resources:
        raise HTTPException(status_code=404, detail="No resource data")
    arc_data = _get_arc_data()
    return await asyncio.to_thread(analyze_migration_ai, resources, arc_data=arc_data, force_refresh=refresh)


@app.get("/api/migration/assessment")
async def migration_assessment():
    """Comprehensive migration assessment with 5R classification, wave planning, and risk scoring."""
    import traceback as _tb
    try:
        data: Optional[DashboardData] = _cache.get("data:*") or _cache.get("data")
        if not data or not data.resources:
            raise HTTPException(status_code=404, detail="No resource data")
        dep_edges = None
        try:
            dep_graph = _cache.get("dependency_graph")
            if dep_graph and hasattr(dep_graph, "edges"):
                dep_edges = [{"source": e.source, "target": e.target} for e in dep_graph.edges]
        except Exception:
            pass
        result = build_migration_assessment(data.resources, dependency_edges=dep_edges)
        return result.model_dump(mode="json")
    except HTTPException:
        raise
    except Exception as exc:
        logger.error("migration_assessment failed: %s\n%s", exc, _tb.format_exc())
        raise HTTPException(status_code=500, detail=str(exc))

@app.get("/api/ai/backup")
async def ai_backup_analysis(refresh: bool = False):
    """AI-powered backup state analysis with recommendations."""
    from services.ai_module_analysis_service import analyze_backup_ai
    resources = _get_resources_list()
    if not resources:
        raise HTTPException(status_code=404, detail="No resource data")
    arc_data = _get_arc_data()
    backup_coverage = None
    if _cache.get("dashboard_data") and hasattr(_cache["dashboard_data"], "backup_coverage"):
        bc = _cache["dashboard_data"].backup_coverage
        backup_coverage = bc.dict() if hasattr(bc, "dict") else bc
    return analyze_backup_ai(resources, arc_data=arc_data, backup_coverage=backup_coverage, force_refresh=refresh)

@app.get("/api/ai/resilience")
async def ai_resilience_analysis(refresh: bool = False):
    """AI-powered resilience analysis of entire Azure estate."""
    from services.ai_module_analysis_service import analyze_resilience_ai
    resources = _get_resources_list()
    if not resources:
        raise HTTPException(status_code=404, detail="No resource data")
    arc_data = _get_arc_data()
    return analyze_resilience_ai(resources, arc_data=arc_data, force_refresh=refresh)

@app.get("/api/bcdr/avs/inventory")
async def bcdr_avs_inventory():
    """Detect AVS resources in the environment and return inventory (no AI)."""
    resources = _get_resources_list()
    if not resources:
        return {"avs_found": False, "private_clouds": [], "related_resources": []}

    resource_dicts = [r if isinstance(r, dict) else r.__dict__ for r in resources]
    private_clouds = []
    related = []
    for r in resource_dicts:
        rtype = (r.get("resource_type") or "").lower()
        rname = (r.get("resource_name") or "").lower()
        if "microsoft.avs" in rtype:
            private_clouds.append({
                "resource_name": r.get("resource_name", ""),
                "resource_type": r.get("resource_type", ""),
                "location": r.get("location", ""),
                "resource_group": r.get("resource_group", ""),
                "subscription_name": r.get("subscription_name", ""),
                "sku": r.get("sku", ""),
                "properties": r.get("properties", {}),
                "tags": r.get("tags", {}),
            })
        elif "vmware" in rname or "avs" in rname or "hcx" in rname:
            related.append({
                "resource_name": r.get("resource_name", ""),
                "resource_type": r.get("resource_type", ""),
                "location": r.get("location", ""),
                "resource_group": r.get("resource_group", ""),
            })
    # Also find ExpressRoute circuits (commonly used with AVS)
    expressroutes = [r for r in resource_dicts if "expressroutecircuit" in (r.get("resource_type") or "").lower()]
    for er in expressroutes:
        related.append({
            "resource_name": er.get("resource_name", ""),
            "resource_type": er.get("resource_type", ""),
            "location": er.get("location", ""),
            "resource_group": er.get("resource_group", ""),
        })
    return {
        "avs_found": len(private_clouds) > 0,
        "private_clouds": private_clouds,
        "related_resources": related,
        "total_avs_nodes": sum(1 for _ in private_clouds),
    }


@app.get("/api/ai/bcdr/avs")
async def ai_bcdr_avs_analysis(refresh: bool = False):
    """AI-powered Azure VMware Solution DR analysis."""
    from services.ai_module_analysis_service import analyze_bcdr_avs
    resources = _get_resources_list()
    if not resources:
        raise HTTPException(status_code=404, detail="No resource data")
    arc_data = _get_arc_data()
    return analyze_bcdr_avs(resources, arc_data=arc_data, force_refresh=refresh)

@app.get("/api/ai/bcdr/deep")
async def ai_bcdr_deep_analysis(refresh: bool = False):
    """Deep AI-powered BCDR analysis covering entire Azure estate with maximum detail."""
    from services.ai_module_analysis_service import analyze_bcdr_deep
    resources = _get_resources_list()
    if not resources:
        raise HTTPException(status_code=404, detail="No resource data")
    arc_data = _get_arc_data()
    return analyze_bcdr_deep(resources, arc_data=arc_data, force_refresh=refresh)


# ── Resource Snapshots API ────────────────────────────────────────────────────

@app.get("/api/snapshots/{resource_id:path}")
async def get_resource_snapshots(resource_id: str, limit: int = 20):
    """Get historical snapshots for a resource."""
    return tagging_svc.get_resource_snapshots(resource_id, limit)


async def health():
    active = get_active_provider()
    dep_graph = _cache.get("dependency_graph")
    return {
        "status":      "ok",
        "ai_enabled":  active != "none",
        "ai_provider": active,
        "demo_mode":   settings_svc.get_value("demo_mode", False),
        "dependency_graph": {
            "nodes": dep_graph.node_count if dep_graph else 0,
            "edges": dep_graph.edge_count if dep_graph else 0,
            "clusters": dep_graph.cluster_count if dep_graph else 0,
        },
        "timestamp":   datetime.now(tz=timezone.utc).isoformat(),
    }


@app.get("/api/openai-deployments")
async def openai_deployments(subscription_id: str, resource_group: str, account_name: str):
    from services.azure_auth import get_credential
    credential = get_credential()
    return get_openai_deployments(credential, subscription_id, resource_group, account_name)


# ── APEX Integration Endpoints ─────────────────────────────────────────────────
# Phase 2 implementation workflow: APEX agents, MCP servers, artifacts

from services.apex_agent_service import get_apex_agent_service
from services.mcp_service import get_mcp_service

# Projects API (extended for APEX workflow)

class ProjectCreate(BaseModel):
    project_name: str
    description: Optional[str] = None
    business_unit: Optional[str] = None
    criticality: Optional[str] = None
    rto_target: Optional[str] = None
    rpo_target: Optional[str] = None
    environment: Optional[str] = None
    dr_tier: Optional[str] = None
    owner: Optional[str] = None

class ProjectUpdate(BaseModel):
    project_name: Optional[str] = None
    description: Optional[str] = None
    business_unit: Optional[str] = None
    criticality: Optional[str] = None
    rto_target: Optional[str] = None
    rpo_target: Optional[str] = None
    environment: Optional[str] = None
    dr_tier: Optional[str] = None
    owner: Optional[str] = None

class AgentExecutionRequest(BaseModel):
    agent_name: str
    user_input: str
    context: Optional[Dict[str, Any]] = None

class PricingRequest(BaseModel):
    resource_type: str
    region: str
    sku: Optional[str] = None

class ArchitecturePricingRequest(BaseModel):
    resources: List[Dict[str, Any]]

class DiagramRequest(BaseModel):
    architecture: Dict[str, Any]
    diagram_type: str = "network"

@app.get("/api/apex/agents")
async def list_apex_agents():
    """List all available APEX agents"""
    agent_svc = get_apex_agent_service()
    return {"agents": agent_svc.list_available_agents()}

@app.post("/api/apex/projects/{project_id}/execute-agent")
async def execute_apex_agent(project_id: str, request: AgentExecutionRequest):
    """Execute an APEX agent for a project"""
    agent_svc = get_apex_agent_service()
    
    result = await agent_svc.execute_agent(
        agent_name=request.agent_name,
        project_id=project_id,
        user_input=request.user_input,
        context=request.context
    )
    
    return result

@app.get("/api/apex/executions/{execution_id}")
async def get_apex_execution_status(execution_id: str):
    """Get status of an APEX agent execution"""
    agent_svc = get_apex_agent_service()
    return agent_svc.get_execution_status(execution_id)

@app.get("/api/apex/projects/{project_id}/executions")
async def list_project_executions(project_id: str):
    """List all agent executions for a project"""
    agent_svc = get_apex_agent_service()
    return {"executions": agent_svc.get_project_executions(project_id)}

@app.get("/api/apex/artifacts/{artifact_id}")
async def get_apex_artifact(artifact_id: str):
    """Get an artifact by ID"""
    agent_svc = get_apex_agent_service()
    artifact = agent_svc.get_artifact(artifact_id)
    
    if not artifact:
        raise HTTPException(status_code=404, detail="Artifact not found")
    
    return artifact

@app.get("/api/apex/executions/{execution_id}/artifacts")
async def get_execution_artifacts(execution_id: str):
    """Get all artifacts for an execution"""
    agent_svc = get_apex_agent_service()
    return {"artifacts": agent_svc.get_execution_artifacts(execution_id)}

# MCP Service Endpoints

@app.post("/api/mcp/pricing")
async def get_azure_pricing(request: PricingRequest):
    """Get Azure pricing for a resource"""
    mcp_svc = get_mcp_service()
    pricing = await mcp_svc.get_azure_pricing(
        request.resource_type,
        request.region,
        request.sku
    )
    return pricing

@app.post("/api/mcp/pricing/architecture")
async def calculate_architecture_cost(request: ArchitecturePricingRequest):
    """Calculate total cost for an architecture"""
    mcp_svc = get_mcp_service()
    architecture = {"resources": request.resources}
    cost_analysis = await mcp_svc.calculate_dr_cost(architecture)
    return cost_analysis

@app.post("/api/mcp/diagram/generate")
async def generate_architecture_diagram(request: DiagramRequest):
    """Generate architecture diagram with Azure icons (draw.io XML)"""
    mcp_svc = get_mcp_service()
    diagram = await mcp_svc.generate_architecture_diagram(
        request.architecture,
        request.diagram_type
    )
    return diagram


class AgentDiagramRequest(BaseModel):
    agent_name: str
    output_data: Dict[str, Any]
    assessment_name: str = "Assessment"


@app.post("/api/mcp/diagram/agent")
async def generate_agent_diagram(request: AgentDiagramRequest):
    """Generate draw.io diagram from APEX agent output with Azure service icons."""
    mcp_svc = get_mcp_service()
    diagram = await mcp_svc.generate_agent_diagram(
        agent_name=request.agent_name,
        output_data=request.output_data,
        assessment_name=request.assessment_name,
    )
    return diagram


class AgentDiagramImageRequest(BaseModel):
    agent_name: str
    output_data: Dict[str, Any]
    assessment_name: str = "Assessment"


@app.post("/api/mcp/diagram/image")
async def generate_agent_diagram_image(request: AgentDiagramImageRequest):
    """Generate a PNG-ready SVG image from APEX agent output with Azure icons.
    Returns SVG that the frontend converts to PNG via canvas."""
    from services.diagram_image_service import DiagramImageService
    from io import BytesIO

    svc = DiagramImageService()
    svg_content = svc.render_from_agent_output(
        agent_name=request.agent_name,
        output_data=request.output_data,
        assessment_name=request.assessment_name,
    )
    return StreamingResponse(
        BytesIO(svg_content.encode("utf-8")),
        media_type="image/svg+xml",
        headers={"Content-Disposition": "attachment; filename=architecture-diagram.svg"},
    )


class DiagramExportRequest(BaseModel):
    xml: str
    format: str = "png"  # svg, png
    width: int = 1400
    height: int = 900


def _xml_escape_svg(text: str) -> str:
    """Escape text for SVG XML."""
    if not text:
        return ""
    return text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace('"', "&quot;")


@app.post("/api/mcp/diagram/export")
async def export_diagram_image(request: DiagramExportRequest):
    """Export draw.io XML diagram to SVG or PNG format.
    Renders Azure service icons, groups, edges, and labels."""
    import xml.etree.ElementTree as ET
    import urllib.parse
    from io import BytesIO

    fmt = request.format.lower()
    if fmt not in ("svg", "png"):
        raise HTTPException(status_code=400, detail="Supported formats: svg, png")

    # Parse draw.io XML
    try:
        root = ET.fromstring(request.xml)
    except ET.ParseError:
        raise HTTPException(status_code=400, detail="Invalid XML")

    # Extract cells from draw.io XML
    cells = root.findall(".//{http://www.w3.org/1999/xhtml}mxCell") or root.findall(".//mxCell")

    # Determine diagram bounds from mxGraphModel
    graph_model = root.find(".//mxGraphModel") or root.find(".//{http://www.w3.org/1999/xhtml}mxGraphModel")
    if graph_model is not None:
        dw = int(graph_model.get("pageWidth", request.width))
        dh = int(graph_model.get("pageHeight", request.height))
    else:
        dw, dh = request.width, request.height

    svg_parts = [
        f'<svg xmlns="http://www.w3.org/2000/svg" xmlns:xlink="http://www.w3.org/1999/xlink" '
        f'width="{dw}" height="{dh}" viewBox="0 0 {dw} {dh}">',
        '<defs>',
        '<style>text { font-family: "Segoe UI", Arial, sans-serif; }</style>',
        '</defs>',
        f'<rect width="{dw}" height="{dh}" fill="#0d1117"/>',
    ]

    # First pass: render groups/containers (cells with children or group styles)
    # Second pass: render nodes, Third pass: edges
    group_cells = []
    node_cells = []
    edge_cells = []

    for cell in cells:
        if cell.get("edge") == "1":
            edge_cells.append(cell)
        elif cell.get("vertex") == "1":
            style = cell.get("style", "")
            # Check if it's a group (has opacity and no image)
            if ("opacity=" in style or "fillColor=#" in style) and "image=" not in style:
                geom = cell.find("mxGeometry") or cell.find("{http://www.w3.org/1999/xhtml}mxGeometry")
                if geom is not None:
                    w = float(geom.get("width", 0))
                    h = float(geom.get("height", 0))
                    if w > 150 and h > 150:
                        group_cells.append(cell)
                        continue
            node_cells.append(cell)

    # Helper to extract style property
    def get_style_prop(style, prop):
        for part in style.split(";"):
            if part.startswith(prop + "="):
                return part[len(prop) + 1:]
        return None

    # Render groups (background containers)
    for cell in group_cells:
        geom = cell.find("mxGeometry") or cell.find("{http://www.w3.org/1999/xhtml}mxGeometry")
        if geom is None:
            continue
        x = float(geom.get("x", 0))
        y = float(geom.get("y", 0))
        w = float(geom.get("width", 200))
        h = float(geom.get("height", 200))
        label = cell.get("value", "")
        style = cell.get("style", "")
        
        fill = get_style_prop(style, "fillColor") or "#1e293b"
        stroke = get_style_prop(style, "strokeColor") or "#334155"
        opacity = get_style_prop(style, "opacity")
        opacity_val = f'opacity="{int(opacity)/100}"' if opacity else 'opacity="0.3"'
        dashed = 'stroke-dasharray="8,4"' if "dashed=1" in style else ""
        
        svg_parts.append(f'<rect x="{x}" y="{y}" width="{w}" height="{h}" rx="12" '
                        f'fill="{fill}" stroke="{stroke}" stroke-width="2" {dashed} {opacity_val}/>')
        if label:
            svg_parts.append(f'<text x="{x + 12}" y="{y + 22}" fill="{stroke}" font-size="13" '
                           f'font-weight="bold">{_xml_escape_svg(label)}</text>')

    # Render nodes (services with icons)
    for cell in node_cells:
        geom = cell.find("mxGeometry") or cell.find("{http://www.w3.org/1999/xhtml}mxGeometry")
        if geom is None:
            continue
        x = float(geom.get("x", 0))
        y = float(geom.get("y", 0))
        w = float(geom.get("width", 64))
        h = float(geom.get("height", 64))
        label = cell.get("value", "")
        style = cell.get("style", "")

        # Check for embedded icon image
        if "image=data:image/svg+xml;base64," in style:
            # Extract the base64 icon data
            img_start = style.find("image=data:image/svg+xml;base64,") + len("image=data:image/svg+xml;base64,")
            img_end = style.find(";", img_start)
            if img_end == -1:
                img_end = len(style)
            b64_data = style[img_start:img_end]
            # URL-decode if needed
            try:
                b64_data = urllib.parse.unquote(b64_data)
            except Exception:
                pass
            
            # Render icon background card
            svg_parts.append(f'<rect x="{x-4}" y="{y-4}" width="{w+8}" height="{h+8}" rx="8" '
                           f'fill="#1a2332" stroke="#2d4a6f" stroke-width="1" opacity="0.8"/>')
            # Render the actual SVG icon as embedded image
            svg_parts.append(f'<image x="{x}" y="{y}" width="{w}" height="{h}" '
                           f'href="data:image/svg+xml;base64,{b64_data}"/>')
            # Render label below icon
            if label:
                svg_parts.append(f'<text x="{x + w/2}" y="{y + h + 16}" text-anchor="middle" '
                               f'fill="#e2e8f0" font-size="11" font-weight="500">{_xml_escape_svg(label)}</text>')
        elif label:
            # Regular labeled box
            fill = get_style_prop(style, "fillColor") or "#1e293b"
            stroke = get_style_prop(style, "strokeColor") or "#475569"
            font_color = get_style_prop(style, "fontColor") or "#e2e8f0"
            font_size = get_style_prop(style, "fontSize") or "12"
            font_style = get_style_prop(style, "fontStyle")
            font_weight = "bold" if font_style == "1" else "normal"
            
            svg_parts.append(f'<rect x="{x}" y="{y}" width="{w}" height="{h}" rx="6" '
                           f'fill="{fill}" stroke="{stroke}" stroke-width="1"/>')
            svg_parts.append(f'<text x="{x + w/2}" y="{y + h/2 + 4}" text-anchor="middle" '
                           f'fill="{font_color}" font-size="{font_size}" font-weight="{font_weight}">'
                           f'{_xml_escape_svg(label)}</text>')

    # Render edges
    for cell in edge_cells:
        style = cell.get("style", "")
        label = cell.get("value", "")
        source_id = cell.get("source")
        target_id = cell.get("target")
        
        stroke_color = get_style_prop(style, "strokeColor") or "#64748b"
        stroke_width = get_style_prop(style, "strokeWidth") or "2"
        dashed = 'stroke-dasharray="6,3"' if "dashed=1" in style else ""
        
        # Try to find source/target geometries for proper edge routing
        source_geom = target_geom = None
        for c in cells:
            cid = c.get("id")
            if cid == source_id:
                g = c.find("mxGeometry") or c.find("{http://www.w3.org/1999/xhtml}mxGeometry")
                if g is not None:
                    source_geom = (float(g.get("x", 0)), float(g.get("y", 0)),
                                  float(g.get("width", 64)), float(g.get("height", 64)))
            if cid == target_id:
                g = c.find("mxGeometry") or c.find("{http://www.w3.org/1999/xhtml}mxGeometry")
                if g is not None:
                    target_geom = (float(g.get("x", 0)), float(g.get("y", 0)),
                                  float(g.get("width", 64)), float(g.get("height", 64)))
        
        if source_geom and target_geom:
            sx = source_geom[0] + source_geom[2] / 2
            sy = source_geom[1] + source_geom[3] / 2
            tx = target_geom[0] + target_geom[2] / 2
            ty = target_geom[1] + target_geom[3] / 2
            
            # Orthogonal routing
            mx = (sx + tx) / 2
            svg_parts.append(f'<path d="M{sx},{sy} L{mx},{sy} L{mx},{ty} L{tx},{ty}" '
                           f'fill="none" stroke="{stroke_color}" stroke-width="{stroke_width}" {dashed} '
                           f'marker-end="url(#arrow)"/>')
            if label:
                svg_parts.append(f'<text x="{mx}" y="{(sy+ty)/2 - 6}" text-anchor="middle" '
                               f'fill="#94a3b8" font-size="10">{_xml_escape_svg(label)}</text>')

    # Add arrowhead marker
    svg_parts.insert(3, '<marker id="arrow" markerWidth="10" markerHeight="7" refX="9" refY="3.5" orient="auto">'
                       '<polygon points="0 0, 10 3.5, 0 7" fill="#64748b"/></marker>')

    svg_parts.append("</svg>")
    svg_content = "\n".join(svg_parts)

    # Always return SVG — frontend handles PNG conversion via canvas
    return StreamingResponse(
        BytesIO(svg_content.encode("utf-8")),
        media_type="image/svg+xml",
        headers={"Content-Disposition": f"attachment; filename=architecture-diagram.svg"},
    )


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Assessment Workflow Endpoints
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

from services.assessment_service import AssessmentService

def get_assessment_service():
    """Get or create Assessment service singleton"""
    if not hasattr(get_assessment_service, "_instance"):
        get_assessment_service._instance = AssessmentService()
    return get_assessment_service._instance


class CreateAssessmentRequest(BaseModel):
    assessment_name: str
    assessment_type: str  # 'service-based' or 'multi-resource'
    service_type: Optional[str] = None
    scope_type: Optional[str] = None
    scope_value: Optional[str] = None
    description: Optional[str] = None
    business_unit: Optional[str] = None
    owner: Optional[str] = None


class ScopeResourcesRequest(BaseModel):
    subscription_ids: List[str]
    selected_resource_ids: Optional[List[str]] = None


@app.get("/api/assessments/services")
async def get_supported_services():
    """Get list of supported Azure service types for assessment"""
    assessment_svc = get_assessment_service()
    return {"services": assessment_svc.get_supported_services()}


@app.get("/api/icons/{resource_type:path}")
async def get_icon_metadata(resource_type: str):
    """Get icon metadata for a specific Azure resource type"""
    from services.azure_icon_service import AzureIconService
    metadata = AzureIconService.get_service_metadata(resource_type)
    return metadata


@app.get("/api/assessments")
async def list_assessments(status: Optional[str] = None):
    """List all assessments"""
    assessment_svc = get_assessment_service()
    assessments = assessment_svc.list_assessments(status=status)
    return {"assessments": assessments, "total": len(assessments)}


@app.get("/api/assessments/active")
async def list_active_assessments():
    """List assessments currently running APEX workflow, with live workflow progress.
    Used by the nav badge to show background progress."""
    assessment_svc = get_assessment_service()
    running = assessment_svc.list_assessments(status="apex-running")
    result = []
    for a in running:
        try:
            wf = assessment_svc.get_latest_apex_workflow(a["assessment_id"])
            a["workflow"] = wf
        except Exception:
            a["workflow"] = None
        result.append(a)
    return {"assessments": result, "count": len(result)}


@app.post("/api/assessments", status_code=201)
async def create_assessment(request: CreateAssessmentRequest):
    """Create a new assessment"""
    assessment_svc = get_assessment_service()
    
    assessment = assessment_svc.create_assessment(
        assessment_name=request.assessment_name,
        assessment_type=request.assessment_type,
        service_type=request.service_type,
        scope_type=request.scope_type,
        scope_value=request.scope_value,
        description=request.description,
        business_unit=request.business_unit,
        owner=request.owner
    )
    
    return assessment


@app.get("/api/assessments/{assessment_id}")
async def get_assessment(assessment_id: str):
    """Get assessment details"""
    assessment_svc = get_assessment_service()
    
    try:
        assessment = assessment_svc.get_assessment(assessment_id)
        return assessment
    except Exception as e:
        raise HTTPException(status_code=404, detail=str(e))


@app.post("/api/assessments/{assessment_id}/discover")
async def discover_assessment_resources(assessment_id: str, request: ScopeResourcesRequest):
    """Discover resources for assessment based on type and scope"""
    assessment_svc = get_assessment_service()
    
    try:
        resources = assessment_svc.discover_resources(
            assessment_id=assessment_id,
            subscription_ids=request.subscription_ids
        )
        return {"resources": resources, "total": len(resources)}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.post("/api/assessments/{assessment_id}/scope")
async def scope_assessment_resources(assessment_id: str, request: dict):
    """Add scoped resources to assessment"""
    assessment_svc = get_assessment_service()
    
    try:
        logger.info(f"Scoping resources for {assessment_id}: {len(request.get('resources', []))} resources")
        result = assessment_svc.scope_resources(
            assessment_id=assessment_id,
            resources=request.get("resources", [])
        )
        return result
    except Exception as e:
        logger.error(f"Error scoping resources: {e}", exc_info=True)
        raise HTTPException(status_code=400, detail=str(e))


@app.post("/api/assessments/{assessment_id}/enrich")
async def enrich_assessment_resources(assessment_id: str):
    """Re-query Resource Graph to enrich scoped resources with detailed config (sku, version, HA mode, etc.)."""
    from services.database import get_raw_connection
    import json as _json

    conn = get_raw_connection()
    try:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT resource_id FROM assessment_resources WHERE assessment_id = ?",
            (assessment_id,)
        )
        resource_ids = [r[0] for r in cursor.fetchall()]
    finally:
        conn.close()

    enriched_count = 0
    errors = []

    for resource_id in resource_ids:
        try:
            # Re-query Resource Graph for the specific resource
            import subprocess, json as _json2
            cmd = [
                "docker", "exec", "zuremap",
                "az", "resource", "show",
                "--ids", resource_id,
                "--output", "json"
            ]
            result = subprocess.run(cmd, capture_output=True, text=True, timeout=15)
            if result.returncode == 0:
                data = _json2.loads(result.stdout)
                # Build enrichment dict
                enrichment = {
                    "sku_name": (data.get("sku") or {}).get("name"),
                    "sku_tier": (data.get("sku") or {}).get("tier"),
                    "kind": data.get("kind"),
                    "config": data.get("properties", {}),
                }
                # Merge into existing resource_metadata
                conn2 = get_raw_connection()
                try:
                    c2 = conn2.cursor()
                    c2.execute(
                        "SELECT resource_metadata FROM assessment_resources WHERE assessment_id=? AND resource_id=?",
                        (assessment_id, resource_id)
                    )
                    existing_meta_row = c2.fetchone()
                    existing_meta = {}
                    if existing_meta_row and existing_meta_row[0]:
                        try:
                            existing_meta = _json2.loads(existing_meta_row[0])
                        except Exception:
                            pass
                    existing_meta.update({k: v for k, v in enrichment.items() if v is not None})
                    c2.execute(
                        "UPDATE assessment_resources SET resource_metadata=? WHERE assessment_id=? AND resource_id=?",
                        (_json2.dumps(existing_meta), assessment_id, resource_id)
                    )
                    conn2.commit()
                    enriched_count += 1
                finally:
                    conn2.close()
        except Exception as ex:
            errors.append({"resource_id": resource_id, "error": str(ex)})

    return {
        "assessment_id": assessment_id,
        "enriched": enriched_count,
        "total": len(resource_ids),
        "errors": errors
    }


@app.post("/api/assessments/{assessment_id}/analyze")
async def run_assessment_analysis(assessment_id: str):
    """Run AI analysis on assessment resources"""
    assessment_svc = get_assessment_service()
    
    try:
        analysis = assessment_svc.run_ai_analysis(assessment_id)
        return analysis
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.post("/api/assessments/{assessment_id}/apex/start")
async def start_assessment_apex_workflow(assessment_id: str):
    """Start APEX sequential agent execution"""
    assessment_svc = get_assessment_service()
    
    try:
        workflow = assessment_svc.start_apex_workflow(assessment_id)
        return workflow
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.get("/api/assessments/apex/workflow/{workflow_id}")
async def get_apex_workflow_status(workflow_id: str):
    """Get APEX workflow execution status"""
    assessment_svc = get_assessment_service()
    
    try:
        status = assessment_svc.get_apex_workflow_status(workflow_id)
        return status
    except Exception as e:
        raise HTTPException(status_code=404, detail=str(e))


@app.get("/api/assessments/{assessment_id}/apex/latest")
async def get_latest_apex_workflow(assessment_id: str):
    """Get latest APEX workflow for an assessment"""
    assessment_svc = get_assessment_service()
    
    try:
        status = assessment_svc.get_latest_apex_workflow(assessment_id)
        if not status:
            raise HTTPException(status_code=404, detail="No workflow found")
        return status
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=404, detail=str(e))


@app.post("/api/assessments/{assessment_id}/apex/resume")
async def resume_assessment_apex_workflow(assessment_id: str):
    """Resume a stale/interrupted APEX workflow from where it left off"""
    assessment_svc = get_assessment_service()
    
    try:
        status = assessment_svc.resume_apex_workflow(assessment_id)
        return status
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.post("/api/assessments/{assessment_id}/report")
async def generate_assessment_report(assessment_id: str):
    """Generate comprehensive assessment report"""
    assessment_svc = get_assessment_service()
    
    try:
        report = assessment_svc.generate_report(assessment_id)
        return report
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.get("/api/assessments/{assessment_id}/agents")
async def get_assessment_agent_executions(assessment_id: str):
    """Return all APEX agent executions for a given assessment."""
    from services.database import get_raw_connection
    import json as _json
    conn = get_raw_connection()
    try:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT * FROM agent_executions WHERE assessment_id = ? ORDER BY started_at",
            (assessment_id,)
        )
        col_names = [d[0] for d in cursor.description]
        executions = []
        for row in cursor.fetchall():
            rec = {col_names[i]: row[i] for i in range(len(col_names))}
            for field in ("input_data", "output_data", "artifacts"):
                if isinstance(rec.get(field), str):
                    try:
                        rec[field] = _json.loads(rec[field])
                    except Exception:
                        pass
            executions.append(rec)
        return {"assessment_id": assessment_id, "executions": executions, "count": len(executions)}
    finally:
        conn.close()


@app.get("/api/assessments/apex/stale")
async def detect_stale_apex_workflows():
    """Detect and mark stale APEX workflows (stuck for >10 min)."""
    assessment_svc = get_assessment_service()
    stale = assessment_svc.detect_and_recover_stale_workflows()
    return {"stale_workflows": stale, "count": len(stale)}


@app.on_event("startup")
async def recover_stale_workflows_on_startup():
    """On server startup, detect workflows that were interrupted by previous shutdown."""
    try:
        assessment_svc = get_assessment_service()
        stale = assessment_svc.detect_and_recover_stale_workflows()
        if stale:
            logger.warning(f"Recovered {len(stale)} stale APEX workflow(s) on startup")
    except Exception as e:
        logger.warning(f"Stale workflow recovery skipped: {e}")


# ═══════════════════════════════════════════════════════════════════════════════
# ON-PREMISES REMOTE DISCOVERY & COLLECTION
# ═══════════════════════════════════════════════════════════════════════════════

@app.get("/api/onprem/prerequisites")
async def onprem_check_prerequisites():
    """Check if this machine can perform remote discovery (AD module, domain, etc.)."""
    from services.onprem_discovery_service import check_prerequisites
    return check_prerequisites()


@app.post("/api/onprem/parse-servers")
async def onprem_parse_servers(request: Request):
    """Smart-parse a server list from any delimited input."""
    from services.onprem_discovery_service import parse_server_list
    body = await request.json()
    raw_input = body.get("input", "")
    return parse_server_list(raw_input)


@app.post("/api/onprem/parse-server-file")
async def onprem_parse_server_file(file: UploadFile = File(...)):
    """Parse server names from an uploaded text file."""
    from services.onprem_discovery_service import parse_server_file
    content = await file.read()
    if len(content) > 5 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="File too large (max 5 MB)")
    return parse_server_file(content, filename=file.filename or "servers.txt")


@app.post("/api/onprem/discover-ad")
async def onprem_discover_ad(request: Request):
    """Discover domain-joined computers via Active Directory."""
    from services.onprem_discovery_service import discover_ad_computers
    body = await request.json() if await request.body() else {}
    return discover_ad_computers(
        ou_filter=body.get("ou_filter", ""),
        name_filter=body.get("name_filter", ""),
        os_filter=body.get("os_filter", ""),
    )


@app.post("/api/onprem/test-connectivity")
async def onprem_test_connectivity(request: Request):
    """Test WinRM/WMI connectivity to a list of servers."""
    from services.onprem_discovery_service import test_connectivity
    body = await request.json()
    servers = body.get("servers", [])
    if not servers:
        raise HTTPException(status_code=400, detail="No servers provided")
    return test_connectivity(servers)


@app.post("/api/onprem/collect-remote")
async def onprem_start_remote_collection(request: Request):
    """Start a background remote data collection job."""
    from services.onprem_discovery_service import start_collection
    body = await request.json()
    servers = body.get("servers", [])
    # Accept both string list ["host1"] and object list [{"host":"host1"}]
    normalized = []
    for s in servers:
        if isinstance(s, str):
            normalized.append(s)
        elif isinstance(s, dict):
            normalized.append(s.get("host") or s.get("hostname") or s.get("name", ""))
    servers = [h for h in normalized if h]
    modules = body.get("modules", {})
    options = body.get("options", {})
    if not servers:
        raise HTTPException(status_code=400, detail="No servers provided")
    if len(servers) > 500:
        raise HTTPException(status_code=400, detail="Maximum 500 servers per collection job")
    return start_collection(servers, modules, options)


@app.get("/api/onprem/collect-remote/status/{job_id}")
async def onprem_collection_status(job_id: str):
    """Get status of a running collection job."""
    from services.onprem_discovery_service import get_collection_status
    return get_collection_status(job_id)


@app.post("/api/onprem/collect-remote/cancel/{job_id}")
async def onprem_cancel_collection(job_id: str):
    """Cancel a running collection job."""
    from services.onprem_discovery_service import cancel_collection
    return cancel_collection(job_id)


@app.get("/api/onprem/collection-jobs")
async def onprem_collection_jobs():
    """Get history of all collection jobs."""
    from services.onprem_discovery_service import get_all_jobs
    return get_all_jobs()


# ═══════════════════════════════════════════════════════════════════════════════
# ON-PREMISES SCHEDULED MONITORING — periodic re-scan of added servers
# ═══════════════════════════════════════════════════════════════════════════════

@app.get("/api/onprem/schedule")
async def onprem_get_schedule():
    """Get the scheduled-monitoring config + live status (next/last run, running)."""
    from services.onprem_scheduler import get_schedule
    return await asyncio.to_thread(get_schedule)


@app.put("/api/onprem/schedule")
async def onprem_update_schedule(request: Request):
    """Create / update the periodic on-prem scan schedule."""
    from services.onprem_scheduler import update_schedule
    body = await request.json()
    if not isinstance(body, dict):
        raise HTTPException(status_code=400, detail="Invalid schedule payload")
    return await asyncio.to_thread(update_schedule, body)


@app.post("/api/onprem/schedule/run-now")
async def onprem_schedule_run_now(request: Request):
    """Trigger an immediate scan using the saved schedule (optional overrides)."""
    from services.onprem_scheduler import run_now
    overrides = None
    try:
        if await request.body():
            data = await request.json()
            if isinstance(data, dict) and data:
                overrides = data
    except Exception:
        overrides = None
    return await asyncio.to_thread(run_now, overrides)


@app.get("/api/onprem/schedule/history")
async def onprem_schedule_history():
    """Recent scheduled/manual scan runs."""
    from services.onprem_scheduler import get_history
    return await asyncio.to_thread(get_history)


# ═══════════════════════════════════════════════════════════════════════════════
# ON-PREMISES DATA COLLECTION ENDPOINTS
# ═══════════════════════════════════════════════════════════════════════════════

@app.get("/api/onprem/summary")
async def onprem_inventory_summary():
    """Get aggregated on-premises inventory summary."""
    from services.onprem_service import get_inventory_summary
    return get_inventory_summary()


@app.get("/api/onprem/servers")
async def onprem_list_servers(batch_id: Optional[str] = None, workload_type: Optional[str] = None):
    """List all on-premises servers, optionally filtered."""
    from services.onprem_service import get_all_servers
    return get_all_servers(batch_id=batch_id, workload_type=workload_type)


@app.get("/api/onprem/servers/{server_id}")
async def onprem_server_detail(server_id: str):
    """Get full detail for a single on-premises server."""
    from services.onprem_service import get_server_detail
    result = get_server_detail(server_id)
    if not result:
        raise HTTPException(status_code=404, detail="Server not found")
    return result


@app.get("/api/onprem/applications")
async def onprem_application_inventory():
    """Cross-server application matrix."""
    from services.onprem_service import get_application_inventory
    return get_application_inventory()


@app.get("/api/onprem/migration-candidates")
async def onprem_migration_candidates():
    """Get servers flagged as migration candidates with recommended targets."""
    from services.onprem_service import get_migration_candidates
    return get_migration_candidates()


@app.get("/api/onprem/inventory")
async def onprem_inventory_filtered(
    workload_type: Optional[str] = None,
    os_filter: Optional[str] = None,
    complexity: Optional[str] = None,
    search: Optional[str] = None,
    has_sql: Optional[bool] = None,
    has_iis: Optional[bool] = None,
    migration_target: Optional[str] = None,
    sort_by: str = "hostname",
    sort_dir: str = "asc",
):
    """Advanced filtered server inventory for migration planning."""
    from services.onprem_service import get_inventory_filtered
    return get_inventory_filtered(
        workload_type=workload_type, os_filter=os_filter,
        complexity=complexity, search=search,
        has_sql=has_sql, has_iis=has_iis,
        migration_target=migration_target,
        sort_by=sort_by, sort_dir=sort_dir,
    )


@app.get("/api/onprem/roles")
async def onprem_role_summary():
    """Server role summary for migration planning dashboard."""
    from services.onprem_service import get_role_summary
    return get_role_summary()


@app.get("/api/onprem/servers/{server_id}/history")
async def onprem_server_scan_history(server_id: str):
    """Scan history for a specific server."""
    from services.onprem_service import get_server_scan_history
    return get_server_scan_history(server_id)


@app.post("/api/onprem/security-findings")
async def onprem_generate_security():
    """Generate security findings from on-prem data and persist them."""
    from services.onprem_service import generate_onprem_security_findings
    from services.security_findings_service import persist_findings
    findings = generate_onprem_security_findings()
    if not findings:
        return {"message": "No security findings generated", "total": 0}
    result = persist_findings(findings, scan_type="onprem")
    return result


@app.get("/api/onprem/ai-context")
async def onprem_ai_context():
    """Get on-prem server context for AI assessment."""
    from services.onprem_service import get_onprem_ai_context
    return get_onprem_ai_context()


@app.get("/api/onprem/batches")
async def onprem_list_batches():
    """List all upload batches."""
    from services.onprem_service import get_upload_batches
    return get_upload_batches()


@app.delete("/api/onprem/batches/{batch_id}")
async def onprem_delete_batch(batch_id: str):
    """Delete an upload batch and its server data."""
    from services.onprem_service import delete_batch
    return delete_batch(batch_id)


@app.post("/api/onprem/upload")
async def onprem_upload(file: UploadFile = File(...)):
    """Upload a ZIP file containing on-premises server collection CSVs."""
    from services.onprem_service import ingest_upload
    if not file.filename or not file.filename.lower().endswith(".zip"):
        raise HTTPException(status_code=400, detail="Only ZIP files are accepted")
    contents = await file.read()
    max_size = 100 * 1024 * 1024  # 100 MB limit
    if len(contents) > max_size:
        raise HTTPException(status_code=413, detail="File too large (max 100 MB)")
    result = ingest_upload(contents, filename=file.filename)
    return result


@app.post("/api/onprem/generate-script")
async def onprem_generate_script(request: Request):
    """Generate a PowerShell collection script based on user options."""
    from services.onprem_script_service import generate_script
    body = await request.json()
    script = generate_script(body)
    return {"script": script, "filename": "Collect-ServerInventory.ps1"}


@app.get("/api/onprem/download-script")
async def onprem_download_script(
    scope: str = "localhost",
    modules: str = "all",
):
    """Download a pre-configured PowerShell collection script."""
    from services.onprem_script_service import generate_script
    from fastapi.responses import PlainTextResponse
    options = {"target_scope": scope}
    if modules != "all":
        all_modules = ["collect_hardware", "collect_os", "collect_applications",
                       "collect_services", "collect_sql", "collect_iis",
                       "collect_security", "collect_certificates", "collect_performance"]
        selected = [m.strip() for m in modules.split(",")]
        for mod in all_modules:
            options[mod] = mod.replace("collect_", "") in selected
    script = generate_script(options)
    return PlainTextResponse(
        content=script,
        media_type="text/plain",
        headers={"Content-Disposition": 'attachment; filename="Collect-ServerInventory.ps1"'}
    )


# ═══════════════════════════════════════════════════════════════════════════════
# ON-PREMISES LDAP / AD INTEGRATION
# ═══════════════════════════════════════════════════════════════════════════════

@app.post("/api/onprem/ldap/test")
async def onprem_ldap_test(config: dict):
    """Test LDAP connection with provided credentials."""
    from services.ldap_service import test_connection
    return test_connection(config)


@app.post("/api/onprem/ldap/discover")
async def onprem_ldap_discover(body: dict):
    """Discover computers from AD via LDAP."""
    from services.ldap_service import discover_computers, get_config_from_settings
    config = body.get("config") or get_config_from_settings()
    filters = body.get("filters", {})
    return discover_computers(config, filters)


@app.get("/api/onprem/ldap/discover")
async def onprem_ldap_discover_get(
    name_filter: str = "", os_filter: str = "",
    ou_filter: str = "",
    server_os_only: bool = True, enabled_only: bool = True
):
    """Discover computers from AD using saved LDAP config."""
    from services.ldap_service import discover_computers, get_config_from_settings, is_configured
    if not is_configured():
        raise HTTPException(status_code=400, detail="LDAP not configured. Set DC host, base DN, and credentials in Settings → On-Premises.")
    config = get_config_from_settings()
    filters = {
        "name_filter": name_filter,
        "os_filter": os_filter,
        "ou_filter": ou_filter,
        "server_os_only": server_os_only,
        "enabled_only": enabled_only,
    }
    return discover_computers(config, filters)


@app.get("/api/onprem/ldap/ous")
async def onprem_ldap_ous():
    """Discover OU structure from AD."""
    from services.ldap_service import discover_ous, get_config_from_settings, is_configured
    if not is_configured():
        raise HTTPException(status_code=400, detail="LDAP not configured.")
    config = get_config_from_settings()
    return discover_ous(config)


@app.get("/api/onprem/ldap/status")
async def onprem_ldap_status():
    """Check if LDAP is configured and test connectivity."""
    from services.ldap_service import is_configured, test_connection, get_config_from_settings
    configured = is_configured()
    result = {"configured": configured, "connected": False, "error": None}
    if configured:
        test = test_connection(get_config_from_settings())
        result["connected"] = test.get("success", False)
        result["error"] = test.get("error")
        result["domain_info"] = test.get("domain_info")
    return result


# ═══════════════════════════════════════════════════════════════════════════════
# ON-PREMISES DISCOVERY ENGINE
# ═══════════════════════════════════════════════════════════════════════════════

@app.get("/api/onprem/engine/status")
async def onprem_engine_status():
    """Get discovery engine status."""
    from services.onprem_engine import get_engine_status
    return get_engine_status()


@app.post("/api/onprem/engine/start")
async def onprem_engine_start(body: dict = {}):
    """Start the discovery engine."""
    from services.onprem_engine import start_engine
    interval = body.get("interval_hours", 0)
    return await start_engine(interval)


@app.post("/api/onprem/engine/stop")
async def onprem_engine_stop():
    """Stop the discovery engine."""
    from services.onprem_engine import stop_engine
    return await stop_engine()


@app.post("/api/onprem/engine/trigger")
async def onprem_engine_trigger():
    """Trigger an immediate discovery cycle."""
    from services.onprem_engine import trigger_now
    return await trigger_now()


# ═══════════════════════════════════════════════════════════════════════════════
# ON-PREMISES CROSS-MODULE BRIDGE (BCDR, Security, Migration)
# ═══════════════════════════════════════════════════════════════════════════════

@app.get("/api/onprem/bridge/bcdr")
async def onprem_bridge_bcdr():
    """On-prem BCDR assessment (backup coverage, DR gaps)."""
    from services.onprem_bridge import get_bcdr_assessment
    return get_bcdr_assessment()


@app.get("/api/onprem/bridge/security")
async def onprem_bridge_security():
    """On-prem security gaps (firewall, AV, updates, certificates)."""
    from services.onprem_bridge import get_security_assessment
    return get_security_assessment()


@app.get("/api/onprem/bridge/migration")
async def onprem_bridge_migration():
    """On-prem migration readiness (sizing, strategy, blockers)."""
    from services.onprem_bridge import get_migration_readiness
    return get_migration_readiness()


@app.get("/api/onprem/bridge/context")
async def onprem_bridge_context():
    """Combined Azure + on-prem context for AI analysis."""
    from services.onprem_bridge import get_ai_context
    return get_ai_context()


@app.get("/api/onprem/ai/analysis")
async def onprem_ai_analysis(refresh: bool = False):
    """AI-powered analysis of on-premises infrastructure for migration and BCDR planning."""
    from services.onprem_service import get_all_servers, get_inventory_summary
    from services.ai_module_analysis_service import analyze_onprem_ai
    servers = get_all_servers()
    if not servers:
        raise HTTPException(status_code=404, detail="No on-premises data. Upload server inventory first.")
    summary = get_inventory_summary()
    return analyze_onprem_ai(servers, summary, force_refresh=refresh)


# ── Monitoring (Azure Monitor) Module ─────────────────────────────────────────

@app.get("/api/monitor/overview")
async def monitor_overview():
    """Monitoring posture rollup: machine agent coverage, resource health, alerts."""
    import asyncio
    from services import monitoring_service
    return await asyncio.to_thread(monitoring_service.get_monitoring_overview)


@app.get("/api/monitor/health")
async def monitor_health():
    """Per-resource Azure Resource Health (Available/Degraded/Unavailable/Unknown)."""
    import asyncio
    from services import monitoring_service
    items = await asyncio.to_thread(monitoring_service.get_resource_health)
    return {"items": items, "total": len(items)}


@app.get("/api/monitor/alerts")
async def monitor_alerts():
    """Fired Azure Monitor alerts (Sev0-Sev4)."""
    import asyncio
    from services import monitoring_service
    items = await asyncio.to_thread(monitoring_service.get_fired_alerts)
    return {"items": items, "total": len(items)}


@app.get("/api/monitor/coverage")
async def monitor_coverage():
    """Per-machine monitoring agent coverage (native VMs + Arc-enabled servers)."""
    import asyncio
    from services import monitoring_service
    items = await asyncio.to_thread(monitoring_service.get_monitoring_coverage)
    return {"items": items, "total": len(items)}


@app.get("/api/monitor/metrics")
async def monitor_metrics(limit: int = 40):
    """Platform metrics (CPU/mem/disk/network) for top resources via Azure Monitor."""
    import asyncio
    from services import monitoring_service
    resources = _get_resources_list()
    items = await asyncio.to_thread(monitoring_service.get_platform_metrics, resources, limit)
    return {"items": items, "total": len(items)}


@app.get("/api/monitor/laperf")
async def monitor_laperf(hours: int = 24):
    """Log Analytics agent performance + heartbeat (Arc/on-prem). Degrades gracefully."""
    import asyncio
    from services import monitoring_service
    return await asyncio.to_thread(monitoring_service.get_la_perf, None, hours)


@app.get("/api/ai/monitoring")
async def ai_monitoring_analysis(refresh: bool = False):
    """AI-powered monitoring & observability analysis across the estate."""
    import asyncio
    from services.ai_module_analysis_service import analyze_monitoring_ai
    from services import monitoring_service
    resources = _get_resources_list()
    if not resources:
        raise HTTPException(status_code=404, detail="No resource data")
    arc_data = _get_arc_data()
    overview = await asyncio.to_thread(monitoring_service.get_monitoring_overview)
    coverage = await asyncio.to_thread(monitoring_service.get_monitoring_coverage)
    alerts = await asyncio.to_thread(monitoring_service.get_fired_alerts)
    monitoring_data = {
        "overview": overview,
        "uncovered_machines": [c for c in coverage if not c.get("agent_installed")][:50],
        "alerts": alerts[:50],
    }
    return await asyncio.to_thread(analyze_monitoring_ai, resources, arc_data, monitoring_data, refresh)


# ── Governance / Identity / Operations / Advisor ──────────────────────────────

@app.get("/api/governance/policy")
async def governance_policy():
    """Azure Policy compliance posture (compliant vs non-compliant, worst policies)."""
    import asyncio
    from services import governance_service
    return await asyncio.to_thread(governance_service.get_policy_compliance)


@app.get("/api/identity/access")
async def identity_access():
    """RBAC posture: role assignments, principal mix, privileged grants."""
    import asyncio
    from services import identity_service
    return await asyncio.to_thread(identity_service.get_access_overview)


@app.get("/api/identity/posture")
async def identity_posture():
    """Full Identity & Access posture: RBAC over-permissioning + Entra ID app-
    registration credential expiry + directory (users/guests) + best-practice findings."""
    import asyncio
    from services import identity_service
    return await asyncio.to_thread(identity_service.get_identity_posture)


@app.get("/api/identity/app-registrations")
async def identity_app_registrations():
    """Entra ID app registrations with secret/certificate expiry (Microsoft Graph)."""
    import asyncio
    from services import graph_service
    return await asyncio.to_thread(graph_service.get_app_registrations)


@app.get("/api/identity/guests")
async def identity_guests():
    """Entra ID guest (external) user accounts (Microsoft Graph)."""
    import asyncio
    from services import graph_service
    return await asyncio.to_thread(graph_service.get_guest_users)


@app.get("/api/operations/service-health")
async def operations_service_health():
    """Azure Service Health events (issues, maintenance, advisories)."""
    import asyncio
    from services import service_health_service
    return await asyncio.to_thread(service_health_service.get_service_health)


@app.get("/api/operations/lifecycle-radar")
async def operations_lifecycle_radar():
    """Retirements & Deprecations radar — fuses Service Health advisories, planned
    maintenance and Advisor recommendations into a dated, inventory-correlated view
    of what Azure is retiring/deprecating and how many of your resources are exposed."""
    import asyncio
    from services import service_health_service
    resources = _get_resources_list()
    return await asyncio.to_thread(service_health_service.get_lifecycle_radar, None, resources)


@app.get("/api/operations/quota")
async def operations_quota():
    """Compute quota usage vs limits per region (capacity planning)."""
    import asyncio
    from services import quota_service
    resources = _get_resources_list()
    return await asyncio.to_thread(quota_service.get_quota_usage, None, resources)


@app.get("/api/advisor")
async def advisor_overview():
    """Azure Advisor recommendations grouped by category + impact."""
    import asyncio
    from services.advisor_service import get_advisor_overview
    return await asyncio.to_thread(get_advisor_overview)


@app.get("/api/ai/governance")
async def ai_governance_analysis(refresh: bool = False, scope: str = None):
    """AI-powered governance, policy & identity analysis."""
    import asyncio
    from services.ai_module_analysis_service import analyze_generic_ai
    from services import governance_service, identity_service
    resources = _get_resources_list()
    if not resources:
        raise HTTPException(status_code=404, detail="No resource data")
    arc_data = _get_arc_data()
    policy = await asyncio.to_thread(governance_service.get_policy_compliance)
    identity = await asyncio.to_thread(identity_service.get_access_overview)
    ctx = {
        "policy_compliance_pct": policy.get("compliance_pct"),
        "non_compliant_resources": policy.get("non_compliant_resources"),
        "top_non_compliant_policies": policy.get("top_non_compliant_policies", [])[:10],
        "policy_exemptions": policy.get("policy_exemptions"),
        "total_role_assignments": identity.get("total_assignments"),
        "privileged_assignments": identity.get("privileged_assignments"),
        "owner_assignments": identity.get("owner_assignments"),
        "by_principal_type": identity.get("by_principal_type"),
    }
    return await asyncio.to_thread(
        analyze_generic_ai, "governance",
        "Azure Governance, Policy and Identity (RBAC) specialist",
        "policy compliance gaps, governance drift, and identity/RBAC over-permissioning risk",
        resources, arc_data, ctx, refresh, scope,
    )


@app.get("/api/ai/advisor")
async def ai_advisor_analysis(refresh: bool = False, category: str = None, scope: str = None):
    """AI-powered prioritization of Azure Advisor recommendations. Optionally
    scoped to a single recommendation category (Cost / Security / Performance /
    HighAvailability / OperationalExcellence) so the AI analyses ONLY the data the
    user filtered to in the UI; each category is cached separately."""
    import asyncio
    from services.ai_module_analysis_service import analyze_generic_ai
    from services.advisor_service import get_advisor_overview
    resources = _get_resources_list()
    if not resources:
        raise HTTPException(status_code=404, detail="No resource data")
    arc_data = _get_arc_data()
    adv = await asyncio.to_thread(get_advisor_overview)
    items = adv.get("items", [])
    cat = (category or "").strip()
    if cat and cat.lower() != "all":
        items = [i for i in items if str(i.get("category", "")).lower() == cat.lower()]
        by_impact: dict = {}
        for i in items:
            k = i.get("impact") or "Unknown"
            by_impact[k] = by_impact.get(k, 0) + 1
        module_key = "advisor_" + cat.lower()
        ctx = {
            "filter_category": cat,
            "total_recommendations": len(items),
            "high_impact": sum(1 for i in items if str(i.get("impact", "")).lower() == "high"),
            "by_impact": by_impact,
            "sample": items[:35],
        }
        focus = (f"Azure Advisor recommendations for the {cat} category ONLY (the input has been filtered by "
                 f"the user to {cat}). Prioritise, deep-dive and remediate only {cat} recommendations against "
                 f"the impacted workloads; do not analyse other Advisor categories.")
    else:
        module_key = "advisor"
        ctx = {
            "filter_category": "all",
            "total_recommendations": adv.get("total"),
            "high_impact": adv.get("high_impact"),
            "by_category": adv.get("by_category"),
            "by_impact": adv.get("by_impact"),
            "sample": items[:25],
        }
        focus = "Azure Advisor recommendations across cost, performance, reliability, security and operational excellence"
    return await asyncio.to_thread(
        analyze_generic_ai, module_key,
        "Azure optimization advisor and Well-Architected reviewer",
        focus,
        resources, arc_data, ctx, refresh, scope,
    )


@app.get("/api/ai/service-health")
async def ai_service_health_analysis(refresh: bool = False, scope: str = None):
    """AI-powered analysis of Azure Service Health events and their workload impact."""
    import asyncio
    from services.ai_module_analysis_service import analyze_generic_ai
    from services import service_health_service
    resources = _get_resources_list()
    if not resources:
        raise HTTPException(status_code=404, detail="No resource data")
    arc_data = _get_arc_data()
    sh = await asyncio.to_thread(service_health_service.get_service_health)
    ctx = {
        "active_events": sh.get("active_events"),
        "service_issues": sh.get("service_issues"),
        "planned_maintenance": sh.get("planned_maintenance"),
        "health_advisories": sh.get("health_advisories"),
        "security_advisories": sh.get("security_advisories"),
        "events_sample": sh.get("items", [])[:30],
    }
    return await asyncio.to_thread(
        analyze_generic_ai, "service_health",
        "Azure Service Health and operational continuity advisor",
        "active Azure Service Health events (issues, planned maintenance, advisories), which workloads/resources are exposed, and concrete mitigations",
        resources, arc_data, ctx, refresh, scope,
    )


@app.get("/api/ai/lifecycle")
async def ai_lifecycle_analysis(refresh: bool = False, scope: str = None):
    """AI-powered Retirements & Deprecations briefing — reads the lifecycle radar
    (Service Health advisories + planned maintenance + Advisor recs correlated to the
    customer's inventory) and prioritises what to act on, by when, and the benefit."""
    import asyncio
    from services.ai_module_analysis_service import analyze_generic_ai
    from services import service_health_service
    resources = _get_resources_list()
    if not resources:
        raise HTTPException(status_code=404, detail="No resource data")
    arc_data = _get_arc_data()
    radar = await asyncio.to_thread(service_health_service.get_lifecycle_radar, None, resources)
    sm = radar.get("summary", {})
    ctx = {
        "total_lifecycle_items": sm.get("total"),
        "retirements": sm.get("retirements"),
        "deprecations": sm.get("deprecations"),
        "upgrades": sm.get("upgrades"),
        "security_or_certificate": sm.get("security"),
        "due_within_30_days": sm.get("due_30"),
        "due_within_90_days": sm.get("due_90"),
        "overdue": sm.get("overdue"),
        "your_resources_exposed": sm.get("exposed_resources"),
        "timeline": radar.get("timeline", [])[:18],
        "by_category": radar.get("by_category"),
        "items_sample": [
            {"source": i["source"], "category": i["category"], "title": i["title"][:160],
             "deadline": i["deadline"], "days_until": i["days_until"],
             "exposed_count": i["exposed_count"], "priority": i["priority"]}
            for i in radar.get("items", [])[:35]
        ],
    }
    return await asyncio.to_thread(
        analyze_generic_ai, "lifecycle",
        "Azure service lifecycle, retirements and deprecations advisor",
        "upcoming Azure retirements, deprecations, SKU/size end-of-support and forced upgrades from Service Health advisories, planned maintenance and Advisor; which of the customer's OWN resources are exposed, the deadline for each, the risk of inaction, and the prioritised remediation (what to migrate/upgrade and by when)",
        resources, arc_data, ctx, refresh, scope,
    )


@app.get("/api/ai/quota")
async def ai_quota_analysis(refresh: bool = False, scope: str = None):
    """AI-powered analysis of Azure quota/capacity headroom and deployment risk."""
    import asyncio
    from services.ai_module_analysis_service import analyze_generic_ai
    from services import quota_service
    resources = _get_resources_list()
    if not resources:
        raise HTTPException(status_code=404, detail="No resource data")
    arc_data = _get_arc_data()
    q = await asyncio.to_thread(quota_service.get_quota_usage, None, resources)
    ctx = {
        "regions": q.get("regions"),
        "strategic_regions": q.get("strategic_regions"),
        "total_quotas": q.get("total_quotas"),
        "near_limit_count": q.get("near_limit_count"),
        "blocked_count": q.get("blocked_count"),
        "total_vcpu_used": q.get("total_vcpu_used"),
        "total_vcpu_limit": q.get("total_vcpu_limit"),
        "vcpu_headroom": q.get("vcpu_headroom"),
        "regions_summary": q.get("regions_summary", [])[:20],
        "blocked_sample": q.get("blocked", [])[:40],
        "near_limit_sample": q.get("near_limit", [])[:30],
        "note": q.get("note"),
    }
    return await asyncio.to_thread(
        analyze_generic_ai, "quota",
        "Azure capacity and quota planning advisor",
        "compute quota/capacity headroom and deployment risk. IMPORTANT: in capacity-restricted regions such as "
        "Qatar Central, VM families with a limit of 0 are BLOCKED and the customer must raise an Azure support "
        "ticket to whitelist quota before any resource can be created. Identify (1) blocked families per region "
        "that need a proactive quota request, (2) families near their limit that will throttle scaling, (3) the "
        "per-region vCPU headroom for the in-scope workloads, and give a prioritised, region-by-region quota-request "
        "plan with the specific families/amounts to request and the business risk of not doing so",
        resources, arc_data, ctx, refresh, scope,
    )


@app.get("/api/ai/insights-dashboard")
async def ai_insights_dashboard():
    """Aggregated latest AI analysis summary per category for the home dashboard.
    Read-only: surfaces whatever has already been generated and never triggers a
    new (expensive) AI analysis."""
    import asyncio
    from services.ai_module_analysis_service import get_ai_insights_summary
    return await asyncio.to_thread(get_ai_insights_summary)


@app.get("/api/ai/executive-briefing")
async def ai_executive_briefing(refresh: bool = False):
    """Cross-category AI executive synthesis (CIO-level briefing) over the latest
    per-module AI summaries — top cross-cutting risks + a unified roadmap."""
    import asyncio
    from services.ai_module_analysis_service import analyze_executive_briefing_ai
    return await asyncio.to_thread(analyze_executive_briefing_ai, refresh)


# ── AI Modernization Assessments (revenue-generating, deep AI) ────────────────

@app.get("/api/assess/sql")
async def assess_sql_estate():
    """SQL footprint (Azure PaaS, IaaS SQL VMs, Arc SQL, on-prem) for the SQL
    modernization assessment."""
    import asyncio
    from services import modernization_assessment_service as _m
    return await asyncio.to_thread(_m.get_sql_estate)


@app.get("/api/ai/sql-modernization")
async def ai_sql_modernization(refresh: bool = False, scope: str = None):
    """AI SQL modernization assessment: on-prem/Arc/IaaS SQL -> Azure SQL Database
    (lightweight) or Azure SQL Managed Instance (full compatibility)."""
    import asyncio
    from services.ai_module_analysis_service import analyze_generic_ai
    from services import modernization_assessment_service as _m
    resources = _get_resources_list()
    if not resources:
        raise HTTPException(status_code=404, detail="No resource data")
    arc_data = _get_arc_data()
    sql = await asyncio.to_thread(_m.get_sql_estate)
    ctx = {
        "iaas_sql_vms": sql.get("iaas_sql_vms"),
        "arc_sql_instances": sql.get("arc_sql_instances"),
        "azure_sql_databases": sql.get("azure_sql_databases"),
        "managed_instances": sql.get("managed_instances"),
        "sql_vm_candidates": sql.get("sql_vm_candidates", [])[:40],
        "onprem_sql": sql.get("onprem_sql", [])[:40],
        "by_kind": sql.get("by_kind"),
    }
    return await asyncio.to_thread(
        analyze_generic_ai, "sql_modernization",
        "Microsoft data platform modernization architect",
        "modernizing SQL Server estates to Azure: map on-prem SQL on Hyper-V/VMware and Arc-enabled SQL "
        "to Azure SQL Database (lightweight/serverless) for app-scoped DBs, and Azure IaaS SQL VMs plus "
        "instance-scoped workloads to Azure SQL Managed Instance (full SQL Server compatibility). Give per-"
        "workload migration paths (target, tier/SKU, compatibility blockers, effort) and the Azure consumption uplift",
        resources, arc_data, ctx, refresh, scope,
    )


@app.get("/api/assess/appservice")
async def assess_appservice_estate():
    """App Service plans + sites for the App Service modernization assessment."""
    import asyncio
    from services import modernization_assessment_service as _m
    return await asyncio.to_thread(_m.get_appservice_estate)


@app.get("/api/ai/appservice")
async def ai_appservice(refresh: bool = False, scope: str = None):
    """AI App Service modernization assessment: plan right-sizing, tier upgrades,
    Functions Flex Consumption, containers, security and reliability."""
    import asyncio
    from services.ai_module_analysis_service import analyze_generic_ai
    from services import modernization_assessment_service as _m
    resources = _get_resources_list()
    if not resources:
        raise HTTPException(status_code=404, detail="No resource data")
    arc_data = _get_arc_data()
    aps = await asyncio.to_thread(_m.get_appservice_estate)
    ctx = {
        "total_plans": aps.get("total_plans"),
        "total_sites": aps.get("total_sites"),
        "function_apps": aps.get("function_apps"),
        "container_apps": aps.get("container_apps"),
        "web_apps": aps.get("web_apps"),
        "by_tier": aps.get("by_tier"),
        "free_basic_plans": aps.get("free_basic_plans"),
        "plans_sample": aps.get("plans", [])[:30],
    }
    return await asyncio.to_thread(
        analyze_generic_ai, "appservice",
        "Azure App Service modernization specialist",
        "Azure App Service modernization: right-size and consolidate App Service plans, move legacy tiers to "
        "Premium v3 / Isolated v2, adopt Functions Flex Consumption and container hosting, enforce HTTPS-only, "
        "managed identity and Key Vault, and improve autoscale/zone-redundancy reliability",
        resources, arc_data, ctx, refresh, scope,
    )


@app.get("/api/ai/vm-performance")
async def ai_vm_performance(refresh: bool = False, scope: str = None):
    """AI VM performance & right-sizing assessment using Azure Monitor metrics."""
    import asyncio
    from services.ai_module_analysis_service import analyze_generic_ai
    from services import monitoring_service
    resources = _get_resources_list()
    if not resources:
        raise HTTPException(status_code=404, detail="No resource data")
    arc_data = _get_arc_data()
    try:
        metrics = await asyncio.to_thread(monitoring_service.get_platform_metrics, resources, 60)
    except Exception:
        metrics = []
    compute = [m for m in metrics if "virtualmachines" in str(m.get("resource_type", "")).lower()][:50]
    ctx = {
        "vm_metrics_sample": compute,
        "vms_measured": len(compute),
        "note": "Azure Monitor 30-day CPU/memory/disk/network utilisation; deallocated VMs included.",
    }
    return await asyncio.to_thread(
        analyze_generic_ai, "vm_performance",
        "Azure performance and capacity engineer",
        "VM performance, utilisation and right-sizing using Azure Monitor metrics (CPU, memory, disk, network): "
        "identify idle/underused VMs to downsize or deallocate, over-utilised VMs to scale up, and SKU/family "
        "recommendations (incl. Arc-enabled servers), with the cost and performance impact per VM",
        resources, arc_data, ctx, refresh, scope,
    )


@app.get("/api/ai/entra")
async def ai_entra(refresh: bool = False, scope: str = None):
    """AI Entra ID & Permissions Management assessment (identity/permission risk)."""
    import asyncio
    from services.ai_module_analysis_service import analyze_generic_ai
    from services import identity_service
    resources = _get_resources_list()
    if not resources:
        raise HTTPException(status_code=404, detail="No resource data")
    arc_data = _get_arc_data()
    try:
        posture = await asyncio.to_thread(identity_service.get_identity_posture)
    except Exception:
        posture = {}
    sm = posture.get("summary", {})
    ctx = {
        "posture_score": posture.get("score"),
        "total_role_assignments": sm.get("total_assignments"),
        "privileged_assignments": sm.get("privileged_assignments"),
        "owner_assignments": sm.get("owner_assignments"),
        "mg_scope_privileged": sm.get("mg_privileged"),
        "service_principal_privileged": sm.get("sp_privileged"),
        "app_registrations": sm.get("app_registrations"),
        "app_secrets_expired": sm.get("apps_expired"),
        "app_secrets_expiring_30d": sm.get("apps_expiring_30"),
        "guest_users": sm.get("guest_users"),
        "graph_available": posture.get("graph_available"),
        "by_principal_type": posture.get("by_principal_type"),
        "by_role": posture.get("by_role"),
        "by_scope_level": posture.get("by_scope_level"),
        "best_practice_findings": posture.get("findings", [])[:14],
        "privileged_sample": posture.get("privileged", [])[:25],
        "expiring_apps_sample": [
            {"app": a["display_name"], "status": a["credential_status"], "days": a["soonest_expiry_days"]}
            for a in (posture.get("app_registrations", {}).get("items", []) if isinstance(posture.get("app_registrations"), dict) else [])[:20]
        ],
    }
    return await asyncio.to_thread(
        analyze_generic_ai, "entra",
        "Microsoft Entra identity and Permissions Management specialist",
        "Entra ID and Permissions Management posture aligned to Microsoft Entra Permissions Management "
        "capabilities: permission creep and unused/over-provisioned permissions (Permission Creep Index style), "
        "privileged and Owner sprawl, guest and service-principal/workload-identity risk, app-registration "
        "secret/certificate EXPIRY and rotation hygiene, least-privilege right-sizing, and PIM / access reviews / "
        "just-in-time access recommendations",
        resources, arc_data, ctx, refresh, scope,
    )


@app.get("/api/ai/waf")
async def ai_waf(refresh: bool = False, scope: str = None):
    """Deep AI Well-Architected Framework assessment across all five pillars."""
    import asyncio
    from services.ai_module_analysis_service import analyze_generic_ai
    resources = _get_resources_list()
    if not resources:
        raise HTTPException(status_code=404, detail="No resource data")
    arc_data = _get_arc_data()
    ctx = {"framework": "Azure Well-Architected Framework",
           "pillars": ["Reliability", "Security", "Cost Optimization", "Operational Excellence", "Performance Efficiency"]}
    return await asyncio.to_thread(
        analyze_generic_ai, "waf",
        "Azure Well-Architected Framework principal reviewer",
        "a deep Well-Architected Framework assessment. Produce ONE category per WAF pillar (Reliability, "
        "Security, Cost Optimization, Operational Excellence, Performance Efficiency), each scored, with "
        "specific findings mapped to the impacted workloads/resources, severity, remediation and the relevant "
        "WAF recommendation, so the customer gets an advanced, exportable pillar-by-pillar review",
        resources, arc_data, ctx, refresh, scope,
    )


@app.get("/api/ai/caf")
async def ai_caf(refresh: bool = False, scope: str = None):
    """Deep AI Cloud Adoption Framework assessment across CAF methodologies."""
    import asyncio
    from services.ai_module_analysis_service import analyze_generic_ai
    resources = _get_resources_list()
    if not resources:
        raise HTTPException(status_code=404, detail="No resource data")
    arc_data = _get_arc_data()
    ctx = {"framework": "Microsoft Cloud Adoption Framework",
           "methodologies": ["Strategy", "Plan", "Ready", "Migrate", "Innovate", "Govern", "Manage", "Secure"]}
    return await asyncio.to_thread(
        analyze_generic_ai, "caf",
        "Microsoft Cloud Adoption Framework strategist",
        "a Cloud Adoption Framework assessment. Produce ONE category per CAF methodology (Strategy, Plan, "
        "Ready, Migrate, Innovate, Govern, Manage, Secure), each with a maturity score, findings tied to the "
        "customer's actual estate/landing-zone signals, gaps, and prioritized next steps to advance adoption",
        resources, arc_data, ctx, refresh, scope,
    )


# ── Update Management Module ──────────────────────────────────────────────────

@app.get("/api/updates/summary")
async def updates_summary():
    """Update Management dashboard KPIs."""
    from services.update_management_service import get_update_summary
    return await get_update_summary()


@app.get("/api/updates/patched")
async def updates_patched(days: int = 30):
    """Machines patched within the given number of days."""
    from services.update_management_service import get_patched_machines
    return await get_patched_machines(days)


@app.get("/api/updates/unpatched")
async def updates_unpatched(days: int = 30):
    """Machines NOT patched within the given number of days."""
    from services.update_management_service import get_unpatched_machines
    return await get_unpatched_machines(days)


@app.get("/api/updates/pending-reboot")
async def updates_pending_reboot():
    """Machines pending reboot after patch installation."""
    from services.update_management_service import get_pending_reboot
    return await get_pending_reboot()


@app.get("/api/updates/rebooted")
async def updates_rebooted(days: int = 30):
    """Machines rebooted after patching within given days."""
    from services.update_management_service import get_rebooted_machines
    return await get_rebooted_machines(days)


@app.get("/api/updates/by-os")
async def updates_by_os():
    """Update stats grouped by OS type."""
    from services.update_management_service import get_updates_by_os
    return await get_updates_by_os()


@app.get("/api/updates/by-subscription")
async def updates_by_subscription():
    """Update stats grouped by subscription."""
    from services.update_management_service import get_updates_by_subscription
    return await get_updates_by_subscription()


@app.get("/api/updates/by-classification")
async def updates_by_classification():
    """Pending updates grouped by classification (Critical/Security/Other)."""
    from services.update_management_service import get_updates_by_classification
    return await get_updates_by_classification()


@app.get("/api/updates/compliance-trend")
async def updates_compliance_trend(days: int = 30):
    """Compliance trend over the given number of days."""
    from services.update_management_service import get_compliance_trend
    return await get_compliance_trend(days)


@app.get("/api/updates/detailed-report")
async def updates_detailed_report(
    subscription_id: str = None,
    resource_group: str = None,
    os_type: str = None,
    machine_type: str = None,
):
    """Full detailed machine update report with filters."""
    from services.update_management_service import get_detailed_report
    return await get_detailed_report(subscription_id, resource_group, os_type, machine_type)


@app.get("/api/updates/filters")
async def updates_filters():
    """Available filter options for Update Management UI."""
    from services.update_management_service import get_filter_options
    return await get_filter_options()


@app.post("/api/updates/refresh")
async def updates_refresh():
    """Force refresh update management data from Azure."""
    from services.update_management_service import refresh_cache
    await refresh_cache()
    return {"status": "refreshed"}


@app.get("/api/updates/resource/{resource_id:path}")
async def get_updates_for_resource(resource_id: str):
    """Patch/update status scoped to a single resource (VM), cross-referenced with the resource cache."""
    try:
        from services.update_management_service import get_update_summary
        summary = await get_update_summary()
        rid_lower = resource_id.lower()

        # Try to find the resource name from the cache
        resource_name = resource_id.split("/")[-1].lower()
        if _cache.get("data"):
            for r in _cache["data"].resources:
                if r.resource_id.lower() == rid_lower or r.resource_id.lower().endswith(rid_lower):
                    resource_name = r.resource_name.lower()
                    break

        # Filter patched/unpatched machines by resource name
        def _matches(machine: dict) -> bool:
            m_id = (machine.get("resource_id") or machine.get("vmId") or "").lower()
            m_name = (machine.get("resource_name") or machine.get("computer") or "").lower()
            return resource_name in m_name or m_name in resource_name or \
                   (m_id and (m_id == rid_lower or m_id.endswith(rid_lower)))

        patched = [m for m in summary.get("patched_machines", []) if _matches(m)]
        unpatched = [m for m in summary.get("unpatched_machines", []) if _matches(m)]

        return {
            "resource_id": resource_id,
            "resource_name": resource_name,
            "patched_machines": patched,
            "unpatched_machines": unpatched,
            "found": len(patched) + len(unpatched) > 0,
        }
    except Exception as exc:
        logger.error("Updates resource fetch failed: %s", exc)
        raise HTTPException(status_code=500, detail=str(exc))


# ── Serve Azure service icons ─────────────────────────────────────────────────
# Serve Microsoft Azure service icons from Icons directory
_ICONS_DIR = pathlib.Path(__file__).parent.parent / "Icons"
if _ICONS_DIR.exists():
    app.mount("/icons", StaticFiles(directory=str(_ICONS_DIR)), name="icons")
    logger.info(f"Serving Azure icons from: {_ICONS_DIR}")

# ── Serve built React frontend ─────────────────────────────────────────────────
# Only active when frontend/dist exists (i.e. after npm run build).
# In dev mode the Vite dev server runs separately on port 5173.

_FRONTEND = pathlib.Path(__file__).parent.parent / "frontend" / "dist"


def _compute_build_id() -> str:
    """Build identifier = the hashed main bundle name Vite emits
    (e.g. 'BnZVfzql' from index-BnZVfzql.js). The SPA captures this at boot and
    polls /api/version; when it changes the tab knows a newer build was deployed
    and can reload — so an open tab never keeps running a stale (possibly buggy)
    bundle after a rebuild."""
    import re as _re
    try:
        html = (_FRONTEND / "index.html").read_text(encoding="utf-8")
        m = _re.search(r"/assets/index-([A-Za-z0-9_\-]+)\.js", html)
        if m:
            return m.group(1)
    except Exception:
        pass
    return "dev"


_BUILD_ID = _compute_build_id()


@app.get("/api/version", include_in_schema=False)
async def api_version():
    """Current frontend build id. Polled by the SPA to auto-reload stale tabs."""
    from fastapi.responses import JSONResponse
    return JSONResponse(
        {"build": _BUILD_ID},
        headers={"Cache-Control": "no-cache, no-store, must-revalidate", "Pragma": "no-cache"},
    )


if _FRONTEND.exists():
    if (_FRONTEND / "assets").exists():
        app.mount("/assets", StaticFiles(directory=str(_FRONTEND / "assets")), name="assets")

    @app.get("/{full_path:path}", include_in_schema=False)
    async def serve_spa(full_path: str):
        # Don't serve SPA for API routes or health check
        if full_path.startswith("api/") or full_path == "health":
            raise HTTPException(status_code=404, detail="Not found")

        # index.html must never be cached — its hashed asset references change on rebuild
        _NO_CACHE = {"Cache-Control": "no-cache, no-store, must-revalidate", "Pragma": "no-cache"}

        requested = (_FRONTEND / full_path).resolve()
        # Block path traversal — only serve files inside the dist directory
        if not str(requested).startswith(str(_FRONTEND.resolve())):
            return FileResponse(str(_FRONTEND / "index.html"), headers=_NO_CACHE)
        if requested.exists() and requested.is_file():
            # Hashed assets (JS/CSS in /assets/) can be cached forever; index.html cannot
            if "assets/" in str(requested):
                return FileResponse(str(requested), headers={"Cache-Control": "public, max-age=31536000, immutable"})
            return FileResponse(str(requested), headers=_NO_CACHE)
        return FileResponse(str(_FRONTEND / "index.html"), headers=_NO_CACHE)
